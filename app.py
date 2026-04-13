from __future__ import annotations

import hashlib
import json
import math
import mimetypes
import os
import random
import smtplib
import sqlite3
import subprocess
import urllib.error
import urllib.request
from contextlib import closing
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from io import BytesIO
from pathlib import Path

from flask import Flask, abort, flash, g, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.exceptions import RequestEntityTooLarge

# Google OAuth (optional — only active when GOOGLE_CLIENT_ID is set)
try:
    from authlib.integrations.flask_client import OAuth as AuthlibOAuth
    _AUTHLIB_AVAILABLE = True
except ImportError:
    _AUTHLIB_AVAILABLE = False


BASE_DIR = Path(__file__).resolve().parent

# ── Data layout (v1.3.0) ────────────────────────────────────────────────
# Operational data (the real lab DB, uploads, exports) and demo data
# (fake users/requests for local dev) live under separate folders so
# demo runs can never corrupt a real deployment. The mode is chosen at
# boot by LAB_SCHEDULER_DEMO_MODE — "1"/"true" → demo; else operational.
DATA_DIR = BASE_DIR / "data"
DATA_OPERATIONAL_DIR = DATA_DIR / "operational"
DATA_DEMO_DIR = DATA_DIR / "demo"

_DEMO_MODE_ENV = os.environ.get("LAB_SCHEDULER_DEMO_MODE", "1").strip().lower()
DEMO_MODE = _DEMO_MODE_ENV in {"1", "true", "yes", "on"}
ORG_NAME = os.environ.get("CATALYST_ORG_NAME", "CATALYST")
ORG_TAGLINE = os.environ.get("CATALYST_ORG_TAGLINE", "Open-source ERP for Research & Operations")
_ACTIVE_DATA_DIR = DATA_DEMO_DIR if DEMO_MODE else DATA_OPERATIONAL_DIR
_ACTIVE_DATA_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = _ACTIVE_DATA_DIR / "lab_scheduler.db"
EXPORT_DIR = _ACTIVE_DATA_DIR / "exports"
UPLOAD_DIR = _ACTIVE_DATA_DIR / "uploads"
STATIC_DIR = BASE_DIR / "static"
INSTRUMENT_IMAGE_DIR = STATIC_DIR / "instrument_images"
ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "xlsx", "csv", "txt"}
POST_COMPLETION_UPLOAD_ROLES = {"super_admin", "instrument_admin", "operator"}
COMMUNICATION_NOTE_TYPES = [
    ("requester_note", "Requester Note", "Question or clarification from the submitting member."),
    ("lab_reply", "Lab Reply", "Single reply from the lab side for coordination."),
    ("operator_note", "Operator Note", "Operational note from the operator or instrument admin."),
    ("final_note", "Final Note", "Final handoff note shared with the requester."),
]
OWNER_EMAILS = {
    email.strip().lower()
    # Override at deploy time with `OWNER_EMAILS=you@example.com` env var.
    # Default: vishva (the demo seed super_admin).
    for email in os.environ.get(
        "OWNER_EMAILS",
        "owner@catalyst.local",
    ).split(",")
    if email.strip()
}
DEMO_ROLE_SWITCHES = {
    "owner": {"label": "Owner", "email": "owner@catalyst.local"},
    "super_admin": {"label": "Super Admin", "email": "dean@catalyst.local"},
    "instrument_admin": {"label": "Instrument Admin", "email": "kondhalkar@catalyst.local"},
    "site_admin": {"label": "Site Admin", "email": "siteadmin@catalyst.local"},
    "operator": {"label": "Operator", "email": "anika@catalyst.local"},
    "member": {"label": "Member", "email": "user1@catalyst.local"},
    "finance": {"label": "Finance", "email": "meera@catalyst.local"},
    "professor": {"label": "Approver", "email": "approver@catalyst.local"},
}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("LAB_SCHEDULER_SECRET_KEY", "lab-scheduler-dev-secret")
app.config["SMTP_HOST"] = os.environ.get("SMTP_HOST", "localhost")
app.config["SMTP_PORT"] = int(os.environ.get("SMTP_PORT", "25"))
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("LAB_SCHEDULER_COOKIE_SECURE", "").lower() in {"1", "true", "yes"}
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=12)
app.config["SESSION_REFRESH_EACH_REQUEST"] = True

# CSRF protection — token machinery is wired up but enforcement is gated
# behind LAB_SCHEDULER_CSRF=1 so we can roll templates and tests over to
# the protected mode in a follow-up wave without forcing an all-or-nothing
# flag day. The meta tag in base.html is always emitted; the JS shim
# auto-injects tokens into forms and fetch() calls.
app.config["WTF_CSRF_ENABLED"] = os.environ.get("LAB_SCHEDULER_CSRF", "").lower() in {"1", "true", "yes"}
app.config["WTF_CSRF_TIME_LIMIT"] = None  # Tokens valid for the session lifetime, not 1 hour
app.config["WTF_CSRF_SSL_STRICT"] = False  # LAN deployment may run plain HTTP

# Demo mode — gates the /demo/switch role-impersonation route and the
# seed_data() demo account inserts. Defaults to ON for development; set
# LAB_SCHEDULER_DEMO_MODE=0 in production to lock both down.
DEMO_MODE = os.environ.get("LAB_SCHEDULER_DEMO_MODE", "1").lower() in {"1", "true", "yes"}

# ── ERP Module System ────────────────────────────────────────────
# Plug-and-play module registry.  To add a new module an AI agent
# only needs to:
#   1. Add a row to MODULE_REGISTRY below
#   2. Add its route(s) in app.py
#   3. Create the template(s)
#   4. (optional) run  scripts/new_module.sh <name>  to scaffold
#
# Set CATALYST_MODULES in .env as a comma-separated list to restrict
# which modules are active.  Default: all modules enabled.
# Example: CATALYST_MODULES=instruments,finance,inbox
#
# Each entry carries everything the nav bar and tooling need:
#   label       — human-readable nav text
#   icon        — emoji shown in nav and admin panels
#   nav_order   — sort key (lower = further left); 0 = hidden from nav
#   description — one-liner for admin / docs
#   nav_endpoint — Flask endpoint name for the primary route
#   nav_active_endpoints — set of endpoints that highlight this nav item
#   nav_badge_key — optional key into nav_pending_counts / inject_globals
#   nav_access   — callable(access_profile, is_owner) -> bool  (access gate)
#   nav_type     — "link" (default), "dropdown" (instruments), "panel" (notifications)
#   children     — list of sub-items for dropdown type (optional)

MODULE_REGISTRY = {
    "instruments": {
        "label": "Instruments",
        "icon": "\U0001f52c",
        "nav_order": 1,
        "description": "Instrument portal",
        "nav_endpoint": "instruments",
        "nav_active_endpoints": {
            "instruments", "instrument_detail", "instrument_history",
            "instrument_calendar", "schedule", "schedule_actions",
            "request_detail", "calendar", "calendar_events", "stats",
            "visualizations", "instrument_visualization", "group_visualization",
        },
        "nav_type": "dropdown",
        # Operational staff nav: owner, super/site admin, instrument_admin, operator, finance_admin.
        # NOT approvers, requesters, faculty — they access instruments via Settings or direct links.
        "nav_access": lambda ap, is_owner: ap.get("_is_operational_nav") or is_owner,
    },
    "finance": {
        "label": "Finance",
        "icon": "\U0001f4b0",
        "nav_order": 2,
        "description": "Grants, invoices, payments",
        "nav_endpoint": "finance_portal",
        "nav_active_endpoints": {"finance_portal", "finance_grants_list", "finance_grant_detail"},
        "nav_access": lambda ap, is_owner: ap.get("can_view_finance_stage") or is_owner,
    },
    "receipts": {
        "label": "Receipts",
        "icon": "\U0001f9fe",
        "nav_order": 0,
        "description": "Expense receipt submission (accessible via /receipts, hidden from nav)",
        "nav_endpoint": "receipts_list",
        "nav_active_endpoints": {"receipts_list", "receipt_new", "receipt_detail", "receipt_review"},
    },
    "inbox": {
        "label": "Inbox",
        "icon": "\U0001f4e8",
        "nav_order": 4,
        "description": "Internal messaging",
        "nav_endpoint": "inbox",
        "nav_active_endpoints": {"inbox"},
        "nav_badge_key": "inbox_unread",
    },
    "notifications": {
        "label": "Notifications",
        "icon": "\U0001f514",
        "nav_order": 5,
        "description": "System notifications",
        "nav_endpoint": "notifications_page",
        "nav_active_endpoints": {"notifications_page"},
        "nav_type": "panel",
        "nav_badge_key": "notice_count",
    },
    "attendance": {
        "label": "Attendance",
        "icon": "\U0001f4cb",
        "nav_order": 6,
        "description": "Daily attendance + leave",
        "nav_endpoint": "attendance_page",
        "nav_active_endpoints": {
            "attendance_page", "leave_request_new",
            "admin_leave_queue", "admin_attendance_calendar",
        },
        # Only lab staff who physically attend — operators + instrument_admins.
        "nav_access": lambda ap, is_owner: ap.get("_is_lab_staff"),
    },
    "todos": {
        "label": "Tasks",
        "icon": "\u2705",
        "nav_order": 7,
        "description": "Assignable task list",
        "nav_endpoint": "todos_page",
        "nav_active_endpoints": {"todos_page", "todo_new", "todo_complete", "todo_delete", "todo_update"},
        "nav_access": lambda ap, is_owner: True,
    },
    "letters": {
        "label": "Letters",
        "icon": "\u2709\ufe0f",
        "nav_order": 8,
        "description": "Create letters on institute letterhead",
        "nav_endpoint": "letters_list",
        "nav_active_endpoints": {"letters_list", "letter_new", "letter_detail", "letter_print"},
        "nav_access": lambda ap, is_owner: ap.get("_is_operational_nav") or is_owner,
    },
    "queue": {
        "label": "Queue",
        "icon": "\U0001f4ca",
        "nav_order": 0,
        "description": "Sample request queue (sub-item of Instruments)",
        "nav_endpoint": "schedule",
        "nav_active_endpoints": {"schedule", "schedule_actions"},
    },
    "calendar": {
        "label": "Calendar",
        "icon": "\U0001f4c5",
        "nav_order": 0,
        "description": "Schedule calendar (sub-item of Instruments)",
        "nav_endpoint": "calendar",
        "nav_active_endpoints": {"calendar", "calendar_events"},
    },
    "stats": {
        "label": "Stats",
        "icon": "\U0001f4c8",
        "nav_order": 0,
        "description": "Analytics dashboard (sub-item of Instruments)",
        "nav_endpoint": "stats",
        "nav_active_endpoints": {"stats", "visualizations", "instrument_visualization", "group_visualization"},
    },
    "vehicles": {
        "label": "Fleet",
        "icon": "\U0001f697",
        "nav_order": 9,
        "description": "Vehicle fleet management",
        "nav_endpoint": "vehicles_list",
        "nav_active_endpoints": {"vehicles_list", "vehicle_detail"},
        "nav_access": lambda ap, is_owner: ap.get("_is_operational_nav") or is_owner,
    },
    "personnel": {
        "label": "Personnel",
        "icon": "\U0001f465",
        "nav_order": 10,
        "description": "Staff & salary management",
        "nav_endpoint": "personnel_list",
        "nav_active_endpoints": {"personnel_list", "personnel_detail", "payroll_view"},
        "nav_access": lambda ap, is_owner: ap.get("_is_operational_nav") or is_owner,
    },
    "admin": {
        "label": "Admin",
        "icon": "\u2699\ufe0f",
        "nav_order": 99,
        "description": "User management + dev tools",
        "nav_endpoint": "dev_panel",
        "nav_active_endpoints": {"dev_panel"},
        "nav_access": lambda ap, is_owner: is_owner,
    },
}

ALL_MODULES = set(MODULE_REGISTRY.keys())

_modules_env = os.environ.get("CATALYST_MODULES", "").strip()
ENABLED_MODULES: set[str] = (
    {m.strip().lower() for m in _modules_env.split(",") if m.strip()}
    if _modules_env
    else ALL_MODULES.copy()  # default: everything on
)


def module_enabled(name: str) -> bool:
    """Check if an ERP module is enabled for this deployment."""
    return name in ENABLED_MODULES


def build_nav_items(user, access_profile, is_owner):
    """Build sorted list of nav item dicts from MODULE_REGISTRY.

    Each item: {key, label, icon, url, active, badge, nav_type, meta}
    Only includes modules that are enabled, have nav_order > 0,
    and pass their access gate.
    """
    from flask import request as _req, url_for as _url_for
    endpoint = _req.endpoint or ""
    items = []
    for key, meta in sorted(MODULE_REGISTRY.items(), key=lambda kv: kv[1].get("nav_order", 50)):
        if meta.get("nav_order", 0) <= 0:
            continue
        if not module_enabled(key):
            continue
        access_fn = meta.get("nav_access")
        if access_fn and not access_fn(access_profile, is_owner):
            continue
        ep = meta.get("nav_endpoint")
        try:
            url = _url_for(ep) if ep else "#"
        except Exception:
            continue  # route not registered yet — skip silently
        items.append({
            "key": key,
            "label": meta["label"],
            "icon": meta["icon"],
            "url": url,
            "active": endpoint in meta.get("nav_active_endpoints", set()),
            "nav_type": meta.get("nav_type", "link"),
            "badge_key": meta.get("nav_badge_key"),
            "meta": meta,
        })
    return items


csrf = CSRFProtect(app)

# ── Google OAuth ────────────────────────────────────────────────
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_ALLOWED_DOMAIN = os.environ.get("GOOGLE_ALLOWED_DOMAIN", "mitwpu.edu.in")  # change this to your institution's domain

_oauth = None
if _AUTHLIB_AVAILABLE and GOOGLE_CLIENT_ID:
    _oauth = AuthlibOAuth(app)
    _oauth.register(
        name="google",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={"scope": "openid email profile"},
    )

# ── SendGrid / External Email ──────────────────────────────────
SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
SENDGRID_FROM = os.environ.get("SENDGRID_FROM", "noreply@catalyst.local")
SENDGRID_DAILY_LIMIT = int(os.environ.get("SENDGRID_DAILY_LIMIT", "100"))


def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def generate_receipt_reference(sample_origin: str) -> str:
    db = get_db()
    first_two = {
        "internal": "17",
        "external": "83",
    }.get((sample_origin or "").strip().lower(), "17")
    while True:
        candidate = f"RCPT-{first_two}{random.randint(1000, 9999)}"
        # v2.0.0 — receipts live in payments, not sample_requests.
        existing = db.execute(
            "SELECT 1 FROM payments WHERE receipt_number = ? LIMIT 1",
            (candidate,),
        ).fetchone()
        if existing is None:
            return candidate


def generate_job_reference(sample_origin: str) -> str:
    db = get_db()
    prefix = {
        "internal": "J2",
        "external": "J4",
    }.get((sample_origin or "").strip().lower(), "J2")
    while True:
        candidate = f"{prefix}{random.randint(100000, 999999)}"
        existing = db.execute(
            "SELECT 1 FROM sample_requests WHERE request_no = ? LIMIT 1",
            (candidate,),
        ).fetchone()
        if existing is None:
            return candidate


def sample_reference_prefix(instrument_name: str | None) -> str:
    token = "".join(ch for ch in (instrument_name or "").upper() if ch.isalnum())
    if len(token) >= 2:
        return token[:2]
    if len(token) == 1:
        return f"{token}X"
    return "SM"


def generate_sample_reference(instrument_name: str | None, sample_origin: str) -> str:
    db = get_db()
    prefix = sample_reference_prefix(instrument_name)
    origin_code = "E" if (sample_origin or "").strip().lower() == "external" else "I"
    while True:
        candidate = f"{prefix}{origin_code}{random.randint(100000, 999999)}"
        existing = db.execute(
            "SELECT 1 FROM sample_requests WHERE sample_ref = ? LIMIT 1",
            (candidate,),
        ).fetchone()
        if existing is None:
            return candidate


def pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def simple_pdf_bytes(title: str, lines: list[str]) -> bytes:
    buffer = BytesIO()
    text_lines = [title] + lines
    y = 770
    commands = ["BT", "/F1 14 Tf", "50 800 Td", f"({pdf_escape(title)}) Tj"]
    for line in lines:
        y -= 22
        commands.extend(["BT", "/F1 11 Tf", f"50 {y} Td", f"({pdf_escape(line)}) Tj", "ET"])
    content = "\n".join(commands).encode("latin-1", errors="replace")
    objects = []
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")
    objects.append(b"2 0 obj << /Type /Pages /Count 1 /Kids [3 0 R] >> endobj\n")
    objects.append(b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n")
    objects.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n")
    objects.append(f"5 0 obj << /Length {len(content)} >> stream\n".encode("latin-1") + content + b"\nendstream endobj\n")
    buffer.write(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(buffer.tell())
        buffer.write(obj)
    xref_start = buffer.tell()
    buffer.write(f"xref\n0 {len(offsets)}\n".encode("latin-1"))
    buffer.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        buffer.write(f"{offset:010d} 00000 n \n".encode("latin-1"))
    buffer.write(
        f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode("latin-1")
    )
    return buffer.getvalue()


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        # WAL journal mode — cheap, concurrent reads during writes, and
        # it survives connection close (it's a file-level setting).
        # Safe to re-apply on every connection; SQLite is a no-op if
        # already in WAL. See PHILOSOPHY §3 — this is the single perf
        # lever the hard-attribute contract permits.
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA synchronous = NORMAL")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_exc: object) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.errorhandler(RequestEntityTooLarge)
def handle_large_upload(_exc):
    flash("Upload too large. Maximum file size is 100 MB.", "error")
    referrer = request.referrer or url_for("index")
    return redirect(referrer)


@app.errorhandler(403)
def handle_forbidden(_exc):
    return render_template("error.html", title="Access Restricted", heading="Access Restricted", message="You do not have permission to view this page.", code=403), 403


@app.errorhandler(404)
def handle_not_found(_exc):
    return render_template("error.html", title="Page Not Found", heading="Page Not Found", message="This page does not exist or is no longer available.", code=404), 404


@app.errorhandler(500)
def handle_server_error(_exc):
    return render_template("error.html", title="Server Error", heading="Something went wrong", message="An internal error occurred. Please try again shortly.", code=500), 500


@app.errorhandler(ValueError)
def handle_value_error(exc):
    """v2.0.1 — catch raw int(request.form[...]) crashes on empty/missing
    form fields. These used to bubble up as 500s; the stability crawler
    flagged ~8 such sites across the request_detail + member admin
    routes. Rather than instrumenting each one, we catch ValueError at
    the Flask layer and redirect to the referrer with a user-friendly
    flash. Real validation bugs still need fixing — this is a safety
    net, not a cure."""
    # Only handle form-parsing errors; let genuine TypeError / other
    # ValueErrors from inside route bodies still produce 500s (those
    # are real bugs the operator should see).
    msg = str(exc)
    if not ("invalid literal" in msg or "base 10" in msg):
        raise exc
    flash("Invalid form input. Please double-check and try again.", "error")
    return redirect(request.referrer or url_for("index"))


def query_all(sql: str, params: tuple = ()) -> list[sqlite3.Row]:
    return get_db().execute(sql, params).fetchall()


def query_one(sql: str, params: tuple = ()) -> sqlite3.Row | None:
    return get_db().execute(sql, params).fetchone()


def execute(sql: str, params: tuple = ()) -> int:
    cur = get_db().execute(sql, params)
    get_db().commit()
    return cur.lastrowid


def generate_member_code(name: str, role: str) -> str:
    """Generate a member code from name and role.
    Format: First 2 letters of name (uppercase) + role abbreviation (2 chars) + zero-padded sequential number
    Role abbreviations: RQ=requester, FA=finance_admin, PA=professor_approver, IA=instrument_admin, OP=operator, SA=site_admin, SU=super_admin
    Example: "Dr. Shah" as requester -> "SHRQ001"
    """
    role_abbreviations = {
        "requester": "RQ",
        "finance_admin": "FA",
        "professor_approver": "PA",
        "instrument_admin": "IA",
        "operator": "OP",
        "site_admin": "SA",
        "super_admin": "SU",
    }
    # Get first 2 letters of name (uppercase)
    name_prefix = (name.strip()[:2] or "XX").upper()
    # Get role abbreviation
    role_abbr = role_abbreviations.get(role, "XX")
    # Count existing users with same role
    count = query_one(
        "SELECT COUNT(*) as cnt FROM users WHERE role = ?",
        (role,)
    )
    seq_num = (count["cnt"] if count else 0) + 1
    # Format: SHRQ001
    return f"{name_prefix}{role_abbr}{seq_num:03d}"


def generate_temp_password() -> str:
    """Generate a short random temporary password for admin-issued resets.

    Policy (DATA_POLICY.md §3, see docs/DATA_POLICY.md):
      - Admins NEVER type passwords directly. They call this helper.
      - The returned string is shown to the admin once via a flash
        message so they can share it out-of-band.
      - The user is expected to change it on first login; the
        `must_change_password` column on `users` gates access until
        they do.
      - 8 characters from a friendly alphabet (no 0/O/1/l/I ambiguity)
        so admins can dictate over the phone if needed.
    """
    import secrets
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(8))


def log_action(actor_id: int | None, entity_type: str, entity_id: int, action: str, payload: dict) -> None:
    log_action_at(actor_id, entity_type, entity_id, action, payload, now_iso())


def log_action_at(actor_id: int | None, entity_type: str, entity_id: int, action: str, payload: dict, created_at: str) -> None:
    db = get_db()
    previous = query_one(
        "SELECT entry_hash FROM audit_logs WHERE entity_type = ? AND entity_id = ? ORDER BY id DESC LIMIT 1",
        (entity_type, entity_id),
    )
    prev_hash = previous["entry_hash"] if previous else ""
    payload_json = json.dumps(payload, sort_keys=True)
    digest = hashlib.sha256(f"{prev_hash}|{entity_type}|{entity_id}|{action}|{payload_json}".encode()).hexdigest()
    db.execute(
        """
        INSERT INTO audit_logs (entity_type, entity_id, action, actor_id, payload_json, prev_hash, entry_hash, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (entity_type, entity_id, action, actor_id, payload_json, prev_hash, digest, created_at),
    )
    db.commit()


def verify_audit_chain(entity_type: str, entity_id: int) -> bool:
    rows = query_all(
        """
        SELECT action, payload_json, prev_hash, entry_hash
        FROM audit_logs
        WHERE entity_type = ? AND entity_id = ?
        ORDER BY id
        """,
        (entity_type, entity_id),
    )
    previous_hash = ""
    for row in rows:
        digest = hashlib.sha256(
            f"{previous_hash}|{entity_type}|{entity_id}|{row['action']}|{row['payload_json']}".encode()
        ).hexdigest()
        if row["prev_hash"] != previous_hash or row["entry_hash"] != digest:
            return False
        previous_hash = row["entry_hash"]
    return True


def send_completion_email(sample_request: sqlite3.Row, results_summary: str) -> tuple[bool, str]:
    msg = EmailMessage()
    msg["Subject"] = f"Lab Result Ready: {sample_request['request_no']}"
    msg["From"] = "noreply@lab.local"
    msg["To"] = sample_request["requester_email"]
    msg.set_content(
        "\n".join(
            [
                f"Hello {sample_request['requester_name']},",
                "",
                f"Your request {sample_request['request_no']} for {sample_request['instrument_name']} is complete.",
                f"Sample: {sample_request['sample_name']}",
                "",
                "Operator remark/result summary:",
                results_summary or "No summary provided.",
                "",
                "Regards,",
                "Lab Facility",
            ]
        )
    )
    try:
        with smtplib.SMTP(app.config["SMTP_HOST"], app.config["SMTP_PORT"], timeout=5) as server:
            server.send_message(msg)
        return True, "Email sent"
    except Exception as exc:
        return False, f"Email not sent: {exc}"


def build_request_status(db: sqlite3.Connection, request_id: int) -> str:
    request_row = db.execute(
        """
        SELECT status, sample_submitted_at, sample_received_at
        FROM sample_requests
        WHERE id = ?
        """,
        (request_id,),
    ).fetchone()
    steps = db.execute(
        "SELECT step_order, status FROM approval_steps WHERE sample_request_id = ? ORDER BY step_order",
        (request_id,),
    ).fetchall()
    if not steps:
        return "submitted"
    if any(step["status"] == "rejected" for step in steps):
        return "rejected"
    for step in steps:
        if step["status"] != "approved":
            return "under_review"
    if request_row and request_row["status"] in {"scheduled", "in_progress", "completed"}:
        return request_row["status"]
    if request_row and request_row["sample_received_at"]:
        return "sample_received"
    if request_row and request_row["sample_submitted_at"]:
        return "sample_submitted"
    return "awaiting_sample_submission"


def request_display_status(status: str) -> str:
    labels = {
        "submitted": "Submitted",
        "under_review": "Under Review",
        "awaiting_sample_submission": "Awaiting Sample Submission",
        "sample_submitted": "Sample Submitted",
        "sample_received": "Sample Received",
        "scheduled": "Scheduled",
        "in_progress": "In Progress",
    }
    return labels.get(status, status.replace("_", " ").title())


def approval_role_label(role: str | None) -> str:
    labels = {
        "finance": "Finance",
        "professor": "Professor",
        "operator": "Operator",
    }
    if not role:
        return "Review"
    return labels.get(role, role.replace("_", " ").title())


# ── Request status state machine ───────────────────────────────────────────
# Defines which status transitions are legal. Any handler that mutates
# `sample_requests.status` should call `assert_status_transition()` first
# to reject invalid moves before they reach the database.
#
# `submitted` is the entry point. `completed`, `rejected`, and `cancelled`
# are terminal — nothing leaves them. `awaiting_sample_submission` is the
# branch the requester drives; the operator path runs from `sample_received`
# onward.
#
# Admin overrides (admin_schedule_override, admin_complete_override) are
# allowed to skip ahead — they pass `force=True` to bypass this check.
REQUEST_STATUS_TRANSITIONS: dict[str, set[str]] = {
    "submitted":                  {"under_review", "rejected", "cancelled"},
    "under_review":               {"awaiting_sample_submission", "rejected", "cancelled"},
    # awaiting → sample_received is the operator quick-receive fast-track:
    # the operator confirms the sample is physically present without
    # waiting for the requester to mark it submitted.
    "awaiting_sample_submission": {"sample_submitted", "sample_received", "rejected", "cancelled"},
    "sample_submitted":           {"sample_received", "rejected"},
    "sample_received":            {"scheduled", "in_progress", "rejected"},
    "scheduled":                  {"in_progress", "scheduled", "completed", "rejected"},
    "in_progress":                {"completed", "rejected"},
    "completed":                  set(),  # terminal
    "rejected":                   set(),  # terminal
    "cancelled":                  set(),  # terminal
}


class InvalidStatusTransition(ValueError):
    pass


def assert_status_transition(current: str | None, target: str, *, force: bool = False) -> None:
    """Raise InvalidStatusTransition if `current → target` is not allowed.

    Pass `force=True` for admin overrides that intentionally skip the
    workflow (admin_schedule_override, admin_complete_override).
    """
    if force:
        return
    current = (current or "").strip()
    if current not in REQUEST_STATUS_TRANSITIONS:
        raise InvalidStatusTransition(f"Unknown current status: {current!r}")
    if target == current:
        return  # idempotent updates (e.g. re-schedule on same status) are fine
    if target not in REQUEST_STATUS_TRANSITIONS[current]:
        raise InvalidStatusTransition(
            f"Cannot transition request from {current!r} to {target!r}. "
            f"Allowed: {sorted(REQUEST_STATUS_TRANSITIONS[current]) or 'terminal'}"
        )


@app.errorhandler(InvalidStatusTransition)
def handle_invalid_status_transition(exc: InvalidStatusTransition):
    flash(f"Invalid status change: {exc}", "error")
    return redirect(request.referrer or url_for("index"))


def request_status_group(status: str | None) -> str:
    status = status or ""
    if status in {"submitted", "under_review", "awaiting_sample_submission", "sample_submitted", "sample_received", "scheduled"}:
        return "Pending"
    if status == "in_progress":
        return "Processing"
    if status == "completed":
        return "Processed"
    if status in {"rejected"}:
        return "Closed"
    return "Open"


def request_status_summary(request_row: sqlite3.Row | dict, approval_steps: list[sqlite3.Row | dict] | None = None) -> str:
    status = (request_row["status"] if request_row else "") or ""
    instrument_mode = instrument_intake_mode(request_row) if request_row else "accepting"
    if status == "submitted":
        return "Request is submitted and waiting for the lab to start accepting new jobs."
    if status == "under_review":
        pending_step = None
        for step in approval_steps or []:
            if step["status"] == "rejected":
                return f"{approval_role_label(step['approver_role'])} rejected this request."
            if step["status"] != "approved":
                pending_step = step
                break
        if pending_step:
            return f"Pending {approval_role_label(pending_step['approver_role']).lower()} approval."
        return "Pending review."
    if status == "awaiting_sample_submission":
        if instrument_mode != "accepting":
            return "Dropoff is on hold right now. You can submit the request, but you cannot drop off the sample until the lab resumes accepting samples."
        return "Approved. Waiting for the member to physically submit the sample."
    if status == "sample_submitted":
        return "Member marked the sample submitted. Operator/admin should confirm receipt."
    if status == "sample_received":
        return "Sample received. Ready for operator scheduling."
    if status == "scheduled":
        scheduled_for = request_row["scheduled_for"] if request_row else None
        if scheduled_for:
            return f"Run scheduled for {format_dt(scheduled_for)}."
        return "Run scheduled."
    if status == "in_progress":
        return "Run is currently being processed."
    if status == "completed":
        completed_at = request_row["completed_at"] if request_row else None
        if completed_at:
            return f"Job completed on {format_dt(completed_at)} and the record is locked."
        return "Job completed and the record is locked."
    if status == "rejected":
        return "Request was rejected and will not be processed."
    return "Request is active."


def request_lifecycle_steps(request_row: sqlite3.Row | dict, approval_steps: list[sqlite3.Row | dict] | None = None) -> list[dict[str, str | bool]]:
    status = (request_row["status"] if request_row else "") or ""
    step_defs = [
        ("created", "Created"),
        ("sample_submitted", "Sample Submitted"),
        ("sample_received", "Sample Received"),
        ("processing", "Processing"),
        ("completed", "Done"),
    ]
    if status in {"rejected"}:
        step_defs[-1] = (status, request_display_status(status))
    rank = {
        "created": 1,
        "submitted": 1,
        "under_review": 1,
        "awaiting_sample_submission": 1,
        "sample_submitted": 2,
        "sample_received": 3,
        "processing": 4,
        "scheduled": 4,
        "in_progress": 4,
        "completed": 5,
        "rejected": 5,
    }
    current_rank = rank.get(status, 1)
    steps = []
    for code, label in step_defs:
        step_rank = rank.get(code, 0)
        if status in {"rejected"} and code == status:
            state = "current"
        elif current_rank > step_rank:
            state = "done"
        elif current_rank == step_rank:
            state = "current"
        else:
            state = "upcoming"
        note = ""
        if code == "created" and request_row and request_row["created_at"]:
            note = format_dt(request_row["created_at"])
        elif code == "sample_submitted" and request_row and request_row["sample_submitted_at"] and current_rank >= 2:
            note = format_dt(request_row["sample_submitted_at"])
        elif code == "sample_received" and request_row and request_row["sample_received_at"] and current_rank >= 3:
            note = format_dt(request_row["sample_received_at"])
        elif code == "processing" and request_row and request_row["scheduled_for"] and current_rank >= 4:
            note = format_dt(request_row["scheduled_for"])
        elif code == "completed" and request_row and request_row["completed_at"] and status == "completed":
            note = format_dt(request_row["completed_at"])
        steps.append({"code": code, "label": label, "state": state, "note": note})
    return steps


def timeline_action_label(action: str) -> str:
    labels = {
        "submitted": "Request created",
        "released_for_review": "Released for review",
        "approval_assigned": "Approver updated",
        "finance_approved": "Finance approved",
        "professor_approved": "Faculty in-charge approved",
        "operator_approved": "Operator approved",
        "finance_rejected": "Finance rejected",
        "professor_rejected": "Faculty in-charge rejected",
        "operator_rejected": "Operator rejected",
        "sample_submitted": "Sample submitted by member",
        "sample_received": "Sample received by lab",
        "dropoff_reopened": "Dropoff reopened",
        "scheduled": "Run scheduled",
        "started": "Run started",
        "completed": "Run completed",
        "rejected": "Request rejected",
        "attachment_uploaded": "File attached",
        "attachment_removed": "File removed",
        "communication_note_saved": "Communication note updated",
        "issue_flagged": "Issue flagged",
        "issue_response_saved": "Issue response added",
        "issue_resolved": "Issue resolved",
        "taken_up_from_board": "Taken up from board",
        "started_from_board": "Started from board",
        "completed_from_board": "Completed from board",
        "operator_reassigned": "Job reassigned",
        "resolved_and_completed": "Job marked done",
        "resolution_saved": "Resolution saved",
        "admin_schedule_override": "Admin schedule override",
        "admin_complete_override": "Admin completion override",
        "status_changed": "Status changed",
        "request_metadata_updated": "Details edited by admin",
    }
    return labels.get(action, action.replace("_", " ").title())


def request_timeline_entries(
    sample_request: sqlite3.Row,
    logs: list[sqlite3.Row],
    attachments: list[sqlite3.Row],
    messages: list[sqlite3.Row],
    message_attachments: dict[int, list[sqlite3.Row]] | None = None,
) -> list[dict]:
    entries: list[dict] = [
        {
            "at": sample_request["created_at"],
            "title": "Request created",
            "detail": f"{sample_request['request_no']} opened for {sample_request['instrument_name']}.",
            "actor": sample_request["originator_name"] or sample_request["requester_name"],
            "kind": "status",
            "scope": "workflow",
        }
    ]
    for log in logs:
        payload = {}
        try:
            payload = json.loads(log["payload_json"] or "{}")
        except Exception:
            payload = {}
        # Skip communication_note_saved — these are already shown as
        # message bubbles from the request_messages table.
        if log["action"] == "communication_note_saved":
            continue
        detail = ""
        if log["action"] == "issue_flagged":
            detail = payload.get("issue_preview", "")
        elif log["action"] == "issue_response_saved":
            detail = payload.get("response_preview", "")
        elif log["action"] == "issue_resolved":
            detail = payload.get("resolution_preview", "") or payload.get("response_preview", "")
        elif log["action"] in {"attachment_uploaded", "attachment_removed"}:
            filename = payload.get("filename", "")
            attachment_type = payload.get("attachment_type", "")
            note_text = payload.get("note", "")
            type_label = attachment_type.replace("_", " ").title() if attachment_type else ""
            parts = [part for part in [filename, type_label] if part]
            detail = " · ".join(parts)
            if note_text:
                detail = f"{detail}\n{note_text}".strip()
        elif log["action"] == "request_metadata_updated":
            parts = []
            if payload.get("title"):
                parts.append(f"title: {payload['title']}")
            if payload.get("sample_name"):
                parts.append(f"sample: {payload['sample_name']} ({payload.get('sample_count', '?')})")
            if payload.get("remarks_preview"):
                parts.append(payload["remarks_preview"])
            detail = " · ".join(parts)
        elif log["action"] in {"scheduled", "taken_up_from_board"} and payload.get("scheduled_for"):
            detail = f"Scheduled for {format_dt(payload['scheduled_for'])}"
        elif payload.get("remarks"):
            detail = payload["remarks"]
        elif payload.get("results_summary"):
            detail = payload["results_summary"]
        elif payload.get("email_status"):
            detail = payload["email_status"]
        entries.append(
            {
                "at": log["created_at"],
                "title": timeline_action_label(log["action"]),
                "detail": detail,
                "actor": (log["actor_name"] if "actor_name" in log.keys() else None) or "System",
                "kind": "event",
                "scope": "conversation" if log["action"] in {"communication_note_saved", "issue_flagged", "issue_response_saved", "issue_resolved"} else "workflow",
            }
        )
    for attachment in attachments:
        if attachment["request_message_id"]:
            continue
        attachment_title = {
            "request_document": "Submission file attached",
            "sample_slip": "Sample slip generated",
            "result_document": "Result file attached",
            "invoice": "Invoice attached",
        }.get(attachment["attachment_type"], "File attached")
        attachment_detail = attachment["note"] or attachment["original_filename"]
        if attachment["attachment_type"] == "request_document" and sample_request["description"]:
            attachment_detail = f"{sample_request['description']}\n{attachment_detail}".strip()
        entries.append(
            {
                "at": attachment["uploaded_at"],
                "title": attachment_title,
                "detail": attachment_detail,
                "actor": attachment["uploaded_by_name"] or "System",
                "kind": "file",
                "scope": "files",
                "files": [{"id": attachment["id"], "original_filename": attachment["original_filename"]}],
            }
        )
    for note in messages:
        note_kind = note["note_kind"] or ""
        note_title = "Comment" if note_kind in {"requester_note", "operator_note"} else note_kind_label(note_kind)
        entries.append(
            {
                "at": note["created_at"],
                "title": note_title,
                "detail": note["message_body"],
                "actor": note["sender_name"],
                "kind": "note",
                "scope": "conversation",
                "side": "left" if note["sender_user_id"] == sample_request["requester_id"] else "right",
                "files": [
                    {"id": attachment["id"], "original_filename": attachment["original_filename"]}
                    for attachment in (message_attachments or {}).get(note["id"], [])
                ],
            }
        )
    entries.sort(key=lambda item: item.get("at") or "")
    seen: set[tuple] = set()
    deduped = []
    for entry in entries:
        signature = (entry.get("at"), entry.get("title"), entry.get("detail"))
        if signature in seen:
            continue
        seen.add(signature)
        deduped.append(entry)
    return deduped


def instrument_timeline_entries(instrument: sqlite3.Row, logs: list[sqlite3.Row]) -> list[dict]:
    entries: list[dict] = []
    for log in logs:
        payload = {}
        try:
            payload = json.loads(log["payload_json"] or "{}")
        except Exception:
            payload = {}
        title = {
            "instrument_created": "Instrument created",
            "instrument_metadata_updated": "Details updated",
            "instrument_operation_updated": "Operation changed",
            "instrument_archived": "Instrument archived",
            "instrument_restored": "Instrument restored",
            "downtime_added": "Downtime added",
        }.get(log["action"], log["action"].replace("_", " ").title())
        detail = ""
        if log["action"] == "instrument_operation_updated":
            detail = intake_mode_label(payload.get("intake_mode"))
        elif log["action"] == "instrument_metadata_updated":
            parts = []
            if payload.get("location"):
                parts.append(payload["location"])
            machine_label = " / ".join(part for part in [payload.get("manufacturer", ""), payload.get("model_number", "")] if part)
            if machine_label:
                parts.append(machine_label)
            detail = " · ".join(parts)
        elif log["action"] == "downtime_added":
            start_label = format_dt(payload.get("start_time"))
            end_label = format_dt(payload.get("end_time"))
            detail = " to ".join(part for part in [start_label, end_label] if part and part != "-")
            if payload.get("reason"):
                detail = f"{detail}\n{payload['reason']}".strip()
        entries.append(
            {
                "at": log["created_at"],
                "title": title,
                "detail": detail,
                "actor": log["actor_name"] or "System",
                "action": log["action"],
            }
        )
    return entries


def format_dt(value: object | None) -> str:
    if value in {None, "", "-"}:
        return "-"
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return "-"
        normalized = text.replace("Z", "+00:00")
        dt = None
        for candidate in (normalized, text):
            try:
                dt = datetime.fromisoformat(candidate)
                break
            except ValueError:
                pass
        if dt is None:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
        if dt is None:
            return text
    return dt.strftime("%d/%m/%Y %H:%M:%S")


def time_ago(value: object | None) -> str:
    """Short humanised 'x ago' hint for a timestamp. '' on parse failure.

    Used by `.row-time-hint` spans in list templates to give a warmth
    cue next to the exact timestamp. Server-side, no JS.
    """
    if value in {None, "", "-"}:
        return ""
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return ""
        dt = None
        for candidate in (text.replace("Z", "+00:00"), text):
            try:
                dt = datetime.fromisoformat(candidate)
                break
            except ValueError:
                pass
        if dt is None:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
        if dt is None:
            return ""
    now = datetime.now(dt.tzinfo) if dt.tzinfo else datetime.now()
    delta = now - dt
    seconds = int(delta.total_seconds())
    if seconds < 0:
        # Future timestamps (scheduled_for) render as "in 3h".
        seconds = -seconds
        prefix, suffix = "in ", ""
    else:
        prefix, suffix = "", " ago"
    if seconds < 45:
        return "in moments" if prefix else "just now"
    if seconds < 3600:
        return f"{prefix}{seconds // 60}m{suffix}"
    if seconds < 86400:
        return f"{prefix}{seconds // 3600}h{suffix}"
    days = seconds // 86400
    if days < 30:
        return f"{prefix}{days}d{suffix}"
    if days < 365:
        return f"{prefix}{days // 30}mo{suffix}"
    return f"{prefix}{days // 365}y{suffix}"


def format_date(value: object | None) -> str:
    if value in {None, "", "-"}:
        return "-"
    if isinstance(value, datetime):
        dt_value = value.date()
    elif isinstance(value, date):
        dt_value = value
    else:
        text = str(value).strip()
        if not text:
            return "-"
        for parser in (
            lambda v: date.fromisoformat(v),
            lambda v: datetime.fromisoformat(v.replace("Z", "+00:00")).date(),
        ):
            try:
                dt_value = parser(text)
                break
            except Exception:
                dt_value = None
        if dt_value is None:
            return text
    return dt_value.strftime("%d/%m/%Y")


def format_duration_short(hours_value: object | None) -> str:
    if hours_value in {None, "", "-"}:
        return "-"
    try:
        total_hours = float(hours_value)
    except (TypeError, ValueError):
        return "-"
    if total_hours < 0:
        return "-"
    total_minutes = int(round(total_hours * 60))
    days, remainder = divmod(total_minutes, 24 * 60)
    hours, minutes = divmod(remainder, 60)
    parts: list[str] = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes and not days:
        parts.append(f"{minutes}m")
    return " ".join(parts) if parts else "0m"


def format_duration_days(hours_value: object | None) -> str:
    if hours_value in {None, "", "-"}:
        return "-"
    try:
        total_days = float(hours_value) / 24.0
    except (TypeError, ValueError):
        return "-"
    if total_days < 0:
        return "-"
    if total_days >= 10:
        return f"{round(total_days):.0f} d"
    if total_days >= 1:
        return f"{total_days:.1f} d"
    return "0 d"


def instrument_intake_mode(instrument_row: sqlite3.Row | dict | None) -> str:
    if not instrument_row:
        return "accepting"
    accepting = int(row_value(instrument_row, "accepting_requests", 1) or 0)
    on_hold = int(row_value(instrument_row, "soft_accept_enabled", 0) or 0)
    if accepting:
        return "accepting"
    if on_hold:
        return "on_hold"
    return "maintenance"


def intake_mode_label(mode: str | None) -> str:
    labels = {
        "accepting": "Accepting",
        "on_hold": "On Hold",
        "maintenance": "Maintenance",
    }
    return labels.get(mode or "", "Accepting")


def intake_mode_flags(mode: str | None) -> tuple[int, int]:
    normalized = (mode or "").strip().lower()
    if normalized == "on_hold":
        return 0, 1
    if normalized == "maintenance":
        return 0, 0
    return 1, 0


def next_instrument_code() -> str:
    row = query_one("SELECT code FROM instruments WHERE code LIKE 'INST-%' ORDER BY id DESC LIMIT 1")
    if row and row["code"]:
        try:
            number = int(str(row["code"]).split("-")[-1]) + 1
        except ValueError:
            number = scoped_instrument_count(current_user()) + 1 if current_user() else 1
    else:
        number = 1
    return f"INST-{number:03d}"


def _load_balance_pick(
    db: sqlite3.Connection,
    candidate_ids: list[int],
    role: str,
    instrument_id: int,
) -> int | None:
    """Round-robin pick among candidate approvers for one (role, instrument).

    The "pool" is `candidate_ids` — every user currently eligible to be
    the default approver for this role on this instrument. If the pool
    has only one member, return them. If it has several, pick the one
    with the fewest currently-pending `approval_steps` rows for this
    role on this instrument, tie-broken by the least-recently-acted-on
    step, tie-broken by lowest user id.

    The net effect: sequential requests against an instrument that has
    two operators / two finance officers / etc. will alternate between
    them so workload stays balanced, while a request whose approval
    chain names an explicit person (see `create_approval_chain()` —
    `cfg["approver_user_id"]` wins over this helper) still routes to
    that specific person unchanged.
    """
    if not candidate_ids:
        return None
    if len(candidate_ids) == 1:
        return candidate_ids[0]
    # The LEFT JOIN only matches approval_steps that are BOTH
    # (a) owned by this user in this role AND (b) still pending AND
    # (c) attached to a sample_request on this instrument. Any step
    # that fails any of those is simply unjoined — it does NOT
    # contaminate the count. That matters: approval_steps sit on a
    # global table and without the instrument scoping a user's work
    # on every instrument would count against their fairness here.
    placeholders = ",".join("?" * len(candidate_ids))
    rows = db.execute(
        f"""
        SELECT u.id AS user_id,
               COUNT(aps.id) AS pending_count,
               MAX(aps.acted_at) AS last_acted_at
        FROM users u
        LEFT JOIN approval_steps aps
               ON aps.approver_user_id = u.id
              AND aps.approver_role = ?
              AND aps.status = 'pending'
              AND EXISTS (
                  SELECT 1 FROM sample_requests sr
                   WHERE sr.id = aps.sample_request_id
                     AND sr.instrument_id = ?
              )
        WHERE u.id IN ({placeholders})
        GROUP BY u.id
        ORDER BY pending_count ASC,
                 (last_acted_at IS NULL) DESC,
                 last_acted_at ASC,
                 u.id ASC
        """,
        (role, instrument_id, *candidate_ids),
    ).fetchall()
    if not rows:
        return candidate_ids[0]
    return rows[0]["user_id"]


def _default_user_for_approval_role(db: sqlite3.Connection, role: str, instrument_id: int) -> int | None:
    """Pick the default approver for (role, instrument).

    Historically this used `ORDER BY u.id LIMIT 1` on each role, which
    meant every new request on an instrument with multiple operators
    (or multiple finance officers, or multiple professors) piled up on
    the lowest-id person in the pool. Now we collect every eligible
    candidate and delegate to `_load_balance_pick()` so workload
    round-robins across the pool.

    This helper only fires when `instrument_approval_config` does NOT
    name a specific `approver_user_id` — explicit per-person routes
    still win (see `create_approval_chain()`).
    """
    if role == "finance":
        rows = db.execute(
            "SELECT id FROM users WHERE role = 'finance_admin' AND active = 1 ORDER BY id"
        ).fetchall()
    elif role == "professor":
        rows = db.execute(
            """
            SELECT u.id FROM users u
            LEFT JOIN instrument_faculty_admins ifa ON ifa.user_id = u.id AND ifa.instrument_id = ?
            WHERE u.active = 1 AND (ifa.instrument_id IS NOT NULL OR u.role IN ('professor_approver', 'super_admin'))
            ORDER BY (ifa.instrument_id IS NOT NULL) DESC, u.id
            """,
            (instrument_id,),
        ).fetchall()
    elif role == "operator":
        rows = db.execute(
            """
            SELECT u.id FROM users u
            JOIN instrument_operators io ON io.user_id = u.id
            WHERE io.instrument_id = ? AND u.active = 1
            ORDER BY u.id
            """,
            (instrument_id,),
        ).fetchall()
    else:
        rows = []
    candidate_ids = [r["id"] for r in rows]
    return _load_balance_pick(db, candidate_ids, role, instrument_id)


def create_approval_chain(db: sqlite3.Connection, request_id: int, instrument_id: int) -> None:
    config_steps = db.execute(
        "SELECT * FROM instrument_approval_config WHERE instrument_id = ? ORDER BY step_order",
        (instrument_id,),
    ).fetchall()
    if config_steps:
        steps = []
        for cfg in config_steps:
            user_id = cfg["approver_user_id"] or _default_user_for_approval_role(db, cfg["approver_role"], instrument_id)
            steps.append((request_id, cfg["step_order"], cfg["approver_role"], user_id))
    else:
        finance_id = _default_user_for_approval_role(db, "finance", instrument_id)
        professor_id = _default_user_for_approval_role(db, "professor", instrument_id)
        operator_id = _default_user_for_approval_role(db, "operator", instrument_id)
        steps = [
            (request_id, 1, "finance", finance_id),
            (request_id, 2, "professor", professor_id),
            (request_id, 3, "operator", operator_id),
        ]
    db.executemany(
        """
        INSERT INTO approval_steps (sample_request_id, step_order, approver_role, approver_user_id, status, remarks)
        VALUES (?, ?, ?, ?, 'pending', '')
        """,
        steps,
    )


def release_submitted_requests_for_instrument(instrument_id: int, actor_id: int) -> int:
    pending_rows = query_all(
        """
        SELECT id
        FROM sample_requests
        WHERE instrument_id = ? AND status = 'submitted'
        ORDER BY id
        """,
        (instrument_id,),
    )
    released = 0
    db = get_db()
    for row in pending_rows:
        existing_steps = query_one(
            "SELECT 1 FROM approval_steps WHERE sample_request_id = ? LIMIT 1",
            (row["id"],),
        )
        if existing_steps is None:
            create_approval_chain(db, row["id"], instrument_id)
        # Source query constrains status to 'submitted' so the transition is always valid.
        assert_status_transition("submitted", "under_review")
        execute(
            "UPDATE sample_requests SET status = 'under_review', remarks = ?, updated_at = ? WHERE id = ?",
            ("Lab is accepting jobs again. Request released into review.", now_iso(), row["id"]),
        )
        execute(
            """
            INSERT INTO request_messages (request_id, sender_user_id, note_kind, message_body, created_at, is_active)
            VALUES (?, ?, 'operator_note', ?, ?, 1)
            """,
            (row["id"], actor_id, "Lab is accepting jobs again. Your request has been released into review.", now_iso()),
        )
        log_action(actor_id, "sample_request", row["id"], "released_for_review", {"reason": "instrument_accepting"})
        write_request_metadata_snapshot(row["id"])
        released += 1
    dropoff_rows = query_all(
        """
        SELECT id
        FROM sample_requests
        WHERE instrument_id = ? AND status = 'awaiting_sample_submission'
        ORDER BY id
        """,
        (instrument_id,),
    )
    for row in dropoff_rows:
        execute(
            """
            INSERT INTO request_messages (request_id, sender_user_id, note_kind, message_body, created_at, is_active)
            VALUES (?, ?, 'operator_note', ?, ?, 1)
            """,
            (row["id"], actor_id, "The lab is accepting samples again. You can now drop off your sample.", now_iso()),
        )
        log_action(actor_id, "sample_request", row["id"], "dropoff_reopened", {"reason": "instrument_accepting"})
        write_request_metadata_snapshot(row["id"])
    return released


def approval_step_is_actionable(step: sqlite3.Row, all_steps: list[sqlite3.Row]) -> bool:
    if step["status"] != "pending":
        return False
    return all(prev["status"] == "approved" for prev in all_steps if prev["step_order"] < step["step_order"])


def nav_pending_counts(user) -> dict[str, int]:
    """W1.4.1 — "what needs my attention right now" counts for the topbar nav.

    Returns a dict keyed by nav item (currently just ``"queue"``) whose
    values are the number of items the current user is expected to act
    on next. The template renders ``<span class="topbar-count-badge">``
    when the value is non-zero, and renders nothing when zero — the nav
    stays visually quiet for idle users and the count appears the instant
    a role gets work queued up.

    Rules:

    * **Approvers** (finance_admin / professor_approver / operator /
      instrument_admin / faculty_in_charge): any ``approval_steps`` row
      assigned to me with status ``pending`` and every earlier step for
      the same request already ``approved`` — i.e. actually actionable
      per ``approval_step_is_actionable``'s definition, computed in SQL.
    * **Requesters**: any of my own requests sitting in
      ``awaiting_sample_submission`` — the state machine is blocked on
      me until I hand samples over.

    Both rules can apply to the same user (an owner who is also a
    requester); the counts add. Empty dict when there is nothing — the
    template uses `{% if counts.queue %}` so an idle user sees no badge
    at all.
    """
    if not user:
        return {}
    db = get_db()
    queue_total = 0
    step_row = db.execute(
        """
        SELECT COUNT(*)
        FROM approval_steps s
        WHERE s.status = 'pending'
          AND s.approver_user_id = ?
          AND NOT EXISTS (
              SELECT 1 FROM approval_steps p
              WHERE p.sample_request_id = s.sample_request_id
                AND p.step_order < s.step_order
                AND p.status != 'approved'
          )
        """,
        (user["id"],),
    ).fetchone()
    if step_row and step_row[0]:
        queue_total += int(step_row[0])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "requester":
        own = db.execute(
            "SELECT COUNT(*) FROM sample_requests "
            "WHERE requester_id = ? AND status = 'awaiting_sample_submission'",
            (user["id"],),
        ).fetchone()[0]
        if own:
            queue_total += int(own)
    return {"queue": queue_total} if queue_total else {}


def safe_token(value: str) -> str:
    sanitized = secure_filename(value or "").replace(".", "_")
    return sanitized or "record"


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_instrument_image(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"png", "jpg", "jpeg"}


def instrument_photo_src(value: str | None) -> str:
    raw = (value or "").strip()
    if not raw:
        return url_for("static", filename="instrument-placeholder.svg")
    if raw.startswith("http://") or raw.startswith("https://") or raw.startswith("data:"):
        return raw
    if raw.startswith("instrument_images/"):
        return url_for("static", filename=raw)
    return raw


def save_instrument_image(instrument_id: int, uploaded_file) -> str:
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename:
        raise ValueError("No image selected.")
    if not allowed_instrument_image(original_filename):
        raise ValueError("Only PNG and JPG images are allowed for instrument photos.")
    sanitized = secure_filename(original_filename)
    extension = sanitized.rsplit(".", 1)[1].lower()
    INSTRUMENT_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    stored_filename = f"instrument_{instrument_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{extension}"
    full_path = INSTRUMENT_IMAGE_DIR / stored_filename
    uploaded_file.save(full_path)
    return f"instrument_images/{stored_filename}"


def request_folder_name(request_row: sqlite3.Row) -> str:
    return f"req_{request_row['id']}_{safe_token(request_row['request_no'])}"


def request_month_bucket(value: object | None) -> str:
    raw = str(value or "").strip()
    if raw:
        candidates = [raw]
        if raw.endswith("Z"):
            candidates.append(raw.replace("Z", "+00:00"))
        for candidate in candidates:
            try:
                return datetime.fromisoformat(candidate).strftime("%b %Y")
            except ValueError:
                continue
    return datetime.utcnow().strftime("%b %Y")


def request_instrument_bucket(request_row: sqlite3.Row) -> str:
    keys = set(request_row.keys())
    instrument_code = ""
    if "instrument_code" in keys:
        instrument_code = str(request_row["instrument_code"] or "").strip()
    if not instrument_code and "instrument_id" in keys:
        row = query_one("SELECT code FROM instruments WHERE id = ?", (request_row["instrument_id"],))
        instrument_code = str(row["code"] or "").strip() if row else ""
    return safe_token(instrument_code or f"instrument-{request_row['instrument_id']}")


def request_storage_root(request_row: sqlite3.Row) -> Path:
    month_bucket = request_month_bucket(request_row["created_at"])
    instrument_bucket = request_instrument_bucket(request_row)
    return UPLOAD_DIR / "requests" / month_bucket / instrument_bucket


def request_folder_path(request_row: sqlite3.Row) -> Path:
    return request_storage_root(request_row) / request_folder_name(request_row)


def request_attachments_path(request_row: sqlite3.Row) -> Path:
    return request_folder_path(request_row) / "attachments"


def ensure_request_folder(request_row: sqlite3.Row) -> Path:
    base_path = request_folder_path(request_row)
    (base_path / "attachments").mkdir(parents=True, exist_ok=True)
    return base_path


def request_metadata_path(request_row: sqlite3.Row) -> Path:
    return request_folder_path(request_row) / "request_metadata.json"


# ── Canonical join shape for `sample_requests` queries ──
# Aliases (load-bearing — every caller depends on these):
#   sr   — sample_requests
#   i    — instruments
#   r    — requester (users)
#   c    — created_by / originator (users)
#   op   — assigned operator (users)
#   recv — received_by operator (users)
#
# Add `REQUEST_ATTACHMENTS_JOIN` after this when you need the attachment
# count / GROUP_CONCAT — it depends on the same `sr` alias and brings in
# `ra` for the request_attachments table.
REQUEST_DETAIL_JOINS = """
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        JOIN users r ON r.id = sr.requester_id
        LEFT JOIN users c ON c.id = sr.created_by_user_id
        LEFT JOIN users op ON op.id = sr.assigned_operator_id
        LEFT JOIN users recv ON recv.id = sr.received_by_operator_id
"""

REQUEST_ATTACHMENTS_JOIN = """
        LEFT JOIN request_attachments ra ON ra.request_id = sr.id AND ra.is_active = 1
"""


def request_snapshot_row(request_id: int) -> sqlite3.Row | None:
    return query_one(
        f"""
        SELECT sr.*, i.name AS instrument_name, i.code AS instrument_code,
               r.name AS requester_name, r.email AS requester_email,
               c.name AS originator_name, c.email AS originator_email, c.role AS originator_role,
               op.name AS operator_name, recv.name AS received_by_name
        {REQUEST_DETAIL_JOINS}
        WHERE sr.id = ?
        """,
        (request_id,),
    )


def write_request_metadata_snapshot(request_id: int) -> None:
    request_row = request_snapshot_row(request_id)
    if request_row is None:
        return
    folder = ensure_request_folder(request_row)
    attachments = get_request_attachments(request_id)
    notes = get_request_notes(request_id)
    issues = get_request_issues(request_id)
    # v2.0.0 — finance state derived from peer aggregates since the
    # legacy columns on sample_requests were dropped.
    _finance_snapshot = computed_finance_for_request(get_db(), request_id)
    approval_steps = query_all(
        """
        SELECT step_order, approver_role, status, remarks, acted_at
        FROM approval_steps
        WHERE sample_request_id = ?
        ORDER BY step_order
        """,
        (request_id,),
    )
    payload = {
        "request_id": request_row["id"],
        "request_no": request_row["request_no"],
        "title": request_row["title"],
        "sample_name": request_row["sample_name"],
        "sample_ref": request_row["sample_ref"],
        "sample_count": request_row["sample_count"],
        "description": request_row["description"],
        "sample_origin": request_row["sample_origin"],
        "priority": request_row["priority"],
        "status": request_row["status"],
        "instrument": {
            "id": request_row["instrument_id"],
            "name": request_row["instrument_name"],
            "code": request_row["instrument_code"],
        },
        "requester": {
            "id": request_row["requester_id"],
            "name": request_row["requester_name"],
            "email": request_row["requester_email"],
        },
        "originator": {
            "id": request_row["created_by_user_id"],
            "name": request_row["originator_name"] or request_row["requester_name"],
            "email": request_row["originator_email"] or request_row["requester_email"],
            "role": request_row["originator_role"] or "requester",
        },
        "originator_note": request_row["originator_note"],
        "operator": {
            "id": request_row["assigned_operator_id"],
            "name": request_row["operator_name"],
        },
        "received_by": {
            "id": request_row["received_by_operator_id"],
            "name": request_row["received_by_name"],
        },
        "finance": {
            "receipt_number": _finance_snapshot["receipt_number"],
            "amount_due": _finance_snapshot["amount_due"],
            "amount_paid": _finance_snapshot["amount_paid"],
            "finance_status": _finance_snapshot["finance_status"],
        },
        "timing": {
            "created_at": request_row["created_at"],
            "updated_at": request_row["updated_at"],
            "submitted_to_lab_at": request_row["submitted_to_lab_at"],
            "sample_submitted_at": request_row["sample_submitted_at"],
            "sample_received_at": request_row["sample_received_at"],
            "scheduled_for": request_row["scheduled_for"],
            "completed_at": request_row["completed_at"],
        },
        "remarks": request_row["remarks"],
        "results_summary": request_row["results_summary"],
        "result_email_status": request_row["result_email_status"],
        "completion_locked": bool(request_row["completion_locked"]),
        "attachments": [
            {
                "id": row["id"],
                "original_filename": row["original_filename"],
                "attachment_type": row["attachment_type"],
                "note": row["note"],
                "uploaded_at": row["uploaded_at"],
                "uploaded_by": row["uploaded_by_name"],
                "relative_path": row["relative_path"],
                "file_size": row["file_size"],
            }
            for row in attachments
        ],
        "communication_notes": [
            {
                "id": row["id"],
                "note_kind": row["note_kind"],
                "sender_name": row["sender_name"],
                "sender_email": row["sender_email"],
                "sender_role": row["sender_role"],
                "message_body": row["message_body"],
                "created_at": row["created_at"],
            }
            for row in notes.values()
        ],
        "issues": [
            {
                "id": row["id"],
                "status": row["status"],
                "issue_message": row["issue_message"],
                "response_message": row["response_message"],
                "created_at": row["created_at"],
                "responded_at": row["responded_at"],
                "resolved_at": row["resolved_at"],
                "created_by_name": row["created_by_name"],
                "responded_by_name": row["responded_by_name"],
                "resolved_by_name": row["resolved_by_name"],
            }
            for row in issues
        ],
        "approval_steps": [dict(row) for row in approval_steps],
        "snapshot_generated_at": now_iso(),
    }
    request_metadata_path(request_row).write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def can_view_request(user: sqlite3.Row, request_row: sqlite3.Row) -> bool:
    profile = user_access_profile(user)
    if profile["can_view_all_requests"]:
        return True
    if request_row["requester_id"] == user["id"]:
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if can_manage_instrument(user["id"], request_row["instrument_id"], user["role"]):
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if can_operate_instrument(user["id"], request_row["instrument_id"], user["role"]):
        return True
    if profile["can_view_finance_stage"]:
        return query_one(
            """
            SELECT 1
            FROM approval_steps aps
            WHERE aps.sample_request_id = ?
              AND aps.approver_role = 'finance'
              AND aps.status = 'pending'
              AND NOT EXISTS (
                SELECT 1
                FROM approval_steps prev
                WHERE prev.sample_request_id = aps.sample_request_id
                  AND prev.step_order < aps.step_order
                  AND prev.status != 'approved'
              )
            """,
            (request_row["id"],),
        ) is not None
    if profile["can_view_professor_stage"]:
        return query_one(
            """
            SELECT 1
            FROM approval_steps aps
            WHERE aps.sample_request_id = ?
              AND aps.approver_role = 'professor'
              AND aps.status = 'pending'
              AND NOT EXISTS (
                SELECT 1
                FROM approval_steps prev
                WHERE prev.sample_request_id = aps.sample_request_id
                  AND prev.step_order < aps.step_order
                  AND prev.status != 'approved'
              )
            """,
            (request_row["id"],),
        ) is not None
    return False


def can_upload_attachment(user: sqlite3.Row, request_row: sqlite3.Row) -> bool:
    profile = user_access_profile(user)
    if profile["can_view_all_requests"]:
        return True
    if request_row["completion_locked"]:
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        return user["role"] in POST_COMPLETION_UPLOAD_ROLES and (
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            can_manage_instrument(user["id"], request_row["instrument_id"], user["role"])
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            or can_operate_instrument(user["id"], request_row["instrument_id"], user["role"])
        )
    if request_row["requester_id"] == user["id"]:
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if can_manage_instrument(user["id"], request_row["instrument_id"], user["role"]):
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if can_operate_instrument(user["id"], request_row["instrument_id"], user["role"]):
        return True
    return False


def can_delete_attachment(user: sqlite3.Row, attachment: sqlite3.Row, request_row: sqlite3.Row) -> bool:
    if user_access_profile(user)["can_view_all_requests"]:
        return True
    if request_row["completion_locked"]:
        return False
    return attachment["uploaded_by_user_id"] == user["id"]


def can_post_message(user: sqlite3.Row, request_row: sqlite3.Row) -> bool:
    return can_view_request(user, request_row)


def note_kind_label(note_kind: str) -> str:
    for key, label, _hint in COMMUNICATION_NOTE_TYPES:
        if key == note_kind:
            return label
    return note_kind.replace("_", " ").title()


def can_edit_request_note(user: sqlite3.Row, request_row: sqlite3.Row, note_kind: str) -> bool:
    if note_kind == "requester_note":
        return request_row["requester_id"] == user["id"] and not request_row["completion_locked"]
    if user_access_profile(user)["can_view_all_requests"]:
        return True
    if note_kind in {"lab_reply", "operator_note", "final_note"}:
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        return can_manage_instrument(user["id"], request_row["instrument_id"], user["role"]) or can_operate_instrument(
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            user["id"], request_row["instrument_id"], user["role"]
        )
    return False


def row_value(row: sqlite3.Row | dict | None, key: str, default=None):
    if row is None:
        return default
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        return default


def request_card_viewer_kind(user: sqlite3.Row, request_row: sqlite3.Row) -> str:
    profile = user_access_profile(user)
    if profile["is_owner"]:
        return "owner"
    if profile["can_view_all_requests"]:
        return "global_admin"
    if profile["can_view_professor_stage"]:
        return "faculty_admin"
    if profile["can_view_finance_stage"]:
        return "finance"
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if can_manage_instrument(user["id"], request_row["instrument_id"], user["role"]):
        return "instrument_admin"
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if can_operate_instrument(user["id"], request_row["instrument_id"], user["role"]):
        return "operator"
    if request_row["requester_id"] == user["id"]:
        return "requester"
    return "viewer"


def request_card_field_allowed(user: sqlite3.Row, request_row: sqlite3.Row, field_name: str) -> bool:
    viewer_kind = request_card_viewer_kind(user, request_row)
    profile = user_access_profile(user)
    always_visible = {
        "request_no",
        "status",
        "instrument",
        "sample_ref",
        "sample_name",
        "sample_count",
        "created_at",
        "completed_at",
        "description",
    }
    if field_name in always_visible:
        return True

    if field_name in set(profile["card_visible_fields"]):
        return True
    field_roles = {
        "requester_identity": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester"},
        "operator_identity": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester"},
        "remarks": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester", "finance"},
        "results_summary": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester"},
        "submitted_documents": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester", "finance"},
        "conversation": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester", "finance"},
        "events": {"owner", "global_admin", "faculty_admin", "instrument_admin", "operator", "requester", "finance"},
    }
    return viewer_kind in field_roles.get(field_name, set())


def request_card_actions(user: sqlite3.Row, request_row: sqlite3.Row) -> dict[str, bool]:
    profile = user_access_profile(user)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_manage = can_manage_instrument(user["id"], request_row["instrument_id"], user["role"])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_operate = can_operate_instrument(user["id"], request_row["instrument_id"], user["role"])
    return {
        "reply": "reply" in set(profile["card_action_fields"]) and can_post_message(user, request_row),
        "upload_attachment": "upload_attachment" in set(profile["card_action_fields"]) and can_upload_attachment(user, request_row),
        "mark_submitted": "mark_submitted" in set(profile["card_action_fields"])
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        and user["role"] == "requester"
        and request_row["requester_id"] == user["id"]
        and request_row["status"] == "awaiting_sample_submission"
        and instrument_intake_mode(request_row) == "accepting",
        "finish_fast": "finish_fast" in set(profile["card_action_fields"]) and (can_manage or can_operate) and request_row["status"] not in {"completed", "rejected"},
        "reassign": "reassign" in set(profile["card_action_fields"]) and (can_manage or can_operate),
        "mark_received": "mark_received" in set(profile["card_action_fields"]) and (can_manage or can_operate) and request_row["status"] == "sample_submitted",
        "update_status": "update_status" in set(profile["card_action_fields"]) and can_manage,
    }


def request_card_attachment_allowed(user: sqlite3.Row, request_row: sqlite3.Row, attachment: sqlite3.Row) -> bool:
    if not can_view_request(user, request_row):
        return False
    if attachment["attachment_type"] in {"request_document", "sample_slip"}:
        return request_card_field_allowed(user, request_row, "submitted_documents")
    return request_card_field_allowed(user, request_row, "events")


def request_card_event_allowed(user: sqlite3.Row, request_row: sqlite3.Row, entry: dict) -> bool:
    if not can_view_request(user, request_row):
        return False
    scope = entry.get("scope", "workflow")
    if scope == "conversation":
        return request_card_field_allowed(user, request_row, "conversation")
    if scope == "files":
        return request_card_field_allowed(user, request_row, "events") or request_card_field_allowed(user, request_row, "submitted_documents")
    return request_card_field_allowed(user, request_row, "events")


def request_card_visible_attachments(user: sqlite3.Row, request_row: sqlite3.Row, attachments: list[sqlite3.Row]) -> list[sqlite3.Row]:
    return [attachment for attachment in attachments if request_card_attachment_allowed(user, request_row, attachment)]


def request_card_visible_timeline(user: sqlite3.Row, request_row: sqlite3.Row, entries: list[dict]) -> list[dict]:
    return [entry for entry in entries if request_card_event_allowed(user, request_row, entry)]


def completion_override_fields(request_row: sqlite3.Row, actor_user_id: int, timestamp_value: str) -> dict[str, object]:
    return {
        "submitted_to_lab_at": request_row["submitted_to_lab_at"] or timestamp_value,
        "sample_submitted_at": request_row["sample_submitted_at"] or timestamp_value,
        "sample_received_at": request_row["sample_received_at"] or timestamp_value,
        "received_by_operator_id": request_row["received_by_operator_id"] or actor_user_id,
        "scheduled_for": request_row["scheduled_for"] or timestamp_value,
        "assigned_operator_id": request_row["assigned_operator_id"] or actor_user_id,
        "completed_at": timestamp_value,
    }


def log_completion_override_events(
    actor_user_id: int,
    request_row: sqlite3.Row,
    completion_fields: dict[str, object],
    timestamp_value: str,
    final_action: str,
    final_payload: dict,
) -> None:
    if not request_row["submitted_to_lab_at"]:
        log_action_at(
            actor_user_id,
            "sample_request",
            request_row["id"],
            "sample_submitted",
            {"remarks": "Backfilled during override completion."},
            timestamp_value,
        )
    if not request_row["sample_received_at"]:
        log_action_at(
            actor_user_id,
            "sample_request",
            request_row["id"],
            "sample_received",
            {"remarks": "Backfilled during override completion."},
            timestamp_value,
        )
    if not request_row["scheduled_for"]:
        log_action_at(
            actor_user_id,
            "sample_request",
            request_row["id"],
            "scheduled",
            {"scheduled_for": completion_fields["scheduled_for"], "remarks": "Backfilled during override completion."},
            timestamp_value,
        )
    log_action_at(actor_user_id, "sample_request", request_row["id"], final_action, final_payload, timestamp_value)


def parse_schedule_day(value: str | None) -> date:
    parsed = parse_date_param(value or "")
    return parsed or datetime.utcnow().date()


def compute_next_schedule_slot(day_value: date, existing_rows: list[sqlite3.Row]) -> datetime:
    slots: list[datetime] = []
    for row in existing_rows:
        scheduled_for = row["scheduled_for"]
        if not scheduled_for:
            continue
        parsed = None
        try:
            parsed = datetime.fromisoformat(str(scheduled_for).replace("Z", "+00:00"))
        except ValueError:
            try:
                parsed = datetime.strptime(str(scheduled_for)[:16], "%Y-%m-%dT%H:%M")
            except ValueError:
                try:
                    parsed = datetime.strptime(str(scheduled_for)[:19], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    parsed = None
        if parsed and parsed.date() == day_value:
            slots.append(parsed.replace(second=0, microsecond=0))
    if not slots:
        return datetime.combine(day_value, datetime.min.time()).replace(hour=9, minute=0)
    return max(slots) + timedelta(minutes=30)


def request_card_policy(user: sqlite3.Row, request_row: sqlite3.Row) -> dict:
    return {
        "viewer_kind": request_card_viewer_kind(user, request_row),
        "fields": {
            "requester_identity": request_card_field_allowed(user, request_row, "requester_identity"),
            "operator_identity": request_card_field_allowed(user, request_row, "operator_identity"),
            "remarks": request_card_field_allowed(user, request_row, "remarks"),
            "results_summary": request_card_field_allowed(user, request_row, "results_summary"),
            "submitted_documents": request_card_field_allowed(user, request_row, "submitted_documents"),
            "conversation": request_card_field_allowed(user, request_row, "conversation"),
            "events": request_card_field_allowed(user, request_row, "events"),
        },
        "actions": request_card_actions(user, request_row),
    }


def can_flag_request_issue(user: sqlite3.Row, request_row: sqlite3.Row) -> bool:
    if request_row["requester_id"] == user["id"]:
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "super_admin":
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    return can_manage_instrument(user["id"], request_row["instrument_id"], user["role"]) or can_operate_instrument(
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        user["id"], request_row["instrument_id"], user["role"]
    )


def can_respond_request_issue(user: sqlite3.Row, request_row: sqlite3.Row) -> bool:
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "super_admin":
        return True
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    return can_manage_instrument(user["id"], request_row["instrument_id"], user["role"]) or can_operate_instrument(
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        user["id"], request_row["instrument_id"], user["role"]
    )


def can_view_user_profile(viewer: sqlite3.Row, target_user: sqlite3.Row) -> bool:
    if viewer["id"] == target_user["id"]:
        return True
    if user_access_profile(viewer)["can_view_user_profiles"] and user_access_profile(viewer)["can_view_all_requests"]:
        return True
    instrument_ids = assigned_instrument_ids(viewer)
    if not instrument_ids:
        return False
    placeholders = ",".join("?" for _ in instrument_ids)
    related_request = query_one(
        f"""
        SELECT 1
        FROM sample_requests sr
        WHERE sr.instrument_id IN ({placeholders})
          AND (
            sr.requester_id = ?
            OR sr.created_by_user_id = ?
            OR sr.assigned_operator_id = ?
            OR sr.received_by_operator_id = ?
          )
        LIMIT 1
        """,
        tuple(instrument_ids) + (target_user["id"], target_user["id"], target_user["id"], target_user["id"]),
    )
    if related_request is not None:
        return True
    related_assignment = query_one(
        f"""
        SELECT 1
        FROM (
          SELECT user_id, instrument_id FROM instrument_admins
          UNION ALL
          SELECT user_id, instrument_id FROM instrument_operators
          UNION ALL
          SELECT user_id, instrument_id FROM instrument_faculty_admins
          UNION ALL
          SELECT user_id, instrument_id FROM instrument_requesters
        ) assignments
        WHERE assignments.user_id = ?
          AND assignments.instrument_id IN ({placeholders})
        LIMIT 1
        """,
        (target_user["id"], *instrument_ids),
    )
    return related_assignment is not None


def can_view_user_profile_id(viewer: sqlite3.Row | None, user_id: int | None) -> bool:
    if viewer is None or not user_id:
        return False
    target_user = query_one("SELECT id, name, email, role, invite_status, active FROM users WHERE id = ?", (user_id,))
    if target_user is None:
        return False
    return can_view_user_profile(viewer, target_user)


def can_view_instrument_history(user: sqlite3.Row, instrument_id: int) -> bool:
    profile = user_access_profile(user)
    if profile["can_view_all_instruments"]:
        return True
    if not profile["can_access_instruments"]:
        return False
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    return can_manage_instrument(user["id"], instrument_id, user["role"]) or can_operate_instrument(user["id"], instrument_id, user["role"])


def can_open_instrument_detail(user: sqlite3.Row, instrument_id: int) -> bool:
    profile = user_access_profile(user)
    if profile["can_view_all_instruments"]:
        return True
    if not profile["can_access_instruments"]:
        return False
    return instrument_id in assigned_instrument_ids(user)


def can_view_group_visualization(user: sqlite3.Row, group_name: str) -> bool:
    if not can_access_stats(user):
        return False
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] in {"super_admin", "site_admin"}:
        return True
    ids = assigned_instrument_ids(user)
    if not ids:
        return False
    placeholders = ",".join("?" for _ in ids)
    row = query_one(
        f"SELECT 1 FROM instruments WHERE id IN ({placeholders}) AND COALESCE(faculty_group, '') = ? LIMIT 1",
        (*ids, group_name),
    )
    return row is not None


def approval_candidate_options(step_role: str, instrument_id: int) -> list[sqlite3.Row]:
    if step_role == "finance":
        return query_all(
            "SELECT id, name, email, role FROM users WHERE active = 1 AND role IN ('finance_admin', 'site_admin', 'super_admin') ORDER BY name"
        )
    if step_role == "professor":
        return query_all(
            """
            SELECT DISTINCT u.id, u.name, u.email, u.role
            FROM users u
            LEFT JOIN instrument_faculty_admins ifa ON ifa.user_id = u.id AND ifa.instrument_id = ?
            WHERE u.active = 1
              AND (
                ifa.instrument_id IS NOT NULL OR
                u.role IN ('professor_approver', 'site_admin', 'super_admin')
              )
            ORDER BY u.name
            """,
            (instrument_id,),
        )
    return query_all(
        """
        SELECT DISTINCT u.id, u.name, u.email, u.role
        FROM users u
        LEFT JOIN instrument_operators io ON io.user_id = u.id AND io.instrument_id = ?
        LEFT JOIN instrument_admins ia ON ia.user_id = u.id AND ia.instrument_id = ?
        LEFT JOIN instrument_faculty_admins ifa ON ifa.user_id = u.id AND ifa.instrument_id = ?
        WHERE u.active = 1
          AND (
            io.instrument_id IS NOT NULL OR
            ia.instrument_id IS NOT NULL OR
            ifa.instrument_id IS NOT NULL OR
            u.role = 'super_admin'
          )
        ORDER BY u.name
        """,
        (instrument_id, instrument_id, instrument_id),
    )


def candidate_allowed_for_step(candidate: sqlite3.Row | None, step_role: str, instrument_id: int) -> bool:
    """True if `candidate` may be assigned to an approval step of type
    `step_role` on `instrument_id`. Uses `user_role_set()` so users with
    multi-role assignments via the `user_roles` junction pass the check
    even when their primary `users.role` column says otherwise. Fixes
    the v1.6.x assign-approver bug where the UI would reject valid
    multi-role admins. (v1.5.0 TODO marker retired at this site.)"""
    if candidate is None:
        return False
    roles = user_role_set(candidate)
    if step_role == "finance":
        return bool(roles & {"finance_admin", "super_admin", "site_admin"})
    if step_role == "professor":
        if roles & {"professor_approver", "super_admin", "site_admin"}:
            return True
        row = query_one(
            "SELECT 1 FROM instrument_faculty_admins WHERE user_id = ? AND instrument_id = ?",
            (candidate["id"], instrument_id),
        )
        return row is not None
    # Operator / instrument_admin step fall-through. Any super- or
    # site-admin is always allowed; anyone else must have an explicit
    # row in the instrument-access tables.
    if roles & {"super_admin", "site_admin"}:
        return True
    primary_role = candidate["role"] if "role" in candidate.keys() else ""
    return can_manage_instrument(candidate["id"], instrument_id, primary_role) or can_operate_instrument(candidate["id"], instrument_id, primary_role)


def attachment_type_choices() -> list[str]:
    return ["request_document", "sample_slip", "result_document", "invoice", "other"]


def get_request_attachments(request_id: int) -> list[sqlite3.Row]:
    return query_all(
        """
        SELECT ra.*, u.name AS uploaded_by_name
        FROM request_attachments ra
        LEFT JOIN users u ON u.id = ra.uploaded_by_user_id
        WHERE ra.request_id = ? AND ra.is_active = 1
        ORDER BY ra.uploaded_at DESC, ra.id DESC
        """,
        (request_id,),
    )


def get_request_notes(request_id: int) -> dict[str, sqlite3.Row]:
    rows = query_all(
        """
        SELECT rm.*, u.name AS sender_name, u.email AS sender_email, u.role AS sender_role
        FROM request_messages rm
        JOIN users u ON u.id = rm.sender_user_id
        WHERE rm.request_id = ? AND rm.is_active = 1
        ORDER BY rm.created_at DESC, rm.id DESC
        """,
        (request_id,),
    )
    grouped: dict[str, sqlite3.Row] = {}
    for row in rows:
        note_kind = row["note_kind"] or "requester_note"
        if note_kind not in grouped:
            grouped[note_kind] = row
    return grouped


def get_request_message_thread(request_id: int) -> list[sqlite3.Row]:
    return query_all(
        """
        SELECT rm.*, u.name AS sender_name, u.email AS sender_email, u.role AS sender_role
        FROM request_messages rm
        JOIN users u ON u.id = rm.sender_user_id
        WHERE rm.request_id = ? AND rm.is_active = 1
        ORDER BY rm.created_at ASC, rm.id ASC
        """,
        (request_id,),
    )


def attachments_by_message_ids(message_ids: list[int]) -> dict[int, list[sqlite3.Row]]:
    if not message_ids:
        return {}
    placeholders = ",".join("?" for _ in message_ids)
    rows = query_all(
        f"""
        SELECT ra.*, u.name AS uploaded_by_name
        FROM request_attachments ra
        LEFT JOIN users u ON u.id = ra.uploaded_by_user_id
        WHERE ra.request_message_id IN ({placeholders}) AND ra.is_active = 1
        ORDER BY ra.uploaded_at ASC, ra.id ASC
        """,
        tuple(message_ids),
    )
    grouped: dict[int, list[sqlite3.Row]] = {message_id: [] for message_id in message_ids}
    for row in rows:
        grouped.setdefault(row["request_message_id"], []).append(row)
    return grouped


def get_request_issues(request_id: int) -> list[sqlite3.Row]:
    return query_all(
        """
        SELECT ri.*,
               creator.name AS created_by_name,
               responder.name AS responded_by_name,
               resolver.name AS resolved_by_name
        FROM request_issues ri
        JOIN users creator ON creator.id = ri.created_by_user_id
        LEFT JOIN users responder ON responder.id = ri.responded_by_user_id
        LEFT JOIN users resolver ON resolver.id = ri.resolved_by_user_id
        WHERE ri.request_id = ?
        ORDER BY CASE WHEN ri.status = 'open' THEN 0 ELSE 1 END, ri.created_at DESC, ri.id DESC
        """,
        (request_id,),
    )


def attachments_by_request_ids(request_ids: list[int]) -> dict[int, list[sqlite3.Row]]:
    if not request_ids:
        return {}
    placeholders = ",".join("?" for _ in request_ids)
    rows = query_all(
        f"""
        SELECT ra.*, u.name AS uploaded_by_name
        FROM request_attachments ra
        LEFT JOIN users u ON u.id = ra.uploaded_by_user_id
        WHERE ra.request_id IN ({placeholders}) AND ra.is_active = 1
        ORDER BY ra.uploaded_at DESC, ra.id DESC
        """,
        tuple(request_ids),
    )
    grouped: dict[int, list[sqlite3.Row]] = {request_id: [] for request_id in request_ids}
    for row in rows:
        grouped.setdefault(row["request_id"], []).append(row)
    return grouped


def attachment_size_label(size: int | None) -> str:
    value = size or 0
    for unit in ["B", "KB", "MB", "GB"]:
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


def save_uploaded_attachment(
    request_row: sqlite3.Row,
    uploaded_file,
    uploaded_by_user_id: int,
    attachment_type: str,
    note: str,
    request_message_id: int | None = None,
) -> sqlite3.Row:
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename:
        raise ValueError("No file selected.")
    if not allowed_file(original_filename):
        raise ValueError("File type not allowed.")
    sanitized = secure_filename(original_filename)
    extension = sanitized.rsplit(".", 1)[1].lower()
    stored_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uploaded_by_user_id}_{sanitized}"
    ensure_request_folder(request_row)
    folder = request_attachments_path(request_row)
    full_path = folder / stored_filename
    uploaded_file.save(full_path)
    relative_path = str(full_path.relative_to(BASE_DIR))
    mime_type = uploaded_file.mimetype or mimetypes.guess_type(sanitized)[0] or "application/octet-stream"
    file_size = full_path.stat().st_size
    attachment_id = execute(
        """
        INSERT INTO request_attachments (
            request_id, user_id, instrument_id, original_filename, stored_filename, relative_path,
            file_extension, mime_type, file_size, uploaded_by_user_id, uploaded_at, attachment_type, note, is_active, request_message_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
        """,
        (
            request_row["id"],
            request_row["requester_id"],
            request_row["instrument_id"],
            original_filename,
            stored_filename,
            relative_path,
            extension,
            mime_type,
            file_size,
            uploaded_by_user_id,
            now_iso(),
            attachment_type,
            note,
            request_message_id,
        ),
    )
    log_action(
        uploaded_by_user_id,
        "sample_request",
        request_row["id"],
        "attachment_uploaded",
        {
            "filename": original_filename,
            "attachment_type": attachment_type,
            "note": note,
            "request_message_id": request_message_id,
        },
    )
    return query_one("SELECT * FROM request_attachments WHERE id = ?", (attachment_id,))


def save_generated_attachment(
    request_row: sqlite3.Row,
    filename: str,
    content_bytes: bytes,
    uploaded_by_user_id: int,
    attachment_type: str,
    note: str,
    mime_type: str = "application/pdf",
) -> sqlite3.Row:
    sanitized = secure_filename(filename)
    extension = sanitized.rsplit(".", 1)[1].lower() if "." in sanitized else "bin"
    stored_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uploaded_by_user_id}_{sanitized}"
    ensure_request_folder(request_row)
    folder = request_attachments_path(request_row)
    full_path = folder / stored_filename
    full_path.write_bytes(content_bytes)
    relative_path = str(full_path.relative_to(BASE_DIR))
    file_size = full_path.stat().st_size
    attachment_id = execute(
        """
        INSERT INTO request_attachments (
            request_id, user_id, instrument_id, original_filename, stored_filename, relative_path,
            file_extension, mime_type, file_size, uploaded_by_user_id, uploaded_at, attachment_type, note, is_active
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (
            request_row["id"],
            request_row["requester_id"],
            request_row["instrument_id"],
            filename,
            stored_filename,
            relative_path,
            extension,
            mime_type,
            file_size,
            uploaded_by_user_id,
            now_iso(),
            attachment_type,
            note,
        ),
    )
    log_action(
        uploaded_by_user_id,
        "sample_request",
        request_row["id"],
        "attachment_uploaded",
        {
            "filename": filename,
            "attachment_type": attachment_type,
            "note": note,
        },
    )
    return query_one("SELECT * FROM request_attachments WHERE id = ?", (attachment_id,))


def generate_sample_slip_pdf(request_row: sqlite3.Row, instrument_name: str, requester_name: str, uploaded_filename: str | None = None) -> bytes:
    lines = [
        f"Sample Number: {request_row['sample_ref']}",
        f"Job Number: {request_row['request_no']}",
        f"Instrument: {instrument_name}",
        f"Requester: {requester_name}",
        f"Sample: {request_row['sample_name']}",
        f"Sample Count: {request_row['sample_count']}",
        f"Created: {format_dt(request_row['created_at'])}",
        "Print this slip and attach it to the physical sample.",
    ]
    if uploaded_filename:
        lines.append(f"Request file: {uploaded_filename}")
    return simple_pdf_bytes("Sample Submission Slip", lines)


def parse_date_param(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def request_filter_values() -> dict[str, str]:
    return {
        "q": request.args.get("q", "").strip(),
        "status": request.args.get("status", "").strip(),
        "status_slice": request.args.get("status_slice", "all").strip() or "all",
        "instrument_id": request.args.get("instrument_id", "").strip(),
        "operator_id": request.args.get("operator_id", "").strip(),
        "requester_id": request.args.get("requester_id", "").strip(),
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "sort": request.args.get("sort", "created_desc").strip() or "created_desc",
    }


def page_value(default: int = 1) -> int:
    raw = request.args.get("page", str(default)).strip()
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return default
    return max(1, value)


def schedule_filter_values() -> dict[str, str]:
    values = request_filter_values()
    values["bucket"] = request.args.get("bucket", "").strip()
    values["period"] = request.args.get("period", "all").strip() or "all"
    return values


def calendar_filter_values() -> dict[str, str]:
    return {
        "instrument_id": request.values.get("instrument_id", "").strip(),
        "operator_id": request.values.get("operator_id", "").strip(),
        "show_scheduled": request.values.get("show_scheduled", "1"),
        "show_in_progress": request.values.get("show_in_progress", "1"),
        "show_completed": request.values.get("show_completed", "0"),
        "show_maintenance": request.values.get("show_maintenance", "1"),
        "view": request.values.get("view", "week").strip() or "week",
        "date": request.values.get("date", "").strip(),
    }


def row_anchor_date(row: sqlite3.Row) -> date | None:
    for field in ("scheduled_for", "completed_at", "sample_received_at", "sample_submitted_at", "created_at"):
        value = row[field] if field in row.keys() else None
        if value:
            return parse_date_param(str(value)[:10])
    return None


def row_matches_period(row: sqlite3.Row, period: str) -> bool:
    if period == "all":
        return True
    anchor = row_anchor_date(row)
    if anchor is None:
        return period == "all"
    today = datetime.utcnow().date()
    if period == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start <= anchor <= end
    if period == "month":
        return anchor.year == today.year and anchor.month == today.month
    return True


def request_history_query(
    where_clauses: list[str] | None = None,
    params: list | None = None,
    filters: dict[str, str] | None = None,
) -> tuple[str, list]:
    clauses = list(where_clauses or [])
    query_params = list(params or [])
    active_filters = filters or {}
    if active_filters.get("q"):
        # v2.0.0 — receipt_number / finance_status removed from
        # sample_requests; receipt search now hits payments via EXISTS.
        clauses.append(
            """
            (
                sr.request_no LIKE ? OR
                COALESCE(sr.sample_ref, '') LIKE ? OR
                sr.title LIKE ? OR
                sr.sample_name LIKE ? OR
                COALESCE(sr.description, '') LIKE ? OR
                COALESCE(sr.priority, '') LIKE ? OR
                COALESCE(sr.remarks, '') LIKE ? OR
                COALESCE(sr.results_summary, '') LIKE ? OR
                COALESCE(i.name, '') LIKE ? OR
                COALESCE(i.code, '') LIKE ? OR
                COALESCE(r.name, '') LIKE ? OR
                COALESCE(c.name, '') LIKE ? OR
                COALESCE(op.name, '') LIKE ? OR
                EXISTS (
                    SELECT 1 FROM payments p
                    JOIN invoices inv ON inv.id = p.invoice_id
                    WHERE inv.request_id = sr.id AND p.receipt_number LIKE ?
                )
            )
            """
        )
        token = f"%{active_filters['q']}%"
        query_params.extend([token] * 14)
    if active_filters.get("status"):
        clauses.append("sr.status = ?")
        query_params.append(active_filters["status"])
    if active_filters.get("instrument_id"):
        clauses.append("sr.instrument_id = ?")
        query_params.append(int(active_filters["instrument_id"]))
    if active_filters.get("operator_id"):
        clauses.append("sr.assigned_operator_id = ?")
        query_params.append(int(active_filters["operator_id"]))
    if active_filters.get("requester_id"):
        clauses.append("sr.requester_id = ?")
        query_params.append(int(active_filters["requester_id"]))
    if active_filters.get("date_from"):
        clauses.append("substr(sr.created_at, 1, 10) >= ?")
        query_params.append(active_filters["date_from"])
    if active_filters.get("date_to"):
        clauses.append("substr(sr.created_at, 1, 10) <= ?")
        query_params.append(active_filters["date_to"])
    sort_key = active_filters.get("sort", "created_desc")
    order_map = {
        "created_desc": "sr.created_at DESC, sr.id DESC",
        "created_asc": "sr.created_at ASC, sr.id ASC",
        "completed_desc": "COALESCE(sr.completed_at, '') DESC, sr.id DESC",
        "scheduled_desc": "COALESCE(sr.scheduled_for, '') DESC, sr.id DESC",
        "instrument_asc": "i.name ASC, sr.created_at DESC",
        "status_asc": "sr.status ASC, sr.created_at DESC",
    }
    order_sql = order_map.get(sort_key, order_map["created_desc"])
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    sql = f"""
        SELECT sr.*, i.name AS instrument_name, i.code AS instrument_code, r.name AS requester_name,
               c.name AS originator_name, c.email AS originator_email, c.role AS originator_role,
               op.name AS operator_name, recv.name AS received_by_name,
               COALESCE(COUNT(ra.id), 0) AS attachment_count,
               GROUP_CONCAT(ra.original_filename, ', ') AS attachment_names
        {REQUEST_DETAIL_JOINS}
        {REQUEST_ATTACHMENTS_JOIN}
        {where_sql}
        GROUP BY sr.id
        ORDER BY {order_sql}
    """
    return sql, query_params


def stats_payload(user: sqlite3.Row | None = None, report_filters: dict[str, str] | None = None) -> dict:
    scoped_user = user or current_user()
    return stats_payload_for_scope(scoped_user, report_filters)


def chart_rows(rows: list[sqlite3.Row], label_key: str, value_key: str, limit: int = 8) -> list[dict]:
    subset = list(rows[:limit])
    max_value = max((row[value_key] for row in subset), default=0) or 1
    output = []
    for row in subset:
        value = row[value_key]
        output.append(
            {
                "label": row[label_key],
                "value": value,
                "width": round((value / max_value) * 100, 1),
            }
        )
    return output


def scoped_stats_filters(user: sqlite3.Row, instrument_id: int | None = None, group_name: str | None = None) -> tuple[list[str], list]:
    clauses, params = request_scope_sql(user, "sr")
    if instrument_id:
        clauses.append("sr.instrument_id = ?")
        params.append(instrument_id)
    if group_name:
        clauses.append("COALESCE(i.faculty_group, '') = ?")
        params.append(group_name)
    return clauses, params


def stats_payload_for_scope(
    user: sqlite3.Row,
    report_filters: dict[str, str] | None = None,
    instrument_id: int | None = None,
    group_name: str | None = None,
) -> dict:
    active_report_filters = report_filters or {}
    if instrument_id is None and active_report_filters.get("instrument_id"):
        try:
            requested_instrument_id = int(active_report_filters["instrument_id"])
        except (TypeError, ValueError):
            requested_instrument_id = None
        if requested_instrument_id and any(row["id"] == requested_instrument_id for row in visible_instruments_for_user(user, active_only=False)):
            instrument_id = requested_instrument_id
    if group_name is None and active_report_filters.get("group_name"):
        group_name = active_report_filters["group_name"]
    db = get_db()
    clauses, params = scoped_stats_filters(user, instrument_id=instrument_id, group_name=group_name)
    metric_clauses, metric_params, report_window = apply_report_window(clauses, params, "sr.completed_at", active_report_filters)
    where_sql = f" AND {' AND '.join(metric_clauses)}" if metric_clauses else ""
    daily = query_all(
        """
        SELECT substr(sr.completed_at, 1, 10) AS bucket, COUNT(*) AS jobs, COALESCE(SUM(sr.sample_count), 0) AS samples
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        WHERE sr.status = 'completed' AND sr.completed_at IS NOT NULL
        GROUP BY substr(completed_at, 1, 10)
        ORDER BY bucket DESC
        """.replace("GROUP BY", f"{where_sql}\n        GROUP BY"),
        tuple(metric_params),
    )
    weekly = query_all(
        """
        SELECT strftime('%Y-W%W', substr(sr.completed_at, 1, 19)) AS bucket, COUNT(*) AS jobs, COALESCE(SUM(sr.sample_count), 0) AS samples
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        WHERE sr.status = 'completed' AND sr.completed_at IS NOT NULL
        GROUP BY strftime('%Y-W%W', substr(sr.completed_at, 1, 19))
        ORDER BY bucket DESC
        """.replace("GROUP BY", f"{where_sql}\n        GROUP BY"),
        tuple(metric_params),
    )
    monthly = query_all(
        """
        SELECT substr(sr.completed_at, 1, 7) AS bucket, COUNT(*) AS jobs, COALESCE(SUM(sr.sample_count), 0) AS samples
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        WHERE sr.status = 'completed' AND sr.completed_at IS NOT NULL
        GROUP BY substr(sr.completed_at, 1, 7)
        ORDER BY bucket DESC
        """.replace("GROUP BY", f"{where_sql}\n        GROUP BY"),
        tuple(metric_params),
    )
    today = datetime.utcnow().date()
    week_start = today - timedelta(days=today.weekday())
    last_week_start = week_start - timedelta(days=7)
    last_week_end = week_start - timedelta(days=1)
    month_start = today.replace(day=1)

    def scoped_completed_total(date_from: str | None, date_to: str | None, field: str) -> float:
        local_clauses = list(clauses)
        local_params = list(params)
        if date_from:
            local_clauses.append("substr(sr.completed_at, 1, 10) >= ?")
            local_params.append(date_from)
        if date_to:
            local_clauses.append("substr(sr.completed_at, 1, 10) <= ?")
            local_params.append(date_to)
        where = f" AND {' AND '.join(local_clauses)}" if local_clauses else ""
        row = query_one(
            f"""
            SELECT COALESCE({field}, 0) AS c
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            WHERE sr.status = 'completed'{where}
            """,
            tuple(local_params),
        )
        return row["c"] if row else 0

    request_clauses, request_params, _ = apply_report_window(clauses, params, "sr.created_at", active_report_filters)
    by_instrument = query_all(
        f"""
        SELECT i.id AS instrument_id, i.name AS instrument_name, i.code AS instrument_code, i.faculty_group,
               COUNT(sr.id) AS total_requests,
               SUM(CASE WHEN sr.status = 'completed' THEN 1 ELSE 0 END) AS completed_jobs,
               COALESCE(SUM(CASE WHEN sr.status = 'completed' THEN sr.sample_count ELSE 0 END), 0) AS completed_samples
        FROM instruments i
        LEFT JOIN sample_requests sr ON sr.instrument_id = i.id
        {'WHERE ' + ' AND '.join(request_clauses) if request_clauses else ''}
        GROUP BY i.id
        ORDER BY completed_jobs DESC, i.name
        """,
        tuple(request_params),
    )
    by_group = query_all(
        f"""
        SELECT COALESCE(NULLIF(i.faculty_group, ''), 'Ungrouped') AS group_name,
               COUNT(sr.id) AS total_requests,
               SUM(CASE WHEN sr.status = 'completed' THEN 1 ELSE 0 END) AS completed_jobs,
               COALESCE(SUM(CASE WHEN sr.status = 'completed' THEN sr.sample_count ELSE 0 END), 0) AS completed_samples
        FROM instruments i
        LEFT JOIN sample_requests sr ON sr.instrument_id = i.id
        {'WHERE ' + ' AND '.join(request_clauses) if request_clauses else ''}
        GROUP BY COALESCE(NULLIF(i.faculty_group, ''), 'Ungrouped')
        ORDER BY completed_jobs DESC, group_name
        """,
        tuple(request_params),
    )
    summary = {
        "completed_jobs": query_one(
            f"SELECT COUNT(*) AS c FROM sample_requests sr JOIN instruments i ON i.id = sr.instrument_id WHERE sr.status = 'completed'{' AND ' + ' AND '.join(metric_clauses) if metric_clauses else ''}",
            tuple(metric_params),
        )["c"],
        "completed_samples": query_one(
            f"SELECT COALESCE(SUM(sr.sample_count), 0) AS c FROM sample_requests sr JOIN instruments i ON i.id = sr.instrument_id WHERE sr.status = 'completed'{' AND ' + ' AND '.join(metric_clauses) if metric_clauses else ''}",
            tuple(metric_params),
        )["c"],
        "avg_samples_per_completed_job": query_one(
            f"SELECT ROUND(COALESCE(AVG(sr.sample_count), 0), 2) AS c FROM sample_requests sr JOIN instruments i ON i.id = sr.instrument_id WHERE sr.status = 'completed'{' AND ' + ' AND '.join(metric_clauses) if metric_clauses else ''}",
            tuple(metric_params),
        )["c"],
        "avg_jobs_per_day": round(sum(row["jobs"] for row in daily) / len(daily), 2) if daily else 0,
        "avg_samples_per_day": round(sum(row["samples"] for row in daily) / len(daily), 2) if daily else 0,
        "avg_jobs_per_week": round(sum(row["jobs"] for row in weekly) / len(weekly), 2) if weekly else 0,
        "avg_samples_per_week": round(sum(row["samples"] for row in weekly) / len(weekly), 2) if weekly else 0,
        "avg_jobs_per_month": round(sum(row["jobs"] for row in monthly) / len(monthly), 2) if monthly else 0,
        "avg_samples_per_month": round(sum(row["samples"] for row in monthly) / len(monthly), 2) if monthly else 0,
        "this_week_jobs": int(scoped_completed_total(week_start.isoformat(), today.isoformat(), "COUNT(*)")),
        "this_week_samples": int(scoped_completed_total(week_start.isoformat(), today.isoformat(), "SUM(sr.sample_count)")),
        "last_week_jobs": int(scoped_completed_total(last_week_start.isoformat(), last_week_end.isoformat(), "COUNT(*)")),
        "last_week_samples": int(scoped_completed_total(last_week_start.isoformat(), last_week_end.isoformat(), "SUM(sr.sample_count)")),
        "month_to_date_jobs": int(scoped_completed_total(month_start.isoformat(), today.isoformat(), "COUNT(*)")),
        "month_to_date_samples": int(scoped_completed_total(month_start.isoformat(), today.isoformat(), "SUM(sr.sample_count)")),
    }
    return {
        "summary": summary,
        "daily": daily,
        "weekly": weekly,
        "monthly": monthly,
        "by_instrument": by_instrument,
        "by_group": by_group,
        "report_window": report_window,
    }


def requester_pulse(user: sqlite3.Row) -> dict | None:
    """One-glance 'state of my samples' summary for requesters.

    Single SQL query with conditional aggregates; returns None for
    non-requester roles so the template skips rendering.
    """
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] != "requester":
        return None
    row = query_one(
        """
        SELECT
          SUM(CASE WHEN sr.status NOT IN ('completed','rejected') THEN 1 ELSE 0 END) AS open_count,
          SUM(CASE WHEN sr.status IN ('submitted','sample_submitted','under_review') THEN 1 ELSE 0 END) AS awaiting_operator_count,
          SUM(CASE WHEN sr.status = 'completed' AND sr.completed_at IS NOT NULL AND julianday('now') - julianday(sr.completed_at) <= 7 THEN 1 ELSE 0 END) AS ready_count,
          MIN(CASE WHEN sr.status NOT IN ('completed','rejected') THEN sr.created_at ELSE NULL END) AS oldest_pending_at
        FROM sample_requests sr
        WHERE sr.requester_id = ?
        """,
        (user["id"],),
    )
    return {
        "open_count": (row["open_count"] if row else 0) or 0,
        "awaiting_operator_count": (row["awaiting_operator_count"] if row else 0) or 0,
        "ready_count": (row["ready_count"] if row else 0) or 0,
        "oldest_pending_at": row["oldest_pending_at"] if row else None,
    }


def dashboard_analytics(user: sqlite3.Row) -> dict:
    weekly_stats = stats_payload(user, {"horizon": "weekly", "date_from": "", "date_to": ""})
    monthly_stats = stats_payload(user, {"horizon": "monthly", "date_from": "", "date_to": ""})
    clauses, params = request_scope_sql(user, "sr")
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    today = datetime.utcnow().date()
    week_start = (today - timedelta(days=today.weekday())).isoformat()
    pending_statuses = ("submitted", "under_review", "awaiting_sample_submission", "sample_submitted", "sample_received", "scheduled", "in_progress")
    pending_row = query_one(
        f"SELECT COUNT(*) AS c FROM sample_requests sr {where_sql} {'AND' if clauses else 'WHERE'} sr.status IN ({','.join('?' for _ in pending_statuses)}) AND substr(sr.created_at,1,10) >= ?",
        (*params, *pending_statuses, week_start),
    )
    avg_row = query_one(
        f"""SELECT AVG((julianday(sr.completed_at) - julianday(COALESCE(sr.sample_received_at, sr.sample_submitted_at, sr.created_at))) * 24.0) AS avg_hours
            FROM sample_requests sr {where_sql} {'AND' if clauses else 'WHERE'} sr.status = 'completed' AND sr.completed_at IS NOT NULL""",
        tuple(params),
    )
    return {
        "this_week_jobs": weekly_stats["summary"]["completed_jobs"],
        "this_week_samples": weekly_stats["summary"]["completed_samples"],
        "this_month_jobs": monthly_stats["summary"]["completed_jobs"],
        "this_month_samples": monthly_stats["summary"]["completed_samples"],
        "weekly_chart": chart_rows(weekly_stats["daily"], "bucket", "jobs", 7),
        "monthly_chart": chart_rows(monthly_stats["weekly"], "bucket", "jobs", 6),
        "pending_this_week": pending_row["c"] if pending_row else 0,
        "avg_return_hours": avg_row["avg_hours"] if avg_row else None,
    }


def report_filter_values() -> dict[str, str]:
    return {
        "horizon": request.values.get("horizon", "monthly").strip() or "monthly",
        "date_from": request.values.get("date_from", "").strip(),
        "date_to": request.values.get("date_to", "").strip(),
        "instrument_id": request.values.get("instrument_id", "").strip(),
        "group_name": request.values.get("group_name", "").strip(),
    }


def resolve_report_window(filters: dict[str, str]) -> tuple[str | None, str | None, str]:
    today = datetime.utcnow().date()
    horizon = filters.get("horizon", "monthly")
    if horizon == "weekly":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start.isoformat(), end.isoformat(), "Weekly"
    if horizon == "monthly":
        start = today.replace(day=1)
        next_month = date(start.year + 1, 1, 1) if start.month == 12 else date(start.year, start.month + 1, 1)
        end = next_month - timedelta(days=1)
        return start.isoformat(), end.isoformat(), "Monthly"
    if horizon == "yearly":
        start = date(today.year, 1, 1)
        end = date(today.year, 12, 31)
        return start.isoformat(), end.isoformat(), "Yearly"
    if horizon == "range":
        return filters.get("date_from") or None, filters.get("date_to") or None, "Custom Range"
    return None, None, "All Time"


def apply_report_window(
    clauses: list[str],
    params: list,
    field_sql: str,
    filters: dict[str, str] | None = None,
) -> tuple[list[str], list, dict[str, str | None]]:
    active_filters = filters or {"horizon": "all"}
    date_from, date_to, label = resolve_report_window(active_filters)
    scoped_clauses = list(clauses)
    scoped_params = list(params)
    if date_from:
        scoped_clauses.append(f"substr({field_sql}, 1, 10) >= ?")
        scoped_params.append(date_from)
    if date_to:
        scoped_clauses.append(f"substr({field_sql}, 1, 10) <= ?")
        scoped_params.append(date_to)
    return scoped_clauses, scoped_params, {"date_from": date_from, "date_to": date_to, "label": label}


def generate_export_workbook(
    user: sqlite3.Row,
    report_filters: dict[str, str] | None = None,
    instrument_id: int | None = None,
    group_name: str | None = None,
    filename_prefix: str = "lab_scheduler_export",
) -> Path:
    EXPORT_DIR.mkdir(exist_ok=True)
    clauses, params = scoped_stats_filters(user, instrument_id=instrument_id, group_name=group_name)
    clauses, params, report_window = apply_report_window(clauses, params, "sr.created_at", report_filters)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    # v2.0.0 — finance columns come from peer aggregates via LEFT JOINs.
    # amount_due from invoices; amount_paid as SUM(payments); finance_status
    # derived at read time; receipt_number from the most recent payment row.
    rows = query_all(
        f"""
        SELECT sr.request_no, sr.status, sr.priority, sr.sample_name, sr.sample_count, sr.sample_origin,
               (SELECT receipt_number FROM payments p
                 JOIN invoices inv ON inv.id = p.invoice_id
                 WHERE inv.request_id = sr.id
                 ORDER BY p.paid_at DESC LIMIT 1) AS receipt_number,
               COALESCE((SELECT SUM(inv.amount_due) FROM invoices inv
                           WHERE inv.request_id = sr.id), 0) AS amount_due,
               COALESCE((SELECT SUM(p.amount) FROM payments p
                           JOIN invoices inv ON inv.id = p.invoice_id
                           WHERE inv.request_id = sr.id), 0) AS amount_paid,
               CASE
                 WHEN NOT EXISTS (SELECT 1 FROM invoices WHERE request_id = sr.id) THEN 'n/a'
                 WHEN COALESCE((SELECT SUM(p.amount) FROM payments p
                                  JOIN invoices inv ON inv.id = p.invoice_id
                                  WHERE inv.request_id = sr.id), 0) = 0 THEN 'pending'
                 WHEN COALESCE((SELECT SUM(p.amount) FROM payments p
                                  JOIN invoices inv ON inv.id = p.invoice_id
                                  WHERE inv.request_id = sr.id), 0)
                      < COALESCE((SELECT SUM(inv.amount_due) FROM invoices inv
                                    WHERE inv.request_id = sr.id), 0) THEN 'partial'
                 ELSE 'paid'
               END AS finance_status,
               sr.created_at,
               sr.sample_submitted_at, sr.sample_received_at, sr.scheduled_for, sr.completed_at,
               sr.remarks, sr.results_summary,
               i.name AS instrument_name, i.faculty_group, r.name AS requester_name, r.email AS requester_email,
               c.name AS originator_name, c.role AS originator_role,
               o.name AS operator_name
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        JOIN users r ON r.id = sr.requester_id
        LEFT JOIN users c ON c.id = sr.created_by_user_id
        LEFT JOIN users o ON o.id = sr.assigned_operator_id
        {where_sql}
        ORDER BY sr.created_at DESC
        """,
        tuple(params),
    )
    stats = stats_payload_for_scope(user, report_filters, instrument_id=instrument_id, group_name=group_name)
    workbook = Workbook()
    schedule_sheet = workbook.active
    schedule_sheet.title = "Schedule Export"
    schedule_sheet.append(
        [
            "Request No",
            "Status",
            "Priority",
            "Instrument",
            "Group",
            "Requester",
            "Requester Email",
            "Originator",
            "Originator Role",
            "Operator",
            "Sample",
            "Sample Count",
            "Origin",
            "Receipt Number",
            "Amount Due",
            "Amount Paid",
            "Finance Status",
            "Created At",
            "Sample Submitted At",
            "Sample Received At",
            "Scheduled For",
            "Completed At",
            "Remarks",
            "Results Summary",
        ]
    )
    for row in rows:
        schedule_sheet.append(
            [
                row["request_no"],
                row["status"],
                row["priority"],
                row["instrument_name"],
                row["faculty_group"] or "",
                row["requester_name"],
                row["requester_email"],
                row["originator_name"] or "",
                row["originator_role"] or "",
                row["operator_name"] or "",
                row["sample_name"],
                row["sample_count"],
                row["sample_origin"],
                row["receipt_number"],
                row["amount_due"],
                row["amount_paid"],
                row["finance_status"],
                row["created_at"],
                row["sample_submitted_at"] or "",
                row["sample_received_at"] or "",
                row["scheduled_for"] or "",
                row["completed_at"] or "",
                row["remarks"],
                row["results_summary"],
            ]
        )

    summary_sheet = workbook.create_sheet("Summary Stats")
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Report Horizon", report_window["label"]])
    summary_sheet.append(["Date From", report_window["date_from"] or "-"])
    summary_sheet.append(["Date To", report_window["date_to"] or "-"])
    for key, value in stats["summary"].items():
        summary_sheet.append([key.replace("_", " ").title(), value])

    for title, dataset in [
        ("Daily Stats", stats["daily"]),
        ("Weekly Stats", stats["weekly"]),
        ("Monthly Stats", stats["monthly"]),
    ]:
        sheet = workbook.create_sheet(title)
        sheet.append(["Period", "Completed Jobs", "Completed Samples"])
        for row in dataset:
            sheet.append([row["bucket"], row["jobs"], row["samples"]])

    instrument_sheet = workbook.create_sheet("Instrument Stats")
    instrument_sheet.append(["Instrument", "Total Requests", "Completed Jobs", "Completed Samples"])
    for row in stats["by_instrument"]:
        instrument_sheet.append([row["instrument_name"], row["total_requests"], row["completed_jobs"], row["completed_samples"]])

    filename = f"{filename_prefix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = EXPORT_DIR / filename
    workbook.save(path)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    scope_label = f"{user['role'] if user['role'] != 'requester' else 'requester-self'} | {report_window['label']}"
    existing = query_one("SELECT id FROM generated_exports WHERE filename = ?", (filename,))
    if existing is None:
        execute(
            "INSERT INTO generated_exports (filename, created_by_user_id, created_at, scope_label) VALUES (?, ?, ?, ?)",
            (filename, user["id"], now_iso(), scope_label),
        )
    return path


def current_user() -> sqlite3.Row | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return query_one("SELECT * FROM users WHERE id = ?", (user_id,))


def is_owner(user: sqlite3.Row | None) -> bool:
    return bool(user and user["email"].strip().lower() in OWNER_EMAILS)


ROLE_ACCESS_PRESETS: dict[str, dict[str, object]] = {
    "requester": {
        "_is_operational_nav": False,
        "_is_lab_staff": False,
        "can_access_instruments": False,
        "can_access_schedule": False,
        "can_access_calendar": False,
        "can_access_stats": False,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": False,
        "can_view_all_instruments": False,
        "can_view_user_profiles": False,
        "can_view_finance_stage": False,
        "can_view_professor_stage": False,
        "can_access_receipts": True,
        "can_review_receipts": False,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment", "mark_submitted"},
    },
    "finance_admin": {
        "_is_operational_nav": True,
        "_is_lab_staff": False,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": True,
        "can_view_all_instruments": True,
        "can_view_user_profiles": False,
        "can_view_finance_stage": True,
        "can_view_professor_stage": False,
        "can_access_receipts": True,
        "can_review_receipts": True,
        "card_visible_fields": {"remarks", "submitted_documents", "conversation", "events", "requester_identity", "finance_data"},
        "card_action_fields": {"reply", "upload_attachment"},
    },
    "professor_approver": {
        "_is_operational_nav": False,
        "_is_lab_staff": False,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": True,
        "can_view_all_instruments": True,
        "can_view_user_profiles": True,
        "can_view_finance_stage": False,
        "can_view_professor_stage": True,
        "can_access_receipts": True,
        "can_review_receipts": False,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment"},
    },
    "faculty_in_charge": {
        "_is_operational_nav": False,
        "_is_lab_staff": False,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": False,
        "can_view_all_instruments": False,
        "can_view_user_profiles": True,
        "can_view_finance_stage": False,
        "can_view_professor_stage": False,
        "can_access_receipts": True,
        "can_review_receipts": False,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment"},
    },
    "instrument_admin": {
        "_is_operational_nav": True,
        "_is_lab_staff": True,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": False,
        "can_view_all_instruments": False,
        "can_view_user_profiles": True,
        "can_view_finance_stage": False,
        "can_view_professor_stage": False,
        "can_access_receipts": True,
        "can_review_receipts": False,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment", "finish_fast", "reassign", "mark_received", "update_status"},
    },
    "operator": {
        "_is_operational_nav": True,
        "_is_lab_staff": True,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": False,
        "can_view_all_instruments": False,
        "can_view_user_profiles": True,
        "can_view_finance_stage": False,
        "can_view_professor_stage": False,
        "can_access_receipts": True,
        "can_review_receipts": False,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment", "finish_fast", "reassign", "mark_received"},
    },
    "site_admin": {
        "_is_operational_nav": True,
        "_is_lab_staff": False,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": False,
        "can_use_role_switcher": False,
        "can_view_all_requests": True,
        "can_view_all_instruments": True,
        "can_view_user_profiles": True,
        "can_view_finance_stage": True,
        "can_view_professor_stage": True,
        "can_access_receipts": True,
        "can_review_receipts": True,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment", "finish_fast", "reassign", "mark_received", "update_status"},
    },
    "super_admin": {
        "_is_operational_nav": True,
        "_is_lab_staff": False,
        "can_access_instruments": True,
        "can_access_schedule": True,
        "can_access_calendar": True,
        "can_access_stats": True,
        "can_manage_members": True,
        "can_use_role_switcher": True,
        "can_view_all_requests": True,
        "can_view_all_instruments": True,
        "can_view_user_profiles": True,
        "can_view_finance_stage": True,
        "can_view_professor_stage": True,
        "can_access_receipts": True,
        "can_review_receipts": True,
        "card_visible_fields": {"remarks", "results_summary", "submitted_documents", "conversation", "events", "requester_identity", "operator_identity"},
        "card_action_fields": {"reply", "upload_attachment", "finish_fast", "reassign", "mark_received", "update_status"},
    },
}


def user_access_profile(user: sqlite3.Row | None) -> dict[str, object]:
    if user is None:
        return {
            "role": None,
            "is_owner": False,
            "assigned_instrument_ids": [],
            "has_instrument_scope": False,
            "can_access_instruments": False,
            "can_access_schedule": False,
            "can_access_calendar": False,
            "can_access_stats": False,
            "can_manage_members": False,
            "can_use_role_switcher": False,
            "can_view_all_requests": False,
            "can_view_all_instruments": False,
            "can_view_user_profiles": False,
            "can_view_finance_stage": False,
            "can_view_professor_stage": False,
            "can_access_receipts": False,
            "can_review_receipts": False,
            "card_visible_fields": set(),
            "card_action_fields": set(),
        }
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    preset = ROLE_ACCESS_PRESETS.get(user["role"], ROLE_ACCESS_PRESETS["requester"])
    instrument_ids = assigned_instrument_ids(user)
    is_owner_user = is_owner(user)
    card_fields = set(preset["card_visible_fields"])
    card_actions = set(preset["card_action_fields"])
    profile = {
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        "role": user["role"],
        "is_owner": is_owner_user,
        "assigned_instrument_ids": instrument_ids,
        "has_instrument_scope": bool(instrument_ids),
        "can_access_instruments": bool(preset["can_access_instruments"] or instrument_ids),
        "can_access_schedule": bool(preset["can_access_schedule"] or instrument_ids),
        "can_access_calendar": bool(preset["can_access_calendar"] or instrument_ids),
        "can_access_stats": bool(preset["can_access_stats"] or instrument_ids),
        "can_manage_members": bool(preset["can_manage_members"] or is_owner_user),
        "can_use_role_switcher": bool(preset["can_use_role_switcher"] or is_owner_user),
        "can_view_all_requests": bool(preset["can_view_all_requests"] or is_owner_user),
        "can_view_all_instruments": bool(preset["can_view_all_instruments"] or is_owner_user),
        "can_view_user_profiles": bool(preset["can_view_user_profiles"] or is_owner_user),
        "can_view_finance_stage": bool(preset["can_view_finance_stage"] or is_owner_user),
        "can_view_professor_stage": bool(preset["can_view_professor_stage"] or is_owner_user),
        "can_access_receipts": bool(preset.get("can_access_receipts", False) or is_owner_user),
        "can_review_receipts": bool(preset.get("can_review_receipts", False) or is_owner_user),
        "_is_operational_nav": bool(preset.get("_is_operational_nav", False) or is_owner_user),
        "_is_lab_staff": bool(preset.get("_is_lab_staff", False)),
        "card_visible_fields": card_fields,
        "card_action_fields": card_actions,
    }
    if is_owner_user:
        profile["can_access_instruments"] = True
        profile["can_access_schedule"] = True
        profile["can_access_calendar"] = True
        profile["can_access_stats"] = True
        profile["can_manage_members"] = True
        profile["can_use_role_switcher"] = True
        profile["can_view_all_requests"] = True
        profile["can_view_all_instruments"] = True
        profile["can_view_user_profiles"] = True
        profile["can_view_finance_stage"] = True
        profile["can_view_professor_stage"] = True
        profile["can_access_receipts"] = True
        profile["can_review_receipts"] = True
    return profile


def can_manage_members(user: sqlite3.Row | None) -> bool:
    return bool(user_access_profile(user)["can_manage_members"])


def can_use_role_switcher(user: sqlite3.Row | None) -> bool:
    return bool(user_access_profile(user)["can_use_role_switcher"])


def can_approve_step(user: sqlite3.Row, step: sqlite3.Row, instrument_id: int) -> bool:
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] in {"super_admin", "site_admin"}:
        return True
    if step["approver_role"] == "finance":
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        return user["role"] == "finance_admin"
    if step["approver_role"] == "professor":
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        return user["role"] in {"professor_approver", "super_admin", "site_admin"}
    if step["approver_role"] == "operator":
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        return can_operate_instrument(user["id"], instrument_id, user["role"])
    return False


def login_required(view):
    @wraps(view)
    def wrapped(**kwargs):
        user = current_user()
        if user is None:
            return redirect(url_for("login"))
        # W1.3.8 password hygiene: a user signed in with a temporary
        # admin-issued password can only reach the change-password
        # page and /logout until they set their own. Allow static-ish
        # endpoints so the change_password page itself renders.
        if row_value(user, "must_change_password", 0):
            allowed = {"change_password", "logout", "static", "api_health_check"}
            if request.endpoint not in allowed:
                return redirect(url_for("change_password"))
        return view(**kwargs)

    return wrapped


def role_required(*roles: str):
    def decorator(view):
        @wraps(view)
        def wrapped(**kwargs):
            user = current_user()
            if user is None:
                return redirect(url_for("login"))
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if user["role"] not in roles:
                abort(403)
            return view(**kwargs)

        return wrapped

    return decorator


def owner_required(view):
    @wraps(view)
    def wrapped(**kwargs):
        user = current_user()
        if user is None:
            return redirect(url_for("login"))
        if not is_owner(user):
            abort(403)
        return view(**kwargs)

    return wrapped


def instrument_access_required(level: str = "view"):
    """Validate `instrument_id` URL param, fetch the instrument, gate access.

    Levels:
      view    — `can_view_instrument_history` (read-only stats / history)
      open    — `can_open_instrument_detail`  (open the detail page)
      manage  — `can_manage_instrument`       (edit settings, downtime)
      operate — `can_operate_instrument`      (act on requests)

    On success the wrapped view receives an additional `instrument`
    keyword argument so it does not need to re-query.

    404 if the instrument does not exist; 403 if the user lacks the
    requested level. Must be applied AFTER `@login_required`.
    """
    def decorator(view):
        @wraps(view)
        def wrapped(instrument_id: int, **kwargs):
            user = current_user()
            instrument = query_one("SELECT * FROM instruments WHERE id = ?", (instrument_id,))
            if instrument is None:
                abort(404)
            if level == "view":
                allowed = can_view_instrument_history(user, instrument_id)
            elif level == "open":
                allowed = can_open_instrument_detail(user, instrument_id)
            elif level == "manage":
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                allowed = can_manage_instrument(user["id"], instrument_id, user["role"])
            elif level == "operate":
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                allowed = can_operate_instrument(user["id"], instrument_id, user["role"])
            else:
                raise ValueError(f"Unknown instrument access level: {level}")
            if not allowed:
                abort(403)
            kwargs["instrument"] = instrument
            return view(instrument_id=instrument_id, **kwargs)
        return wrapped
    return decorator


def can_manage_instrument(user_id: int, instrument_id: int, role: str) -> bool:
    if role in {"super_admin", "site_admin"}:
        return True
    for table in ("instrument_admins", "instrument_operators", "instrument_faculty_admins"):
        row = query_one(
            f"SELECT 1 FROM {table} WHERE user_id = ? AND instrument_id = ?",
            (user_id, instrument_id),
        )
        if row is not None:
            return True
    return False


def can_operate_instrument(user_id: int, instrument_id: int, role: str) -> bool:
    if can_manage_instrument(user_id, instrument_id, role):
        return True
    row = query_one(
        "SELECT 1 FROM instrument_operators WHERE user_id = ? AND instrument_id = ?",
        (user_id, instrument_id),
    )
    return row is not None


def assigned_instrument_ids(user: sqlite3.Row) -> list[int]:
    """Return the instrument IDs the user has any role on.

    Cached in Flask `g` per request, keyed by user id, because it
    fires 3–5 times on a single dashboard or schedule render
    (dashboard fan-out, calendar filter, scope SQL, template
    visibility checks). The DB issues 3 table scans + a UNION on
    every uncached call, so the cache saves real work.

    Cache is invalidated automatically at the end of every request
    via Flask's `g` lifecycle. Outside a request context the function
    just runs the query directly.
    """
    try:
        cache = g.setdefault("_assigned_instrument_ids", {})
    except RuntimeError:
        cache = None  # outside request context (CLI, scripts, tests)

    user_id = user["id"]
    if cache is not None and user_id in cache:
        return cache[user_id]

    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] in {"super_admin", "site_admin", "professor_approver"}:
        rows = query_all("SELECT id FROM instruments ORDER BY id")
        result = [row["id"] for row in rows]
    else:
        rows = query_all(
            """
            SELECT instrument_id FROM instrument_admins WHERE user_id = ?
            UNION
            SELECT instrument_id FROM instrument_operators WHERE user_id = ?
            UNION
            SELECT instrument_id FROM instrument_faculty_admins WHERE user_id = ?
            UNION
            SELECT instrument_id FROM instrument_requesters WHERE user_id = ?
            ORDER BY instrument_id
            """,
            (user_id, user_id, user_id, user_id),
        )
        result = [row["instrument_id"] for row in rows]

    if cache is not None:
        cache[user_id] = result
    return result


def request_assignment_candidates(sample_request: sqlite3.Row) -> list[sqlite3.Row]:
    """Return users eligible to be assigned as the operator on a request.

    Eligibility = anyone who can operate the request's instrument
    (instrument_operators ∪ instrument_admins for that instrument).
    Used by quick_assign and bulk_assign on /schedule.
    """
    instrument_id = sample_request["instrument_id"]
    return query_all(
        """
        SELECT u.id, u.name
        FROM users u
        JOIN (
            SELECT user_id FROM instrument_operators WHERE instrument_id = ?
            UNION
            SELECT user_id FROM instrument_admins   WHERE instrument_id = ?
        ) eligible ON eligible.user_id = u.id
        WHERE u.active = 1
        ORDER BY u.name COLLATE NOCASE
        """,
        (instrument_id, instrument_id),
    )


def visible_instruments_for_user(user: sqlite3.Row, active_only: bool = True) -> list[sqlite3.Row]:
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    role = user["role"]
    status_clause = "WHERE status = 'active'" if active_only else ""
    if role in {"super_admin", "site_admin", "professor_approver"}:
        return query_all(
            f"SELECT id, name, code, status FROM instruments {status_clause} ORDER BY name"
        )
    instrument_ids = assigned_instrument_ids(user)
    if instrument_ids:
        placeholders = ",".join("?" for _ in instrument_ids)
        status_and = "AND status = 'active'" if active_only else ""
        return query_all(
            f"SELECT id, name, code, status FROM instruments WHERE id IN ({placeholders}) {status_and} ORDER BY name",
            tuple(instrument_ids),
        )
    return []


def has_instrument_area_access(user: sqlite3.Row | None) -> bool:
    return bool(user_access_profile(user)["can_access_instruments"])


def sync_instrument_assignments(table: str, instrument_id: int, user_ids: list[int]) -> None:
    allowed = {"instrument_admins", "instrument_operators", "instrument_faculty_admins", "instrument_requesters"}
    if table not in allowed:
        raise ValueError("Unsupported assignment table")
    db = get_db()
    db.execute(f"DELETE FROM {table} WHERE instrument_id = ?", (instrument_id,))
    for user_id in sorted(set(user_ids)):
        db.execute(f"INSERT INTO {table} (user_id, instrument_id) VALUES (?, ?)", (user_id, instrument_id))
    db.commit()


def request_scope_sql(user: sqlite3.Row, alias: str = "sr") -> tuple[list[str], list]:
    clauses: list[str] = []
    params: list = []
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] in {"super_admin", "site_admin", "professor_approver"}:
        return clauses, params
    instrument_ids = assigned_instrument_ids(user)
    if instrument_ids:
        placeholders = ",".join("?" for _ in instrument_ids)
        clauses.append(f"{alias}.instrument_id IN ({placeholders})")
        params.extend(instrument_ids)
        return clauses, params
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "requester":
        clauses.append(f"{alias}.requester_id = ?")
        params.append(user["id"])
        return clauses, params
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "finance_admin":
        clauses.append(f"{alias}.status = 'under_review'")
        clauses.append(
            f"EXISTS (SELECT 1 FROM approval_steps aps WHERE aps.sample_request_id = {alias}.id AND aps.approver_role = 'finance')"
        )
        return clauses, params
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "professor_approver":
        clauses.append(f"{alias}.status = 'under_review'")
        clauses.append(
            f"EXISTS (SELECT 1 FROM approval_steps aps WHERE aps.sample_request_id = {alias}.id AND aps.approver_role = 'professor')"
        )
        return clauses, params
    clauses.append("1 = 0")
    return clauses, params


def scoped_instrument_count(user: sqlite3.Row) -> int:
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] in {"super_admin", "site_admin", "professor_approver"}:
        return query_one("SELECT COUNT(*) AS c FROM instruments WHERE status = 'active'")["c"]
    instrument_ids = assigned_instrument_ids(user)
    if instrument_ids:
        placeholders = ",".join("?" for _ in instrument_ids)
        row = query_one(f"SELECT COUNT(*) AS c FROM instruments WHERE status = 'active' AND id IN ({placeholders})", tuple(instrument_ids))
        return row["c"] if row else 0
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "requester":
        return query_one("SELECT COUNT(DISTINCT instrument_id) AS c FROM sample_requests WHERE requester_id = ?", (user["id"],))["c"]
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] in {"finance_admin", "professor_approver"}:
        clauses, params = request_scope_sql(user, "sr")
        row = query_one(
            f"SELECT COUNT(DISTINCT sr.instrument_id) AS c FROM sample_requests sr {'WHERE ' + ' AND '.join(clauses) if clauses else ''}",
            tuple(params),
        )
        return row["c"] if row else 0
    return 0


def can_access_stats(user: sqlite3.Row | None) -> bool:
    return bool(user_access_profile(user)["can_access_stats"])


def can_access_schedule(user: sqlite3.Row | None) -> bool:
    return bool(user_access_profile(user)["can_access_schedule"])


def can_access_calendar(user: sqlite3.Row | None) -> bool:
    return bool(user_access_profile(user)["can_access_calendar"])


def init_db() -> None:
    db = sqlite3.connect(DB_PATH)
    # Pin WAL mode at schema-creation time so every DB CATALYST ever
    # bootstraps is born in WAL. Also kick foreign_keys on for the
    # init path itself (get_db() handles runtime connections).
    db.execute("PRAGMA journal_mode = WAL")
    db.execute("PRAGMA synchronous = NORMAL")
    db.execute("PRAGMA foreign_keys = ON")
    with closing(db.cursor()) as cur:
        cur.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL,
                invited_by INTEGER,
                invite_status TEXT NOT NULL DEFAULT 'active',
                active INTEGER NOT NULL DEFAULT 1,
                member_code TEXT DEFAULT NULL
            );

            CREATE TABLE IF NOT EXISTS instruments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                code TEXT UNIQUE NOT NULL,
                category TEXT NOT NULL,
                location TEXT NOT NULL,
                daily_capacity INTEGER NOT NULL DEFAULT 3,
                status TEXT NOT NULL DEFAULT 'active',
                notes TEXT NOT NULL DEFAULT '',
                office_info TEXT NOT NULL DEFAULT '',
                faculty_group TEXT NOT NULL DEFAULT '',
                manufacturer TEXT NOT NULL DEFAULT '',
                model_number TEXT NOT NULL DEFAULT '',
                capabilities_summary TEXT NOT NULL DEFAULT '',
                machine_photo_url TEXT NOT NULL DEFAULT '',
                reference_links TEXT NOT NULL DEFAULT '',
                instrument_description TEXT NOT NULL DEFAULT '',
                accepting_requests INTEGER NOT NULL DEFAULT 1,
                soft_accept_enabled INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS instrument_admins (
                user_id INTEGER NOT NULL,
                instrument_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, instrument_id),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS instrument_operators (
                user_id INTEGER NOT NULL,
                instrument_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, instrument_id),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            );

            -- v2.2.0 — requesters assigned to specific instruments.
            -- A requester with a row here gets notifications for that
            -- instrument's noticeboard and appears in the instrument's
            -- team view. Requesters WITHOUT rows can still submit to
            -- any instrument — this is opt-in subscription, not a gate.
            -- v2.2.3 — per-user read state for notices. A row here means
            -- the user has seen the notice. Absence = unread.
            CREATE TABLE IF NOT EXISTS notice_reads (
                user_id INTEGER NOT NULL,
                notice_id INTEGER NOT NULL,
                read_at TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (user_id, notice_id),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (notice_id) REFERENCES notices(id) ON DELETE CASCADE
            );

            -- v2.3.0 — Attendance + Leave tracking
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'present',
                check_in TEXT,
                check_out TEXT,
                notes TEXT NOT NULL DEFAULT '',
                marked_by_user_id INTEGER,
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (marked_by_user_id) REFERENCES users(id) ON DELETE SET NULL,
                UNIQUE (user_id, date)
            );
            CREATE INDEX IF NOT EXISTS idx_attendance_user ON attendance(user_id, date);
            CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date);

            CREATE TABLE IF NOT EXISTS leave_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                leave_type TEXT NOT NULL DEFAULT 'casual',
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                reason TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'pending',
                approved_by_user_id INTEGER,
                approved_at TEXT,
                rejection_reason TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (approved_by_user_id) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_leave_user ON leave_requests(user_id);
            CREATE INDEX IF NOT EXISTS idx_leave_status ON leave_requests(status);
            CREATE INDEX IF NOT EXISTS idx_leave_dates ON leave_requests(start_date, end_date);

            CREATE TABLE IF NOT EXISTS reporting_structure (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                manager_id INTEGER NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (manager_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(user_id)
            );

            CREATE TABLE IF NOT EXISTS leave_balances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                leave_type TEXT NOT NULL,
                balance REAL NOT NULL DEFAULT 0,
                year INTEGER NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE (user_id, leave_type, year)
            );

            CREATE TABLE IF NOT EXISTS instrument_requesters (
                user_id INTEGER NOT NULL,
                instrument_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, instrument_id),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS instrument_faculty_admins (
                user_id INTEGER NOT NULL,
                instrument_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, instrument_id),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            );

            -- v2.2.0 — instrument maintenance + calibration log.
            -- Every maintenance event, calibration, service call is a
            -- row here. The instrument detail page shows these as an
            -- event-widget timeline. NABL auditors check this table.
            CREATE TABLE IF NOT EXISTS instrument_maintenance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument_id INTEGER NOT NULL,
                event_type TEXT NOT NULL DEFAULT 'maintenance',
                title TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                performed_by_user_id INTEGER,
                performed_at TEXT NOT NULL DEFAULT '',
                next_due_at TEXT,
                cost REAL NOT NULL DEFAULT 0,
                certificate_number TEXT NOT NULL DEFAULT '',
                attachment_filename TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE,
                FOREIGN KEY (performed_by_user_id) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_inst_maint_instrument ON instrument_maintenance(instrument_id);
            CREATE INDEX IF NOT EXISTS idx_inst_maint_type ON instrument_maintenance(event_type);
            CREATE INDEX IF NOT EXISTS idx_inst_maint_next_due ON instrument_maintenance(next_due_at);

            CREATE TABLE IF NOT EXISTS sample_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_no TEXT UNIQUE NOT NULL,
                sample_ref TEXT UNIQUE,
                requester_id INTEGER NOT NULL,
                created_by_user_id INTEGER NOT NULL,
                originator_note TEXT NOT NULL DEFAULT '',
                instrument_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                sample_name TEXT NOT NULL,
                sample_count INTEGER NOT NULL DEFAULT 1,
                description TEXT NOT NULL,
                sample_origin TEXT NOT NULL DEFAULT 'internal',
                -- v2.0.0 — legacy finance columns removed.
                -- Money lives in invoices + payments (peer aggregates).
                priority TEXT NOT NULL DEFAULT 'normal',
                status TEXT NOT NULL DEFAULT 'submitted',
                submitted_to_lab_at TEXT,
                sample_submitted_at TEXT,
                sample_received_at TEXT,
                sample_dropoff_note TEXT NOT NULL DEFAULT '',
                received_by_operator_id INTEGER,
                assigned_operator_id INTEGER,
                scheduled_for TEXT,
                remarks TEXT NOT NULL DEFAULT '',
                results_summary TEXT NOT NULL DEFAULT '',
                result_email_status TEXT NOT NULL DEFAULT '',
                result_email_sent_at TEXT,
                completion_locked INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                completed_at TEXT,
                FOREIGN KEY (requester_id) REFERENCES users(id),
                FOREIGN KEY (created_by_user_id) REFERENCES users(id),
                FOREIGN KEY (instrument_id) REFERENCES instruments(id),
                FOREIGN KEY (received_by_operator_id) REFERENCES users(id),
                FOREIGN KEY (assigned_operator_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS approval_steps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sample_request_id INTEGER NOT NULL,
                step_order INTEGER NOT NULL,
                approver_role TEXT NOT NULL,
                approver_user_id INTEGER,
                status TEXT NOT NULL DEFAULT 'pending',
                remarks TEXT NOT NULL DEFAULT '',
                acted_at TEXT,
                FOREIGN KEY (sample_request_id) REFERENCES sample_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (approver_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entity_type TEXT NOT NULL,
                entity_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                actor_id INTEGER,
                payload_json TEXT NOT NULL,
                prev_hash TEXT NOT NULL,
                entry_hash TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (actor_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS request_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                instrument_id INTEGER NOT NULL,
                original_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                relative_path TEXT NOT NULL,
                file_extension TEXT NOT NULL,
                mime_type TEXT NOT NULL,
                file_size INTEGER NOT NULL DEFAULT 0,
                uploaded_by_user_id INTEGER NOT NULL,
                uploaded_at TEXT NOT NULL,
                attachment_type TEXT NOT NULL DEFAULT 'other',
                note TEXT NOT NULL DEFAULT '',
                is_active INTEGER NOT NULL DEFAULT 1,
                request_message_id INTEGER,
                FOREIGN KEY (request_id) REFERENCES sample_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (instrument_id) REFERENCES instruments(id),
                FOREIGN KEY (uploaded_by_user_id) REFERENCES users(id),
                FOREIGN KEY (request_message_id) REFERENCES request_messages(id)
            );

            CREATE TABLE IF NOT EXISTS request_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                sender_user_id INTEGER NOT NULL,
                note_kind TEXT NOT NULL DEFAULT 'requester_note',
                message_body TEXT NOT NULL,
                created_at TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (request_id) REFERENCES sample_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (sender_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS request_issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                created_by_user_id INTEGER NOT NULL,
                issue_message TEXT NOT NULL,
                response_message TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'open',
                created_at TEXT NOT NULL,
                responded_at TEXT,
                responded_by_user_id INTEGER,
                resolved_at TEXT,
                resolved_by_user_id INTEGER,
                FOREIGN KEY (request_id) REFERENCES sample_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (created_by_user_id) REFERENCES users(id),
                FOREIGN KEY (responded_by_user_id) REFERENCES users(id),
                FOREIGN KEY (resolved_by_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS instrument_downtime (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument_id INTEGER NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                reason TEXT NOT NULL,
                created_by_user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE,
                FOREIGN KEY (created_by_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS generated_exports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT UNIQUE NOT NULL,
                created_by_user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                scope_label TEXT NOT NULL,
                FOREIGN KEY (created_by_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS instrument_approval_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument_id INTEGER NOT NULL,
                step_order INTEGER NOT NULL,
                approver_role TEXT NOT NULL,
                approver_user_id INTEGER,
                UNIQUE(instrument_id, step_order),
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE,
                FOREIGN KEY (approver_user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS announcements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                body TEXT NOT NULL,
                priority TEXT NOT NULL DEFAULT 'info',
                created_by_user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (created_by_user_id) REFERENCES users(id)
            );

            -- W1.3.7 multi-role users (additive). `users.role` remains
            -- the canonical "primary role" (display + topbar). This
            -- table layers additional roles on top without breaking
            -- any existing permission check. Every user gets at least
            -- one row here mirroring users.role via the backfill in
            -- init_db().
            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER NOT NULL,
                role TEXT NOT NULL,
                granted_at TEXT NOT NULL DEFAULT '',
                granted_by_user_id INTEGER,
                PRIMARY KEY (user_id, role),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (granted_by_user_id) REFERENCES users(id)
            );

            -- W1.3.6 instrument groups (additive). A group bundles
            -- instruments for bulk assignment in the user-admin
            -- assignment matrix. Does NOT replace `instruments.category`
            -- — that stays as the free-text taxonomy. Groups are
            -- admin-curated and act as grant shortcuts.
            CREATE TABLE IF NOT EXISTS instrument_group (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS instrument_group_member (
                group_id INTEGER NOT NULL,
                instrument_id INTEGER NOT NULL,
                PRIMARY KEY (group_id, instrument_id),
                FOREIGN KEY (group_id) REFERENCES instrument_group(id) ON DELETE CASCADE,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            );

            -- v1.6.0 — noticeboard. Site-wide, instrument-scoped, or
            -- role-scoped announcements that land on every user's home
            -- page in a sticky tile. scope = 'site' shows to everyone;
            -- scope = 'instrument' + scope_target = <instrument code>
            -- shows to users with access to that instrument; scope =
            -- 'role' + scope_target = <role> shows only to users who
            -- hold that role (per user_roles junction). expires_at is
            -- optional — null means never expires.
            CREATE TABLE IF NOT EXISTS notices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scope TEXT NOT NULL,
                scope_target TEXT,
                severity TEXT NOT NULL DEFAULT 'info',
                subject TEXT NOT NULL,
                body TEXT NOT NULL DEFAULT '',
                author_id INTEGER,
                created_at TEXT NOT NULL DEFAULT '',
                expires_at TEXT,
                FOREIGN KEY (author_id) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_notices_scope   ON notices(scope, scope_target);
            CREATE INDEX IF NOT EXISTS idx_notices_expires ON notices(expires_at);

            -- v1.6.2 — user-to-user direct messages. Separate from
            -- notices (which are broadcast / scope-filtered). A
            -- message has exactly one sender and one recipient.
            -- read_at NULL = unread; set on first view. Threading
            -- and multi-recipient are deliberately deferred to a
            -- later tag — one table, one row per message.
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sender_id    INTEGER NOT NULL,
                recipient_id INTEGER NOT NULL,
                subject      TEXT NOT NULL,
                body         TEXT NOT NULL DEFAULT '',
                sent_at      TEXT NOT NULL,
                read_at      TEXT,
                FOREIGN KEY (sender_id)    REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (recipient_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_messages_recipient ON messages(recipient_id, read_at);
            CREATE INDEX IF NOT EXISTS idx_messages_sender    ON messages(sender_id);
            CREATE INDEX IF NOT EXISTS idx_messages_sent_at   ON messages(sent_at);

            -- v1.7.0 — Grants + budgets. Every external sample_request
            -- can be attached to a grant (optional FK on the request).
            -- total_budget is in the same currency as sample_requests
            -- amount_due/amount_paid (₹). spend is derived via SUM at
            -- read time — no denormalized column that could drift.
            CREATE TABLE IF NOT EXISTS grants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                sponsor TEXT NOT NULL DEFAULT '',
                pi_user_id INTEGER,
                total_budget REAL NOT NULL DEFAULT 0,
                start_date TEXT,
                end_date TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (pi_user_id) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_grants_status ON grants(status);
            CREATE INDEX IF NOT EXISTS idx_grants_pi     ON grants(pi_user_id);

            -- v2.0.0-alpha.1 — Domain split. sample_requests stops being
            -- the single source of truth for money + project context.
            -- Four new peer aggregates land alongside it:
            --
            --   Project   → research container ("why is this work
            --               happening"). Groups many requests.
            --   Invoice   → billing document for a request. A request
            --               may have multiple invoices over time
            --               (reissue, credit note). Money owed lives
            --               here, NOT on sample_requests.
            --   Payment   → a single money event against an invoice.
            --               Enables partial payments without the
            --               amount_paid-as-column hack.
            --   GrantAlloc→ many-to-many between grants and projects.
            --               Grants fund projects, projects contain
            --               requests, requests bill invoices.
            --
            -- In alpha.1 the new tables exist but the old columns on
            -- sample_requests stay the primary read source. Dual-write
            -- helpers keep both in sync. alpha.2 flips finance reads to
            -- the new tables; alpha.3 stops writing the old columns;
            -- beta.1 drops them. Each step is a separate tag, each
            -- reversible up to beta.1.
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                pi_user_id INTEGER,
                description TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'active',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (pi_user_id) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_projects_status ON projects(status);
            CREATE INDEX IF NOT EXISTS idx_projects_pi     ON projects(pi_user_id);

            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                project_id INTEGER,
                amount_due REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                issued_at TEXT NOT NULL DEFAULT '',
                due_at TEXT,
                notes TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (request_id) REFERENCES sample_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_invoices_request ON invoices(request_id);
            CREATE INDEX IF NOT EXISTS idx_invoices_project ON invoices(project_id);
            CREATE INDEX IF NOT EXISTS idx_invoices_status  ON invoices(status);

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                method TEXT NOT NULL DEFAULT 'unspecified',
                receipt_number TEXT NOT NULL DEFAULT '',
                paid_at TEXT NOT NULL DEFAULT '',
                recorded_by_user_id INTEGER,
                notes TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE,
                FOREIGN KEY (recorded_by_user_id) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_payments_invoice ON payments(invoice_id);
            CREATE INDEX IF NOT EXISTS idx_payments_paid_at ON payments(paid_at);

            CREATE TABLE IF NOT EXISTS grant_allocations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                grant_id INTEGER NOT NULL,
                project_id INTEGER NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                allocated_at TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (grant_id) REFERENCES grants(id) ON DELETE CASCADE,
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
                UNIQUE (grant_id, project_id)
            );
            CREATE INDEX IF NOT EXISTS idx_grant_allocs_grant   ON grant_allocations(grant_id);
            CREATE INDEX IF NOT EXISTS idx_grant_allocs_project ON grant_allocations(project_id);

            CREATE TABLE IF NOT EXISTS grant_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                grant_id INTEGER NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                expense_type TEXT NOT NULL DEFAULT 'equipment',
                receipt_number TEXT NOT NULL DEFAULT '',
                recorded_by_user_id INTEGER,
                recorded_at TEXT NOT NULL,
                notes TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (grant_id) REFERENCES grants(id) ON DELETE CASCADE,
                FOREIGN KEY (recorded_by_user_id) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_grant_expenses_grant ON grant_expenses(grant_id);

            -- v2.1.0 currency support for grant expenses
            -- original_amount: amount in foreign currency (NULL = INR)
            -- original_currency: ISO 4217 code (NULL = INR)
            -- exchange_rate: multiplier to convert to INR (NULL = 1.0)
            -- expense_date: date the purchase was made
            -- "amount" column remains the INR-equivalent for all aggregation

            CREATE TABLE IF NOT EXISTS grant_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                grant_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                role TEXT NOT NULL DEFAULT 'member',
                added_by_user_id INTEGER,
                added_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (grant_id) REFERENCES grants(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE (grant_id, user_id)
            );
            CREATE INDEX IF NOT EXISTS idx_grant_members_grant ON grant_members(grant_id);
            CREATE INDEX IF NOT EXISTS idx_grant_members_user  ON grant_members(user_id);

            -- sample_requests gains a project_id FK for forward-lookup.
            -- Nullable in alpha.1 because the backfill stitches it after
            -- the fact; NOT NULL would block the tables-exist-before-
            -- backfill-runs sequence. The real contract (every request
            -- must belong to a project) is enforced at dual-write time
            -- and by tests/test_domain_split.py invariants.
            """
        )
        cur.execute("PRAGMA table_info(sample_requests)")
        _sr_cols_after_v20 = {row[1] for row in cur.fetchall()}
        if "project_id" not in _sr_cols_after_v20:
            cur.execute(
                "ALTER TABLE sample_requests ADD COLUMN project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL"
            )
        columns = {row[1] for row in cur.execute("PRAGMA table_info(sample_requests)").fetchall()}
        if "created_by_user_id" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN created_by_user_id INTEGER")
        if "sample_ref" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN sample_ref TEXT")
        if "originator_note" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN originator_note TEXT NOT NULL DEFAULT ''")
        cur.execute("UPDATE sample_requests SET created_by_user_id = requester_id WHERE created_by_user_id IS NULL")
        if "submitted_to_lab_at" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN submitted_to_lab_at TEXT")
        if "sample_submitted_at" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN sample_submitted_at TEXT")
        if "sample_received_at" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN sample_received_at TEXT")
        if "sample_dropoff_note" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN sample_dropoff_note TEXT NOT NULL DEFAULT ''")
        if "received_by_operator_id" not in columns:
            cur.execute("ALTER TABLE sample_requests ADD COLUMN received_by_operator_id INTEGER")
        # v2.0.0 — legacy grant_id ALTER removed. Dropped in
        # _drop_legacy_finance_columns() which runs after init_db.
        attachment_columns = {row[1] for row in cur.execute("PRAGMA table_info(request_attachments)").fetchall()}
        if "note" not in attachment_columns:
            cur.execute("ALTER TABLE request_attachments ADD COLUMN note TEXT NOT NULL DEFAULT ''")
        if "request_message_id" not in attachment_columns:
            cur.execute("ALTER TABLE request_attachments ADD COLUMN request_message_id INTEGER")
        request_message_columns = {row[1] for row in cur.execute("PRAGMA table_info(request_messages)").fetchall()}
        if "note_kind" not in request_message_columns:
            cur.execute("ALTER TABLE request_messages ADD COLUMN note_kind TEXT NOT NULL DEFAULT 'requester_note'")
        instrument_columns = {row[1] for row in cur.execute("PRAGMA table_info(instruments)").fetchall()}
        if "office_info" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN office_info TEXT NOT NULL DEFAULT ''")
        if "faculty_group" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN faculty_group TEXT NOT NULL DEFAULT ''")
        if "manufacturer" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN manufacturer TEXT NOT NULL DEFAULT ''")
        if "model_number" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN model_number TEXT NOT NULL DEFAULT ''")
        if "capabilities_summary" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN capabilities_summary TEXT NOT NULL DEFAULT ''")
        if "machine_photo_url" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN machine_photo_url TEXT NOT NULL DEFAULT ''")
        if "reference_links" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN reference_links TEXT NOT NULL DEFAULT ''")
        if "instrument_description" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN instrument_description TEXT NOT NULL DEFAULT ''")
        if "accepting_requests" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN accepting_requests INTEGER NOT NULL DEFAULT 1")
        if "soft_accept_enabled" not in instrument_columns:
            cur.execute("ALTER TABLE instruments ADD COLUMN soft_accept_enabled INTEGER NOT NULL DEFAULT 0")

        # Grant type (internal/DST/external) — mirrors instrument categories
        grant_cols = {col[1] for col in cur.execute("PRAGMA table_info(grants)").fetchall()}
        if "grant_type" not in grant_cols:
            try:
                cur.execute("ALTER TABLE grants ADD COLUMN grant_type TEXT NOT NULL DEFAULT 'internal'")
            except Exception:
                pass
        if "department" not in grant_cols:
            try:
                cur.execute("ALTER TABLE grants ADD COLUMN department TEXT NOT NULL DEFAULT ''")
            except Exception:
                pass
        if "portfolio_manager_id" not in grant_cols:
            try:
                cur.execute("ALTER TABLE grants ADD COLUMN portfolio_manager_id INTEGER REFERENCES users(id)")
            except Exception:
                pass
        if "granted_to" not in grant_cols:
            try:
                cur.execute("ALTER TABLE grants ADD COLUMN granted_to TEXT NOT NULL DEFAULT ''")
            except Exception:
                pass
        if "administered_by_user_id" not in grant_cols:
            try:
                cur.execute("ALTER TABLE grants ADD COLUMN administered_by_user_id INTEGER REFERENCES users(id)")
            except Exception:
                pass

        # Pricing & payment instructions per instrument (Form Control panel)
        inst_cols = {col[1] for col in cur.execute("PRAGMA table_info(instruments)").fetchall()}
        for col_name, col_sql in [
            ("price_per_sample", "ALTER TABLE instruments ADD COLUMN price_per_sample TEXT NOT NULL DEFAULT ''"),
            ("payment_instructions", "ALTER TABLE instruments ADD COLUMN payment_instructions TEXT NOT NULL DEFAULT ''"),
            ("payment_proof_note", "ALTER TABLE instruments ADD COLUMN payment_proof_note TEXT NOT NULL DEFAULT ''"),
            ("default_grant_id", "ALTER TABLE instruments ADD COLUMN default_grant_id INTEGER REFERENCES grants(id)"),
        ]:
            if col_name not in inst_cols:
                try:
                    cur.execute(col_sql)
                except Exception:
                    pass

        # Maintenance log — charge to grant
        maint_cols = {col[1] for col in cur.execute("PRAGMA table_info(instrument_maintenance)").fetchall()}
        if "grant_id" not in maint_cols:
            try:
                cur.execute("ALTER TABLE instrument_maintenance ADD COLUMN grant_id INTEGER REFERENCES grants(id)")
            except Exception:
                pass

        # Migrate: add member_code column to users if it doesn't exist
        user_columns = {col[1] for col in cur.execute("PRAGMA table_info(users)").fetchall()}
        if "member_code" not in user_columns:
            try:
                cur.execute("ALTER TABLE users ADD COLUMN member_code TEXT DEFAULT NULL")
            except:
                pass  # Column already exists or other issue, skip

        # W1.3.8 — Password hygiene. Admin-created / admin-reset
        # users must change their password on first login. Flag
        # defaults to 0 so existing seed rows keep working; only
        # new admin-issued temporary passwords set this to 1.
        if "must_change_password" not in user_columns:
            try:
                cur.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER NOT NULL DEFAULT 0")
            except:
                pass

        # Phase 6 W6.1 — Indexes for hot query paths.
        # Every query that filters by status, instrument, requester, or
        # joins approval_steps/audit_logs/request_attachments hits these.
        cur.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_sr_status              ON sample_requests(status);
            CREATE INDEX IF NOT EXISTS idx_sr_instrument_status   ON sample_requests(instrument_id, status);
            CREATE INDEX IF NOT EXISTS idx_sr_requester           ON sample_requests(requester_id);
            CREATE INDEX IF NOT EXISTS idx_sr_assigned_operator   ON sample_requests(assigned_operator_id);
            CREATE INDEX IF NOT EXISTS idx_sr_created_at          ON sample_requests(created_at);
            CREATE INDEX IF NOT EXISTS idx_sr_completed_at        ON sample_requests(completed_at);
            CREATE INDEX IF NOT EXISTS idx_sr_scheduled_for       ON sample_requests(scheduled_for);

            CREATE INDEX IF NOT EXISTS idx_steps_request          ON approval_steps(sample_request_id);
            CREATE INDEX IF NOT EXISTS idx_steps_request_order    ON approval_steps(sample_request_id, step_order);
            CREATE INDEX IF NOT EXISTS idx_steps_status           ON approval_steps(status);

            CREATE INDEX IF NOT EXISTS idx_attach_request         ON request_attachments(request_id);
            CREATE INDEX IF NOT EXISTS idx_attach_active          ON request_attachments(request_id, is_active);

            CREATE INDEX IF NOT EXISTS idx_messages_request       ON request_messages(request_id);

            CREATE INDEX IF NOT EXISTS idx_issues_request         ON request_issues(request_id);
            CREATE INDEX IF NOT EXISTS idx_issues_status          ON request_issues(status);

            CREATE INDEX IF NOT EXISTS idx_audit_entity           ON audit_logs(entity_type, entity_id);
            CREATE INDEX IF NOT EXISTS idx_audit_created          ON audit_logs(created_at);

            CREATE INDEX IF NOT EXISTS idx_inst_operators_inst    ON instrument_operators(instrument_id);
            CREATE INDEX IF NOT EXISTS idx_inst_admins_inst       ON instrument_admins(instrument_id);
            CREATE INDEX IF NOT EXISTS idx_inst_faculty_inst      ON instrument_faculty_admins(instrument_id);

            CREATE INDEX IF NOT EXISTS idx_downtime_inst          ON instrument_downtime(instrument_id);
            CREATE INDEX IF NOT EXISTS idx_downtime_active        ON instrument_downtime(is_active);

            -- W1.3.6 / W1.3.7 indexes
            CREATE INDEX IF NOT EXISTS idx_user_roles_user        ON user_roles(user_id);
            CREATE INDEX IF NOT EXISTS idx_user_roles_role        ON user_roles(role);
            CREATE INDEX IF NOT EXISTS idx_igm_group              ON instrument_group_member(group_id);
            CREATE INDEX IF NOT EXISTS idx_igm_instrument         ON instrument_group_member(instrument_id);
            """
        )

        # Phase 1: Inbox mailbox enhancement — threading, soft-delete, attachments
        for col_sql in [
            "ALTER TABLE messages ADD COLUMN parent_message_id INTEGER REFERENCES messages(id)",
            "ALTER TABLE messages ADD COLUMN deleted_by_sender INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE messages ADD COLUMN deleted_by_recipient INTEGER NOT NULL DEFAULT 0",
            # Phase 1b: permanent-hide (audit-safe delete from Deleted folder)
            "ALTER TABLE messages ADD COLUMN permanently_hidden INTEGER NOT NULL DEFAULT 0",
        ]:
            try:
                cur.execute(col_sql)
            except Exception:
                pass

        # v2.1.0: currency support for grant expenses
        for col_sql in [
            "ALTER TABLE grant_expenses ADD COLUMN original_amount REAL",
            "ALTER TABLE grant_expenses ADD COLUMN original_currency TEXT",
            "ALTER TABLE grant_expenses ADD COLUMN exchange_rate REAL",
            "ALTER TABLE grant_expenses ADD COLUMN expense_date TEXT",
        ]:
            try:
                cur.execute(col_sql)
            except Exception:
                pass

        cur.execute("""
            CREATE TABLE IF NOT EXISTS message_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                message_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                stored_name TEXT NOT NULL,
                size_bytes INTEGER NOT NULL DEFAULT 0,
                uploaded_at TEXT NOT NULL,
                FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE
            )
        """)

        # Phase 2: Instrument form control panel
        cur.execute("""
            CREATE TABLE IF NOT EXISTS instrument_custom_fields (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument_id INTEGER NOT NULL,
                field_label TEXT NOT NULL,
                field_type TEXT NOT NULL DEFAULT 'text',
                field_options TEXT NOT NULL DEFAULT '',
                is_required INTEGER NOT NULL DEFAULT 0,
                display_order INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS request_custom_field_values (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                custom_field_id INTEGER NOT NULL,
                field_value TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (request_id) REFERENCES sample_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (custom_field_id) REFERENCES instrument_custom_fields(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS instrument_inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument_id INTEGER NOT NULL,
                item_name TEXT NOT NULL,
                category TEXT NOT NULL DEFAULT 'consumable',
                quantity INTEGER NOT NULL DEFAULT 0,
                minimum_quantity INTEGER NOT NULL DEFAULT 0,
                unit TEXT NOT NULL DEFAULT 'units',
                unit_cost REAL NOT NULL DEFAULT 0,
                last_restocked_at TEXT,
                notes TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE
            )
        """)
        # ── Communications upgrade: mailing lists + email templates ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mailing_lists (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                scope TEXT NOT NULL DEFAULT 'site',
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (created_by_user_id) REFERENCES users(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mailing_list_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                list_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                FOREIGN KEY (list_id) REFERENCES mailing_lists(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id),
                UNIQUE(list_id, user_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS instrument_email_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument_id INTEGER NOT NULL,
                event_type TEXT NOT NULL,
                subject_template TEXT NOT NULL DEFAULT '',
                body_template TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (instrument_id) REFERENCES instruments(id) ON DELETE CASCADE,
                UNIQUE(instrument_id, event_type)
            )
        """)
        for col_sql in [
            "ALTER TABLE instrument_approval_config ADD COLUMN notify_submitter INTEGER NOT NULL DEFAULT 0",
        ]:
            try:
                cur.execute(col_sql)
            except Exception:
                pass

        # W1.3.7 backfill — every existing user gets a user_roles row
        # mirroring their primary users.role. Idempotent: only inserts
        # the pair (user_id, role) if it's not already present.
        cur.execute(
            """
            INSERT OR IGNORE INTO user_roles (user_id, role, granted_at)
            SELECT id, role, COALESCE(NULLIF('', ''), datetime('now'))
              FROM users
             WHERE role IS NOT NULL AND role != ''
            """
        )

        # W1.3.6 backfill — seed one instrument_group per distinct
        # instruments.category if no groups exist yet, then populate
        # the member table. This turns the free-text category taxonomy
        # into first-class groupings without touching the category
        # column (which stays as the canonical label).
        existing_groups = cur.execute(
            "SELECT COUNT(*) FROM instrument_group"
        ).fetchone()[0]
        if existing_groups == 0:
            categories = cur.execute(
                "SELECT DISTINCT category FROM instruments "
                "WHERE category IS NOT NULL AND category != '' "
                "ORDER BY category"
            ).fetchall()
            for (cat,) in categories:
                cur.execute(
                    "INSERT OR IGNORE INTO instrument_group "
                    "(name, description, created_at) VALUES (?, ?, datetime('now'))",
                    (cat, f"Auto-seeded from category '{cat}'"),
                )
                group_row = cur.execute(
                    "SELECT id FROM instrument_group WHERE name = ?", (cat,)
                ).fetchone()
                if group_row:
                    gid = group_row[0]
                    cur.execute(
                        "INSERT OR IGNORE INTO instrument_group_member "
                        "(group_id, instrument_id) "
                        "SELECT ?, id FROM instruments WHERE category = ?",
                        (gid, cat),
                    )

        # ── v1.1.0 — System notifications, budget rules, email queue, OAuth, todos ──
        cur.execute("""
            CREATE TABLE IF NOT EXISTS system_notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                category TEXT NOT NULL DEFAULT 'general',
                title TEXT NOT NULL,
                body TEXT NOT NULL DEFAULT '',
                href TEXT NOT NULL DEFAULT '',
                source_type TEXT NOT NULL DEFAULT '',
                source_id INTEGER,
                is_read INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sysnotif_user ON system_notifications(user_id, is_read)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sysnotif_created ON system_notifications(created_at)")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS budget_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                grant_id INTEGER,
                warn_utilization_pct REAL DEFAULT 80.0,
                block_utilization_pct REAL DEFAULT 100.0,
                require_receipt INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY (grant_id) REFERENCES grants(id) ON DELETE CASCADE
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS email_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                to_address TEXT NOT NULL,
                subject TEXT NOT NULL,
                body_html TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'queued',
                attempts INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT '',
                sent_at TEXT
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_todos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                assigned_by_user_id INTEGER,
                title TEXT NOT NULL,
                body TEXT NOT NULL DEFAULT '',
                priority TEXT NOT NULL DEFAULT 'normal',
                status TEXT NOT NULL DEFAULT 'open',
                due_date TEXT,
                created_at TEXT NOT NULL DEFAULT '',
                completed_at TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (assigned_by_user_id) REFERENCES users(id) ON DELETE SET NULL
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_user_todos_user ON user_todos(user_id, status)")
        # v1.3.1 — add category column to user_todos
        try:
            cur.execute("ALTER TABLE user_todos ADD COLUMN category TEXT NOT NULL DEFAULT 'general'")
        except Exception:
            pass  # column already exists

        cur.execute("""
            CREATE TABLE IF NOT EXISTS expense_receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                submitted_by_user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                amount REAL NOT NULL DEFAULT 0,
                currency TEXT NOT NULL DEFAULT 'INR',
                category TEXT NOT NULL DEFAULT 'general',
                receipt_date TEXT NOT NULL DEFAULT '',
                receipt_image_path TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'pending',
                reviewed_by_user_id INTEGER,
                reviewer_note TEXT NOT NULL DEFAULT '',
                reviewed_at TEXT,
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (submitted_by_user_id) REFERENCES users(id),
                FOREIGN KEY (reviewed_by_user_id) REFERENCES users(id)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_user ON expense_receipts(submitted_by_user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_receipts_status ON expense_receipts(status)")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS letters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                author_user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                body TEXT NOT NULL DEFAULT '',
                recipient TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'draft',
                letterhead TEXT NOT NULL DEFAULT 'default',
                created_at TEXT NOT NULL DEFAULT '',
                updated_at TEXT,
                FOREIGN KEY (author_user_id) REFERENCES users(id)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_letters_author ON letters(author_user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_letters_status ON letters(status)")

        # ── Vehicle / Fleet module ──────────────────────────────
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS vehicles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                registration_no TEXT UNIQUE NOT NULL,
                vehicle_type TEXT NOT NULL DEFAULT 'car',
                assigned_driver_user_id INTEGER,
                status TEXT NOT NULL DEFAULT 'active',
                purchase_date TEXT,
                purchase_cost REAL NOT NULL DEFAULT 0,
                insurance_expiry TEXT,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (assigned_driver_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS vehicle_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vehicle_id INTEGER NOT NULL,
                log_type TEXT NOT NULL DEFAULT 'fuel',
                amount REAL NOT NULL DEFAULT 0,
                description TEXT NOT NULL DEFAULT '',
                odometer_km REAL,
                logged_by_user_id INTEGER NOT NULL,
                log_date TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (vehicle_id) REFERENCES vehicles(id),
                FOREIGN KEY (logged_by_user_id) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_veh_logs_vehicle ON vehicle_logs(vehicle_id);
            CREATE INDEX IF NOT EXISTS idx_veh_logs_type ON vehicle_logs(log_type);
        """)

        # ── Personnel / Salary module ─────────────────────────────
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS salary_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL UNIQUE,
                monthly_salary REAL NOT NULL DEFAULT 0,
                bank_account TEXT NOT NULL DEFAULT '',
                bank_name TEXT NOT NULL DEFAULT '',
                ifsc_code TEXT NOT NULL DEFAULT '',
                pan_number TEXT NOT NULL DEFAULT '',
                aadhar_number TEXT NOT NULL DEFAULT '',
                join_date TEXT,
                designation TEXT NOT NULL DEFAULT '',
                department TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_salary_config_user ON salary_config(user_id);

            CREATE TABLE IF NOT EXISTS salary_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                month TEXT NOT NULL,
                year INTEGER NOT NULL,
                base_salary REAL NOT NULL DEFAULT 0,
                days_worked INTEGER NOT NULL DEFAULT 0,
                days_in_month INTEGER NOT NULL DEFAULT 30,
                deductions REAL NOT NULL DEFAULT 0,
                bonus REAL NOT NULL DEFAULT 0,
                net_pay REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                paid_at TEXT,
                paid_by_user_id INTEGER,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (paid_by_user_id) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_salary_payments_user ON salary_payments(user_id);
            CREATE INDEX IF NOT EXISTS idx_salary_payments_period ON salary_payments(year, month);
        """)

        # Additive column migrations for v1.1.0
        for col_sql in [
            "ALTER TABLE invoices ADD COLUMN invoice_number TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE invoices ADD COLUMN description TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE invoices ADD COLUMN created_by_user_id INTEGER REFERENCES users(id)",
            "ALTER TABLE invoices ADD COLUMN grant_id INTEGER REFERENCES grants(id)",
            "ALTER TABLE users ADD COLUMN google_id TEXT",
            "ALTER TABLE users ADD COLUMN avatar_url TEXT NOT NULL DEFAULT ''",
        ]:
            try:
                cur.execute(col_sql)
            except Exception:
                pass

        db.commit()
    db.close()
    seed_data()
    # Second pass of the W1.3.7 backfill — the first pass runs before
    # seed_data() creates demo users, so user_roles starts empty.
    # Re-run the same INSERT OR IGNORE after seeding so every freshly-
    # seeded user gets their junction row. Idempotent: no-op if
    # already populated. This is the v1.5.0 acceptance gate fix.
    _backfill_user_roles()
    # v1.6.0 — seed demo notices so the public demo shows the
    # NOTICEBOARD tile populated. Idempotent: skips if any notice row
    # already exists.
    _seed_demo_notices()
    # v1.6.2 — seed demo messages between canonical personas so the
    # inbox tour shows real traffic. Idempotent, same gate as notices.
    _seed_demo_messages()
    # v1.7.0 — seed demo grants + link some external requests to them
    # so the /finance/grants portal renders populated budgets + spend.
    _seed_demo_grants()
    _seed_demo_vehicles()
    # v2.0.0 — one-shot pre-drop migration. Reads legacy finance
    # columns if present, hands them to sync_request_to_peer_aggregates,
    # then drops the columns. Second run is a no-op.
    _backfill_domain_split()
    _drop_legacy_finance_columns()
    # v2.0.2 — sweep stale debug/trace logs older than 7 days.
    _purge_stale_logs()


def _purge_stale_logs(retention_days: int = 7) -> None:
    """v2.0.2 — delete stale debug/trace log files older than N days.

    Runs once at init_db time. Touches ONLY file-based debug traces.
    Every DB-level log is off-limits because of tamper-evidence:

      - audit_logs TABLE has prev_hash + entry_hash forming a chain.
        Deleting rows breaks the chain and destroys the tamper-evident
        property. NEVER auto-purged. If an operator genuinely needs to
        trim old audit rows, they must do it manually through a proper
        archive-table migration that preserves the hash chain — not
        through this function.

      - mailing_list_send_events TABLE is audit-first append-only for
        external communications. Compliance artifact. Never purged.

      - data/**, uploads/** are user data and attachment files linked
        from sample_requests rows. Never auto-purged.

      - logs/server.log, logs/cloudflared.log are live Flask output
        with open file descriptors. Must be rotated via logrotate or
        launchd's newsyslog, not Python truncation, because truncating
        mid-run produces a null gap between existing and new writes.

    Purged (file-based only):
      - catalyst_log.json                — debug overlay dump
      - logs/wave-*.log               — old wave-run captures
      - reports/*_log.json            — stale crawler report files
                                        (each wave run overwrites the
                                        latest, so anything older than
                                        retention_days is dead weight)

    Safety: only deletes files whose mtime is older than the cutoff.
    Current files are untouched. Idempotent: nothing to delete → no-op.
    Errors on individual files are swallowed — partial cleanup is
    better than no cleanup.
    """
    import time
    cutoff = time.time() - retention_days * 86400
    base = Path(__file__).resolve().parent

    singletons = [base / "catalyst_log.json"]
    globs = [
        base.glob("logs/wave-*.log"),
        base.glob("reports/*_log.json"),
    ]

    for p in singletons:
        try:
            if p.exists() and p.stat().st_mtime < cutoff:
                p.unlink()
        except OSError:
            pass
    for g in globs:
        for p in g:
            try:
                if p.exists() and p.stat().st_mtime < cutoff:
                    p.unlink()
            except OSError:
                pass


def _drop_legacy_finance_columns() -> None:
    """v2.0.0 — the irreversible beat. Drops amount_due, amount_paid,
    finance_status, receipt_number, grant_id from sample_requests.

    Runs AFTER _backfill_domain_split has moved the values into the
    peer aggregates. Idempotent via PRAGMA table_info — if the column
    is already gone, skip it.

    SQLite >= 3.35 supports DROP COLUMN natively. Python 3.11+ ships
    with 3.35+, so this is safe on every supported interpreter.
    """
    db = sqlite3.connect(DB_PATH)
    try:
        sr_cols = {row[1] for row in db.execute(
            "PRAGMA table_info(sample_requests)"
        ).fetchall()}
        for col in ("amount_due", "amount_paid", "finance_status",
                    "receipt_number", "grant_id"):
            if col in sr_cols:
                try:
                    db.execute(f"ALTER TABLE sample_requests DROP COLUMN {col}")
                except sqlite3.OperationalError:
                    # Column can't be dropped (index / FK dependency) —
                    # leave it. The runtime never reads it anyway.
                    pass
        db.commit()
    finally:
        db.close()


def _seed_demo_vehicles() -> None:
    """Seed 2 demo vehicles + 3 drivers. Idempotent."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        existing = db.execute("SELECT COUNT(*) AS c FROM vehicles").fetchone()["c"]
        if existing > 0:
            return
        now = now_iso()
        cur = db.cursor()
        # Vehicles
        cur.execute(
            "INSERT INTO vehicles (name, code, registration_number, make, model, vehicle_type, fuel_type, color, status, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("Innova", "VEH-001", "MH 12 MB 0350", "Toyota", "Innova Crysta", "suv", "diesel", "White", "active", now),
        )
        innova_id = cur.lastrowid
        cur.execute(
            "INSERT INTO vehicles (name, code, registration_number, make, model, vehicle_type, fuel_type, color, status, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("Camry", "VEH-002", "MH 12 3339", "Toyota", "Camry", "sedan", "petrol", "Silver", "active", now),
        )
        camry_id = cur.lastrowid
        # Drivers — look up by name first, fall back to standalone entries
        driver_names = ["Balaji Phunde", "Balu Patil", "Mangesh Ghule"]
        for i, dname in enumerate(driver_names):
            user_row = db.execute("SELECT id FROM users WHERE name = ?", (dname,)).fetchone()
            uid = user_row["id"] if user_row else None
            # Assign first two drivers to Innova, third to Camry
            vid = innova_id if i < 2 else camry_id
            is_primary = 1 if i in (0, 2) else 0
            cur.execute(
                "INSERT INTO vehicle_drivers (vehicle_id, user_id, driver_name, is_primary, assigned_at) VALUES (?, ?, ?, ?, ?)",
                (vid, uid, dname, is_primary, now),
            )
        db.commit()
    except Exception:
        pass
    finally:
        db.close()


def _backfill_domain_split() -> None:
    """v2.0.0 — one-shot pre-drop migration.

    The only path that ever reads the legacy finance columns on
    sample_requests. Runs exactly once per existing DB: reads the
    pre-v2.0 values, hands them to sync_request_to_peer_aggregates
    via explicit args, then the caller (init_db) drops the columns.

    After the DROP, this function is a no-op — the SELECT finds no
    legacy columns and bails out immediately. Fresh databases never
    create the columns in the first place, so backfill has nothing
    to read and exits clean.
    """
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        tables = {row[0] for row in db.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()}
        required = {"projects", "invoices", "payments", "grant_allocations"}
        if not required.issubset(tables):
            return

        # Check if legacy columns still exist. If they don't (fresh
        # v2.0 DB or post-drop second run), nothing to backfill.
        sr_cols = {row[1] for row in db.execute(
            "PRAGMA table_info(sample_requests)"
        ).fetchall()}
        legacy_cols = {"amount_due", "amount_paid", "finance_status",
                       "receipt_number", "grant_id"}
        if not legacy_cols.issubset(sr_cols):
            return

        rows = db.execute(
            """
            SELECT id, amount_due, amount_paid, finance_status,
                   receipt_number, grant_id
              FROM sample_requests
             WHERE amount_due > 0
                OR grant_id IS NOT NULL
                OR finance_status NOT IN ('n/a', '')
            """
        ).fetchall()
        for r in rows:
            sync_request_to_peer_aggregates(
                db,
                r["id"],
                amount_due=float(r["amount_due"] or 0),
                amount_paid=float(r["amount_paid"] or 0),
                finance_status=r["finance_status"],
                receipt_number=r["receipt_number"],
                grant_id=r["grant_id"],
            )
        db.commit()
    finally:
        db.close()


def computed_finance_for_request(
    db: sqlite3.Connection,
    request_id: int,
) -> dict:
    """v2.0.0 — derive finance state for a request from the peer
    aggregates. Returns the same shape the legacy columns used to
    expose so form-prefill call sites can swap transparently:

        {amount_due, amount_paid, finance_status, receipt_number}

    Empty/zero defaults when no invoice exists yet (e.g. internal
    request with no billing)."""
    row = db.execute(
        """
        SELECT inv.id AS invoice_id,
               COALESCE(inv.amount_due, 0) AS amount_due,
               COALESCE((SELECT SUM(p.amount) FROM payments p WHERE p.invoice_id = inv.id), 0) AS amount_paid,
               (SELECT receipt_number FROM payments
                 WHERE invoice_id = inv.id
                 ORDER BY paid_at DESC LIMIT 1) AS receipt_number
          FROM invoices inv
         WHERE inv.request_id = ?
         LIMIT 1
        """,
        (request_id,),
    ).fetchone()
    if row is None:
        return {"amount_due": 0.0, "amount_paid": 0.0,
                "finance_status": "n/a", "receipt_number": ""}
    amount_due = float(row["amount_due"] or 0)
    amount_paid = float(row["amount_paid"] or 0)
    if amount_paid <= 0:
        status = "pending" if amount_due > 0 else "n/a"
    elif amount_paid < amount_due:
        status = "partial"
    else:
        status = "paid"
    return {
        "amount_due": amount_due,
        "amount_paid": amount_paid,
        "finance_status": status,
        "receipt_number": row["receipt_number"] or "",
    }


def sync_request_to_peer_aggregates(
    db: sqlite3.Connection,
    request_id: int,
    *,
    amount_due: float | None = None,
    amount_paid: float | None = None,
    finance_status: str | None = None,
    receipt_number: str | None = None,
    grant_id: int | None = None,
) -> None:
    """v2.0.0 — per-request mitosis sync, explicit-args edition.

    The legacy finance columns on sample_requests are dropped in v2.0,
    so the helper cannot fall back to reading them. Callers MUST pass
    the authoritative values via keyword args.

    Called from:
      - new_request POST           (after INSERT)
      - resolve_sample / admin_complete_override / complete / finish_now
      - _seed_demo_grants          (to create grant_allocations rows)
      - _backfill_domain_split     (one-shot pre-drop migration)

    Invariants (per-request):
      - Exactly one Project exists with code PROJ-LEGACY-<request_no>.
      - sample_requests.project_id is stitched to that project.
      - Exactly one Invoice exists iff amount_due > 0.
      - SUM(payments.amount) == amount_paid.
      - If grant_id is set, grant_allocations row links grant → project.

    Idempotent: safe to call twice. Does NOT commit — caller owns the
    transaction boundary.
    """
    tables = {row[0] for row in db.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()}
    if not {"projects", "invoices", "payments", "grant_allocations"}.issubset(tables):
        return

    r = db.execute(
        """
        SELECT id, request_no, requester_id, title, project_id, created_at
          FROM sample_requests
         WHERE id = ?
        """,
        (request_id,),
    ).fetchone()
    if r is None:
        return

    # Caller values take precedence; missing args default to zero/empty.
    amount_due = float(amount_due if amount_due is not None else 0)
    amount_paid = float(amount_paid if amount_paid is not None else 0)
    finance_status = finance_status or "n/a"
    legacy_receipt = (receipt_number or "").strip()
    legacy_grant_id = grant_id

    req_id = r["id"]
    req_no = r["request_no"]
    now = datetime.utcnow().isoformat(timespec="seconds")
    created_at = r["created_at"] or now

    # Short-circuit: nothing to sync when there's no billing signal
    # and no pre-existing project stitch.
    if amount_due == 0 and amount_paid == 0 and legacy_grant_id is None and r["project_id"] is None:
        return

    # 1. Project — one per request, idempotent via unique code.
    project_code = f"PROJ-LEGACY-{req_no}"
    proj_row = db.execute(
        "SELECT id FROM projects WHERE code = ?", (project_code,)
    ).fetchone()
    if proj_row:
        project_id = proj_row["id"]
    else:
        cur = db.execute(
            """
            INSERT INTO projects (code, name, pi_user_id, description, status, created_at)
            VALUES (?, ?, ?, ?, 'active', ?)
            """,
            (
                project_code,
                (r["title"] or req_no)[:200],
                r["requester_id"],
                f"Auto-created by v2.0.0-alpha.3 runtime sync from {req_no}.",
                created_at,
            ),
        )
        project_id = cur.lastrowid

    # 2. Stitch sample_requests.project_id if still NULL.
    if r["project_id"] is None:
        db.execute(
            "UPDATE sample_requests SET project_id = ? WHERE id = ?",
            (project_id, req_id),
        )

    # 3. Invoice — one per request iff amount_due > 0.
    invoice_id = None
    if amount_due > 0:
        inv_row = db.execute(
            "SELECT id, amount_due FROM invoices WHERE request_id = ?", (req_id,)
        ).fetchone()
        if inv_row:
            invoice_id = inv_row["id"]
            # If the legacy amount_due was edited upward, update the
            # invoice to match. Downward edits are NOT applied — that
            # would imply a credit, which humans handle.
            if float(inv_row["amount_due"] or 0) < amount_due:
                db.execute(
                    "UPDATE invoices SET amount_due = ? WHERE id = ?",
                    (amount_due, invoice_id),
                )
        else:
            status_map = {
                "paid": "paid",
                "partial": "partial",
                "pending": "pending",
                "n/a": "pending",
                "": "pending",
            }
            inv_status = status_map.get(finance_status or "", "pending")
            cur = db.execute(
                """
                INSERT INTO invoices
                    (request_id, project_id, amount_due, status, issued_at, due_at, notes)
                VALUES (?, ?, ?, ?, ?, NULL, ?)
                """,
                (
                    req_id,
                    project_id,
                    amount_due,
                    inv_status,
                    created_at,
                    "Auto-created by v2.0.0-alpha.3 runtime sync.",
                ),
            )
            invoice_id = cur.lastrowid

        # 4. Payment — top-up delta so SUM(payments) == amount_paid.
        if amount_paid > 0 and invoice_id is not None:
            current_paid = db.execute(
                "SELECT COALESCE(SUM(amount), 0) AS s FROM payments WHERE invoice_id = ?",
                (invoice_id,),
            ).fetchone()["s"]
            delta = amount_paid - float(current_paid or 0)
            if delta > 0.005:  # tolerance: half a paisa
                receipt = legacy_receipt or f"LEGACY-{req_id}"
                db.execute(
                    """
                    INSERT INTO payments
                        (invoice_id, amount, method, receipt_number, paid_at, recorded_by_user_id, notes)
                    VALUES (?, ?, 'legacy', ?, ?, NULL, ?)
                    """,
                    (
                        invoice_id,
                        delta,
                        receipt,
                        now,
                        "Auto-created by v2.0.0-alpha.3 runtime sync (top-up delta).",
                    ),
                )

    # 5. Grant allocation — legacy grant FK → project link.
    if legacy_grant_id is not None:
        alloc_row = db.execute(
            "SELECT id FROM grant_allocations WHERE grant_id = ? AND project_id = ?",
            (legacy_grant_id, project_id),
        ).fetchone()
        if not alloc_row:
            db.execute(
                """
                INSERT INTO grant_allocations
                    (grant_id, project_id, amount, allocated_at, notes)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    legacy_grant_id,
                    project_id,
                    amount_due,
                    created_at,
                    f"Auto-created by v2.0.0-alpha.3 runtime sync from {req_no}.",
                ),
            )


# ── v2.0.0-alpha.1 dual-write helpers ────────────────────────────
# These are the forward-path for code that wants to write to the new
# aggregates. In alpha.1 they are called alongside legacy column
# writes; alpha.3 will flip the legacy calls off. They are pure
# functions over a db connection — no Flask globals, so tests can
# drive them against a tmp DB.

def create_invoice_for_request(
    db: sqlite3.Connection,
    request_id: int,
    amount_due: float,
    project_id: Optional[int] = None,
    status: str = "pending",
    notes: str = "",
) -> int:
    """Create an invoice for a request. Returns the new invoice id.

    Does not touch sample_requests.amount_due — that is the caller's
    responsibility during the dual-write window. alpha.3 will drop the
    column-side write."""
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    if project_id is None:
        row = db.execute(
            "SELECT project_id FROM sample_requests WHERE id = ?", (request_id,)
        ).fetchone()
        project_id = row["project_id"] if row and row["project_id"] is not None else None
    cur = db.execute(
        """
        INSERT INTO invoices (request_id, project_id, amount_due, status, issued_at, due_at, notes)
        VALUES (?, ?, ?, ?, ?, NULL, ?)
        """,
        (request_id, project_id, float(amount_due), status, now_iso, notes),
    )
    return cur.lastrowid


def record_payment(
    db: sqlite3.Connection,
    invoice_id: int,
    amount: float,
    receipt_number: str = "",
    method: str = "unspecified",
    recorded_by_user_id: Optional[int] = None,
    notes: str = "",
) -> int:
    """Record a payment against an invoice. Returns the new payment id.

    Multiple calls against the same invoice are valid — this is how
    partial payments work in the new model. Total paid for an invoice
    is SUM(payments.amount) WHERE invoice_id = ?."""
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    cur = db.execute(
        """
        INSERT INTO payments
            (invoice_id, amount, method, receipt_number, paid_at, recorded_by_user_id, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (invoice_id, float(amount), method, receipt_number, now_iso, recorded_by_user_id, notes),
    )
    return cur.lastrowid



def _seed_demo_grants() -> None:
    """Seed 3 demo grants and attach a few existing external requests
    to them so the finance portal's grants page shows real spend.
    DEMO_MODE only, idempotent."""
    if not DEMO_MODE:
        return
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        existing = db.execute("SELECT COUNT(*) AS c FROM grants").fetchone()["c"]
        if existing:
            return
        admin_row = db.execute("SELECT id FROM users WHERE email = 'admin@lab.local'").fetchone()
        approver_row = db.execute("SELECT id FROM users WHERE email = 'approver@lab.local'").fetchone()
        pi_admin = admin_row["id"] if admin_row else None
        pi_approver = approver_row["id"] if approver_row else None
        now_iso = datetime.utcnow().isoformat(timespec="seconds")
        seeds = [
            ("DST-2026-001", "Nano-Materials Characterization", "Department of Science & Technology",
             pi_approver, 2500000.0, "2026-01-01", "2027-03-31", "active",
             "Phase 1 — FESEM + XRD + AFM characterization of functional coatings."),
            ("CEFIPRA-2026-07", "Ceramic Phase Analysis Collaboration", "Indo-French Centre for Promotion of Advanced Research (CEFIPRA)",
             pi_admin, 1800000.0, "2026-02-15", "2028-02-14", "active",
             "Bilateral grant — XRD + Raman access for cross-institute samples."),
            ("INST-INT-24", "Internal Materials Seed Grant", "Internal",
             pi_admin, 450000.0, "2026-04-01", "2026-09-30", "active",
             "Small internal seed — supports pilot runs before external funding applications."),
        ]
        for code, name, sponsor, pi, total, start, end, status, notes in seeds:
            db.execute(
                """
                INSERT INTO grants (code, name, sponsor, pi_user_id, total_budget, start_date, end_date, status, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (code, name, sponsor, pi, total, start, end, status, notes, now_iso),
            )
        # v2.0.0 — attach external sample_requests to grants via the
        # sync helper, which creates a grant_allocations row instead
        # of writing to the now-dropped sample_requests.grant_id column.
        grant_ids = [r["id"] for r in db.execute("SELECT id FROM grants ORDER BY id").fetchall()]
        if grant_ids:
            external = db.execute(
                """
                SELECT sr.id FROM sample_requests sr
                 WHERE sr.sample_origin = 'external'
                   AND NOT EXISTS (
                     SELECT 1 FROM grant_allocations ga
                      WHERE ga.project_id = sr.project_id
                   )
                 ORDER BY sr.id
                """
            ).fetchall()
            for idx, row in enumerate(external):
                gid = grant_ids[idx % len(grant_ids)]
                sync_request_to_peer_aggregates(db, row["id"], grant_id=gid)
        db.commit()
    except sqlite3.OperationalError:
        pass
    finally:
        db.close()


def _seed_demo_messages() -> None:
    """Seed a handful of demo direct messages so the inbox tour on the
    public demo shows real traffic. DEMO_MODE only, idempotent."""
    if not DEMO_MODE:
        return
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        existing = db.execute("SELECT COUNT(*) AS c FROM messages").fetchone()["c"]
        if existing:
            return
        def uid(email: str) -> int | None:
            row = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
            return row["id"] if row else None
        admin = uid("admin@lab.local")
        operator = uid("operator@lab.local")
        requester = uid("requester@lab.local")
        approver = uid("approver@lab.local")
        now = datetime.utcnow()
        seeds = [
            # (sender, recipient, subject, body, sent_offset_minutes, read)
            (admin, operator,
             "Welcome aboard",
             "Hi — welcome to the CATALYST live demo. This is a direct message from the facility admin. You'll see sitewide and role-scoped announcements in the NOTICEBOARD tile on your home page, but this channel is for one-to-one back-and-forth. Ping me anytime.",
             120, False),
            (operator, admin,
             "Re: FESEM maintenance window",
             "Noted on the Monday 09:00-13:00 downtime. I've shifted the two high-priority samples from Monday morning to Sunday's afternoon slot. Will confirm with the requesters directly.",
             95, False),
            (approver, requester,
             "Your recent sample request — approved",
             "Hi — I've approved your latest request on the FESEM queue. It should move into the operator's schedule later today. Let me know if you need anything else on the write-up side.",
             40, False),
            (admin, requester,
             "Quick survey link",
             "When you get a moment, we're collecting feedback on the new noticeboard + quick actions UI. Two minutes, no login required — reply to this message with any thoughts.",
             15, True),
        ]
        for sender, recipient, subj, body, offset_min, is_read in seeds:
            if not sender or not recipient:
                continue
            sent_at = (now - timedelta(minutes=offset_min)).isoformat(timespec="seconds")
            read_at = (now - timedelta(minutes=offset_min - 5)).isoformat(timespec="seconds") if is_read else None
            db.execute(
                """
                INSERT INTO messages (sender_id, recipient_id, subject, body, sent_at, read_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (sender, recipient, subj, body, sent_at, read_at),
            )
        db.commit()
    except sqlite3.OperationalError:
        pass
    finally:
        db.close()


def _seed_demo_notices() -> None:
    """Seed 3 demo notices so the public demo shows an active
    NOTICEBOARD. Only runs in DEMO_MODE and only on an empty table."""
    if not DEMO_MODE:
        return
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        existing = db.execute("SELECT COUNT(*) AS c FROM notices").fetchone()["c"]
        if existing:
            return
        admin_id_row = db.execute(
            "SELECT id FROM users WHERE email = 'admin@lab.local'"
        ).fetchone()
        admin_id = admin_id_row["id"] if admin_id_row else None
        demo_notices = [
            (
                "role", "operator", "warning",
                "FESEM preventive maintenance — Monday 09:00–13:00",
                "Scheduled downtime for annual calibration. Queue items in that window will auto-shift. Operators: flag urgent samples that must run before maintenance by Friday EOD.",
                admin_id,
            ),
            (
                "site", None, "info",
                "New grant application deadline: April 30",
                "DST-SERB grant applications for the next funding cycle close April 30. PIs should submit instrument usage projections to the finance office by April 25. Contact finance@catalyst.local for the budget template.",
                admin_id,
            ),
            (
                "role", "requester", "info",
                "Sample submission guidelines updated",
                "Please ensure all external samples include a signed purchase order or payment receipt. Instruments now show payment instructions before the submission form — check the pricing section when selecting your instrument.",
                admin_id,
            ),
        ]
        now_iso = datetime.utcnow().isoformat(timespec="seconds")
        for scope, target, severity, subject, body, author in demo_notices:
            db.execute(
                """
                INSERT INTO notices
                    (scope, scope_target, severity, subject, body,
                     author_id, created_at, expires_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, NULL)
                """,
                (scope, target, severity, subject, body, author, now_iso),
            )
        db.commit()
    except sqlite3.OperationalError:
        # notices table may not exist yet on a v1.5.x DB — graceful
        pass
    finally:
        db.close()


def _backfill_user_roles() -> None:
    """Populate `user_roles` with one row per `users.role`.
    Idempotent: safe to call repeatedly."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        db.execute(
            """
            INSERT OR IGNORE INTO user_roles (user_id, role, granted_at)
            SELECT id, role, datetime('now')
              FROM users
             WHERE role IS NOT NULL AND role != ''
            """
        )
        db.commit()
    finally:
        db.close()


def seed_data() -> None:
    if not DEMO_MODE:
        # Production deployment — never seed demo accounts. The first
        # super_admin must be created manually via init_db + a one-off
        # script that inserts a real account.
        return
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row

    # Unified demo seed: one simple password ("12345") for every seeded
    # account, and role-named emails so a public tester logging in via
    # portfolio site → /login?demo=1 instantly sees which role
    # they're exercising. DEMO_MODE-only path; operational never hits
    # this because seed_data() returns early when DEMO_MODE is off.
    DEMO_PASSWORD = "12345"
    demo_pw_hash = generate_password_hash(DEMO_PASSWORD, method="pbkdf2:sha256")

    # Rebind every seeded persona's password on every boot so existing
    # demo DBs catch up without a re-seed.
    _persona_emails = (
        "owner@catalyst.local", "dean@catalyst.local", "kondhalkar@catalyst.local",
        "siteadmin@catalyst.local", "anika@catalyst.local", "ravi@catalyst.local",
        "chetan@catalyst.local", "meera@catalyst.local", "suresh@catalyst.local",
        "approver@catalyst.local",
        "satyajeetn",
        "user1@catalyst.local", "user2@catalyst.local", "user3@catalyst.local",
        "user4@catalyst.local", "user5@catalyst.local",
    )
    db.executemany(
        "UPDATE users SET password_hash = ? WHERE email = ?",
        [(demo_pw_hash, e) for e in _persona_emails],
    )
    db.commit()

    existing = db.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"]
    if existing:
        db.close()
        return

    # ── Core user roster ─────────────────────────────────────────
    # Owner + super_admin + admins + operators + finance + generic users.
    # Password for ALL demo accounts: "12345"
    core_users = [
        # Owner (god-view via OWNER_EMAILS env var)
        ("Facility Owner", "owner@catalyst.local",      "super_admin"),
        # Dean — super admin
        ("Dean Rao",             "dean@catalyst.local",          "super_admin"),
        # Kondhalkar — admin across many instruments
        ("Prof. Kondhalkar",     "kondhalkar@catalyst.local",    "instrument_admin"),
        # Site admin
        ("Site Admin",           "siteadmin@catalyst.local",         "site_admin"),
        # Operators
        ("Operator Anika",       "anika@catalyst.local",         "operator"),
        ("Operator Ravi",        "ravi@catalyst.local",          "operator"),
        ("Operator Chetan",      "chetan@catalyst.local",        "operator"),
        # Finance operators
        ("Finance Meera",        "meera@catalyst.local",         "finance_admin"),
        ("Finance Suresh",       "suresh@catalyst.local",        "finance_admin"),
        # Approver
        ("Prof. Approver",       "approver@catalyst.local",      "professor_approver"),
        # Developer — Satyajeet
        ("Satyajeet Nagargoje",  "satyajeetn",                   "site_admin"),
        # Generic user accounts (User 1–5)
        ("User One",             "user1@catalyst.local",         "requester"),
        ("User Two",             "user2@catalyst.local",         "requester"),
        ("User Three",           "user3@catalyst.local",         "requester"),
        ("User Four",            "user4@catalyst.local",         "requester"),
        ("User Five",            "user5@catalyst.local",         "requester"),
    ]
    for name, email, role in core_users:
        db.execute(
            "INSERT OR IGNORE INTO users (name, email, password_hash, role, invite_status) VALUES (?, ?, ?, ?, 'active')",
            (name, email, demo_pw_hash, role),
        )

    # ── v2.0.3 — Demo instrument inventory ──
    # 21 instruments across the imaging, spectroscopy, mechanical-test,
    # and battery-fab clusters. Numbering preserved for backwards
    # compat: INST-001 = FESEM (smoke_test fixture), INST-002 = ICP-MS,
    # INST-003 = XRD. New instruments INST-004 onward are appended.
    instruments = [
        # ── Imaging cluster ────────────────────────────────────────
        ("FESEM", "INST-001", "Microscopy", "Facility Bay A — Imaging Hall", 4,
         "Sub-nanometre imaging, FEG source, 50 eV–30 kV, mag 10x to 1,000,000x, in-chamber plasma cleaner, EDS elemental mapping.",
         "Facility Bay A · Office A-201", "Imaging & Nanoscale", "TESCAN", "S8152",
         "Field-emission SEM for nanoscale surface morphology, fracture analysis, particle sizing, and elemental mapping (EDS).",
         "", "", "Field Emission Scanning Electron Microscope — high-resolution imaging at the nanoscale.", 1, 0),
        ("Polarizing Optical Microscope", "INST-009", "Microscopy", "Facility Bay A — Imaging Hall", 6,
         "3.6 W LED Köhler illumination, 360° rotatable stage (0.1 mm vernier), 4×/10×/20×/40× objectives, polarizer 0–90°, integrated LCD capture.",
         "Facility Bay A · Office A-204", "Imaging & Nanoscale", "OPTIKA", "B-510POL",
         "Polarized-light microscopy for liquid crystals, polymers, thin-film texture, birefringence and grain-boundary analysis.",
         "", "", "Polarizing Optical Microscope for anisotropic and transparent materials.", 1, 0),

        # ── Spectroscopy / Diffraction cluster ─────────────────────
        ("ICP-MS", "INST-002", "Spectroscopy", "Facility Bay B — Analytical", 3,
         "SHIMADZU ICPMS-2040 LF, low argon (11 L/min), trace ppt detection, isotope ratio capable, semi-quant overview, liquid + solid sample intro.",
         "Facility Bay B · Office B-104", "Analytical Chemistry", "SHIMADZU", "ICPMS-2040 LF",
         "Inductively Coupled Plasma Mass Spectrometry for trace elemental quantification (ppt limits), full periodic table.",
         "", "", "ICP-MS for trace and ultra-trace multi-element analysis in liquids and digests.", 1, 1),
        ("XRD", "INST-003", "Diffraction", "Facility Bay B — Analytical", 5,
         "Empyrean DY3280 with 1Der detector, vertical goniometer, GIXRD capable, large reference-pattern library.",
         "Facility Bay B · Office B-110", "Analytical Chemistry", "MALVERN PANALYTICAL", "Empyrean-DY3280",
         "X-Ray Diffractometer for phase ID, crystal structure, thin-film GIXRD on powders and solids.",
         "", "", "X-Ray Diffractometer for crystalline phase analysis.", 1, 0),
        ("Raman Spectrometer", "INST-004", "Spectroscopy", "Facility Bay B — Analytical", 4,
         "JASCO NRS-4500 confocal micro-Raman, 532/785 nm lasers, 50–8000 cm⁻¹, ~1 cm⁻¹ resolution, 2D/3D mapping, dual spatial filter.",
         "Facility Bay B · Office B-112", "Analytical Chemistry", "JASCO", "NRS-4500",
         "Confocal micro-Raman for molecular vibrations, chemical mapping, polymorphism, carbon-material characterization.",
         "", "", "Confocal micro-Raman imaging and spectroscopy system.", 1, 0),
        ("Particle / Zeta Size Analyser", "INST-005", "Light Scattering", "Facility Bay B — Analytical", 6,
         "Zetasizer Advance, DLS + ELS + multi-angle DLS, 0.3 nm – 10 µm particle size, 3.8 nm – 100 µm zeta range, 3–12 µL sample, NIBS optics.",
         "Facility Bay B · Office B-114", "Colloids & Nanomaterials", "MALVERN PANALYTICAL", "Zetasizer Advance",
         "Particle/Zeta size analyser for colloids, nanoparticles, proteins, exosomes, liposomes — DLS + ELS workflows.",
         "", "", "Particle size + zeta potential characterization for nanoparticles and colloids.", 1, 0),
        ("UV-Visible / UV-DRS", "INST-011", "Spectroscopy", "Facility Bay B — Analytical", 6,
         "LABINDIA UV 3200 + UV 3092 DRS module. Double-beam, deuterium + tungsten-halogen, 190–800 nm, transmission + reflectance for solids/powders/liquids.",
         "Facility Bay B · Office B-116", "Analytical Chemistry", "LABINDIA", "UV 3200 + UV 3092",
         "UV-Visible spectrophotometer with Diffuse Reflectance accessory — band gap, photoinitiator quant, degradation studies.",
         "", "", "UV-Vis with DRS for liquid + solid characterization.", 1, 0),
        ("UV-VIS-NIR Spectrophotometer", "INST-012", "Spectroscopy", "Facility Bay B — Analytical", 5,
         "SHIMADZU UV-3600i Plus, 185–3300 nm, 3-detector (PMT/InGaAs/PbS), stray light <0.00005% at 220 nm, ±0.08 nm UV/Vis accuracy.",
         "Facility Bay B · Office B-118", "Analytical Chemistry", "SHIMADZU", "UV-3600i Plus",
         "Research-grade UV-VIS-NIR spectrophotometer — band gap, AR coatings, optical fibers, biological NIR analysis.",
         "", "", "UV-VIS-NIR Spectrophotometer with three-detector full-range system.", 1, 0),

        # ── Surface / Mechanical micro-test cluster ────────────────
        ("Nanoindenter", "INST-006", "Mechanical Micro-test", "Facility Bay C — Surface Lab", 4,
         "INDUSTRON NG-80, max load 10 mN, load resolution 5 nN, displacement resolution 1 nm, load + displacement controlled modes.",
         "Facility Bay C · Office C-201", "Mechanical Surfaces", "INDUSTRON", "NG-80",
         "Nanoindenter for hardness, elastic modulus, fracture toughness on metals, polymers, ceramics, coatings, thin films.",
         "", "", "Nanoindenter — probes nanomechanical properties of bulk and thin-film samples.", 1, 0),
        ("Surface Profiler", "INST-007", "Surface Metrology", "Facility Bay C — Surface Lab", 5,
         "BRUKER Dektak Pro, max load 6 mg, scan length 3000 µm, surface roughness + thin-film thickness + residual stress.",
         "Facility Bay C · Office C-205", "Surface Science", "BRUKER", "Dektak Pro",
         "Stylus surface profiler — average roughness, peak-to-valley, thin-film thickness, residual stress in films.",
         "", "", "Surface Profiler for micron and nano-scale surface integrity measurements.", 1, 0),
        ("Tribometer", "INST-008", "Tribology", "Facility Bay C — Surface Lab", 3,
         "DUCOM POD-4.0 Pin-on-Disc, room temp – 900 °C, max sample 60 mm dia × 2 mm thick, stainless steel ball counter material.",
         "Facility Bay C · Office C-208", "Tribology", "DUCOM", "POD-4.0",
         "Tribometer — friction coefficient, wear morphology, reciprocatory + rotary wear, room temp to 900 °C.",
         "", "", "Pin-on-Disc Tribometer for tribological and wear behaviour characterization.", 1, 0),

        # ── Battery fabrication cluster ────────────────────────────
        ("Battery Fabrication System", "INST-010", "Energy Storage", "Facility Bay D — Battery Lab", 3,
         "Coin-cell (CR2032) fabrication line: slurry mixer, doctor-blade coater, vacuum oven, calender, electrode puncher, Ar glove box (<0.01 ppm H₂O/O₂), multi-channel battery tester.",
         "Facility Bay D · Office D-110", "Energy Materials", "Multi-vendor", "CR2032 coin-cell line",
         "Complete coin-cell battery fabrication facility for Li-ion + Na-ion half/full cells with full electrochemical testing.",
         "", "", "End-to-end coin-cell battery fabrication and electrochemical testing.", 1, 0),

        # ── NABL ISO/IEC 17025:2017 mechanical testing line ────────
        ("Universal Testing Machine — 100 kN", "INST-013", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 3,
         "0–100 kN load range. NABL accredited per IS 1608 (Part 1), ASTM E8/E8M, ASTM D3039/D3039M.",
         "Facility Bay E · Office E-101", "Mechanical Testing (NABL)", "—", "UTM-100kN",
         "Universal testing machine — 100 kN range for tensile/compression on metals and composites.",
         "", "", "100 kN UTM — NABL accredited tensile testing.", 1, 0),
        ("Universal Testing Machine — 5 kN", "INST-017", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 4,
         "0–5 kN load range for low-load tensile/peel/film tests. NABL accredited per ASTM E345.",
         "Facility Bay E · Office E-102", "Mechanical Testing (NABL)", "—", "UTM-5kN",
         "Low-load UTM for thin films, foils, and elastomers per ASTM E345.",
         "", "", "5 kN UTM — NABL accredited low-load tensile testing.", 1, 0),
        ("Universal Testing Machine — 1000 kN", "INST-021", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 2,
         "0–1000 kN heavy-load UTM. NABL accredited per IS 1608 (Part 1).",
         "Facility Bay E · Office E-103", "Mechanical Testing (NABL)", "—", "UTM-1000kN",
         "Heavy-load UTM for structural members, reinforcement bars, large composites.",
         "", "", "1000 kN UTM — NABL accredited heavy-load tensile testing.", 1, 0),
        ("Hardness Testing — Rockwell", "INST-014", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 6,
         "Rockwell hardness scales A/B/C/D/E/F/G/H/K. NABL accredited per IS 1586 (Part 1).",
         "Facility Bay E · Office E-110", "Mechanical Testing (NABL)", "—", "Rockwell",
         "Rockwell hardness on metals and alloys per IS 1586.",
         "", "", "Rockwell hardness tester — NABL accredited.", 1, 0),
        ("Hardness Testing — Vickers / Brinell", "INST-015", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 5,
         "Vickers and Brinell macrohardness. NABL accredited per IS 1500 (Part 1).",
         "Facility Bay E · Office E-111", "Mechanical Testing (NABL)", "—", "Vickers/Brinell",
         "Vickers + Brinell hardness on metals, alloys, and case-hardened materials.",
         "", "", "Vickers / Brinell hardness tester — NABL accredited.", 1, 0),
        ("Hardness Testing — Micro-Vickers", "INST-016", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 6,
         "Micro-Vickers indentation for thin sections, coatings, weld zones. NABL accredited per IS 1501 (Part 1) and ISO 6507-1.",
         "Facility Bay E · Office E-112", "Mechanical Testing (NABL)", "—", "Micro-Vickers",
         "Micro-Vickers hardness for thin sections, coatings, and microstructure-level mapping.",
         "", "", "Micro-Vickers hardness tester — NABL accredited.", 1, 0),
        ("Microscope RV 3", "INST-018", "Metallography", "Facility Bay E — NABL Mechanical", 5,
         "Metallographic microscope per ASTM E112 grain-size standard. NABL accredited.",
         "Facility Bay E · Office E-114", "Mechanical Testing (NABL)", "—", "RV 3",
         "Metallographic microscope for grain-size determination per ASTM E112.",
         "", "", "Metallographic microscope — NABL accredited grain size analysis.", 1, 0),
        ("Axial Computerized Fatigue Testing Machine", "INST-019", "Fatigue Testing", "Facility Bay E — NABL Mechanical", 2,
         "Computerized axial fatigue rig. NABL accredited per ASTM D3479 / D3479M.",
         "Facility Bay E · Office E-116", "Mechanical Testing (NABL)", "—", "Axial Fatigue Rig",
         "Axial fatigue testing on composites and metals per ASTM D3479.",
         "", "", "Axial computerized fatigue testing — NABL accredited.", 1, 0),
        ("Compression Testing Machine", "INST-020", "Mechanical Testing", "Facility Bay E — NABL Mechanical", 4,
         "Compression testing for cement, concrete, ceramics, and rigid foams. NABL accredited per IS 516 (Part 1/Sec 1).",
         "Facility Bay E · Office E-118", "Mechanical Testing (NABL)", "—", "CTM",
         "Compression testing per IS 516 — concrete, ceramics, rigid materials.",
         "", "", "Compression testing machine — NABL accredited.", 1, 0),
    ]
    for name, code, category, location, cap, notes, office_info, faculty_group, manufacturer, model_number, capabilities_summary, machine_photo_url, reference_links, instrument_description, accepting_requests, soft_accept_enabled in instruments:
        db.execute(
            """
            INSERT OR IGNORE INTO instruments (name, code, category, location, daily_capacity, notes, office_info, faculty_group, manufacturer, model_number, capabilities_summary, machine_photo_url, reference_links, instrument_description, accepting_requests, soft_accept_enabled)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (name, code, category, location, cap, notes, office_info, faculty_group, manufacturer, model_number, capabilities_summary, machine_photo_url, reference_links, instrument_description, accepting_requests, soft_accept_enabled),
        )

    # (No additional personnel block — all users defined in core_users above)

    assignments = [
        # ── Kondhalkar is admin on most instruments ───────────────
        ("kondhalkar@catalyst.local", "INST-001", "admin"),
        ("kondhalkar@catalyst.local", "INST-002", "admin"),
        ("kondhalkar@catalyst.local", "INST-003", "admin"),
        ("kondhalkar@catalyst.local", "INST-004", "admin"),
        ("kondhalkar@catalyst.local", "INST-005", "admin"),
        ("kondhalkar@catalyst.local", "INST-006", "admin"),
        ("kondhalkar@catalyst.local", "INST-007", "admin"),
        ("kondhalkar@catalyst.local", "INST-008", "admin"),
        ("kondhalkar@catalyst.local", "INST-009", "admin"),
        ("kondhalkar@catalyst.local", "INST-010", "admin"),
        ("kondhalkar@catalyst.local", "INST-011", "admin"),
        ("kondhalkar@catalyst.local", "INST-012", "admin"),
        ("kondhalkar@catalyst.local", "INST-013", "admin"),

        # ── Approver as faculty on imaging + spectroscopy ─────────
        ("approver@catalyst.local",  "INST-001", "faculty"),
        ("approver@catalyst.local",  "INST-002", "faculty"),
        ("approver@catalyst.local",  "INST-003", "faculty"),
        ("approver@catalyst.local",  "INST-004", "faculty"),
        ("approver@catalyst.local",  "INST-005", "faculty"),

        # ── Operators spread across instruments ───────────────────
        # Anika: imaging cluster
        ("anika@catalyst.local",     "INST-001", "operator"),
        ("anika@catalyst.local",     "INST-009", "operator"),
        # Ravi: spectroscopy cluster
        ("ravi@catalyst.local",      "INST-002", "operator"),
        ("ravi@catalyst.local",      "INST-003", "operator"),
        ("ravi@catalyst.local",      "INST-004", "operator"),
        ("ravi@catalyst.local",      "INST-011", "operator"),
        # Chetan: surface + mechanical + battery
        ("chetan@catalyst.local",    "INST-005", "operator"),
        ("chetan@catalyst.local",    "INST-006", "operator"),
        ("chetan@catalyst.local",    "INST-007", "operator"),
        ("chetan@catalyst.local",    "INST-008", "operator"),
        ("chetan@catalyst.local",    "INST-010", "operator"),
        ("chetan@catalyst.local",    "INST-012", "operator"),
        ("chetan@catalyst.local",    "INST-013", "operator"),
    ]
    for email, code, kind in assignments:
        _u = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        _i = db.execute("SELECT id FROM instruments WHERE code = ?", (code,)).fetchone()
        if not _u or not _i:
            continue
        user_id = _u[0]
        inst_id = _i[0]
        if kind == "admin":
            table = "instrument_admins"
        elif kind == "operator":
            table = "instrument_operators"
        else:
            table = "instrument_faculty_admins"
        db.execute(f"INSERT OR IGNORE INTO {table} (user_id, instrument_id) VALUES (?, ?)", (user_id, inst_id))

    # No demo requests — clean slate for production demo
    demo_requests = []
    for req_no, requester_email, inst_code, title, sample_name, sample_count, sample_origin, receipt_number, amount_due, amount_paid, finance_status, status, operator_email, scheduled, sample_submitted_at, sample_received_at in demo_requests:
        requester_id = db.execute("SELECT id FROM users WHERE email = ?", (requester_email,)).fetchone()[0]
        instrument_id = db.execute("SELECT id FROM instruments WHERE code = ?", (inst_code,)).fetchone()[0]
        operator_id = None
        if operator_email:
            operator_id = db.execute("SELECT id FROM users WHERE email = ?", (operator_email,)).fetchone()[0]
        created = now_iso()
        # v2.0.0 — legacy finance columns removed. Sync helper below
        # pipes values into invoices/payments.
        cur = db.execute(
            """
            INSERT OR IGNORE INTO sample_requests
            (request_no, requester_id, created_by_user_id, instrument_id, title, sample_name, sample_count, description, sample_origin,
             priority, status, sample_submitted_at, sample_received_at,
             received_by_operator_id, assigned_operator_id, scheduled_for,
             remarks, results_summary, result_email_status, result_email_sent_at, completion_locked, created_at, updated_at, completed_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                req_no,
                requester_id,
                requester_id,
                instrument_id,
                title,
                sample_name,
                sample_count,
                f"Demo request for {title}.",
                sample_origin,
                "normal",
                status,
                sample_submitted_at,
                sample_received_at,
                operator_id if sample_received_at else None,
                operator_id,
                scheduled,
                "Initial seeded remark." if status != "under_review" else "",
                "Completed with report uploaded." if status == "completed" else "",
                "Seeded as already emailed." if status == "completed" else "",
                created if status == "completed" else None,
                1 if status == "completed" else 0,
                created,
                created,
                created if status == "completed" else None,
            ),
        )
        request_id = cur.lastrowid
        if not request_id:
            existing_request = db.execute("SELECT id FROM sample_requests WHERE request_no = ?", (req_no,)).fetchone()
            if existing_request is None:
                continue
            request_id = existing_request["id"]
        # v2.0.0 — feed finance values into peer aggregates.
        sync_request_to_peer_aggregates(
            db, request_id,
            amount_due=amount_due, amount_paid=amount_paid,
            finance_status=finance_status, receipt_number=receipt_number,
        )
        create_approval_chain(db, request_id, instrument_id)
        if status in {"awaiting_sample_submission", "sample_submitted", "sample_received", "scheduled", "completed"}:
            db.execute(
                "UPDATE approval_steps SET status = 'approved', acted_at = ?, remarks = 'Seeded approval complete' WHERE sample_request_id = ?",
                (created, request_id),
            )
        elif status == "under_review":
            db.execute(
                "UPDATE approval_steps SET status = 'pending' WHERE sample_request_id = ?",
                (request_id,),
            )

    seeded_rows = db.execute("SELECT id, requester_id, status FROM sample_requests").fetchall()
    for row in seeded_rows:
        payload_json = json.dumps({"status": row["status"]}, sort_keys=True)
        digest = hashlib.sha256(
            f"|sample_request|{row['id']}|seeded|{payload_json}".encode()
        ).hexdigest()
        db.execute(
            """
            INSERT INTO audit_logs (entity_type, entity_id, action, actor_id, payload_json, prev_hash, entry_hash, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("sample_request", row["id"], "seeded", row["requester_id"], payload_json, "", digest, now_iso()),
        )

    db.commit()
    db.close()


ROLE_DISPLAY_NAMES = {
    "super_admin": "Super Admin",
    "site_admin": "Site Admin",
    "instrument_admin": "Operational Admin",
    "faculty_in_charge": "Faculty Admin",
    "operator": "Operator",
    "professor_approver": "Approver",
    "finance_admin": "Finance Admin",
    "requester": "Lab Member",
}

ROLE_NEXT_ACTIONS = {
    "super_admin": "Full facility control. Start at Stats or Users.",
    "site_admin": "Manage users and site-wide settings. Start at Admin → Users.",
    "instrument_admin": "Run your instruments. Start at Instruments or Schedule.",
    "faculty_in_charge": "Oversee your instruments. Start at Instruments.",
    "operator": "Handle today's queue. Start at Schedule.",
    "professor_approver": "Approve pending requests. Start at Schedule → Under Review.",
    "finance_admin": "Clear finance approvals. Start at Schedule → Under Review.",
    "requester": "Submit and track your samples. Start at New Request.",
}


def role_display_name(role: str | None) -> str:
    return ROLE_DISPLAY_NAMES.get(role or "", (role or "").replace("_", " ").title())


def role_next_action(role: str | None) -> str:
    return ROLE_NEXT_ACTIONS.get(role or "", "")


# ---------------------------------------------------------------------------
# W1.3.7 — multi-role helpers. `users.role` stays the primary role (display,
# topbar, `current_role_display`). The user_roles table layers additional
# roles for permission checks. `user_role_set(user)` always includes the
# primary role, so existing single-role lookups are a subset.
# ---------------------------------------------------------------------------


def user_role_set(user: sqlite3.Row | None) -> frozenset[str]:
    """Return every role assigned to `user`, primary + additional."""
    if not user:
        return frozenset()
    roles: set[str] = set()
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    primary = user["role"] if "role" in user.keys() else None
    if primary:
        roles.add(primary)
    try:
        extra = query_all(
            "SELECT role FROM user_roles WHERE user_id = ?", (user["id"],)
        )
        for row in extra:
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if row["role"]:
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                roles.add(row["role"])
    except sqlite3.OperationalError:
        # Table may not exist on first run before init_db; degrade gracefully.
        pass
    return frozenset(roles)


def user_has_role(user: sqlite3.Row | None, role: str) -> bool:
    return role in user_role_set(user)


def grant_user_role(user_id: int, role: str, granted_by: int | None = None) -> None:
    """Idempotent role grant — safe to call repeatedly."""
    execute(
        "INSERT OR IGNORE INTO user_roles (user_id, role, granted_at, granted_by_user_id) "
        "VALUES (?, ?, datetime('now'), ?)",
        (user_id, role, granted_by),
    )


def revoke_user_role(user_id: int, role: str) -> None:
    """Remove a role from a user. Does not touch users.role — callers
    that want to change the primary role must update that column too."""
    execute(
        "DELETE FROM user_roles WHERE user_id = ? AND role = ?",
        (user_id, role),
    )


# ---------------------------------------------------------------------------
# W1.3.6 — instrument group helpers. Groups are admin-curated bundles used
# as grant shortcuts in the user-admin assignment matrix. They do not change
# any instrument-access check — `instrument_admins` / `instrument_operators`
# / `instrument_faculty_admins` stay authoritative.
# ---------------------------------------------------------------------------


def instrument_groups_all() -> list[sqlite3.Row]:
    try:
        return query_all(
            "SELECT id, name, description FROM instrument_group ORDER BY name"
        )
    except sqlite3.OperationalError:
        return []


def instrument_group_member_ids(group_id: int) -> list[int]:
    try:
        rows = query_all(
            "SELECT instrument_id FROM instrument_group_member WHERE group_id = ?",
            (group_id,),
        )
        return [int(r["instrument_id"]) for r in rows]
    except sqlite3.OperationalError:
        return []


def _recent_combined_notifications(user) -> list[dict]:
    """Return latest 5 combined admin notices + system notifications for the nav dropdown."""
    if not user:
        return []
    items: list[dict] = []
    for n in active_notices_for_user(user)[:5]:
        items.append({
            "id": n["id"],
            "kind": "notice",
            "subject": n["subject"],
            "severity": n.get("severity", "info"),
            "scope_label": n.get("scope_label", ""),
            "href": "",
            "is_read": n.get("is_read", False),
            "created_at": n.get("created_at", ""),
        })
    try:
        for s in query_all(
            "SELECT id, title, href, category, is_read, created_at FROM system_notifications WHERE user_id = ? ORDER BY created_at DESC LIMIT 5",
            (user["id"],),
        ):
            row = dict(s)
            items.append({
                "id": row["id"],
                "kind": "system",
                "subject": row.get("title", ""),
                "severity": "info",
                "scope_label": row.get("category", "system"),
                "href": row.get("href", ""),
                "is_read": bool(row.get("is_read", 0)),
                "created_at": row.get("created_at", ""),
            })
    except sqlite3.OperationalError:
        pass
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return items[:5]


@app.context_processor
def inject_globals():
    user = current_user()
    access_profile = user_access_profile(user)
    support_admin_email = sorted(OWNER_EMAILS)[0] if OWNER_EMAILS else "owner@catalyst.local"
    V = "requester finance_admin professor_approver faculty_in_charge operator instrument_admin site_admin super_admin"
    # Instruments for nav hover dropdown (only if user has instrument area access)
    nav_instruments = []
    nav_instruments_truncated = False
    if user and access_profile["can_access_instruments"]:
        _all_nav = query_all(
            "SELECT id, name, code, accepting_requests, soft_accept_enabled FROM instruments WHERE status = 'active' ORDER BY name"
        )
        nav_instruments = _all_nav[:15]
        nav_instruments_truncated = len(_all_nav) > 15
    return {
        "V": V,
        "current_user": user,
        "demo_mode": DEMO_MODE,
        "org_name": ORG_NAME,
        "org_tagline": ORG_TAGLINE,
        "module_enabled": module_enabled,
        "access_profile_user": access_profile,
        "role_display_name": role_display_name,
        "role_next_action": role_next_action,
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        "current_role_display": role_display_name(user["role"]) if user else "",
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        "current_role_hint": role_next_action(user["role"]) if user else "",
        "current_role_set": user_role_set(user),
        "user_has_role": lambda role: user_has_role(user, role),
        "instrument_groups_all": instrument_groups_all,
        "support_admin_email": support_admin_email,
        "nav_instruments": nav_instruments,
        "nav_instruments_truncated": nav_instruments_truncated,
        "nav_pending_counts": nav_pending_counts(user),
        "nav_notice_count": (unread_notice_count(user) + unread_system_notification_count(user)) if user else 0,
        "recent_notices": _recent_combined_notifications(user),
        "nav_inbox_unread": unread_message_count(user) if user else 0,
        "enabled_modules_nav": build_nav_items(user, access_profile, is_owner(user)),
        "MODULE_REGISTRY": MODULE_REGISTRY,
        "google_oauth_enabled": bool(_oauth),
        "timedelta": timedelta,
        "is_owner_user": is_owner(user),
        "can_manage_members_user": bool(access_profile["can_manage_members"]),
        "can_use_role_switcher_user": bool(access_profile["can_use_role_switcher"]),
        "can_access_schedule_user": bool(access_profile["can_access_schedule"]),
        "can_access_calendar_user": bool(access_profile["can_access_calendar"]),
        "can_access_stats_user": bool(access_profile["can_access_stats"]),
        "request_display_status": request_display_status,
        "request_status_group": request_status_group,
        "request_status_summary": request_status_summary,
        "request_lifecycle_steps": request_lifecycle_steps,
        "format_dt": format_dt,
        "time_ago": time_ago,
        "format_date": format_date,
        "format_duration_short": format_duration_short,
        "format_duration_days": format_duration_days,
        "instrument_intake_mode": instrument_intake_mode,
        "intake_mode_label": intake_mode_label,
        "has_instrument_area_access_user": bool(access_profile["can_access_instruments"]),
        "can_open_instrument_detail_id_user": lambda instrument_id: can_open_instrument_detail(user, instrument_id),
        "can_view_user_profile_id_user": lambda target_user_id: can_view_user_profile_id(user, target_user_id),
        "instrument_photo_src": instrument_photo_src,
        "request_card_policy_user": lambda request_row: request_card_policy(user, request_row),
        "request_card_can_view_field_user": lambda request_row, field_name: request_card_field_allowed(user, request_row, field_name),
    }


# ─── v1.6.1 Role-specific dashboard panels ────────────────────────────
# Two compact tiles above existing content: "Your Action Items" (things
# needing THIS user's attention right now) and "At a Glance" (summary
# stats scoped to role). Max 5 items each.


def _dashboard_action_items(user: sqlite3.Row) -> list[dict]:
    """Return up to 5 actionable items for the logged-in user's role.

    Each item: {label, detail, href, age, icon}
    """
    roles = user_role_set(user)
    db = get_db()
    items: list[dict] = []

    # ── Approvers: pending approval steps assigned to their role ──
    if "professor_approver" in roles:
        rows = query_all(
            """
            SELECT sr.id, sr.request_no, sr.sample_name, sr.created_at,
                   i.name AS instrument_name
            FROM approval_steps aps
            JOIN sample_requests sr ON sr.id = aps.sample_request_id
            JOIN instruments i ON i.id = sr.instrument_id
            WHERE aps.approver_role = 'professor' AND aps.status = 'pending'
              AND sr.status = 'under_review'
            ORDER BY sr.created_at ASC
            LIMIT 5
            """,
        )
        for r in rows:
            items.append({
                "icon": "stamp",
                "label": f"{r['request_no']} — {r['sample_name'] or 'Untitled'}",
                "detail": r["instrument_name"],
                "href": url_for("request_detail", request_id=r["id"]),
                "age": r["created_at"],
            })

    # ── Finance admin: pending finance clearances ──
    if "finance_admin" in roles:
        rows = query_all(
            """
            SELECT sr.id, sr.request_no, sr.sample_name, sr.created_at,
                   i.name AS instrument_name
            FROM approval_steps aps
            JOIN sample_requests sr ON sr.id = aps.sample_request_id
            JOIN instruments i ON i.id = sr.instrument_id
            WHERE aps.approver_role = 'finance' AND aps.status = 'pending'
              AND sr.status = 'under_review'
            ORDER BY sr.created_at ASC
            LIMIT 5
            """,
        )
        for r in rows:
            items.append({
                "label": f"{r['request_no']} — {r['sample_name'] or 'Untitled'}",
                "icon": "banknote",
                "detail": f"Finance clearance · {r['instrument_name']}",
                "href": url_for("request_detail", request_id=r["id"]),
                "age": r["created_at"],
            })

    # ── Operator: samples awaiting processing (received/scheduled/in_progress assigned to them) ──
    if "operator" in roles:
        rows = query_all(
            """
            SELECT sr.id, sr.request_no, sr.sample_name, sr.status, sr.created_at,
                   i.name AS instrument_name
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            WHERE sr.assigned_operator_id = ?
              AND sr.status IN ('sample_received', 'scheduled', 'in_progress')
            ORDER BY
              CASE sr.status
                WHEN 'in_progress' THEN 1
                WHEN 'scheduled' THEN 2
                WHEN 'sample_received' THEN 3
              END,
              sr.created_at ASC
            LIMIT 5
            """,
            (user["id"],),
        )
        for r in rows:
            items.append({
                "label": f"{r['request_no']} — {r['sample_name'] or 'Untitled'}",
                "icon": "flask",
                "detail": f"{r['status'].replace('_', ' ').title()} · {r['instrument_name']}",
                "href": url_for("request_detail", request_id=r["id"]),
                "age": r["created_at"],
            })

    # ── Instrument admin: pending sample_submitted for their instruments ──
    if "instrument_admin" in roles:
        inst_ids = assigned_instrument_ids(user)
        if inst_ids:
            ph = ",".join("?" for _ in inst_ids)
            rows = query_all(
                f"""
                SELECT sr.id, sr.request_no, sr.sample_name, sr.created_at,
                       i.name AS instrument_name
                FROM sample_requests sr
                JOIN instruments i ON i.id = sr.instrument_id
                WHERE sr.instrument_id IN ({ph})
                  AND sr.status = 'sample_submitted'
                ORDER BY sr.created_at ASC
                LIMIT 5
                """,
                tuple(inst_ids),
            )
            for r in rows:
                items.append({
                    "label": f"{r['request_no']} — {r['sample_name'] or 'Untitled'}",
                    "icon": "inbox",
                    "detail": f"Needs intake · {r['instrument_name']}",
                    "href": url_for("request_detail", request_id=r["id"]),
                    "age": r["created_at"],
                })

    # ── Super admin / owner: pending approvals site-wide ──
    if roles & {"super_admin", "site_admin"} or is_owner(user):
        if not items:  # avoid duplicates if also an approver
            rows = query_all(
                """
                SELECT sr.id, sr.request_no, sr.sample_name, sr.created_at,
                       aps.approver_role, i.name AS instrument_name
                FROM approval_steps aps
                JOIN sample_requests sr ON sr.id = aps.sample_request_id
                JOIN instruments i ON i.id = sr.instrument_id
                WHERE aps.status = 'pending' AND sr.status = 'under_review'
                ORDER BY sr.created_at ASC
                LIMIT 5
                """,
            )
            for r in rows:
                items.append({
                    "label": f"{r['request_no']} — {r['sample_name'] or 'Untitled'}",
                    "icon": "clock",
                    "detail": f"Awaiting {r['approver_role']} · {r['instrument_name']}",
                    "href": url_for("request_detail", request_id=r["id"]),
                    "age": r["created_at"],
                })

    # ── Requester: their requests needing action from them ──
    if "requester" in roles:
        rows = query_all(
            """
            SELECT sr.id, sr.request_no, sr.sample_name, sr.status, sr.created_at,
                   i.name AS instrument_name
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            WHERE sr.requester_id = ?
              AND sr.status IN ('draft', 'revision_requested')
            ORDER BY sr.created_at DESC
            LIMIT 5
            """,
            (user["id"],),
        )
        for r in rows:
            items.append({
                "label": f"{r['request_no']} — {r['sample_name'] or 'Untitled'}",
                "icon": "edit",
                "detail": f"{r['status'].replace('_', ' ').title()} · {r['instrument_name']}",
                "href": url_for("request_detail", request_id=r["id"]),
                "age": r["created_at"],
            })

    return items[:5]


def _dashboard_at_a_glance(user: sqlite3.Row) -> list[dict]:
    """Role-scoped summary stats. Each item: {value, label, href, tone}."""
    roles = user_role_set(user)
    db = get_db()
    stats: list[dict] = []

    # ── Super admin / owner: site-wide overview ──
    if roles & {"super_admin", "site_admin"} or is_owner(user):
        pending_approvals = db.execute(
            "SELECT COUNT(*) FROM approval_steps WHERE status = 'pending'"
        ).fetchone()[0]
        active_requests = db.execute(
            "SELECT COUNT(*) FROM sample_requests WHERE status NOT IN ('completed', 'rejected', 'draft')"
        ).fetchone()[0]
        stats.append({"value": active_requests, "label": "Active requests", "href": url_for("schedule"), "tone": "active"})
        stats.append({"value": pending_approvals, "label": "Pending approvals", "href": url_for("schedule") + "?pending_me=1", "tone": "wait"})

    # ── Instrument admin: queue depth for their instruments ──
    if "instrument_admin" in roles:
        inst_ids = assigned_instrument_ids(user)
        if inst_ids:
            ph = ",".join("?" for _ in inst_ids)
            queued = db.execute(
                f"SELECT COUNT(*) FROM sample_requests WHERE instrument_id IN ({ph}) AND status IN ('sample_submitted','sample_received','scheduled','in_progress')",
                tuple(inst_ids),
            ).fetchone()[0]
            submitted = db.execute(
                f"SELECT COUNT(*) FROM sample_requests WHERE instrument_id IN ({ph}) AND status = 'sample_submitted'",
                tuple(inst_ids),
            ).fetchone()[0]
            stats.append({"value": queued, "label": "In queue", "href": url_for("schedule"), "tone": "active"})
            stats.append({"value": submitted, "label": "Awaiting intake", "href": url_for("schedule") + "?bucket=sample_submitted", "tone": "wait"})

    # ── Operator: today's work ──
    if "operator" in roles:
        my_active = db.execute(
            "SELECT COUNT(*) FROM sample_requests WHERE assigned_operator_id = ? AND status IN ('sample_received','scheduled','in_progress')",
            (user["id"],),
        ).fetchone()[0]
        my_today = db.execute(
            "SELECT COUNT(*) FROM sample_requests WHERE assigned_operator_id = ? AND status IN ('sample_received','scheduled','in_progress') AND date(scheduled_for) = date('now')",
            (user["id"],),
        ).fetchone()[0]
        stats.append({"value": my_active, "label": "My active samples", "href": url_for("schedule") + "?today=1", "tone": "active"})
        stats.append({"value": my_today, "label": "Scheduled today", "href": url_for("schedule") + "?today=1", "tone": "wait"})

    # ── Professor approver: pending / recent decisions ──
    if "professor_approver" in roles:
        pending = db.execute(
            """SELECT COUNT(*) FROM approval_steps aps
               JOIN sample_requests sr ON sr.id = aps.sample_request_id
               WHERE aps.approver_role = 'professor' AND aps.status = 'pending'
                 AND sr.status = 'under_review'"""
        ).fetchone()[0]
        recent_decided = db.execute(
            """SELECT COUNT(*) FROM approval_steps aps
               WHERE aps.approver_role = 'professor' AND aps.status IN ('approved','rejected')
                 AND aps.acted_at >= date('now', '-7 days')"""
        ).fetchone()[0]
        stats.append({"value": pending, "label": "Awaiting your approval", "href": url_for("schedule") + "?pending_me=1", "tone": "wait"})
        stats.append({"value": recent_decided, "label": "Decided (7d)", "href": url_for("schedule") + "?pending_me=1", "tone": "active"})

    # ── Finance admin ──
    if "finance_admin" in roles:
        pending_fin = db.execute(
            """SELECT COUNT(*) FROM approval_steps aps
               JOIN sample_requests sr ON sr.id = aps.sample_request_id
               WHERE aps.approver_role = 'finance' AND aps.status = 'pending'
                 AND sr.status = 'under_review'"""
        ).fetchone()[0]
        stats.append({"value": pending_fin, "label": "Pending clearances", "href": url_for("finance_portal"), "tone": "wait"})

    # ── Requester: their request breakdown ──
    if "requester" in roles:
        row = db.execute(
            """SELECT
                 SUM(CASE WHEN status NOT IN ('completed','rejected') THEN 1 ELSE 0 END) AS active,
                 SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) AS done
               FROM sample_requests WHERE requester_id = ?""",
            (user["id"],),
        ).fetchone()
        stats.append({"value": (row[0] or 0), "label": "Active requests", "href": url_for("schedule") + "?mine=1", "tone": "active"})
        stats.append({"value": (row[1] or 0), "label": "Completed", "href": url_for("schedule") + "?mine=1", "tone": "completed"})

    return stats


@app.route("/hub")
def hub():
    """Public landing page — project directory + AI agent infrastructure reference."""
    from datetime import datetime as _dt
    return render_template("_hub.html", machine_time=_dt.now().strftime("%Y-%m-%d %H:%M"))


@app.route("/")
@login_required
def index():
    user = current_user()
    db = get_db()
    recent_page = page_value(int(request.args.get("recent_page", "1") or 1))
    clauses, params = request_scope_sql(user, "sr")
    where_sql = f" AND {' AND '.join(clauses)}" if clauses else ""
    queue_statuses = ("sample_submitted", "sample_received", "scheduled", "in_progress")
    counts = {
        "instruments": scoped_instrument_count(user),
        "open_requests": (
            db.execute(
                f"SELECT COUNT(*) FROM sample_requests sr WHERE status NOT IN ('completed', 'rejected'){where_sql}",
                tuple(params),
            ).fetchone()[0]
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if user["role"] == "requester" and not has_instrument_area_access(user)
            else db.execute(
                f"SELECT COUNT(*) FROM sample_requests sr WHERE status IN ({','.join('?' for _ in queue_statuses)}){where_sql}",
                tuple(queue_statuses) + tuple(params),
            ).fetchone()[0]
        ),
        "completed": db.execute(f"SELECT COUNT(*) FROM sample_requests sr WHERE status = 'completed'{where_sql}", tuple(params)).fetchone()[0],
        "samples_done": db.execute(
            f"SELECT COALESCE(SUM(sample_count), 0) FROM sample_requests sr WHERE status = 'completed'{where_sql}",
            tuple(params),
        ).fetchone()[0],
    }
    recent_per_page = 5
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "requester" and not has_instrument_area_access(user):
        recent_total = db.execute(
            "SELECT COUNT(*) FROM sample_requests sr WHERE sr.requester_id = ?",
            (user["id"],),
        ).fetchone()[0]
        recent_total_pages = max(1, math.ceil(recent_total / recent_per_page))
        recent_page = min(recent_page, recent_total_pages)
        requests_rows = query_all(
            """
            SELECT sr.*, i.name AS instrument_name, u.name AS operator_name,
                   c.name AS originator_name, c.role AS originator_role
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            LEFT JOIN users c ON c.id = sr.created_by_user_id
            LEFT JOIN users u ON u.id = sr.assigned_operator_id
            WHERE sr.requester_id = ?
            ORDER BY sr.created_at DESC
            LIMIT ? OFFSET ?
            """,
            (user["id"], recent_per_page, (recent_page - 1) * recent_per_page),
        )
    else:
        queue_where_clauses = list(clauses) + ["sr.status IN ('sample_submitted', 'sample_received', 'scheduled', 'in_progress', 'completed')"]
        queue_where_sql = "WHERE " + " AND ".join(queue_where_clauses)
        recent_total = db.execute(
            f"""
            SELECT COUNT(*)
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            JOIN users r ON r.id = sr.requester_id
            LEFT JOIN users c ON c.id = sr.created_by_user_id
            LEFT JOIN users u ON u.id = sr.assigned_operator_id
            {queue_where_sql}
            """,
            tuple(params),
        ).fetchone()[0]
        recent_total_pages = max(1, math.ceil(recent_total / recent_per_page))
        recent_page = min(recent_page, recent_total_pages)
        requests_rows = query_all(
            """
            SELECT sr.*, i.name AS instrument_name, u.name AS operator_name, r.name AS requester_name,
                   c.name AS originator_name, c.role AS originator_role
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            JOIN users r ON r.id = sr.requester_id
            LEFT JOIN users c ON c.id = sr.created_by_user_id
            LEFT JOIN users u ON u.id = sr.assigned_operator_id
            """
            + queue_where_sql
            + """
            ORDER BY
              CASE sr.status
                WHEN 'sample_submitted' THEN 1
                WHEN 'sample_received' THEN 2
                WHEN 'scheduled' THEN 3
                WHEN 'in_progress' THEN 4
                WHEN 'completed' THEN 5
                ELSE 6
              END,
              sr.created_at DESC
            LIMIT ? OFFSET ?
            """,
            tuple(params) + (recent_per_page, (recent_page - 1) * recent_per_page),
        )
    instruments = query_all("SELECT * FROM instruments ORDER BY name")
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] == "requester" and not has_instrument_area_access(user):
        instruments = []
    instrument_fifo_queue: list[dict] = []
    instrument_fifo_total: int = 0
    pending_receipt_lookup_rows: list[sqlite3.Row] = []
    quick_intake_rows: list[sqlite3.Row] = []
    quick_intake_total: int = 0
    if has_instrument_area_access(user):
        fifo_rows = query_all(
            f"""
            SELECT sr.id, sr.request_no, sr.sample_name, sr.status, sr.created_at,
                   sr.sample_submitted_at, sr.sample_received_at, sr.scheduled_for,
                   sr.assigned_operator_id,
                   i.id AS instrument_id, i.name AS instrument_name,
                   r.name AS requester_name
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            JOIN users r ON r.id = sr.requester_id
            WHERE sr.status IN ('sample_submitted', 'sample_received', 'scheduled', 'in_progress'){where_sql}
            ORDER BY i.name,
                     COALESCE(sr.sample_submitted_at, sr.sample_received_at, sr.created_at) ASC,
                     sr.id ASC
            """,
            tuple(params),
        )
        grouped_fifo: dict[tuple[int, str], list[sqlite3.Row]] = {}
        for row in fifo_rows:
            key = (row["instrument_id"], row["instrument_name"])
            grouped_fifo.setdefault(key, []).append(row)
        _all_fifo = [
            {
                "instrument_id": instrument_id,
                "instrument_name": instrument_name,
                "rows": rows[:5],
            }
            for (instrument_id, instrument_name), rows in grouped_fifo.items()
        ]
        instrument_fifo_queue = _all_fifo[:9]  # Dashboard shows max 9 cards
        instrument_fifo_total = len(_all_fifo)
        pending_receipt_lookup_rows = [
            row for row in fifo_rows if row["status"] == "sample_submitted"
        ][:5]
        # Quick Intake cards: up to 6 rows across all active statuses
        quick_intake_rows = list(fifo_rows)[:6]
        quick_intake_total = len(fifo_rows)
    dashboard_metrics = dashboard_analytics(user) if can_access_stats(user) else None
    req_pulse = requester_pulse(user)
    profile = user_access_profile(user)
    can_operate_queue = bool(
        {"reassign", "mark_received"} & set(profile["card_action_fields"])
    )
    operators = query_all(
        "SELECT id, name FROM users WHERE role IN ('operator','instrument_admin','super_admin') ORDER BY name"
    ) if has_instrument_area_access(user) and can_operate_queue else []
    # Upcoming downtime across all instruments (for dashboard)
    upcoming_downtime_all = []
    if has_instrument_area_access(user):
        upcoming_downtime_all = query_all(
            """
            SELECT idt.*, i.name AS instrument_name, i.code AS instrument_code
            FROM instrument_downtime idt
            JOIN instruments i ON i.id = idt.instrument_id
            WHERE idt.is_active = 1 AND idt.end_time >= datetime('now')
            ORDER BY idt.start_time ASC
            LIMIT 5
            """,
        )

    # ── Attendance streak (consecutive present days, up to last 7) ──
    attendance_streak = 0
    if module_enabled("attendance"):
        try:
            streak_rows = db.execute(
                "SELECT date FROM attendance WHERE user_id = ? AND status = 'present' AND date >= date('now', '-7 days') ORDER BY date DESC",
                (user["id"],),
            ).fetchall()
            if streak_rows:
                from datetime import date as _date
                today = _date.today()
                for i, r in enumerate(streak_rows):
                    expected = (today - timedelta(days=i)).isoformat()
                    if r[0] == expected:
                        attendance_streak += 1
                    else:
                        break
        except Exception:
            attendance_streak = 0

    # ── Cross-module dashboard tiles: Fleet Status + Payroll Due ──
    dash_fleet_status = None
    dash_payroll_due = None
    if is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"}):
        if module_enabled("vehicles"):
            _today_iso = date.today().isoformat()
            _30d = (date.today() + timedelta(days=30)).isoformat()
            _month_start = date.today().replace(day=1).isoformat()
            dash_fleet_status = {
                "active_vehicles": (query_one("SELECT COUNT(*) AS c FROM vehicles WHERE status = 'active'") or {"c": 0})["c"],
                "fuel_this_month": (query_one("SELECT COALESCE(SUM(amount), 0) AS total FROM vehicle_logs WHERE log_type = 'fuel' AND log_date >= ?", (_month_start,)) or {"total": 0})["total"],
                "insurance_expiring": (query_one("SELECT COUNT(*) AS c FROM vehicles WHERE insurance_expiry IS NOT NULL AND insurance_expiry <= ? AND insurance_expiry >= ?", (_30d, _today_iso)) or {"c": 0})["c"],
            }
        if module_enabled("personnel"):
            _today = date.today()
            _cur_year = _today.year
            _cur_month = f"{_today.month:02d}"
            dash_payroll_due = {
                "unpaid_count": (query_one(
                    "SELECT COUNT(*) AS c FROM salary_config sc WHERE sc.user_id NOT IN "
                    "(SELECT sp.user_id FROM salary_payments sp WHERE sp.year = ? AND sp.month = ? AND sp.status = 'paid')",
                    (_cur_year, _cur_month)) or {"c": 0})["c"],
                "pending_amount": (query_one(
                    "SELECT COALESCE(SUM(net_pay), 0) AS total FROM salary_payments WHERE year = ? AND month = ? AND status = 'pending'",
                    (_cur_year, _cur_month)) or {"total": 0})["total"],
            }

    return render_template(
        "dashboard.html",
        counts=counts,
        requests=requests_rows,
        recent_page=recent_page,
        recent_total_pages=recent_total_pages,
        instruments=instruments,
        instrument_fifo_queue=instrument_fifo_queue,
        instrument_fifo_total=instrument_fifo_total if has_instrument_area_access(user) else 0,
        pending_receipt_lookup_rows=pending_receipt_lookup_rows if can_operate_queue else [],
        quick_intake_rows=quick_intake_rows if can_operate_queue else [],
        quick_intake_total=quick_intake_total if can_operate_queue else 0,
        operators=operators,
        can_operate_queue=can_operate_queue,
        role_switches=DEMO_ROLE_SWITCHES,
        dashboard_metrics=dashboard_metrics,
        req_pulse=req_pulse,
        upcoming_downtime=upcoming_downtime_all,
        active_notices=active_notices_for_user(user),
        quick_actions=quick_actions_for_user(user),
        dash_action_items=_dashboard_action_items(user),
        dash_at_a_glance=_dashboard_at_a_glance(user),
        attendance_streak=attendance_streak,
        dash_fleet_status=dash_fleet_status,
        dash_payroll_due=dash_payroll_due,
    )


# ─── v1.6.0 Noticeboard + Quick Actions ───────────────────────────────
# New capability on the home page: site-wide / role-scoped / instrument-
# scoped notices land in a sticky NOTICEBOARD tile; per-role QUICK
# ACTIONS tile answers "why is the user here and what do they do next?"
# in two clicks. Ferrari dashboard aesthetic — terminal, monospaced,
# earned-pixels only.


def active_notices_for_user(user: sqlite3.Row | None) -> list[dict]:
    """Every active notice visible to `user`, newest-first, ordered by
    severity (critical → warning → info) then by created_at desc.
    Filters out expired notices and notices the user lacks scope for."""
    if not user:
        return []
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    try:
        rows = query_all(
            """
            SELECT id, scope, scope_target, severity, subject, body,
                   author_id, created_at, expires_at
              FROM notices
             WHERE (expires_at IS NULL OR expires_at = '' OR expires_at > ?)
             ORDER BY
                CASE severity
                    WHEN 'critical' THEN 0
                    WHEN 'warning'  THEN 1
                    ELSE 2
                END,
                created_at DESC
            """,
            (now_iso,),
        )
    except sqlite3.OperationalError:
        # Table not yet created on a pre-v1.6.0 DB. Graceful.
        return []
    role_set = user_role_set(user)
    # Instrument codes the user has any kind of access to (admin /
    # operator / faculty). Cheap single-query fetch.
    try:
        acc_rows = query_all(
            """
            SELECT DISTINCT i.code FROM instruments i
            LEFT JOIN instrument_admins    ia ON ia.instrument_id = i.id
            LEFT JOIN instrument_operators io ON io.instrument_id = i.id
            LEFT JOIN instrument_faculty_admins fa ON fa.instrument_id = i.id
            LEFT JOIN instrument_requesters ir ON ir.instrument_id = i.id
            WHERE ia.user_id = ? OR io.user_id = ? OR fa.user_id = ? OR ir.user_id = ?
            """,
            (user["id"], user["id"], user["id"], user["id"]),
        )
        user_instrument_codes = {r["code"] for r in acc_rows}
    except sqlite3.OperationalError:
        user_instrument_codes = set()
    # Read state — which notices has this user already seen?
    try:
        read_ids = {r["notice_id"] for r in query_all(
            "SELECT notice_id FROM notice_reads WHERE user_id = ?",
            (user["id"],),
        )}
    except sqlite3.OperationalError:
        read_ids = set()
    visible: list[dict] = []
    for row in rows:
        scope = row["scope"]
        target = row["scope_target"] or ""
        if scope == "site":
            keep = True
        elif scope == "role":
            keep = target in role_set
        elif scope == "instrument":
            keep = target in user_instrument_codes
        elif scope == "mailing_list":
            try:
                member_ids = {r["user_id"] for r in query_all(
                    "SELECT user_id FROM mailing_list_members WHERE list_id = ?", (int(target),)
                )}
                keep = user["id"] in member_ids
            except (sqlite3.OperationalError, ValueError):
                keep = False
        else:
            keep = False
        if keep:
            d = dict(row)
            # Friendly scope label for the tile
            if scope == "site":
                d["scope_label"] = "site-wide"
            elif scope == "role":
                d["scope_label"] = f"role · {target}"
            elif scope == "instrument":
                d["scope_label"] = f"instrument · {target}"
            elif scope == "mailing_list":
                d["scope_label"] = f"mailing list · {target}"
            else:
                d["scope_label"] = scope
            d["is_read"] = row["id"] in read_ids
            visible.append(d)
    return visible


def unread_notice_count(user: sqlite3.Row | None) -> int:
    """Count of active notices the user hasn't read yet."""
    notices = active_notices_for_user(user)
    return sum(1 for n in notices if not n.get("is_read"))


def quick_actions_for_user(user: sqlite3.Row | None) -> list[dict]:
    """Per-role "why are you here and what's your next two clicks?"
    Returns a small list of dicts with {label, href, eyebrow, accent}.
    Order matters — first item is the primary CTA. Ferrari-dashboard
    minimal, never more than 4 actions."""
    if not user:
        return []
    roles = user_role_set(user)
    actions: list[dict] = []
    # Requester — "submit a request" is the headline
    if "requester" in roles or "faculty_in_charge" in roles:
        actions.append({
            "label": "New sample request",
            "href": url_for("new_request"),
            "eyebrow": "SUBMIT",
            "accent": "primary",
        })
        actions.append({
            "label": "My requests",
            "href": url_for("schedule") + "?mine=1",
            "eyebrow": "STATUS",
            "accent": "ghost",
        })
    # Operator — "today's queue" is the headline
    if "operator" in roles:
        actions.append({
            "label": "Today's queue",
            "href": url_for("schedule") + "?today=1",
            "eyebrow": "WORK",
            "accent": "primary",
        })
        actions.append({
            "label": "My instruments",
            "href": url_for("instruments"),
            "eyebrow": "FLEET",
            "accent": "ghost",
        })
    # Approver / finance — "pending my approval" is the headline
    if "professor_approver" in roles or "finance_admin" in roles:
        actions.append({
            "label": "Pending my approval",
            "href": url_for("schedule") + "?pending_me=1",
            "eyebrow": "REVIEW",
            "accent": "primary",
        })
    # Finance-specific: direct portal card for anyone who can see it
    if roles & {"finance_admin", "super_admin", "site_admin"} or is_owner(user):
        actions.append({
            "label": "Finance portal",
            "href": url_for("finance_portal"),
            "eyebrow": "BILLING",
            "accent": "ghost",
        })
    # Instrument / site / super admin — admin surface
    if roles & {"instrument_admin", "site_admin", "super_admin"}:
        actions.append({
            "label": "Queue",
            "href": url_for("schedule"),
            "eyebrow": "OPS",
            "accent": "primary",
        })
        actions.append({
            "label": "Dev panel",
            "href": url_for("dev_panel"),
            "eyebrow": "CONTROL",
            "accent": "ghost",
        })
    # Cap at 4 actions so the tile stays tight; de-dupe by label.
    seen = set()
    uniq = []
    for a in actions:
        if a["label"] in seen:
            continue
        seen.add(a["label"])
        uniq.append(a)
    # v1.6.2 — everyone gets an "Inbox" quick action, pinned last so
    # the primary role CTA stays in slot 1. Badge hint shows unread
    # count inline when the user has any.
    try:
        unread = unread_message_count(user)
    except Exception:
        unread = 0
    uniq.append({
        "label": "Inbox" if unread == 0 else f"Inbox · {unread}",
        "href": url_for("inbox"),
        "eyebrow": "MESSAGES",
        "accent": "ghost" if unread == 0 else "primary",
    })
    # v2.2.1 — Notifications quick action
    uniq.append({
        "label": "Notifications",
        "href": url_for("notifications_page"),
        "eyebrow": "ALERTS",
        "accent": "ghost",
    })
    return uniq[:6]


# ─── v1.6.2 user-to-user messaging helpers ──────────────────────────
# Same single-source-of-truth pattern as notices: the `messages` table
# is the one write surface, and every read (inbox list, detail view,
# unread count badge, home-page preview tile) hits those same rows.


def unread_message_count(user: sqlite3.Row | None) -> int:
    """Return the number of unread messages addressed to `user`.
    Used for the Quick Actions badge + the home-page inbox preview."""
    if not user:
        return 0
    try:
        row = query_one(
            "SELECT COUNT(*) AS c FROM messages WHERE recipient_id = ? AND read_at IS NULL",
            (user["id"],),
        )
        return int(row["c"]) if row else 0
    except sqlite3.OperationalError:
        return 0


def send_completion_inbox_message(operator_id: int, sample_request: sqlite3.Row | dict) -> None:
    """Send a system inbox message to the requester when a request is completed.
    Called from every code path that marks a request as completed."""
    try:
        recipient_id = sample_request["user_id"]
        request_no = sample_request["request_no"]
        instrument_name = sample_request.get("instrument_name") or ""
        sample_name = sample_request.get("sample_name") or ""
        subject = f"Request {request_no}: Completed"
        body = (
            f"Your request {request_no} for {sample_name} on {instrument_name} "
            f"has been completed. Results are attached to the request."
        )
        execute(
            "INSERT INTO messages (sender_id, recipient_id, subject, body, sent_at, read_at) VALUES (?, ?, ?, ?, ?, NULL)",
            (operator_id, recipient_id, subject, body, now_iso()),
        )
    except Exception:
        pass  # best-effort; don't break the completion flow



# ─── v1.6.1 Admin Notices — sitewide messaging write surface ──────────
# Single source of truth is the `notices` table. Home page tile reads
# from it via `active_notices_for_user`; this admin UI writes into it.
# Owner / site_admin / super_admin only.


NOTICE_SCOPES = ("site", "role", "instrument", "mailing_list")
NOTICE_SEVERITIES = ("info", "warning", "critical")
NOTICE_ROLE_TARGETS = (
    "super_admin", "site_admin", "instrument_admin",
    "professor_approver", "finance_admin", "faculty_in_charge",
    "operator", "requester",
)


def _user_can_post_notice(user: sqlite3.Row | None) -> bool:
    """Gate notice posting to owner / site_admin / super_admin. All
    three are legit 'site messaging authority' roles."""
    if not user:
        return False
    roles = user_role_set(user)
    return bool(roles & {"super_admin", "site_admin"}) or is_owner(user)


@app.route("/admin/notices", methods=["GET"])
@login_required
def admin_notices():
    """List every notice + render the compose form. Read access
    implicit for any logged-in user so they can see what's on the
    board; write is gated via `_user_can_post_notice` in the POST
    handlers below."""
    user = current_user()
    rows = query_all(
        """
        SELECT n.id, n.scope, n.scope_target, n.severity, n.subject,
               n.body, n.created_at, n.expires_at,
               u.name AS author_name
          FROM notices n
          LEFT JOIN users u ON u.id = n.author_id
         ORDER BY
            CASE n.severity
                WHEN 'critical' THEN 0
                WHEN 'warning'  THEN 1
                ELSE 2
            END,
            n.created_at DESC
        """
    )
    # Instrument codes for the scope picker dropdown
    instrument_rows = query_all(
        "SELECT code, name FROM instruments WHERE status = 'active' ORDER BY code"
    )
    can_post = _user_can_post_notice(user)
    mailing_lists = query_all("SELECT id, name FROM mailing_lists ORDER BY name")
    return render_template(
        "admin_notices.html",
        notices=rows,
        instruments=instrument_rows,
        scopes=NOTICE_SCOPES,
        severities=NOTICE_SEVERITIES,
        role_targets=NOTICE_ROLE_TARGETS,
        can_post=can_post,
        mailing_lists=mailing_lists,
    )


@app.route("/admin/notices/new", methods=["POST"])
@login_required
def admin_notices_new():
    """Create a new notice. Validates scope + severity + required
    subject, persists with the authenticated user as author."""
    user = current_user()
    if not _user_can_post_notice(user):
        abort(403)
    scope = (request.form.get("scope") or "site").strip()
    scope_target = (request.form.get("scope_target") or "").strip() or None
    severity = (request.form.get("severity") or "info").strip()
    subject = (request.form.get("subject") or "").strip()
    body = (request.form.get("body") or "").strip()
    expires_at = (request.form.get("expires_at") or "").strip() or None
    if scope not in NOTICE_SCOPES:
        flash("Invalid scope.", "error")
        return redirect(url_for("admin_notices"))
    if severity not in NOTICE_SEVERITIES:
        flash("Invalid severity.", "error")
        return redirect(url_for("admin_notices"))
    if not subject:
        flash("Subject is required.", "error")
        return redirect(url_for("admin_notices"))
    # Scope target validation: site scope must have no target; role
    # and instrument scopes must have a recognized target.
    if scope == "site":
        scope_target = None
    elif scope == "role":
        if scope_target not in NOTICE_ROLE_TARGETS:
            flash("Invalid role target.", "error")
            return redirect(url_for("admin_notices"))
    elif scope == "instrument":
        codes = {r["code"] for r in query_all("SELECT code FROM instruments")}
        if scope_target not in codes:
            flash("Invalid instrument target.", "error")
            return redirect(url_for("admin_notices"))
    elif scope == "mailing_list":
        scope_target = (request.form.get("scope_target_mailing_list") or "").strip()
        valid_ids = {str(r["id"]) for r in query_all("SELECT id FROM mailing_lists")}
        if scope_target not in valid_ids:
            flash("Invalid mailing list.", "error")
            return redirect(url_for("admin_notices"))
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    notice_id = execute(
        """
        INSERT INTO notices
            (scope, scope_target, severity, subject, body,
             author_id, created_at, expires_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (scope, scope_target, severity, subject, body,
         user["id"], now_iso, expires_at),
    )
    # v2.0.1 — audit log (was silently missing pre-v2.0.1).
    log_action(user["id"], "notice", notice_id, "notice_created",
               {"scope": scope, "scope_target": scope_target,
                "severity": severity, "subject": subject[:200]})
    flash(f"Notice posted: {subject[:60]}", "success")
    return redirect(url_for("admin_notices"))


@app.route("/admin/notices/<int:notice_id>/delete", methods=["POST"])
@login_required
def admin_notices_delete(notice_id: int):
    """Delete a notice. Same write gate as posting."""
    user = current_user()
    if not _user_can_post_notice(user):
        abort(403)
    # v2.0.1 — audit log BEFORE delete so we capture the subject.
    notice_row = query_one("SELECT subject, scope FROM notices WHERE id = ?", (notice_id,))
    execute("DELETE FROM notices WHERE id = ?", (notice_id,))
    log_action(user["id"], "notice", notice_id, "notice_deleted",
               {"subject": (notice_row["subject"] if notice_row else "")[:200],
                "scope": notice_row["scope"] if notice_row else None})
    flash(f"Notice #{notice_id} removed.", "success")
    return redirect(url_for("admin_notices"))


# ─── Mailing list management routes ──────────────────────────────────
@app.route("/admin/mailing-lists", methods=["GET"])
@login_required
def admin_mailing_lists():
    user = current_user()
    if not _user_can_post_notice(user):
        abort(403)
    lists = query_all(
        """SELECT ml.*, u.name AS creator_name,
                  (SELECT COUNT(*) FROM mailing_list_members WHERE list_id = ml.id) AS member_count
             FROM mailing_lists ml
             LEFT JOIN users u ON u.id = ml.created_by_user_id
             ORDER BY ml.name"""
    )
    all_users = query_all("SELECT id, name, email FROM users WHERE active = 1 ORDER BY name")
    return render_template("admin_mailing_lists.html", lists=lists, all_users=all_users)


@app.route("/admin/mailing-lists/new", methods=["POST"])
@login_required
def admin_mailing_list_create():
    user = current_user()
    if not _user_can_post_notice(user):
        abort(403)
    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()
    if not name:
        flash("List name is required.", "error")
        return redirect(url_for("admin_mailing_lists"))
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    list_id = execute(
        "INSERT INTO mailing_lists (name, description, created_by_user_id, created_at) VALUES (?, ?, ?, ?)",
        (name, description, user["id"], now_iso),
    )
    member_ids = request.form.getlist("member_ids")
    db = get_db()
    for uid in member_ids:
        try:
            db.execute("INSERT OR IGNORE INTO mailing_list_members (list_id, user_id) VALUES (?, ?)", (list_id, int(uid)))
        except (ValueError, TypeError):
            pass
    db.commit()
    log_action(user["id"], "mailing_list", list_id, "mailing_list_created", {"name": name})
    flash(f"Mailing list '{name}' created.", "success")
    return redirect(url_for("admin_mailing_lists"))


@app.route("/admin/mailing-lists/<int:list_id>/delete", methods=["POST"])
@login_required
def admin_mailing_list_delete(list_id: int):
    user = current_user()
    if not _user_can_post_notice(user):
        abort(403)
    execute("DELETE FROM mailing_lists WHERE id = ?", (list_id,))
    log_action(user["id"], "mailing_list", list_id, "mailing_list_deleted", {})
    flash("Mailing list deleted.", "success")
    return redirect(url_for("admin_mailing_lists"))


# ─── v1.6.2 User-to-user messaging routes ─────────────────────────────
# Directed messages, one sender + one recipient per row. Separate from
# notices (broadcast). Single source of truth: the `messages` table.


@app.route("/notifications/mark-read", methods=["POST"])
@login_required
def notification_mark_read():
    """Mark a notice or system notification as read. Or mark all read."""
    user = current_user()
    notice_id = request.form.get("notice_id", "").strip()
    if notice_id == "all":
        # Mark all admin notices read
        notices = active_notices_for_user(user)
        for n in notices:
            if not n.get("is_read"):
                execute(
                    "INSERT OR IGNORE INTO notice_reads (user_id, notice_id, read_at) VALUES (?, ?, ?)",
                    (user["id"], n["id"], now_iso()),
                )
        # Mark all system notifications read
        try:
            execute(
                "UPDATE system_notifications SET is_read = 1 WHERE user_id = ? AND is_read = 0",
                (user["id"],),
            )
        except sqlite3.OperationalError:
            pass
    elif notice_id.startswith("sys_"):
        # System notification
        sys_id = int(notice_id[4:])
        try:
            execute(
                "UPDATE system_notifications SET is_read = 1 WHERE id = ? AND user_id = ?",
                (sys_id, user["id"]),
            )
        except sqlite3.OperationalError:
            pass
    elif notice_id:
        execute(
            "INSERT OR IGNORE INTO notice_reads (user_id, notice_id, read_at) VALUES (?, ?, ?)",
            (user["id"], int(notice_id), now_iso()),
        )
    return redirect(url_for("notifications_page"))


@app.route("/notifications", methods=["GET"])
@login_required
def notifications_page():
    """Unified notification feed: admin notices + system notifications,
    newest first. Supports ?cat= filter (all, request, finance, instrument, admin)."""
    user = current_user()
    cat = request.args.get("cat", "all").strip().lower()
    # Admin notices — normalise into unified shape
    admin_notices = active_notices_for_user(user)
    unified: list[dict] = []
    for n in admin_notices:
        unified.append({
            "id": n["id"],
            "kind": "notice",
            "category": "admin",
            "title": n["subject"],
            "body": n.get("body", ""),
            "href": "",
            "severity": n.get("severity", "info"),
            "scope_label": n.get("scope_label", ""),
            "is_read": n.get("is_read", False),
            "created_at": n.get("created_at", ""),
            "author_name": n.get("author_name", ""),
        })
    # System notifications
    try:
        sys_rows = query_all(
            "SELECT * FROM system_notifications WHERE user_id = ? ORDER BY created_at DESC LIMIT 50",
            (user["id"],),
        )
    except sqlite3.OperationalError:
        sys_rows = []
    for s in sys_rows:
        row = dict(s)
        unified.append({
            "id": row["id"],
            "kind": "system",
            "category": row.get("category", "request"),
            "title": row.get("title", ""),
            "body": row.get("body", ""),
            "href": row.get("href", ""),
            "severity": "info",
            "scope_label": row.get("category", "system"),
            "is_read": bool(row.get("is_read", 0)),
            "created_at": row.get("created_at", ""),
            "author_name": "",
        })
    for item in unified:
        item["href"] = item.get("href") or url_for("notifications_page", cat=item.get("category", "all"))
        item["kind_label"] = "System" if item.get("kind") == "system" else "Notice"
        if item.get("category") == "request":
            item["action_label"] = "Open request"
        elif item.get("category") == "finance":
            item["action_label"] = "Open finance"
        elif item.get("category") == "instrument":
            item["action_label"] = "Open instrument"
        elif item.get("category") == "admin":
            item["action_label"] = "Review notice"
        else:
            item["action_label"] = "Open"
    # Surface unread items first, then sort newest first inside each group.
    unified.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    unified.sort(key=lambda x: 1 if x.get("is_read") else 0)
    # Filter by category if not "all"
    if cat and cat != "all":
        unified = [n for n in unified if n["category"] == cat]
    # Category counts for tab badges
    all_count = len(admin_notices) + len(sys_rows)
    cat_counts = {"all": all_count}
    for n in admin_notices:
        cat_counts["admin"] = cat_counts.get("admin", 0) + 1
    for s in sys_rows:
        c = dict(s).get("category", "request")
        cat_counts[c] = cat_counts.get(c, 0) + 1
    unread_count = sum(1 for n in unified if not n.get("is_read"))
    actionable_count = sum(1 for n in unified if n.get("href"))
    recent_unread = [n for n in unified if not n.get("is_read")][:3]
    return render_template(
        "notifications.html",
        notices=unified,
        active_cat=cat,
        cat_counts=cat_counts,
        unread_count=unread_count,
        actionable_count=actionable_count,
        recent_unread=recent_unread,
    )


def _save_message_attachments(message_id: int, files, now_iso: str):
    """Save uploaded files to disk and insert message_attachments rows."""
    import time as _time
    for f in files:
        if not f or not f.filename:
            continue
        if not allowed_file(f.filename):
            continue
        original = secure_filename(f.filename)
        stored = f"{int(_time.time())}_{original}"
        dest_dir = UPLOAD_DIR / "messages" / str(message_id)
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / stored
        f.save(str(dest))
        size = dest.stat().st_size
        execute(
            """INSERT INTO message_attachments
               (message_id, filename, stored_name, size_bytes, uploaded_at)
               VALUES (?, ?, ?, ?, ?)""",
            (message_id, original, stored, size, now_iso),
        )


@app.route("/inbox", methods=["GET"])
@login_required
def inbox():
    """List messages for the current user. Supports folder query param:
    inbox (default), sent, deleted.

    Owner (is_owner) can view any user's inbox via ?user_id=<id> — useful
    for admin oversight. The target user's name is passed to the template
    so the UI can show "Viewing <name>'s inbox".
    """
    user = current_user()
    folder = request.args.get("folder", "inbox")
    if folder not in ("inbox", "sent", "deleted"):
        folder = "inbox"

    # Owner can view another user's mailbox
    target_user_id = user["id"]
    viewing_as = None
    raw_uid = request.args.get("user_id", "")
    if raw_uid and is_owner(user):
        try:
            target_user_id = int(raw_uid)
            if target_user_id != user["id"]:
                target = query_one("SELECT name FROM users WHERE id = ?", (target_user_id,))
                viewing_as = target["name"] if target else f"User #{target_user_id}"
        except (ValueError, TypeError):
            pass

    # permanently_hidden messages never appear in any folder
    if folder == "inbox":
        rows = query_all(
            """
            SELECT m.id, m.subject, m.body, m.sent_at, m.read_at,
                   u.id AS sender_id, u.name AS sender_name, u.email AS sender_email
              FROM messages m
              LEFT JOIN users u ON u.id = m.sender_id
             WHERE m.recipient_id = ? AND m.deleted_by_recipient = 0
                   AND m.permanently_hidden = 0
             ORDER BY
                CASE WHEN m.read_at IS NULL THEN 0 ELSE 1 END,
                m.sent_at DESC
            """,
            (target_user_id,),
        )
    elif folder == "sent":
        rows = query_all(
            """
            SELECT m.id, m.subject, m.body, m.sent_at, m.read_at,
                   u.id AS sender_id, u.name AS sender_name, u.email AS sender_email
              FROM messages m
              LEFT JOIN users u ON u.id = m.recipient_id
             WHERE m.sender_id = ? AND m.deleted_by_sender = 0
                   AND m.permanently_hidden = 0
             ORDER BY m.sent_at DESC
            """,
            (target_user_id,),
        )
    else:  # deleted
        rows = query_all(
            """
            SELECT m.id, m.subject, m.body, m.sent_at, m.read_at,
                   u.id AS sender_id, u.name AS sender_name, u.email AS sender_email
              FROM messages m
              LEFT JOIN users u ON u.id = m.sender_id
             WHERE ((m.recipient_id = ? AND m.deleted_by_recipient = 1)
                OR (m.sender_id = ? AND m.deleted_by_sender = 1))
                AND m.permanently_hidden = 0
             ORDER BY m.sent_at DESC
            """,
            (target_user_id, target_user_id),
        )

    unread = sum(1 for r in rows if not r["read_at"])
    folder_counts = {
        "inbox": query_one(
            """
            SELECT COUNT(*) AS c
              FROM messages
             WHERE recipient_id = ? AND deleted_by_recipient = 0
               AND permanently_hidden = 0
            """,
            (target_user_id,),
        )["c"],
        "sent": query_one(
            """
            SELECT COUNT(*) AS c
              FROM messages
             WHERE sender_id = ? AND deleted_by_sender = 0
               AND permanently_hidden = 0
            """,
            (target_user_id,),
        )["c"],
        "deleted": query_one(
            """
            SELECT COUNT(*) AS c
              FROM messages
             WHERE ((recipient_id = ? AND deleted_by_recipient = 1)
                 OR (sender_id = ? AND deleted_by_sender = 1))
               AND permanently_hidden = 0
            """,
            (target_user_id, target_user_id),
        )["c"],
    }
    notice_rows = active_notices_for_user(user)
    try:
        sys_notice_rows = query_all(
            """
            SELECT id, title, category, href, is_read, created_at
              FROM system_notifications
             WHERE user_id = ?
             ORDER BY created_at DESC
             LIMIT 5
            """,
            (user["id"],),
        )
    except sqlite3.OperationalError:
        sys_notice_rows = []
    inbox_notifications = []
    for notice in notice_rows[:3]:
        inbox_notifications.append({
            "title": notice.get("subject", "Notice"),
            "category": "admin",
            "href": url_for("notifications_page", cat="admin"),
            "is_read": bool(notice.get("is_read")),
        })
    for row in sys_notice_rows[:3]:
        inbox_notifications.append({
            "title": row["title"],
            "category": row["category"] or "request",
            "href": row["href"] or url_for("notifications_page", cat=row["category"] or "all"),
            "is_read": bool(row["is_read"]),
        })
    inbox_notifications.sort(key=lambda item: item["is_read"])
    notification_unread = sum(1 for item in inbox_notifications if not item["is_read"])
    return render_template(
        "inbox.html",
        messages=rows,
        unread=unread,
        folder=folder,
        viewing_as=viewing_as,
        target_user_id=target_user_id if viewing_as else None,
        folder_counts=folder_counts,
        inbox_notifications=inbox_notifications[:4],
        notification_unread=notification_unread,
    )


@app.route("/messages/<int:message_id>", methods=["GET"])
@login_required
def message_detail(message_id: int):
    """Show a single message. Auto-marks it read on first view if the
    viewer is the recipient."""
    user = current_user()
    row = query_one(
        """
        SELECT m.id, m.subject, m.body, m.sent_at, m.read_at,
               m.sender_id, m.recipient_id,
               su.name AS sender_name, su.email AS sender_email,
               ru.name AS recipient_name, ru.email AS recipient_email
          FROM messages m
          LEFT JOIN users su ON su.id = m.sender_id
          LEFT JOIN users ru ON ru.id = m.recipient_id
         WHERE m.id = ?
        """,
        (message_id,),
    )
    if not row:
        abort(404)
    # Sender, recipient, or owner (read-only oversight) may read.
    is_own_message = user["id"] in (row["sender_id"], row["recipient_id"])
    if not is_own_message and not is_owner(user):
        abort(403)
    # Auto-mark-read on first view when the viewer is the recipient.
    # Owner viewing someone else's mail does NOT mark it read.
    if is_own_message and user["id"] == row["recipient_id"] and not row["read_at"]:
        now_iso = datetime.utcnow().isoformat(timespec="seconds")
        execute(
            "UPDATE messages SET read_at = ? WHERE id = ?",
            (now_iso, message_id),
        )
        row = dict(row)
        row["read_at"] = now_iso

    # Fetch attachments for this message
    attachments = query_all(
        "SELECT id, filename, stored_name, size_bytes, uploaded_at FROM message_attachments WHERE message_id = ?",
        (message_id,),
    )

    # Fetch parent message if this is a reply
    parent_message = None
    parent_id = query_one("SELECT parent_message_id FROM messages WHERE id = ?", (message_id,))
    if parent_id and parent_id["parent_message_id"]:
        parent_message = query_one(
            """
            SELECT m.id, m.subject, u.name AS sender_name
              FROM messages m
              LEFT JOIN users u ON u.id = m.sender_id
             WHERE m.id = ?
            """,
            (parent_id["parent_message_id"],),
        )

    return render_template(
        "message_detail.html",
        message=row,
        attachments=attachments,
        parent_message=parent_message,
        is_own_message=is_own_message,
    )


@app.route("/messages/new", methods=["GET"])
@login_required
def message_compose():
    """Compose form. `?to=<user_id>` pre-fills the recipient picker.

    5000-user scaling note: unbounded `SELECT … FROM users` does not
    fit in a browser dropdown. We cap at 200 rows using a relevance
    ranking — recent conversation partners first, then alphabetical.
    A `?q=<search>` query param narrows the list by name/email
    substring for when the target isn't in the first 200.
    """
    user = current_user()
    to_user_id = request.args.get("to", type=int)
    q = (request.args.get("q") or "").strip()
    RECIPIENT_CAP = 200
    params: list = [user["id"], user["id"], user["id"]]
    where_extra = ""
    if q:
        where_extra = " AND (LOWER(u.name) LIKE ? OR LOWER(u.email) LIKE ?)"
        like = f"%{q.lower()}%"
        params.extend([like, like])
    # Recency ranking: MAX(last message exchanged with this user) desc,
    # NULLs sort last. Covers both directions of the conversation.
    options = query_all(
        f"""
        SELECT u.id, u.name, u.email, u.role,
               (SELECT MAX(m.sent_at)
                  FROM messages m
                 WHERE (m.sender_id = ? AND m.recipient_id = u.id)
                    OR (m.sender_id = u.id AND m.recipient_id = ?)) AS last_chat
          FROM users u
         WHERE u.active = 1 AND u.id != ?{where_extra}
         ORDER BY CASE WHEN last_chat IS NULL THEN 1 ELSE 0 END,
                  last_chat DESC,
                  u.name
         LIMIT {RECIPIENT_CAP}
        """,
        tuple(params),
    )
    # Force-include the preselected user if it's not already in the
    # capped list (e.g. they're alphabetically after Z and you clicked
    # Reply to them). Small additional fetch, bounded at 1 row.
    if to_user_id and not any(o["id"] == to_user_id for o in options):
        pinned = query_one(
            "SELECT id, name, email, role, NULL AS last_chat FROM users WHERE id = ? AND active = 1",
            (to_user_id,),
        )
        if pinned:
            options = [pinned] + list(options)
    return render_template(
        "message_compose.html",
        options=options,
        preselected_to=to_user_id,
        recipient_cap=RECIPIENT_CAP,
        search_q=q,
    )


@app.route("/messages/new", methods=["POST"])
@login_required
def message_send():
    """Persist a new message. Sender = current user, recipient from
    form. Minimal validation: recipient must exist + be active,
    subject required, body optional."""
    user = current_user()
    recipient_id = request.form.get("recipient_id", type=int)
    subject = (request.form.get("subject") or "").strip()
    body = (request.form.get("body") or "").strip()
    if not recipient_id:
        flash("Pick a recipient.", "error")
        return redirect(url_for("message_compose"))
    if not subject:
        flash("Subject is required.", "error")
        return redirect(url_for("message_compose", to=recipient_id))
    recipient = query_one(
        "SELECT id, name FROM users WHERE id = ? AND active = 1",
        (recipient_id,),
    )
    if not recipient:
        flash("Recipient not found.", "error")
        return redirect(url_for("message_compose"))
    if recipient["id"] == user["id"]:
        flash("You can't send a message to yourself.", "error")
        return redirect(url_for("message_compose"))
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    message_id = execute(
        """
        INSERT INTO messages (sender_id, recipient_id, subject, body, sent_at, read_at)
        VALUES (?, ?, ?, ?, ?, NULL)
        """,
        (user["id"], recipient["id"], subject, body, now_iso),
    )
    # Handle file attachments
    _save_message_attachments(message_id, request.files.getlist("attachments"), now_iso)
    flash(f"Message sent to {recipient['name']}.", "success")
    return redirect(url_for("inbox"))


@app.route("/messages/<int:message_id>/reply", methods=["POST"])
@login_required
def message_reply(message_id: int):
    """Reply to an existing message. Sets parent_message_id for threading."""
    user = current_user()
    original = query_one(
        "SELECT id, sender_id, recipient_id, subject FROM messages WHERE id = ?",
        (message_id,),
    )
    if not original:
        abort(404)
    if user["id"] not in (original["sender_id"], original["recipient_id"]):
        abort(403)
    # Reply goes to the other party
    if user["id"] == original["sender_id"]:
        reply_to_id = original["recipient_id"]
    else:
        reply_to_id = original["sender_id"]
    subject = (request.form.get("subject") or "").strip()
    if not subject:
        orig_subj = original["subject"] or ""
        subject = orig_subj if orig_subj.lower().startswith("re:") else f"Re: {orig_subj}"
    body = (request.form.get("body") or "").strip()
    now_iso = datetime.utcnow().isoformat(timespec="seconds")
    new_id = execute(
        """INSERT INTO messages (sender_id, recipient_id, subject, body, sent_at, read_at, parent_message_id)
           VALUES (?, ?, ?, ?, ?, NULL, ?)""",
        (user["id"], reply_to_id, subject, body, now_iso, message_id),
    )
    _save_message_attachments(new_id, request.files.getlist("attachments"), now_iso)
    flash("Reply sent.", "success")
    return redirect(url_for("message_detail", message_id=new_id))


@app.route("/messages/<int:message_id>/delete", methods=["POST"])
@login_required
def message_delete(message_id: int):
    """Soft-delete a message for the current user.

    First delete from Inbox/Sent → moves to Deleted folder.
    Second delete from Deleted folder → sets permanently_hidden=1
    (hidden from all UI views but the DB record is preserved for
    audit — the user asked for this explicitly).
    """
    user = current_user()
    msg = query_one(
        "SELECT sender_id, recipient_id, deleted_by_sender, deleted_by_recipient FROM messages WHERE id = ?",
        (message_id,),
    )
    if not msg:
        abort(404)
    is_recipient = user["id"] == msg["recipient_id"]
    is_sender = user["id"] == msg["sender_id"]
    if not (is_recipient or is_sender):
        abort(403)

    # If already in the Deleted folder (soft-deleted), this is a
    # permanent hide. The record stays in the DB for audit.
    already_deleted = (
        (is_recipient and msg["deleted_by_recipient"])
        or (is_sender and msg["deleted_by_sender"])
    )
    if already_deleted:
        execute("UPDATE messages SET permanently_hidden = 1 WHERE id = ?", (message_id,))
        log_action(user["id"], "message", message_id, "message_permanently_hidden", {"subject": msg["subject"][:80] if msg.get("subject") else ""})
        flash("Message permanently removed.", "success")
    else:
        if is_recipient:
            execute("UPDATE messages SET deleted_by_recipient = 1 WHERE id = ?", (message_id,))
        elif is_sender:
            execute("UPDATE messages SET deleted_by_sender = 1 WHERE id = ?", (message_id,))
        log_action(user["id"], "message", message_id, "message_deleted", {"by": "recipient" if is_recipient else "sender"})
        flash("Message moved to Deleted.", "success")
    return redirect(url_for("inbox", folder="deleted" if already_deleted else "inbox"))


@app.route("/messages/<int:message_id>/attachment/<int:att_id>", methods=["GET"])
@login_required
def message_attachment(message_id: int, att_id: int):
    """Serve a message attachment file. Only sender/recipient may access."""
    user = current_user()
    msg = query_one(
        "SELECT sender_id, recipient_id FROM messages WHERE id = ?",
        (message_id,),
    )
    if not msg:
        abort(404)
    if user["id"] not in (msg["sender_id"], msg["recipient_id"]):
        abort(403)
    att = query_one(
        "SELECT filename, stored_name FROM message_attachments WHERE id = ? AND message_id = ?",
        (att_id, message_id),
    )
    if not att:
        abort(404)
    file_path = UPLOAD_DIR / "messages" / str(message_id) / att["stored_name"]
    if not file_path.exists():
        abort(404)
    return send_file(str(file_path), as_attachment=True, download_name=att["filename"])


@app.route("/finance")
@login_required
def finance_portal():
    """v2.0.0-alpha.2 — Finance portal reads from the new peer aggregates.

    The v1.7.0 implementation read SUM(amount_due) / SUM(amount_paid) /
    finance_status straight off sample_requests. alpha.1 added the peer
    tables (invoices, payments, projects, grant_allocations) and
    dual-wrote them via _backfill_domain_split(). alpha.2 (this step)
    flips the read path: every number on /finance now comes from
    invoices + payments, with finance_status derived at read time from
    SUM(payments.amount) vs invoices.amount_due — no denormalized status
    column that could drift.

    The legacy columns on sample_requests are still being written, so
    this tag is reversible: if a bug surfaces on the new path, revert
    the route bodies and /finance returns to reading the old columns.
    alpha.3 will stop writing the legacy columns; beta.1 drops them.

    Row shape into the template is preserved — `outstanding` and
    `recently_paid` rows still expose `amount_due`, `amount_paid`,
    `finance_status`, `receipt_number`, so finance.html does not change.

    Gated to finance_admin / super_admin / site_admin / owner.
    """
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)
    roles = user_role_set(user)
    can_edit = _user_can_edit_finance(user)
    # KPIs — headline row. Computed over invoices joined to external
    # sample_requests, with paid-sum derived from payments.
    kpi_row = query_one(
        """
        SELECT
          COALESCE(SUM(inv.amount_due), 0) AS total_owed,
          COALESCE(SUM(COALESCE(p.paid, 0)), 0) AS total_paid,
          COUNT(CASE WHEN COALESCE(p.paid, 0) < inv.amount_due THEN 1 END) AS pending_count,
          COUNT(CASE WHEN COALESCE(p.paid, 0) >= inv.amount_due AND inv.amount_due > 0 THEN 1 END) AS paid_count
        FROM invoices inv
        JOIN sample_requests sr ON sr.id = inv.request_id
        LEFT JOIN (
          SELECT invoice_id, SUM(amount) AS paid
            FROM payments
           GROUP BY invoice_id
        ) p ON p.invoice_id = inv.id
        WHERE sr.sample_origin = 'external'
        """
    )
    kpis = dict(kpi_row) if kpi_row else {
        "total_owed": 0, "total_paid": 0, "pending_count": 0, "paid_count": 0,
    }
    kpis["outstanding"] = max(0, (kpis["total_owed"] or 0) - (kpis["total_paid"] or 0))
    kpis["outstanding_fmt"] = "₹{:,.0f}".format(kpis["outstanding"])
    kpis["total_owed_fmt"] = "₹{:,.0f}".format(kpis["total_owed"] or 0)
    kpis["total_paid_fmt"] = "₹{:,.0f}".format(kpis["total_paid"] or 0)
    kpis["collection_rate"] = int(
        (kpis["total_paid"] or 0) * 100 / (kpis["total_owed"] or 1)
    ) if (kpis["total_owed"] or 0) > 0 else 0

    # By-instrument aggregation — subquery per metric keeps the join
    # graph shallow and makes each column independently auditable.
    by_instrument = query_all(
        """
        SELECT
          i.code, i.name,
          (SELECT COUNT(*) FROM sample_requests WHERE instrument_id = i.id) AS total_requests,
          COALESCE((
            SELECT SUM(inv.amount_due)
              FROM invoices inv
              JOIN sample_requests sr ON sr.id = inv.request_id
             WHERE sr.instrument_id = i.id AND sr.sample_origin = 'external'
          ), 0) AS owed,
          COALESCE((
            SELECT SUM(p.amount)
              FROM payments p
              JOIN invoices inv ON inv.id = p.invoice_id
              JOIN sample_requests sr ON sr.id = inv.request_id
             WHERE sr.instrument_id = i.id AND sr.sample_origin = 'external'
          ), 0) AS paid
        FROM instruments i
        WHERE i.status = 'active'
          AND (SELECT COUNT(*) FROM sample_requests WHERE instrument_id = i.id) > 0
        ORDER BY owed DESC
        LIMIT 20
        """
    )

    # Outstanding list — invoices with SUM(payments) < amount_due.
    # finance_status is derived on the fly; receipt_number taken from
    # the most recent payment row on that invoice (NULL for pure-pending).
    outstanding = query_all(
        """
        SELECT
          sr.id, sr.request_no, sr.title, sr.created_at,
          inv.amount_due,
          COALESCE(p.paid, 0) AS amount_paid,
          CASE
            WHEN COALESCE(p.paid, 0) = 0 THEN 'pending'
            WHEN COALESCE(p.paid, 0) < inv.amount_due THEN 'partial'
            ELSE 'paid'
          END AS finance_status,
          (SELECT receipt_number FROM payments
            WHERE invoice_id = inv.id
            ORDER BY paid_at DESC LIMIT 1) AS receipt_number,
          i.code AS instrument_code, i.name AS instrument_name,
          u.name AS requester_name
        FROM invoices inv
        JOIN sample_requests sr ON sr.id = inv.request_id
        JOIN instruments i ON i.id = sr.instrument_id
        LEFT JOIN users u ON u.id = sr.requester_id
        LEFT JOIN (
          SELECT invoice_id, SUM(amount) AS paid FROM payments GROUP BY invoice_id
        ) p ON p.invoice_id = inv.id
        WHERE sr.sample_origin = 'external'
          AND COALESCE(p.paid, 0) < inv.amount_due
        ORDER BY sr.created_at DESC
        LIMIT 30
        """
    )

    # Recently paid — invoices where SUM(payments) >= amount_due > 0.
    recently_paid = query_all(
        """
        SELECT
          sr.id, sr.request_no, sr.title, sr.created_at,
          inv.amount_due,
          COALESCE(p.paid, 0) AS amount_paid,
          (SELECT receipt_number FROM payments
            WHERE invoice_id = inv.id
            ORDER BY paid_at DESC LIMIT 1) AS receipt_number,
          i.code AS instrument_code
        FROM invoices inv
        JOIN sample_requests sr ON sr.id = inv.request_id
        JOIN instruments i ON i.id = sr.instrument_id
        LEFT JOIN (
          SELECT invoice_id, SUM(amount) AS paid FROM payments GROUP BY invoice_id
        ) p ON p.invoice_id = inv.id
        WHERE sr.sample_origin = 'external'
          AND inv.amount_due > 0
          AND COALESCE(p.paid, 0) >= inv.amount_due
        ORDER BY sr.created_at DESC
        LIMIT 15
        """
    )
    # ── Format amounts as ₹ strings for stat_blob display ──
    def _fmt(v):
        return "₹{:,.0f}".format(v or 0)

    kpis["outstanding_fmt"] = _fmt(kpis["outstanding"])
    kpis["total_owed_fmt"] = _fmt(kpis["total_owed"])
    kpis["total_paid_fmt"] = _fmt(kpis["total_paid"])
    total_owed = kpis["total_owed"] or 0
    collection_rate = "{:.0f}".format(
        (kpis["total_paid"] or 0) * 100 / total_owed
    ) if total_owed > 0 else "—"

    # ── By-instrument: pre-format amounts ──
    by_instrument_fmt = []
    for row in by_instrument:
        r = dict(row)
        r["owed_fmt"] = _fmt(r["owed"])
        r["paid_fmt"] = _fmt(r["paid"])
        r["outstanding_amt"] = (r["owed"] or 0) - (r["paid"] or 0)
        r["outstanding_fmt"] = _fmt(r["outstanding_amt"])
        by_instrument_fmt.append(r)

    # ── Outstanding / recently paid: pre-format amounts ──
    outstanding_fmt = []
    for row in outstanding:
        r = dict(row)
        r["balance_fmt"] = _fmt((r["amount_due"] or 0) - (r["amount_paid"] or 0))
        r["amount_due_fmt"] = _fmt(r["amount_due"])
        r["amount_paid_fmt"] = _fmt(r["amount_paid"])
        outstanding_fmt.append(r)

    recently_paid_fmt = []
    for row in recently_paid:
        r = dict(row)
        r["amount_paid_fmt"] = _fmt(r["amount_paid"])
        recently_paid_fmt.append(r)

    # ── Grant summary for sidebar tile ──
    grant_summary = query_one(
        """
        SELECT
          COUNT(*) AS total_grants,
          COUNT(CASE WHEN status = 'active' THEN 1 END) AS active_grants,
          COALESCE(SUM(total_budget), 0) AS total_budget,
          COALESCE(SUM(CASE WHEN status = 'active' THEN total_budget END), 0) AS active_budget
        FROM grants
        """
    )
    grant_kpis = dict(grant_summary) if grant_summary else {
        "total_grants": 0, "active_grants": 0, "total_budget": 0, "active_budget": 0,
    }
    grant_kpis["active_budget_fmt"] = _fmt(grant_kpis["active_budget"])

    # ── Cross-module KPIs: Fleet costs + Salary outflow ──
    vehicle_spend = query_one("SELECT COALESCE(SUM(amount), 0) AS total FROM vehicle_logs") or {"total": 0}
    salary_outflow = query_one("SELECT COALESCE(SUM(net_pay), 0) AS total FROM salary_payments WHERE status = 'paid'") or {"total": 0}

    return render_template(
        "finance.html",
        kpis=kpis,
        collection_rate=collection_rate,
        by_instrument=by_instrument_fmt,
        outstanding=outstanding_fmt,
        recently_paid=recently_paid_fmt,
        grant_kpis=grant_kpis,
        can_edit_finance=_user_can_edit_finance(user),
        vehicle_spend=vehicle_spend,
        salary_outflow=salary_outflow,
    )


def _user_can_view_finance(user: sqlite3.Row | None) -> bool:
    """Finance portal access. Operators can view revenue for their
    instruments (read-only); finance_admin/site_admin/super_admin/owner
    get full access including grants and editing."""
    if not user:
        return False
    roles = user_role_set(user)
    return bool(roles & {"finance_admin", "super_admin", "site_admin", "operator", "instrument_admin"}) or is_owner(user)


def _user_can_edit_finance(user: sqlite3.Row | None) -> bool:
    """Full finance editing (grants, invoices). Excludes operators."""
    if not user:
        return False
    roles = user_role_set(user)
    return bool(roles & {"finance_admin", "super_admin", "site_admin"}) or is_owner(user)


def check_budget(grant_id, amount):
    """Returns (allowed: bool, warning: str|None).

    Queries budget_rules + current utilization for the given grant.
    Blocks if over block_utilization_pct, warns if over warn_utilization_pct.
    """
    grant = query_one("SELECT id, total_budget FROM grants WHERE id = ?", (grant_id,))
    if not grant or not grant["total_budget"] or grant["total_budget"] <= 0:
        return True, None  # no budget cap — always allowed

    total_budget = grant["total_budget"]

    # Current spend: sum of payments on invoices linked to this grant
    # + sum of grant_expenses
    inv_spend = query_one(
        """SELECT COALESCE(SUM(p.amount), 0) AS s
           FROM payments p
           JOIN invoices inv ON inv.id = p.invoice_id
           WHERE inv.grant_id = ?""",
        (grant_id,),
    )
    exp_spend = query_one(
        "SELECT COALESCE(SUM(amount), 0) AS s FROM grant_expenses WHERE grant_id = ?",
        (grant_id,),
    )
    current_spend = (inv_spend["s"] if inv_spend else 0) + (exp_spend["s"] if exp_spend else 0)
    projected = current_spend + amount
    utilization_pct = (projected / total_budget) * 100

    # Load budget rules (fall back to defaults)
    rules = query_one("SELECT * FROM budget_rules WHERE grant_id = ?", (grant_id,))
    warn_pct = rules["warn_utilization_pct"] if rules else 80.0
    block_pct = rules["block_utilization_pct"] if rules else 100.0

    if utilization_pct >= block_pct:
        return False, (
            f"Budget BLOCKED: this would bring grant utilization to "
            f"{utilization_pct:.1f}% (limit {block_pct:.0f}%). "
            f"Current spend: {current_spend:,.0f}, new amount: {amount:,.0f}, "
            f"budget: {total_budget:,.0f}."
        )
    if utilization_pct >= warn_pct:
        return True, (
            f"Budget WARNING: grant utilization will reach {utilization_pct:.1f}% "
            f"(warn threshold {warn_pct:.0f}%). "
            f"Current spend: {current_spend:,.0f}, new amount: {amount:,.0f}, "
            f"budget: {total_budget:,.0f}."
        )
    return True, None


def _next_invoice_number():
    """Generate next invoice number in INV-YYYY-NNNN format."""
    import datetime
    year = datetime.datetime.now().year
    prefix = f"INV-{year}-"
    row = query_one(
        "SELECT invoice_number FROM invoices WHERE invoice_number LIKE ? ORDER BY invoice_number DESC LIMIT 1",
        (f"{prefix}%",),
    )
    if row and row["invoice_number"]:
        try:
            seq = int(row["invoice_number"].split("-")[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


@app.route("/finance/invoices")
@login_required
def finance_invoices_list():
    """Invoice list page with filters."""
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)

    # Filter params
    status_filter = request.args.get("status", "all")
    grant_filter = request.args.get("grant_id", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    where_clauses = []
    params = []

    if status_filter and status_filter != "all":
        if status_filter in ("pending", "partial", "paid", "void"):
            where_clauses.append("derived_status = ?")
            params.append(status_filter)

    if grant_filter:
        try:
            where_clauses.append("inv.grant_id = ?")
            params.append(int(grant_filter))
        except (ValueError, TypeError):
            pass

    instrument_filter = request.args.get("instrument_id", "")
    if instrument_filter:
        try:
            where_clauses.append("sr.instrument_id = ?")
            params.append(int(instrument_filter))
        except (ValueError, TypeError):
            pass

    if date_from:
        where_clauses.append("inv.issued_at >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("inv.issued_at <= ?")
        params.append(date_to + "T23:59:59")

    where_sql = ""
    # Build the query with derived status
    base_sql = """
        SELECT
          inv.id, inv.invoice_number, inv.amount_due, inv.status AS raw_status,
          inv.issued_at, inv.notes, inv.description, inv.grant_id,
          sr.request_no, sr.title AS request_title,
          g.name AS grant_name, g.code AS grant_code,
          COALESCE(p.paid, 0) AS amount_paid,
          CASE
            WHEN inv.status = 'void' THEN 'void'
            WHEN COALESCE(p.paid, 0) = 0 THEN 'pending'
            WHEN COALESCE(p.paid, 0) < inv.amount_due THEN 'partial'
            ELSE 'paid'
          END AS derived_status
        FROM invoices inv
        LEFT JOIN sample_requests sr ON sr.id = inv.request_id
        LEFT JOIN grants g ON g.id = inv.grant_id
        LEFT JOIN (
          SELECT invoice_id, SUM(amount) AS paid FROM payments GROUP BY invoice_id
        ) p ON p.invoice_id = inv.id
    """

    if where_clauses:
        # Wrap to filter on derived_status
        full_sql = f"SELECT * FROM ({base_sql}) AS sub WHERE {' AND '.join(where_clauses)} ORDER BY issued_at DESC LIMIT 200"
    else:
        full_sql = f"SELECT * FROM ({base_sql}) AS sub ORDER BY issued_at DESC LIMIT 200"

    # SQLite doesn't support AS sub syntax; use a CTE instead
    cte_sql = f"""
        WITH sub AS ({base_sql})
        SELECT * FROM sub
        {"WHERE " + " AND ".join(where_clauses) if where_clauses else ""}
        ORDER BY issued_at DESC
        LIMIT 200
    """

    invoices = query_all(cte_sql, tuple(params))

    # Grants for filter dropdown
    grants_list = query_all("SELECT id, code, name FROM grants ORDER BY name")

    return render_template(
        "finance_invoices.html",
        invoices=invoices,
        grants_list=grants_list,
        status_filter=status_filter,
        grant_filter=grant_filter,
        date_from=date_from,
        date_to=date_to,
        can_edit_finance=_user_can_edit_finance(user),
    )


@app.route("/finance/invoices/new", methods=["GET", "POST"])
@login_required
def finance_invoice_new():
    """Create a new invoice."""
    user = current_user()
    if not _user_can_edit_finance(user):
        abort(403)

    grants_list = query_all("SELECT id, code, name FROM grants WHERE status = 'active' ORDER BY name")
    requests_list = query_all(
        "SELECT id, request_no, title FROM sample_requests WHERE sample_origin = 'external' ORDER BY created_at DESC LIMIT 100"
    )

    if request.method == "POST":
        request_id = request.form.get("request_id", "").strip()
        description = request.form.get("description", "").strip()
        amount_str = request.form.get("amount_due", "0").strip()
        grant_id_str = request.form.get("grant_id", "").strip()
        notes = request.form.get("notes", "").strip()

        try:
            amount_due = float(amount_str)
        except (ValueError, TypeError):
            flash("Invalid amount.", "error")
            return render_template(
                "finance_invoice_form.html",
                grants_list=grants_list,
                requests_list=requests_list,
                mode="new",
            )

        if amount_due <= 0:
            flash("Amount must be positive.", "error")
            return render_template(
                "finance_invoice_form.html",
                grants_list=grants_list,
                requests_list=requests_list,
                mode="new",
            )

        grant_id = int(grant_id_str) if grant_id_str else None
        req_id = int(request_id) if request_id else None

        # If no request selected, we still need a request_id for FK.
        # Create a placeholder or require one.
        if not req_id:
            flash("Please select a sample request.", "error")
            return render_template(
                "finance_invoice_form.html",
                grants_list=grants_list,
                requests_list=requests_list,
                mode="new",
            )

        # Budget check
        if grant_id:
            allowed, warning = check_budget(grant_id, amount_due)
            if not allowed:
                flash(warning, "error")
                return render_template(
                    "finance_invoice_form.html",
                    grants_list=grants_list,
                    requests_list=requests_list,
                    mode="new",
                )
            if warning:
                flash(warning, "warning")

        invoice_number = _next_invoice_number()
        inv_id = execute(
            """INSERT INTO invoices
               (request_id, project_id, amount_due, status, issued_at, due_at, notes,
                invoice_number, description, created_by_user_id, grant_id)
               VALUES (?, NULL, ?, 'pending', ?, NULL, ?, ?, ?, ?, ?)""",
            (req_id, amount_due, now_iso(), notes, invoice_number, description, user["id"], grant_id),
        )

        log_action(user["id"], "invoice", inv_id, "invoice_created", {
            "invoice_number": invoice_number, "amount_due": amount_due,
            "request_id": req_id, "grant_id": grant_id,
        })

        # Notify requester
        sr = query_one("SELECT requester_id, request_no FROM sample_requests WHERE id = ?", (req_id,))
        if sr:
            notify(
                sr["requester_id"], "finance",
                f"Invoice {invoice_number} created",
                f"Amount: {amount_due:,.2f} for request {sr['request_no']}.",
                href=url_for("finance_invoice_detail", invoice_id=inv_id),
                source_type="invoice", source_id=inv_id,
            )

        flash(f"Invoice {invoice_number} created successfully.", "success")
        return redirect(url_for("finance_invoice_detail", invoice_id=inv_id))

    return render_template(
        "finance_invoice_form.html",
        grants_list=grants_list,
        requests_list=requests_list,
        mode="new",
    )


@app.route("/finance/invoices/<int:invoice_id>")
@login_required
def finance_invoice_detail(invoice_id):
    """Invoice detail with payment timeline."""
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)

    inv = query_one(
        """SELECT inv.*, sr.request_no, sr.title AS request_title, sr.requester_id,
                  g.name AS grant_name, g.code AS grant_code, g.total_budget AS grant_budget,
                  u.name AS created_by_name
           FROM invoices inv
           LEFT JOIN sample_requests sr ON sr.id = inv.request_id
           LEFT JOIN grants g ON g.id = inv.grant_id
           LEFT JOIN users u ON u.id = inv.created_by_user_id
           WHERE inv.id = ?""",
        (invoice_id,),
    )
    if not inv:
        abort(404)

    payments_list = query_all(
        """SELECT p.*, u.name AS recorded_by_name
           FROM payments p
           LEFT JOIN users u ON u.id = p.recorded_by_user_id
           WHERE p.invoice_id = ?
           ORDER BY p.paid_at DESC""",
        (invoice_id,),
    )

    total_paid = sum(p["amount"] for p in payments_list)
    remaining = max(0, inv["amount_due"] - total_paid)
    derived_status = "void" if inv["status"] == "void" else (
        "paid" if total_paid >= inv["amount_due"] and inv["amount_due"] > 0
        else "partial" if total_paid > 0
        else "pending"
    )

    # Budget utilization for the grant (if linked)
    budget_util = None
    if inv["grant_id"]:
        grant_spend = query_one(
            """SELECT COALESCE(SUM(p.amount), 0) AS s
               FROM payments p JOIN invoices i ON i.id = p.invoice_id
               WHERE i.grant_id = ?""",
            (inv["grant_id"],),
        )
        exp_spend = query_one(
            "SELECT COALESCE(SUM(amount), 0) AS s FROM grant_expenses WHERE grant_id = ?",
            (inv["grant_id"],),
        )
        total_spend = (grant_spend["s"] if grant_spend else 0) + (exp_spend["s"] if exp_spend else 0)
        if inv["grant_budget"] and inv["grant_budget"] > 0:
            budget_util = {
                "spent": total_spend,
                "budget": inv["grant_budget"],
                "pct": (total_spend / inv["grant_budget"]) * 100,
            }

    return render_template(
        "finance_invoice_detail.html",
        inv=inv,
        payments=payments_list,
        total_paid=total_paid,
        remaining=remaining,
        derived_status=derived_status,
        budget_util=budget_util,
        can_edit_finance=_user_can_edit_finance(user),
    )


@app.route("/finance/invoices/<int:invoice_id>/pay", methods=["POST"])
@login_required
def finance_invoice_pay(invoice_id):
    """Record a payment against an invoice."""
    user = current_user()
    if not _user_can_edit_finance(user):
        abort(403)

    inv = query_one("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
    if not inv:
        abort(404)
    if inv["status"] == "void":
        flash("Cannot record payment on a voided invoice.", "error")
        return redirect(url_for("finance_invoice_detail", invoice_id=invoice_id))

    amount_str = request.form.get("amount", "0").strip()
    method = request.form.get("method", "bank_transfer").strip()
    receipt_number = request.form.get("receipt_number", "").strip()
    notes = request.form.get("notes", "").strip()

    try:
        amount = float(amount_str)
    except (ValueError, TypeError):
        flash("Invalid payment amount.", "error")
        return redirect(url_for("finance_invoice_detail", invoice_id=invoice_id))

    if amount <= 0:
        flash("Payment amount must be positive.", "error")
        return redirect(url_for("finance_invoice_detail", invoice_id=invoice_id))

    # Budget check if grant linked
    if inv["grant_id"]:
        allowed, warning = check_budget(inv["grant_id"], amount)
        if not allowed:
            flash(warning, "error")
            return redirect(url_for("finance_invoice_detail", invoice_id=invoice_id))
        if warning:
            flash(warning, "warning")

    pay_id = execute(
        """INSERT INTO payments (invoice_id, amount, method, receipt_number, paid_at, recorded_by_user_id, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (invoice_id, amount, method, receipt_number, now_iso(), user["id"], notes),
    )

    # Check if fully paid — update invoice status
    total_paid_row = query_one(
        "SELECT COALESCE(SUM(amount), 0) AS s FROM payments WHERE invoice_id = ?",
        (invoice_id,),
    )
    total_paid = total_paid_row["s"] if total_paid_row else 0
    new_status = "paid" if total_paid >= inv["amount_due"] else "partial" if total_paid > 0 else "pending"
    execute("UPDATE invoices SET status = ? WHERE id = ?", (new_status, invoice_id))

    log_action(user["id"], "payment", pay_id, "payment_recorded", {
        "invoice_id": invoice_id, "amount": amount, "method": method,
        "receipt_number": receipt_number, "new_invoice_status": new_status,
    })

    # Notify requester
    sr = query_one(
        "SELECT sr.requester_id, sr.request_no FROM sample_requests sr JOIN invoices inv ON inv.request_id = sr.id WHERE inv.id = ?",
        (invoice_id,),
    )
    if sr:
        notify(
            sr["requester_id"], "finance",
            f"Payment of {amount:,.2f} recorded",
            f"Against invoice {inv['invoice_number'] or invoice_id} for request {sr['request_no']}. Status: {new_status}.",
            href=url_for("finance_invoice_detail", invoice_id=invoice_id),
            source_type="payment", source_id=pay_id,
        )

    flash(f"Payment of {amount:,.2f} recorded. Invoice is now {new_status}.", "success")
    return redirect(url_for("finance_invoice_detail", invoice_id=invoice_id))


@app.route("/finance/invoices/<int:invoice_id>/void", methods=["POST"])
@login_required
def finance_invoice_void(invoice_id):
    """Void an invoice. Only finance_admin/owner."""
    user = current_user()
    if not _user_can_edit_finance(user):
        abort(403)

    inv = query_one("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
    if not inv:
        abort(404)

    execute("UPDATE invoices SET status = 'void' WHERE id = ?", (invoice_id,))
    log_action(user["id"], "invoice", invoice_id, "invoice_voided", {
        "invoice_number": inv["invoice_number"],
    })

    # Notify requester
    sr = query_one(
        "SELECT sr.requester_id, sr.request_no FROM sample_requests sr JOIN invoices inv ON inv.request_id = sr.id WHERE inv.id = ?",
        (invoice_id,),
    )
    if sr:
        notify(
            sr["requester_id"], "finance",
            f"Invoice {inv['invoice_number'] or invoice_id} voided",
            f"For request {sr['request_no']}.",
            href=url_for("finance_invoice_detail", invoice_id=invoice_id),
            source_type="invoice", source_id=invoice_id,
        )

    flash("Invoice voided.", "success")
    return redirect(url_for("finance_invoice_detail", invoice_id=invoice_id))


@app.route("/finance/spend")
@login_required
def finance_spend():
    """Unified spend view: combine invoice payments + grant_expenses."""
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)

    grant_filter = request.args.get("grant_id", "")

    grant_where = ""
    params_payments = []
    params_expenses = []

    if grant_filter:
        try:
            gid = int(grant_filter)
            grant_where_payments = "WHERE inv.grant_id = ?"
            grant_where_expenses = "WHERE ge.grant_id = ?"
            params_payments = [gid]
            params_expenses = [gid]
        except (ValueError, TypeError):
            grant_where_payments = ""
            grant_where_expenses = ""
    else:
        grant_where_payments = ""
        grant_where_expenses = ""

    # Invoice payments
    payment_rows = query_all(
        f"""SELECT
              p.id, p.amount, p.method, p.receipt_number, p.paid_at AS event_date,
              p.notes, 'payment' AS entry_type,
              inv.id AS invoice_id, inv.invoice_number, inv.grant_id,
              g.name AS grant_name, g.code AS grant_code,
              sr.request_no, u.name AS recorded_by_name
            FROM payments p
            JOIN invoices inv ON inv.id = p.invoice_id
            LEFT JOIN grants g ON g.id = inv.grant_id
            LEFT JOIN sample_requests sr ON sr.id = inv.request_id
            LEFT JOIN users u ON u.id = p.recorded_by_user_id
            {grant_where_payments}
            ORDER BY p.paid_at DESC
            LIMIT 300""",
        tuple(params_payments),
    )

    # Grant expenses
    expense_rows = query_all(
        f"""SELECT
              ge.id, ge.amount, ge.expense_type AS method, ge.receipt_number,
              ge.recorded_at AS event_date, ge.notes, 'expense' AS entry_type,
              NULL AS invoice_id, '' AS invoice_number, ge.grant_id,
              g.name AS grant_name, g.code AS grant_code,
              ge.description AS request_no, u.name AS recorded_by_name
            FROM grant_expenses ge
            LEFT JOIN grants g ON g.id = ge.grant_id
            LEFT JOIN users u ON u.id = ge.recorded_by_user_id
            {grant_where_expenses}
            ORDER BY ge.recorded_at DESC
            LIMIT 300""",
        tuple(params_expenses),
    )

    # Approved expense receipts
    receipt_rows = query_all(
        """SELECT er.id, 'receipt' AS entry_type, er.title AS request_no,
              er.amount, er.category AS method, '' AS receipt_number,
              er.created_at AS event_date, '' AS notes,
              NULL AS invoice_id, '' AS invoice_number, NULL AS grant_id,
              '' AS grant_name, '' AS grant_code,
              u.name AS recorded_by_name
            FROM expense_receipts er
            JOIN users u ON u.id = er.submitted_by_user_id
            WHERE er.status = 'approved'
            ORDER BY er.created_at DESC
            LIMIT 300"""
    )

    # Merge and sort by date descending
    all_entries = list(payment_rows) + list(expense_rows) + list(receipt_rows)
    all_entries.sort(key=lambda r: r["event_date"] or "", reverse=True)

    grants_list = query_all("SELECT id, code, name FROM grants ORDER BY name")

    return render_template(
        "finance_spend.html",
        entries=all_entries[:300],
        grants_list=grants_list,
        grant_filter=grant_filter,
        can_edit_finance=_user_can_edit_finance(user),
    )


@app.route("/finance/grants", methods=["GET", "POST"])
@login_required
def finance_grants_list():
    """v2.0.0-alpha.2 — Grants inventory reads via the new peer graph.

    Old path: spend = SUM(sample_requests.amount_paid) WHERE grant_id = g.id
    New path: grants → grant_allocations → projects → sample_requests →
              invoices → payments

    The graph is longer but the semantics are the same — every row that
    used to be reachable via sample_requests.grant_id is now reachable
    via the allocation table after the alpha.1 backfill. The domain
    split test suite locks this invariant."""
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)

    if request.method == "POST":
        if not _user_can_edit_finance(user):
            abort(403)
        code = request.form.get("code", "").strip()
        name = request.form.get("name", "").strip()
        if not code or not name:
            flash("Code and name are required.", "error")
            return redirect(url_for("finance_grants_list"))
        sponsor = request.form.get("sponsor", "").strip()
        total_budget = float(request.form.get("total_budget", 0) or 0)
        start_date = request.form.get("start_date", "").strip() or None
        end_date = request.form.get("end_date", "").strip() or None
        grant_type = request.form.get("grant_type", "external").strip()
        department = request.form.get("department", "").strip()
        notes = request.form.get("notes", "").strip()
        new_id = execute(
            """INSERT INTO grants (code, name, sponsor, total_budget, start_date, end_date,
               grant_type, department, notes, status, pi_user_id, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)""",
            (code, name, sponsor, total_budget, start_date, end_date,
             grant_type, department, notes, user["id"], now_iso()),
        )
        get_db().commit()
        log_action(user["id"], "grant", new_id, "grant_created", {"code": code, "name": name, "budget": total_budget})
        flash(f"Grant {code} created.", "success")
        return redirect(url_for("finance_grant_detail", grant_id=new_id))

    rows = query_all(
        """
        SELECT
          g.id, g.code, g.name, g.sponsor, g.total_budget,
          g.start_date, g.end_date, g.status, g.notes,
          pi.name AS pi_name, pi.email AS pi_email,
          COALESCE((
            SELECT SUM(p.amount)
              FROM payments p
              JOIN invoices inv ON inv.id = p.invoice_id
              JOIN sample_requests sr ON sr.id = inv.request_id
              JOIN grant_allocations ga ON ga.project_id = sr.project_id
             WHERE ga.grant_id = g.id
          ), 0) AS spend_paid,
          COALESCE((
            SELECT SUM(inv.amount_due)
              FROM invoices inv
              JOIN sample_requests sr ON sr.id = inv.request_id
              JOIN grant_allocations ga ON ga.project_id = sr.project_id
             WHERE ga.grant_id = g.id
          ), 0) AS spend_billed,
          COALESCE((
            SELECT COUNT(DISTINCT sr.id)
              FROM sample_requests sr
              JOIN grant_allocations ga ON ga.project_id = sr.project_id
             WHERE ga.grant_id = g.id
          ), 0) AS sample_count
        FROM grants g
        LEFT JOIN users pi ON pi.id = g.pi_user_id
        ORDER BY
          CASE g.status WHEN 'active' THEN 0 ELSE 1 END,
          g.end_date
        """
    )
    totals_row = query_one(
        """
        SELECT
          (SELECT COALESCE(SUM(total_budget), 0) FROM grants) AS total_budget,
          (SELECT COUNT(*) FROM grants) AS grant_count,
          (SELECT COALESCE(SUM(p.amount), 0)
             FROM payments p
             JOIN invoices inv ON inv.id = p.invoice_id
             JOIN sample_requests sr ON sr.id = inv.request_id
             JOIN grant_allocations ga ON ga.project_id = sr.project_id) AS total_paid
        """
    )
    totals = dict(totals_row) if totals_row else {
        "total_budget": 0, "total_paid": 0, "grant_count": 0,
    }
    totals["total_remaining"] = max(0, (totals["total_budget"] or 0) - (totals["total_paid"] or 0))
    totals["total_budget_fmt"] = "₹{:,.0f}".format(totals["total_budget"] or 0)
    totals["total_paid_fmt"] = "₹{:,.0f}".format(totals["total_paid"] or 0)
    totals["total_remaining_fmt"] = "₹{:,.0f}".format(totals["total_remaining"])
    return render_template(
        "finance_grants.html",
        grants=rows,
        totals=totals,
    )


@app.route("/finance/grants/<int:grant_id>", methods=["GET", "POST"])
@login_required
def finance_grant_detail(grant_id: int):
    """v2.0.0-alpha.2 — Single-grant drill-down via the peer graph.

    Walks grant → grant_allocations → projects → sample_requests →
    invoices → payments. amount_due / amount_paid / finance_status /
    receipt_number are derived per row; the template iteration
    variable name (`inv`) is unchanged.

    POST: edit grant metadata (name, sponsor, PI, budget, dates,
    grant_type, department, notes). Gated to finance editors only."""
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)

    if request.method == "POST":
        if not _user_can_edit_finance(user):
            abort(403)
        action = request.form.get("action", "")
        if action == "update_grant_metadata":
            db = get_db()
            portfolio_manager_id = request.form.get("portfolio_manager_id", "").strip()
            administered_by_user_id = request.form.get("administered_by_user_id", "").strip()
            db.execute(
                """UPDATE grants SET
                    code = ?, name = ?, sponsor = ?, grant_type = ?, department = ?,
                    total_budget = ?, start_date = ?, end_date = ?,
                    notes = ?, status = ?,
                    portfolio_manager_id = ?,
                    administered_by_user_id = ?
                 WHERE id = ?""",
                (
                    request.form.get("code", "").strip(),
                    request.form.get("name", "").strip(),
                    request.form.get("sponsor", "").strip(),
                    request.form.get("grant_type", "internal").strip(),
                    request.form.get("department", "").strip(),
                    float(request.form.get("total_budget", "0") or 0),
                    request.form.get("start_date", "").strip(),
                    request.form.get("end_date", "").strip(),
                    request.form.get("notes", "").strip(),
                    request.form.get("status", "active").strip(),
                    int(portfolio_manager_id) if portfolio_manager_id else None,
                    int(administered_by_user_id) if administered_by_user_id else None,
                    grant_id,
                ),
            )
            db.commit()
            log_action(user["id"], "grant", grant_id, "grant_metadata_updated", {})
            flash("Grant metadata updated.", "success")
            return redirect(url_for("finance_grant_detail", grant_id=grant_id))
        elif action == "add_grant_member":
            member_user_id = request.form.get("member_user_id", "").strip()
            member_role = request.form.get("member_role", "member").strip()
            if member_role not in ("member", "viewer", "admin"):
                member_role = "member"
            if member_user_id:
                db = get_db()
                try:
                    db.execute(
                        """INSERT INTO grant_members (grant_id, user_id, role, added_by_user_id, added_at)
                           VALUES (?, ?, ?, ?, ?)""",
                        (grant_id, int(member_user_id), member_role, user["id"], now_iso()),
                    )
                    db.commit()
                    log_action(user["id"], "grant", grant_id, "grant_member_added", {"member_user_id": int(member_user_id), "role": member_role})
                    flash("Member added.", "success")
                except Exception:
                    flash("User is already a member of this grant.", "warning")
            return redirect(url_for("finance_grant_detail", grant_id=grant_id))
        elif action == "remove_grant_member":
            member_id = request.form.get("member_id", "").strip()
            if member_id:
                db = get_db()
                db.execute("DELETE FROM grant_members WHERE id = ? AND grant_id = ?", (int(member_id), grant_id))
                db.commit()
                log_action(user["id"], "grant", grant_id, "grant_member_removed", {"member_id": int(member_id)})
                flash("Member removed.", "success")
            return redirect(url_for("finance_grant_detail", grant_id=grant_id))
    grant = query_one(
        """
        SELECT g.*, pi.name AS pi_name, pi.email AS pi_email,
               pm.name AS portfolio_manager_name,
               adm.name AS administrator_name
          FROM grants g
          LEFT JOIN users pi ON pi.id = g.pi_user_id
          LEFT JOIN users pm ON pm.id = g.portfolio_manager_id
          LEFT JOIN users adm ON adm.id = g.administered_by_user_id
         WHERE g.id = ?
        """,
        (grant_id,),
    )
    if not grant:
        abort(404)
    # Candidates for Portfolio Manager / Administered By dropdowns
    admin_candidates = query_all(
        "SELECT id, name FROM users WHERE role IN ('finance_admin','super_admin','site_admin','instrument_admin') ORDER BY name"
    )
    grant_members = query_all(
        """SELECT gm.id, gm.role, gm.added_at, u.id AS user_id, u.name, u.email,
                  ab.name AS added_by_name
             FROM grant_members gm
             JOIN users u ON u.id = gm.user_id
             LEFT JOIN users ab ON ab.id = gm.added_by_user_id
            WHERE gm.grant_id = ?
            ORDER BY gm.role, u.name""",
        (grant_id,),
    )
    all_users = query_all("SELECT id, name, email FROM users ORDER BY name")
    charged = query_all(
        """
        SELECT
          sr.id, sr.request_no, sr.title, sr.sample_name, sr.created_at, sr.sample_origin,
          COALESCE(inv.amount_due, 0) AS amount_due,
          COALESCE((
            SELECT SUM(p.amount) FROM payments p WHERE p.invoice_id = inv.id
          ), 0) AS amount_paid,
          CASE
            WHEN inv.id IS NULL THEN 'n/a'
            WHEN COALESCE((SELECT SUM(amount) FROM payments WHERE invoice_id = inv.id), 0) = 0 THEN 'pending'
            WHEN COALESCE((SELECT SUM(amount) FROM payments WHERE invoice_id = inv.id), 0) < inv.amount_due THEN 'partial'
            ELSE 'paid'
          END AS finance_status,
          (SELECT receipt_number FROM payments
            WHERE invoice_id = inv.id
            ORDER BY paid_at DESC LIMIT 1) AS receipt_number,
          i.code AS instrument_code, i.name AS instrument_name,
          u.name AS requester_name
        FROM grant_allocations ga
        JOIN sample_requests sr ON sr.project_id = ga.project_id
        LEFT JOIN invoices inv ON inv.request_id = sr.id
        JOIN instruments i ON i.id = sr.instrument_id
        LEFT JOIN users u ON u.id = sr.requester_id
        WHERE ga.grant_id = ?
        ORDER BY sr.created_at DESC
        LIMIT 200
        """,
        (grant_id,),
    )
    total_paid = sum(r["amount_paid"] or 0 for r in charged)
    total_billed = sum(r["amount_due"] or 0 for r in charged)
    remaining = max(0, (grant["total_budget"] or 0) - total_paid)
    percent_used = int((total_paid / grant["total_budget"] * 100) if grant["total_budget"] else 0)

    # Pre-compute formatted amounts
    budget_fmt = "₹{:,.0f}".format(grant["total_budget"] or 0)
    spent_fmt = "₹{:,.0f}".format(total_paid)
    remaining_fmt = "₹{:,.0f}".format(remaining)
    billed_fmt = "₹{:,.0f}".format(total_billed)

    # Budget utilization class for bar color
    if percent_used >= 100:
        budget_util_class = "budget-bar-red"
    elif percent_used >= 80:
        budget_util_class = "budget-bar-amber"
    else:
        budget_util_class = "budget-bar-green"

    # Expenses (non-sample charges)
    expenses = query_all(
        """SELECT ge.*, u.name AS recorder_name
           FROM grant_expenses ge
           LEFT JOIN users u ON u.id = ge.recorded_by_user_id
           WHERE ge.grant_id = ?
           ORDER BY ge.recorded_at DESC""",
        (grant_id,),
    )
    total_expenses = sum(e["amount"] or 0 for e in expenses)
    expenses_fmt = "₹{:,.0f}".format(total_expenses)

    return render_template(
        "finance_grant_detail.html",
        grant=grant,
        charged=charged,
        total_paid=total_paid,
        total_billed=total_billed,
        remaining=remaining,
        percent_used=percent_used,
        budget_fmt=budget_fmt,
        spent_fmt=spent_fmt,
        remaining_fmt=remaining_fmt,
        billed_fmt=billed_fmt,
        utilization_pct=percent_used,
        budget_util_class=budget_util_class,
        expenses=expenses,
        total_expenses=total_expenses,
        expenses_fmt=expenses_fmt,
        can_edit_finance=_user_can_edit_finance(user),
        admin_candidates=admin_candidates,
        grant_members=grant_members,
        all_users=all_users,
        today=__import__("datetime").date.today().isoformat(),
    )


@app.route("/finance/grants/<int:grant_id>/expenses", methods=["GET", "POST"])
@login_required
def finance_grant_expenses(grant_id: int):
    """Grant expenses — non-sample charges (equipment, reagents, vendors)."""
    user = current_user()
    if not _user_can_view_finance(user):
        abort(403)
    grant = query_one(
        "SELECT g.*, pi.name AS pi_name FROM grants g LEFT JOIN users pi ON pi.id = g.pi_user_id WHERE g.id = ?",
        (grant_id,),
    )
    if not grant:
        abort(404)

    if request.method == "POST":
        if not _user_can_edit_finance(user):
            abort(403)
        db = get_db()
        raw_amount = float(request.form.get("amount", "0") or 0)
        currency = request.form.get("currency", "INR").strip().upper()
        exchange_rate_str = request.form.get("exchange_rate", "").strip()
        expense_date = request.form.get("expense_date", "").strip() or now_iso()[:10]

        # Compute INR amount
        if currency and currency != "INR" and exchange_rate_str:
            exchange_rate = float(exchange_rate_str)
            inr_amount = round(raw_amount * exchange_rate, 2)
            original_amount = raw_amount
            original_currency = currency
        else:
            inr_amount = raw_amount
            original_amount = None
            original_currency = None
            exchange_rate = None

        db.execute(
            """INSERT INTO grant_expenses
               (grant_id, description, amount, expense_type, receipt_number,
                recorded_by_user_id, recorded_at, notes,
                original_amount, original_currency, exchange_rate, expense_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                grant_id,
                request.form.get("description", "").strip(),
                inr_amount,
                request.form.get("expense_type", "equipment").strip(),
                request.form.get("receipt_number", "").strip(),
                user["id"],
                now_iso(),
                request.form.get("notes", "").strip(),
                original_amount,
                original_currency,
                exchange_rate,
                expense_date,
            ),
        )
        db.commit()
        log_action(user["id"], "grant", grant_id, "expense_recorded", {
            "description": request.form.get("description", "").strip(),
            "amount": inr_amount,
            "original": f"{original_currency} {original_amount}" if original_currency else None,
            "rate": exchange_rate,
        })
        flash("Expense recorded.", "success")
        return redirect(url_for("finance_grant_expenses", grant_id=grant_id))

    expenses = query_all(
        """SELECT ge.*, u.name AS recorder_name
           FROM grant_expenses ge
           LEFT JOIN users u ON u.id = ge.recorded_by_user_id
           WHERE ge.grant_id = ?
           ORDER BY ge.recorded_at DESC""",
        (grant_id,),
    )
    total_expenses = sum(e["amount"] or 0 for e in expenses)
    from datetime import date as _date
    return render_template(
        "finance_grant_expenses.html",
        grant=grant,
        expenses=expenses,
        total_expenses=total_expenses,
        can_edit_finance=_user_can_edit_finance(user),
        today=_date.today().isoformat(),
    )


@app.route("/finance/grants/<int:grant_id>/form-control", methods=["GET", "POST"])
@login_required
def finance_grant_form_control(grant_id: int):
    """Grant form-control — approval config, budget rules, overview."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"finance_admin", "super_admin"} or is_owner(user)):
        abort(403)
    grant = query_one(
        "SELECT g.*, pi.name AS pi_name FROM grants g LEFT JOIN users pi ON pi.id = g.pi_user_id WHERE g.id = ?",
        (grant_id,),
    )
    if not grant:
        abort(404)

    if request.method == "POST":
        action = request.form.get("action", "")
        db = get_db()

        if action == "save_budget_rules":
            # Store budget rules as grant metadata columns (future: separate table)
            warn_pct = float(request.form.get("warn_threshold_pct", "80") or 80)
            freeze_pct = float(request.form.get("freeze_threshold_pct", "100") or 100)
            require_receipt = 1 if request.form.get("require_receipt") else 0
            db.execute(
                """UPDATE grants SET notes = ? WHERE id = ?""",
                (
                    f"budget_warn_pct={warn_pct};freeze_pct={freeze_pct};require_receipt={require_receipt}",
                    grant_id,
                ),
            )
            db.commit()
            log_action(user["id"], "grant", grant_id, "budget_rules_updated", {
                "warn_pct": warn_pct, "freeze_pct": freeze_pct,
            })
            flash("Budget rules saved.", "success")
            return redirect(url_for("finance_grant_form_control", grant_id=grant_id))

        abort(400)

    # Parse budget rules from notes (simple key=value pairs)
    budget_rules = {"warn_threshold_pct": 80, "freeze_threshold_pct": 100, "require_receipt": 0}
    if grant["notes"] and "budget_warn_pct=" in (grant["notes"] or ""):
        for part in grant["notes"].split(";"):
            if "=" in part:
                k, v = part.split("=", 1)
                k = k.strip()
                if k == "budget_warn_pct":
                    budget_rules["warn_threshold_pct"] = float(v)
                elif k == "freeze_pct":
                    budget_rules["freeze_threshold_pct"] = float(v)
                elif k == "require_receipt":
                    budget_rules["require_receipt"] = int(v)

    expense_count = query_one(
        "SELECT COUNT(*) AS cnt FROM grant_expenses WHERE grant_id = ?", (grant_id,),
    )
    sample_count = query_one(
        """SELECT COUNT(DISTINCT sr.id) AS cnt
           FROM sample_requests sr
           JOIN grant_allocations ga ON ga.project_id = sr.project_id
           WHERE ga.grant_id = ?""",
        (grant_id,),
    )
    return render_template(
        "finance_grant_form_control.html",
        grant=grant,
        budget_rules=budget_rules,
        expense_count=(expense_count["cnt"] if expense_count else 0),
        sample_count=(sample_count["cnt"] if sample_count else 0),
    )


@app.route("/api/health-check")
def api_health_check():
    """Lightweight healthcheck endpoint — no auth required."""
    db = get_db()
    try:
        db.execute("SELECT 1")
        db_ok = True
    except Exception:
        db_ok = False
    return jsonify({"ok": db_ok, "status": "healthy" if db_ok else "degraded"})


@app.route("/sitemap")
@login_required
def sitemap():
    user = current_user()
    access_profile = user_access_profile(user)
    db = get_db()

    # --- Build Apple Settings-style sections ---
    sections = []

    # Core section — always visible
    core_items = [
        {"label": "Home", "hint": "Dashboard and overview", "type": "link", "href": url_for("index")},
        {"label": "New Request", "hint": "Submit a new sample request", "type": "link", "href": url_for("new_request")},
        {"label": "Notifications", "hint": "All active notices", "type": "link", "href": url_for("notifications_page")},
        {"label": "Inbox", "hint": "Direct messages", "type": "link", "href": url_for("inbox")},
        {"label": "Attendance & Leave", "hint": "Mark attendance, apply for leave", "type": "link", "href": url_for("attendance_page")},
        {"label": "My Profile", "hint": "View and manage your account", "type": "link", "href": url_for("user_profile", user_id=user["id"])},
    ]
    core_info = [
        {"label": "Logged in as", "type": "text", "value": user["name"]},
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        {"label": "Role", "type": "badge", "value": user["role"].replace("_", " ").title(), "status": "submitted"},
        {"label": "Username", "type": "text", "value": user["email"]},
    ]
    sections.append({
        "key": "core",
        "title": "General",
        "icon": "⚙",
        "groups": [
            {"title": "Navigation", "entries": core_items},
            {"title": "Account", "entries": core_info},
        ],
    })

    # Operations section — if user has instrument/schedule access
    if access_profile["can_access_instruments"] or access_profile["can_access_schedule"]:
        ops_items = []
        if access_profile["can_access_instruments"]:
            instrument_count = db.execute("SELECT COUNT(*) FROM instruments WHERE status = 'active'").fetchone()[0]
            ops_items.append({"label": "Instruments", "hint": f"{instrument_count} active instruments", "type": "link", "href": url_for("instruments")})
        if access_profile["can_access_schedule"]:
            open_count = db.execute("SELECT COUNT(*) FROM sample_requests WHERE status NOT IN ('completed', 'rejected')").fetchone()[0]
            ops_items.append({"label": "Job Queue", "hint": f"{open_count} open jobs", "type": "link", "href": url_for("schedule")})
            ops_items.append({"label": "Completed Jobs", "hint": "View processed history", "type": "link", "href": url_for("schedule", bucket="completed")})
        if access_profile.get("can_view_finance_stage") or is_owner(user):
            ops_items.append({"label": "Finance Portal", "hint": "Billing, invoices, grants", "type": "link", "href": url_for("finance_portal")})
        sections.append({
            "key": "operations",
            "title": "Operations",
            "icon": "⚡",
            "groups": [{"title": "Workspace", "entries": ops_items}],
        })

    # Reporting section
    if access_profile["can_access_calendar"] or access_profile["can_access_stats"]:
        report_items = []
        if access_profile["can_access_calendar"]:
            report_items.append({"label": "Calendar", "hint": "Weekly schedule and downtime", "type": "link", "href": url_for("calendar")})
            report_items.append({"label": "Subscribe Calendar (.ics)", "hint": "Add to Google Calendar / Apple Calendar / Outlook", "type": "link", "href": url_for("calendar_ics")})
        if access_profile["can_access_stats"]:
            report_items.append({"label": "Statistics", "hint": "Operations control dashboard", "type": "link", "href": url_for("stats")})
            report_items.append({"label": "Data Export", "hint": "Generate Excel reports", "type": "link", "href": url_for("visualizations")})
        sections.append({
            "key": "reporting",
            "title": "Reporting",
            "icon": "📊",
            "groups": [{"title": "Views", "entries": report_items}],
        })

    # Administration section — admins only
    if access_profile["can_manage_members"]:
        user_count = db.execute("SELECT COUNT(*) FROM users WHERE active = 1").fetchone()[0]
        invited_count = db.execute("SELECT COUNT(*) FROM users WHERE invite_status = 'invited'").fetchone()[0]
        admin_items = [
            {"label": "User Management", "hint": f"{user_count} active, {invited_count} pending invites", "type": "link", "href": url_for("admin_users")},
            {"label": "Notices", "hint": "Post site-wide announcements", "type": "link", "href": url_for("admin_notices")},
            {"label": "Calibrations Due", "hint": "NABL compliance — upcoming calibrations", "type": "link", "href": url_for("admin_calibrations_upcoming")},
            {"label": "Leave Queue", "hint": "Approve/reject leave requests", "type": "link", "href": url_for("admin_leave_queue")},
            {"label": "Attendance Overview", "hint": "Daily attendance for all staff", "type": "link", "href": url_for("admin_attendance_calendar")},
        ]
        system_items = [
            {"label": "Server Port", "type": "text", "value": "5055"},
            {"label": "Database", "type": "text", "value": "SQLite (local)"},
        ]
        sections.append({
            "key": "admin",
            "title": "Administration",
            "icon": "🔧",
            "groups": [
                {"title": "Manage", "entries": admin_items},
                {"title": "System", "entries": system_items},
            ],
        })

    return render_template("sitemap.html", sections=sections, title="Settings")


@app.route("/docs")
@login_required
def docs():
    """Render project documentation + progress bar from PROJECT.md."""
    import re

    project_path = os.path.join(os.path.dirname(__file__), "docs", "PROJECT.md")
    readme_path = os.path.join(os.path.dirname(__file__), "README.md")

    project_content = ""
    readme_content = ""
    try:
        with open(project_path, "r") as f:
            project_content = f.read()
    except FileNotFoundError:
        project_content = "_PROJECT.md not found._"
    try:
        with open(readme_path, "r") as f:
            readme_content = f.read()
    except FileNotFoundError:
        readme_content = ""

    # Parse progress table from README
    progress_phases = []
    for line in readme_content.split("\n"):
        m = re.match(r"\|\s*(.+?)\s*\|\s*(Done|In Progress|Planned|Future)\s*\|.*?(\d+)%\s*\|", line)
        if m:
            progress_phases.append({
                "name": m.group(1).strip(),
                "status": m.group(2).strip(),
                "pct": int(m.group(3)),
            })

    # Simple markdown-to-HTML (headings, bold, code blocks, lists, tables, links)
    def md_to_html(md):
        import html as html_mod

        def esc(text):
            """Escape HTML entities to prevent XSS."""
            return html_mod.escape(str(text))

        def inline_fmt(text):
            """Apply inline markdown formatting to already-escaped text."""
            text = esc(text)
            text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)
            text = re.sub(r"`(.+?)`", r"<code>\1</code>", text)
            # Links: only allow http/https/mailto, not javascript:
            def safe_link(m):
                label, href = m.group(1), m.group(2)
                if href.lower().startswith(("http://", "https://", "mailto:", "/", "#")):
                    return f'<a href="{href}">{label}</a>'
                return f"{label} ({href})"
            text = re.sub(r"\[(.+?)\]\((.+?)\)", safe_link, text)
            return text

        lines = md.split("\n")
        html_parts = []
        in_code = False
        in_table = False
        in_list = False
        for line in lines:
            # Code blocks
            if line.strip().startswith("```"):
                if in_code:
                    html_parts.append("</code></pre>")
                    in_code = False
                else:
                    lang = esc(line.strip()[3:].strip())
                    html_parts.append(f'<pre><code class="lang-{lang}">')
                    in_code = True
                continue
            if in_code:
                html_parts.append(esc(line))
                continue
            # Close table/list if needed
            if in_table and not line.strip().startswith("|"):
                html_parts.append("</tbody></table>")
                in_table = False
            if in_list and not re.match(r"^\s*[-*]\s", line) and not re.match(r"^\s*\d+\.\s", line) and line.strip():
                html_parts.append("</ul>")
                in_list = False
            # Blank lines
            if not line.strip():
                continue
            # Headings
            hm = re.match(r"^(#{1,6})\s+(.*)", line)
            if hm:
                level = len(hm.group(1))
                text = inline_fmt(hm.group(2).strip())
                html_parts.append(f"<h{level}>{text}</h{level}>")
                continue
            # Tables
            if line.strip().startswith("|"):
                cells = [c.strip() for c in line.strip().strip("|").split("|")]
                if all(re.match(r"^[-:]+$", c) for c in cells):
                    continue  # separator row
                if not in_table:
                    html_parts.append('<table class="doc-table"><thead><tr>')
                    for c in cells:
                        html_parts.append(f"<th>{esc(c)}</th>")
                    html_parts.append("</tr></thead><tbody>")
                    in_table = True
                else:
                    html_parts.append("<tr>")
                    for c in cells:
                        html_parts.append(f"<td>{esc(c)}</td>")
                    html_parts.append("</tr>")
                continue
            # Lists
            lm = re.match(r"^\s*[-*]\s+(.*)", line)
            if lm:
                if not in_list:
                    html_parts.append("<ul>")
                    in_list = True
                item_text = esc(lm.group(1))
                # Checkboxes
                item_text = item_text.replace("[x]", "&#9745;").replace("[ ]", "&#9744;")
                html_parts.append(f"<li>{item_text}</li>")
                continue
            # Paragraph — escape then apply inline formatting
            html_parts.append(f"<p>{inline_fmt(line)}</p>")
        if in_table:
            html_parts.append("</tbody></table>")
        if in_list:
            html_parts.append("</ul>")
        if in_code:
            html_parts.append("</code></pre>")
        return "\n".join(html_parts)

    project_html = md_to_html(project_content)

    return render_template(
        "docs.html",
        title="Documentation",
        progress_phases=progress_phases,
        project_html=project_html,
    )


@app.route("/requests/<int:request_id>/quick-receive", methods=["POST"])
@login_required
def quick_receive_request(request_id: int):
    user = current_user()
    sample_request = query_one("SELECT * FROM sample_requests WHERE id = ?", (request_id,))
    if sample_request is None:
        return jsonify({"ok": False, "error": "Request not found."}), 404
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_manage = can_manage_instrument(user["id"], sample_request["instrument_id"], user["role"])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_operate = can_operate_instrument(user["id"], sample_request["instrument_id"], user["role"])
    if not (can_operate or can_manage):
        return jsonify({"ok": False, "error": "Forbidden."}), 403
    if sample_request["status"] not in {"sample_submitted", "awaiting_sample_submission"}:
        return jsonify({"ok": False, "error": "Sample cannot be marked received from its current state."}), 400
    assert_status_transition(sample_request["status"], "sample_received")
    execute(
        """
        UPDATE sample_requests
        SET status = 'sample_received', sample_received_at = ?, received_by_operator_id = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (now_iso(), user["id"], now_iso(), request_id),
    )
    log_action(user["id"], "sample_request", request_id, "sample_received", {"remarks": ""})
    return jsonify({"ok": True, "request_id": request_id, "new_status": "sample_received"})


@app.route("/demo/switch/<role_key>")
@login_required
def demo_switch_role(role_key: str):
    if not DEMO_MODE:
        abort(404)  # Pretend the route doesn't exist in production.
    user = current_user()
    if not can_use_role_switcher(user):
        abort(403)
    target = DEMO_ROLE_SWITCHES.get(role_key)
    if not target:
        abort(404)
    target_user = query_one("SELECT * FROM users WHERE email = ? AND active = 1", (target["email"],))
    if target_user is None:
        flash("Demo role account not found.", "error")
        return redirect(url_for("index"))
    session["user_id"] = target_user["id"]
    flash(f"Switched to {target['label']} view.", "success")
    return redirect(url_for("index"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        user = query_one("SELECT * FROM users WHERE email = ? AND active = 1", (email,))
        if user and user["invite_status"] == "active" and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session.permanent = True
            log_action(user["id"], "auth", user["id"], "login", {"ip": request.remote_addr or ""})
            # W1.3.8 password hygiene: if the admin issued a temporary
            # password (create_user or reset_password), force the user
            # to set their own before they can use the app.
            if row_value(user, "must_change_password", 0):
                flash(
                    f"Welcome {user['name']}. You are signed in with a temporary password — "
                    f"please choose a new one to continue.",
                    "success",
                )
                return redirect(url_for("change_password"))
            flash(f"Signed in as {user['name']}.", "success")
            return redirect(url_for("index"))
        # Log failed attempt — use user id if email matched but password wrong
        failed_uid = user["id"] if user else None
        log_action(failed_uid, "auth", failed_uid or 0, "login_failed", {"email": email, "ip": request.remote_addr or ""})
        flash("Invalid login.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    uid = session.get("user_id")
    if uid:
        log_action(uid, "auth", uid, "logout", {})
    session.clear()
    return redirect(url_for("login"))


# ── Google OAuth routes ─────────────────────────────────────────
@app.route("/auth/google")
def auth_google():
    """Redirect to Google OAuth consent screen."""
    if not _oauth:
        flash("Google sign-in is not configured on this server.", "error")
        return redirect(url_for("login"))
    redirect_uri = url_for("auth_google_callback", _external=True)
    return _oauth.google.authorize_redirect(redirect_uri)


@app.route("/auth/google/callback")
def auth_google_callback():
    """Handle Google OAuth callback — match by email to existing user."""
    if not _oauth:
        abort(404)
    try:
        token = _oauth.google.authorize_access_token()
        userinfo = token.get("userinfo") or _oauth.google.userinfo()
    except Exception:
        flash("Google sign-in failed. Please try again.", "error")
        return redirect(url_for("login"))

    email = (userinfo.get("email") or "").strip().lower()
    if not email:
        flash("Could not retrieve email from Google.", "error")
        return redirect(url_for("login"))

    # Domain restriction
    if GOOGLE_ALLOWED_DOMAIN:
        domain = email.split("@")[-1] if "@" in email else ""
        if domain != GOOGLE_ALLOWED_DOMAIN:
            flash(f"Only @{GOOGLE_ALLOWED_DOMAIN} accounts are allowed.", "error")
            return redirect(url_for("login"))

    user = query_one("SELECT * FROM users WHERE email = ? AND active = 1", (email,))
    if not user:
        flash("No CATALYST account found for this email. Contact your departmental secretary to get access.", "error")
        return redirect(url_for("login"))

    # Link google_id if not yet linked
    google_id = userinfo.get("sub", "")
    avatar_url = userinfo.get("picture", "")
    if google_id and not row_value(user, "google_id", ""):
        execute("UPDATE users SET google_id = ?, avatar_url = ? WHERE id = ?",
                (google_id, avatar_url, user["id"]))
    elif avatar_url:
        execute("UPDATE users SET avatar_url = ? WHERE id = ?", (avatar_url, user["id"]))

    session.clear()
    session["user_id"] = user["id"]
    session.permanent = True
    flash(f"Signed in as {user['name']} via Google.", "success")
    return redirect(url_for("index"))


# ── System notification helper ──────────────────────────────────
def notify(user_id: int, category: str, title: str, body: str = "",
           href: str = "", source_type: str = "", source_id: int | None = None):
    """Create a system notification for a user."""
    execute(
        "INSERT INTO system_notifications (user_id, category, title, body, href, source_type, source_id, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, category, title, body, href, source_type, source_id, now_iso()),
    )


def unread_system_notification_count(user) -> int:
    """Count of unread system notifications for a user."""
    row = query_one(
        "SELECT COUNT(*) AS c FROM system_notifications WHERE user_id = ? AND is_read = 0",
        (user["id"],),
    )
    return row["c"] if row else 0


# ── Todo / Task routes ────────────────────────────────────────────

_TODO_CATEGORIES = ("general", "instrument", "finance", "admin", "personal")

def _can_assign_todo_to(assigner, target_user_id: int) -> bool:
    """Check if assigner can assign a todo/task to target_user_id."""
    if not assigner:
        return False
    role = assigner["role"]
    # super_admin / owner / site_admin can assign to anyone
    if is_owner(assigner) or role in ("super_admin", "site_admin"):
        return True
    # instrument_admin can assign to operators/faculty on their instruments
    if role == "instrument_admin":
        managed = {r["user_id"] for r in query_all(
            "SELECT user_id FROM instrument_operators WHERE instrument_id IN (SELECT instrument_id FROM instrument_admins WHERE user_id = ?)",
            (assigner["id"],),
        )}
        managed |= {r["user_id"] for r in query_all(
            "SELECT user_id FROM instrument_requesters WHERE instrument_id IN (SELECT instrument_id FROM instrument_admins WHERE user_id = ?)",
            (assigner["id"],),
        )}
        return target_user_id in managed
    # finance_admin can assign to other finance users
    if role == "finance_admin":
        finance_ids = {r["id"] for r in query_all(
            "SELECT id FROM users WHERE role = 'finance_admin' AND active = 1"
        )}
        return target_user_id in finance_ids
    return False


def _assignable_users_for(user):
    """Return list of users this person can assign tasks to."""
    if is_owner(user) or user["role"] in ("super_admin", "site_admin"):
        return query_all(
            "SELECT id, name, email FROM users WHERE active = 1 AND invite_status = 'active' AND id != ? ORDER BY name",
            (user["id"],),
        )
    if user["role"] == "instrument_admin":
        rows = query_all(
            """SELECT DISTINCT u.id, u.name, u.email FROM users u
               WHERE u.active = 1 AND u.invite_status = 'active' AND u.id != ? AND (
                 u.id IN (SELECT user_id FROM instrument_operators WHERE instrument_id IN (SELECT instrument_id FROM instrument_admins WHERE user_id = ?))
                 OR u.id IN (SELECT user_id FROM instrument_requesters WHERE instrument_id IN (SELECT instrument_id FROM instrument_admins WHERE user_id = ?))
               ) ORDER BY u.name""",
            (user["id"], user["id"], user["id"]),
        )
        return rows
    if user["role"] == "finance_admin":
        return query_all(
            "SELECT id, name, email FROM users WHERE role = 'finance_admin' AND active = 1 AND id != ? ORDER BY name",
            (user["id"],),
        )
    return []


_TODO_ORDER_SQL = """ORDER BY
    CASE t.status WHEN 'open' THEN 0 WHEN 'in_progress' THEN 1 ELSE 2 END,
    CASE t.priority WHEN 'urgent' THEN 0 WHEN 'high' THEN 1 WHEN 'normal' THEN 2 WHEN 'low' THEN 3 ELSE 4 END,
    t.due_date ASC NULLS LAST, t.created_at DESC"""


@app.route("/todos", methods=["GET"])
@login_required
def todos_page():
    """Task & Todo hub with three tabs: my todos, assigned to me, assigned by me."""
    user = current_user()
    tab = request.args.get("tab", "my")
    if tab not in ("my", "assigned", "by_me"):
        tab = "my"

    if tab == "my":
        items = query_all(
            f"""SELECT t.*, u.name AS assigner_name, u2.name AS assignee_name
               FROM user_todos t
               LEFT JOIN users u ON u.id = t.assigned_by_user_id
               LEFT JOIN users u2 ON u2.id = t.user_id
               WHERE t.user_id = ? AND (t.assigned_by_user_id IS NULL OR t.assigned_by_user_id = ?)
               {_TODO_ORDER_SQL}""",
            (user["id"], user["id"]),
        )
    elif tab == "assigned":
        items = query_all(
            f"""SELECT t.*, u.name AS assigner_name, u2.name AS assignee_name
               FROM user_todos t
               LEFT JOIN users u ON u.id = t.assigned_by_user_id
               LEFT JOIN users u2 ON u2.id = t.user_id
               WHERE t.user_id = ? AND t.assigned_by_user_id IS NOT NULL AND t.assigned_by_user_id != ?
               {_TODO_ORDER_SQL}""",
            (user["id"], user["id"]),
        )
    else:  # by_me
        items = query_all(
            f"""SELECT t.*, u.name AS assigner_name, u2.name AS assignee_name
               FROM user_todos t
               LEFT JOIN users u ON u.id = t.assigned_by_user_id
               LEFT JOIN users u2 ON u2.id = t.user_id
               WHERE t.assigned_by_user_id = ? AND t.user_id != ?
               {_TODO_ORDER_SQL}""",
            (user["id"], user["id"]),
        )

    # Tab counts for badges
    my_count = query_one(
        "SELECT COUNT(*) AS c FROM user_todos WHERE user_id = ? AND (assigned_by_user_id IS NULL OR assigned_by_user_id = ?) AND status != 'done'",
        (user["id"], user["id"]),
    )["c"]
    assigned_count = query_one(
        "SELECT COUNT(*) AS c FROM user_todos WHERE user_id = ? AND assigned_by_user_id IS NOT NULL AND assigned_by_user_id != ? AND status != 'done'",
        (user["id"], user["id"]),
    )["c"]
    by_me_count = query_one(
        "SELECT COUNT(*) AS c FROM user_todos WHERE assigned_by_user_id = ? AND user_id != ? AND status != 'done'",
        (user["id"], user["id"]),
    )["c"]

    assignable_users = _assignable_users_for(user)

    todo_counts = {
        "open": sum(1 for i in items if i["status"] == "open"),
        "in_progress": sum(1 for i in items if i["status"] == "in_progress"),
        "done": sum(1 for i in items if i["status"] == "done"),
        "total": len(items),
    }
    return render_template(
        "todos.html",
        items=items,
        tab=tab,
        my_count=my_count,
        assigned_count=assigned_count,
        by_me_count=by_me_count,
        assignable_users=assignable_users,
        categories=_TODO_CATEGORIES,
        todo_counts=todo_counts,
    )


@app.route("/todos/new", methods=["POST"])
@login_required
def todo_new():
    """Create a new todo or assigned task."""
    user = current_user()
    title = request.form.get("title", "").strip()
    if not title:
        flash("Title is required.", "error")
        return redirect(url_for("todos_page"))
    body = request.form.get("body", "").strip()
    priority = request.form.get("priority", "normal").strip()
    if priority not in ("normal", "high", "urgent"):
        priority = "normal"
    due_date = request.form.get("due_date", "").strip() or None
    category = request.form.get("category", "general").strip()
    if category not in _TODO_CATEGORIES:
        category = "general"
    assign_to = request.form.get("assign_to", "").strip()
    if assign_to:
        target_id = int(assign_to)
        if target_id != user["id"] and not _can_assign_todo_to(user, target_id):
            flash("You cannot assign tasks to that user.", "error")
            return redirect(url_for("todos_page"))
    else:
        target_id = user["id"]
    execute(
        "INSERT INTO user_todos (user_id, assigned_by_user_id, title, body, priority, due_date, category, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (target_id, user["id"], title, body, priority, due_date, category, now_iso()),
    )
    # Notify assignee if it's someone else
    if target_id != user["id"]:
        try:
            notify(target_id, "task", f"New task: {title}",
                   f"Assigned by {user['name']}", url_for("todos_page", tab="assigned"),
                   "user_todo", None)
        except Exception:
            pass
    flash("Task created.", "success")
    return redirect(url_for("todos_page"))


@app.route("/todos/<int:todo_id>/update", methods=["POST"])
@login_required
def todo_update(todo_id: int):
    """Update status or fields of a todo/task."""
    user = current_user()
    todo = query_one("SELECT * FROM user_todos WHERE id = ?", (todo_id,))
    if not todo or (todo["user_id"] != user["id"] and todo["assigned_by_user_id"] != user["id"]):
        abort(403)
    status = request.form.get("status", "").strip()
    if status in ("open", "in_progress", "done"):
        completed_at = now_iso() if status == "done" else None
        execute(
            "UPDATE user_todos SET status = ?, completed_at = ? WHERE id = ?",
            (status, completed_at, todo_id),
        )
    title = request.form.get("title", "").strip()
    if title:
        execute("UPDATE user_todos SET title = ? WHERE id = ?", (title, todo_id))
    body = request.form.get("body")
    if body is not None:
        execute("UPDATE user_todos SET body = ? WHERE id = ?", (body.strip(), todo_id))
    flash("Task updated.", "success")
    return redirect(url_for("todos_page", tab=request.form.get("tab", "my")))


@app.route("/todos/<int:todo_id>/complete", methods=["POST"])
@login_required
def todo_complete(todo_id: int):
    """Mark a todo as done."""
    user = current_user()
    todo = query_one("SELECT * FROM user_todos WHERE id = ?", (todo_id,))
    if not todo or (todo["user_id"] != user["id"] and todo["assigned_by_user_id"] != user["id"]):
        abort(403)
    execute(
        "UPDATE user_todos SET status = 'done', completed_at = ? WHERE id = ?",
        (now_iso(), todo_id),
    )
    flash("Task completed.", "success")
    tab = request.form.get("tab", "my")
    return redirect(url_for("todos_page", tab=tab))


@app.route("/todos/<int:todo_id>/delete", methods=["POST"])
@login_required
def todo_delete(todo_id: int):
    """Delete a todo item."""
    user = current_user()
    todo = query_one("SELECT * FROM user_todos WHERE id = ?", (todo_id,))
    if not todo or (todo["user_id"] != user["id"] and todo["assigned_by_user_id"] != user["id"]):
        abort(403)
    execute("DELETE FROM user_todos WHERE id = ?", (todo_id,))
    flash("Task deleted.", "success")
    tab = request.form.get("tab", "my")
    return redirect(url_for("todos_page", tab=tab))


# ── Letters module ─────────────────────────────────────────────────────

@app.route("/letters", methods=["GET"])
@login_required
def letters_list():
    """List user's letters. Admins see all."""
    user = current_user()
    ap = user_access_profile(user)
    if is_owner(user) or user["role"] in ("super_admin", "site_admin"):
        letters = query_all(
            "SELECT l.*, u.name AS author_name FROM letters l "
            "JOIN users u ON u.id = l.author_user_id "
            "ORDER BY l.created_at DESC"
        )
    else:
        letters = query_all(
            "SELECT l.*, u.name AS author_name FROM letters l "
            "JOIN users u ON u.id = l.author_user_id "
            "WHERE l.author_user_id = ? ORDER BY l.created_at DESC",
            (user["id"],),
        )
    return render_template("letters.html", title="Letters", letters=letters)


@app.route("/letters/new", methods=["GET", "POST"])
@login_required
def letter_new():
    """Create a new letter."""
    user = current_user()
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        recipient = request.form.get("recipient", "").strip()
        body = request.form.get("body", "").strip()
        letterhead = request.form.get("letterhead", "default").strip()
        if not title:
            flash("Title is required.", "error")
            return redirect(url_for("letter_new"))
        ts = now_iso()
        lid = execute(
            "INSERT INTO letters (author_user_id, title, body, recipient, letterhead, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user["id"], title, body, recipient, letterhead, ts, ts),
        )
        flash("Letter created.", "success")
        return redirect(url_for("letter_detail", letter_id=lid))
    prefill_recipient = request.args.get("recipient", "")
    return render_template("letter_form.html", title="New Letter", letter=None, prefill_recipient=prefill_recipient)


@app.route("/letters/<int:letter_id>", methods=["GET"])
@login_required
def letter_detail(letter_id: int):
    """View a letter."""
    user = current_user()
    letter = query_one("SELECT * FROM letters WHERE id = ?", (letter_id,))
    if not letter:
        abort(404)
    if letter["author_user_id"] != user["id"] and not (is_owner(user) or user["role"] in ("super_admin", "site_admin")):
        abort(403)
    author = query_one("SELECT name FROM users WHERE id = ?", (letter["author_user_id"],))
    return render_template("letter_detail.html", title=letter["title"], letter=letter, author_name=author["name"] if author else "Unknown")


@app.route("/letters/<int:letter_id>/print", methods=["GET"])
@login_required
def letter_print(letter_id: int):
    """Render letter on letterhead for printing."""
    user = current_user()
    letter = query_one("SELECT * FROM letters WHERE id = ?", (letter_id,))
    if not letter:
        abort(404)
    if letter["author_user_id"] != user["id"] and not (is_owner(user) or user["role"] in ("super_admin", "site_admin")):
        abort(403)
    author = query_one("SELECT name FROM users WHERE id = ?", (letter["author_user_id"],))
    return render_template("_letter_print.html", letter=letter, author_name=author["name"] if author else "Unknown")


@app.route("/letters/<int:letter_id>/update", methods=["POST"])
@login_required
def letter_update(letter_id: int):
    """Edit a letter."""
    user = current_user()
    letter = query_one("SELECT * FROM letters WHERE id = ?", (letter_id,))
    if not letter:
        abort(404)
    if letter["author_user_id"] != user["id"] and not is_owner(user):
        abort(403)
    title = request.form.get("title", "").strip()
    recipient = request.form.get("recipient", "").strip()
    body = request.form.get("body", "").strip()
    letterhead = request.form.get("letterhead", "default").strip()
    status = request.form.get("status", letter["status"]).strip()
    execute(
        "UPDATE letters SET title = ?, recipient = ?, body = ?, letterhead = ?, status = ?, updated_at = ? WHERE id = ?",
        (title, recipient, body, letterhead, status, now_iso(), letter_id),
    )
    flash("Letter updated.", "success")
    return redirect(url_for("letter_detail", letter_id=letter_id))


# ── External email queue helper ─────────────────────────────────
def queue_external_email(to: str, subject: str, body_html: str):
    """Queue an email for SendGrid delivery. Secretary use only."""
    execute(
        "INSERT INTO email_queue (to_address, subject, body_html, created_at) VALUES (?, ?, ?, ?)",
        (to, subject, body_html, now_iso()),
    )


def flush_email_queue() -> int:
    """Send queued emails via SendGrid. Returns count sent. Call from cron or on-demand."""
    if not SENDGRID_API_KEY:
        return 0
    today = date.today().isoformat()
    sent_today = query_one(
        "SELECT COUNT(*) AS c FROM email_queue WHERE status = 'sent' AND sent_at LIKE ?",
        (f"{today}%",),
    )
    remaining = SENDGRID_DAILY_LIMIT - (sent_today["c"] if sent_today else 0)
    if remaining <= 0:
        return 0
    queued = query_all(
        "SELECT * FROM email_queue WHERE status = 'queued' ORDER BY created_at LIMIT ?",
        (remaining,),
    )
    sent = 0
    for msg in queued:
        try:
            import urllib.request
            payload = json.dumps({
                "personalizations": [{"to": [{"email": msg["to_address"]}]}],
                "from": {"email": SENDGRID_FROM},
                "subject": msg["subject"],
                "content": [{"type": "text/html", "value": msg["body_html"]}],
            }).encode()
            req = urllib.request.Request(
                "https://api.sendgrid.com/v3/mail/send",
                data=payload,
                headers={
                    "Authorization": f"Bearer {SENDGRID_API_KEY}",
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            urllib.request.urlopen(req, timeout=10)
            execute("UPDATE email_queue SET status = 'sent', sent_at = ? WHERE id = ?",
                    (now_iso(), msg["id"]))
            sent += 1
        except Exception as exc:
            execute("UPDATE email_queue SET attempts = attempts + 1 WHERE id = ?", (msg["id"],))
            app.logger.warning(f"SendGrid send failed for queue#{msg['id']}: {exc}")
    return sent


@app.route("/instruments", methods=["GET", "POST"])
@login_required
def instruments():
    user = current_user()
    if not user_access_profile(user)["can_access_instruments"]:
        abort(403)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_add_instrument = is_owner(user) or user["role"] in {"super_admin", "site_admin"}
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "create_instrument":
            if not can_add_instrument:
                abort(403)
            name = request.form.get("new_name", "").strip()
            category = request.form.get("new_category", "").strip()
            location = request.form.get("new_location", "").strip()
            if not name or not category or not location:
                flash("Name, category, and location are required to add an instrument.", "error")
                return redirect(url_for("instruments"))
            code = request.form.get("new_code", "").strip() or next_instrument_code()
            daily_capacity = max(1, min(50, int(request.form.get("new_daily_capacity", "3") or 3)))
            execute(
                """
                INSERT INTO instruments (name, code, category, location, daily_capacity, status, notes, office_info, faculty_group, manufacturer, model_number, capabilities_summary, machine_photo_url, reference_links, instrument_description, accepting_requests, soft_accept_enabled)
                VALUES (?, ?, ?, ?, ?, 'active', '', '', '', '', '', '', '', '', '', 1, 0)
                """,
                (name, code, category, location, daily_capacity),
            )
            new_instrument = query_one("SELECT id FROM instruments WHERE code = ?", (code,))
            if new_instrument:
                log_action(user["id"], "instrument", new_instrument["id"], "instrument_created", {"name": name, "code": code})
                flash(f"{name} added to the instrument list.", "success")
                return redirect(url_for("instrument_detail", instrument_id=new_instrument["id"]))
            flash("Instrument was created, but could not be reopened immediately.", "success")
            return redirect(url_for("instruments"))
        abort(400)
    instrument_filter = ""
    params: list = []
    if not user_access_profile(user)["can_view_all_instruments"]:
        ids = assigned_instrument_ids(user)
        if not ids:
            rows = []
            return render_template("instruments.html", instruments=rows, visible_links={})
        placeholders = ",".join("?" for _ in ids)
        instrument_filter = f"WHERE i.id IN ({placeholders})"
        params.extend(ids)
    rows = query_all(
        f"""
        SELECT i.*,
               GROUP_CONCAT(DISTINCT a.name) AS admins,
               GROUP_CONCAT(DISTINCT o.name) AS operators,
               GROUP_CONCAT(DISTINCT f.name) AS faculty_in_charge
        FROM instruments i
        LEFT JOIN instrument_admins ia ON ia.instrument_id = i.id
        LEFT JOIN users a ON a.id = ia.user_id
        LEFT JOIN instrument_operators io ON io.instrument_id = i.id
        LEFT JOIN users o ON o.id = io.user_id
        LEFT JOIN instrument_faculty_admins ifa ON ifa.instrument_id = i.id
        LEFT JOIN users f ON f.id = ifa.user_id
        {instrument_filter}
        GROUP BY i.id
        ORDER BY COALESCE(i.category, ''), i.name
        """,
        tuple(params),
    )
    turnaround_filter = ""
    if instrument_filter:
        turnaround_filter = instrument_filter.replace("WHERE i.id", "WHERE sr.instrument_id")
    turnaround_rows = query_all(
        f"""
        SELECT
            sr.instrument_id,
            AVG(
                julianday(sr.completed_at) - julianday(
                    COALESCE(sr.sample_received_at, sr.sample_submitted_at, sr.created_at)
                )
            ) * 24.0 AS avg_return_hours
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        {turnaround_filter}
        {' AND ' if turnaround_filter else 'WHERE '}sr.status = 'completed'
          AND sr.completed_at IS NOT NULL
          AND COALESCE(sr.sample_received_at, sr.sample_submitted_at, sr.created_at) IS NOT NULL
        GROUP BY sr.instrument_id
        """,
        tuple(params),
    )
    turnaround_map = {row["instrument_id"]: row["avg_return_hours"] for row in turnaround_rows}
    active_instruments = [row for row in rows if row["status"] == "active"]
    archived_instruments = [row for row in rows if row["status"] == "archived"]
    visible_links = {}
    for instrument in rows:
        visible_links[instrument["id"]] = can_open_instrument_detail(user, instrument["id"])
    return render_template(
        "instruments.html",
        instruments=active_instruments,
        archived_instruments=archived_instruments,
        visible_links=visible_links,
        turnaround_map=turnaround_map,
        can_add_instrument=can_add_instrument,
        suggested_new_code=next_instrument_code(),
        can_view_archived=bool(user_access_profile(user)["can_access_instruments"]),
    )


@app.route("/instruments/<int:instrument_id>", methods=["GET", "POST"])
@login_required
def instrument_detail(instrument_id: int):
    user = current_user()
    if not user_access_profile(user)["can_access_instruments"]:
        abort(403)
    if not can_open_instrument_detail(user, instrument_id):
        abort(403)

    instrument = query_one(
        """
        SELECT i.*,
               GROUP_CONCAT(DISTINCT a.name) AS admins,
               GROUP_CONCAT(DISTINCT o.name) AS operators,
               GROUP_CONCAT(DISTINCT f.name) AS faculty_in_charge
        FROM instruments i
        LEFT JOIN instrument_admins ia ON ia.instrument_id = i.id
        LEFT JOIN users a ON a.id = ia.user_id
        LEFT JOIN instrument_operators io ON io.instrument_id = i.id
        LEFT JOIN users o ON o.id = io.user_id
        LEFT JOIN instrument_faculty_admins ifa ON ifa.instrument_id = i.id
        LEFT JOIN users f ON f.id = ifa.user_id
        WHERE i.id = ?
        GROUP BY i.id
        """,
        (instrument_id,),
    )
    if instrument is None:
        abort(404)

    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_edit = user["role"] in {"super_admin", "site_admin"} or can_manage_instrument(user["id"], instrument_id, user["role"])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_edit_assignments = is_owner(user) or user["role"] in {"super_admin", "site_admin"}
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_archive_instrument = user["role"] == "super_admin"
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_restore_instrument = user["role"] == "super_admin"
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_add_downtime_here = user["role"] == "super_admin" or can_manage_instrument(user["id"], instrument_id, user["role"])
    if request.method == "POST":
        action = request.form.get("action", "update_metadata")
        if action == "update_metadata":
            if not can_edit:
                abort(403)
            code = request.form.get("code", instrument["code"]).strip() or instrument["code"]
            category = request.form.get("category", instrument["category"] or "").strip()
            location = request.form.get("location", instrument["location"]).strip() or instrument["location"]
            manufacturer = request.form.get("manufacturer", instrument["manufacturer"]).strip()
            model_number = request.form.get("model_number", instrument["model_number"]).strip()
            machine_photo_url = instrument["machine_photo_url"]
            uploaded_photo = request.files.get("machine_photo_file")
            if uploaded_photo and (uploaded_photo.filename or "").strip():
                try:
                    machine_photo_url = save_instrument_image(instrument_id, uploaded_photo)
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("instrument_detail", instrument_id=instrument_id))
            link_values = []
            for idx in range(1, 6):
                value = request.form.get(f"reference_link_{idx}", "").strip()
                if value:
                    link_values.append(value)
            reference_links = instrument["reference_links"]
            notes = request.form.get("notes", instrument["notes"]).strip()
            prior_accepting = int(instrument["accepting_requests"] or 0)
            intake_mode = request.form.get("intake_mode", "accepting").strip()
            accepting_requests, soft_accept_enabled = intake_mode_flags(intake_mode)
            if can_edit_assignments:
                instrument_admin_ids = [row["user_id"] for row in query_all("SELECT user_id FROM instrument_admins WHERE instrument_id = ?", (instrument_id,))]
                operator_ids = [int(value) for value in request.form.getlist("operator_ids") if value.strip()]
                faculty_admin_ids = [int(value) for value in request.form.getlist("faculty_admin_ids") if value.strip()]
            else:
                instrument_admin_ids = [row["user_id"] for row in query_all("SELECT user_id FROM instrument_admins WHERE instrument_id = ?", (instrument_id,))]
                operator_ids = [row["user_id"] for row in query_all("SELECT user_id FROM instrument_operators WHERE instrument_id = ?", (instrument_id,))]
                faculty_admin_ids = [row["user_id"] for row in query_all("SELECT user_id FROM instrument_faculty_admins WHERE instrument_id = ?", (instrument_id,))]
            execute(
                """
                UPDATE instruments
                SET code = ?, category = ?, location = ?, manufacturer = ?, model_number = ?, machine_photo_url = ?, reference_links = ?, notes = ?, accepting_requests = ?, soft_accept_enabled = ?
                WHERE id = ?
                """,
                (code, category, location, manufacturer, model_number, machine_photo_url, reference_links, notes, accepting_requests, soft_accept_enabled, instrument_id),
            )
            log_action(
                user["id"],
                "instrument",
                instrument_id,
                "instrument_metadata_updated",
                {
                    "code": code,
                    "location": location,
                    "manufacturer": manufacturer,
                    "model_number": model_number,
                    "machine_photo_url": machine_photo_url,
                    "reference_links": reference_links,
                    "notes": notes,
                    "intake_mode": intake_mode,
                },
            )
            if can_edit_assignments:
                sync_instrument_assignments("instrument_admins", instrument_id, instrument_admin_ids)
                sync_instrument_assignments("instrument_operators", instrument_id, operator_ids)
                sync_instrument_assignments("instrument_faculty_admins", instrument_id, faculty_admin_ids)
            released_count = 0
            if not prior_accepting and accepting_requests:
                released_count = release_submitted_requests_for_instrument(instrument_id, user["id"])
            flash(
                "Instrument page updated." + (f" Released {released_count} queued request(s) into review." if released_count else ""),
                "success",
            )
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        if action == "update_operation":
            if not can_edit:
                abort(403)
            prior_accepting = int(instrument["accepting_requests"] or 0)
            intake_mode = request.form.get("intake_mode", "accepting").strip()
            event_at = datetime.utcnow().isoformat(timespec="seconds")
            accepting_requests, soft_accept_enabled = intake_mode_flags(intake_mode)
            execute(
                "UPDATE instruments SET accepting_requests = ?, soft_accept_enabled = ? WHERE id = ?",
                (accepting_requests, soft_accept_enabled, instrument_id),
            )
            released_count = 0
            if not prior_accepting and accepting_requests:
                released_count = release_submitted_requests_for_instrument(instrument_id, user["id"])
            log_action(
                user["id"],
                "instrument",
                instrument_id,
                "instrument_operation_updated",
                {"intake_mode": intake_mode},
            )
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify(
                    {
                        "ok": True,
                        "intake_mode": intake_mode,
                        "label": intake_mode_label(intake_mode),
                        "released_count": released_count,
                        "actor_name": user["name"],
                        "event_title": "Operation changed",
                        "event_detail": intake_mode_label(intake_mode),
                        "event_at": format_dt(event_at),
                    }
                )
            flash(
                "Instrument operation updated." + (f" Released {released_count} queued request(s) into review." if released_count else ""),
                "success",
            )
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        if action == "archive_instrument":
            if not can_archive_instrument:
                abort(403)
            confirm_code = request.form.get("confirm_code", "").strip()
            if confirm_code != instrument["code"]:
                flash("Enter the exact instrument code to archive this instrument.", "error")
                return redirect(url_for("instrument_detail", instrument_id=instrument_id))
            execute("UPDATE instruments SET status = 'archived' WHERE id = ?", (instrument_id,))
            log_action(user["id"], "instrument", instrument_id, "instrument_archived", {"name": instrument["name"], "code": instrument["code"]})
            flash("Instrument archived.", "success")
            return redirect(url_for("instruments"))
        if action == "restore_instrument":
            if not can_restore_instrument:
                abort(403)
            execute("UPDATE instruments SET status = 'active' WHERE id = ?", (instrument_id,))
            log_action(user["id"], "instrument", instrument_id, "instrument_restored", {"name": instrument["name"], "code": instrument["code"]})
            flash("Instrument restored to active inventory.", "success")
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        if action == "add_downtime":
            if not can_add_downtime_here:
                abort(403)
            start_time = request.form.get("start_time", "").strip()
            end_time = request.form.get("end_time", "").strip()
            reason = request.form.get("reason", "").strip()
            if not start_time or not end_time or end_time <= start_time:
                flash("Downtime end must be after start.", "error")
                return redirect(url_for("instrument_detail", instrument_id=instrument_id))
            execute(
                """
                INSERT INTO instrument_downtime (instrument_id, start_time, end_time, reason, created_by_user_id, created_at, is_active)
                VALUES (?, ?, ?, ?, ?, ?, 1)
                """,
                (instrument_id, start_time, end_time, reason, user["id"], now_iso()),
            )
            log_action(user["id"], "instrument", instrument_id, "downtime_added", {"start_time": start_time, "end_time": end_time, "reason": reason})
            flash("Downtime block added.", "success")
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        if action == "save_approval_config":
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if not (is_owner(user) or user["role"] == "super_admin"):
                abort(403)
            db = get_db()
            # Preserve notify_submitter settings before deleting rows
            old_notify = {}
            for row in db.execute(
                "SELECT step_order, approver_role, notify_submitter FROM instrument_approval_config WHERE instrument_id = ?",
                (instrument_id,),
            ).fetchall():
                old_notify[(row["step_order"], row["approver_role"])] = row["notify_submitter"]
            db.execute("DELETE FROM instrument_approval_config WHERE instrument_id = ?", (instrument_id,))
            step_order = 1
            valid_roles = {"finance", "professor", "operator"}
            for idx in range(1, 7):
                role = request.form.get(f"step_role_{idx}", "").strip()
                if not role or role not in valid_roles:
                    continue
                user_id_raw = request.form.get(f"step_user_{idx}", "").strip()
                approver_user_id = int(user_id_raw) if user_id_raw else None
                notify = old_notify.get((step_order, role), 0)
                db.execute(
                    "INSERT INTO instrument_approval_config (instrument_id, step_order, approver_role, approver_user_id, notify_submitter) VALUES (?, ?, ?, ?, ?)",
                    (instrument_id, step_order, role, approver_user_id, notify),
                )
                step_order += 1
            db.commit()
            log_action(user["id"], "instrument", instrument_id, "approval_config_updated", {"step_count": step_order - 1})
            flash("Approval sequence saved.", "success")
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        if action == "add_inventory_item":
            if not can_edit:
                abort(403)
            item_name = request.form.get("item_name", "").strip()
            inv_category = request.form.get("inv_category", "consumable").strip()
            quantity = int(request.form.get("quantity", 0) or 0)
            minimum_quantity = int(request.form.get("minimum_quantity", 0) or 0)
            unit = request.form.get("unit", "units").strip() or "units"
            unit_cost = float(request.form.get("unit_cost", 0) or 0)
            inv_notes = request.form.get("inv_notes", "").strip()
            if not item_name:
                flash("Item name is required.", "error")
                return redirect(url_for("instrument_detail", instrument_id=instrument_id))
            execute(
                """
                INSERT INTO instrument_inventory
                    (instrument_id, item_name, category, quantity, minimum_quantity, unit, unit_cost, last_restocked_at, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (instrument_id, item_name, inv_category, quantity, minimum_quantity, unit, unit_cost, now_iso(), inv_notes),
            )
            log_action(user["id"], "instrument", instrument_id, "inventory_item_added", {"item_name": item_name, "quantity": quantity})
            flash(f"Inventory item '{item_name}' added.", "success")
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        if action == "update_finance":
            if not can_edit:
                abort(403)
            grant_id = request.form.get("default_grant_id", "").strip()
            execute("UPDATE instruments SET default_grant_id = ? WHERE id = ?",
                    (int(grant_id) if grant_id else None, instrument_id))
            log_action(user["id"], "instrument", instrument_id, "default_grant_updated", {"grant_id": int(grant_id) if grant_id else None})
            flash("Default grant updated.", "success")
            return redirect(url_for("instrument_detail", instrument_id=instrument_id))
        abort(400)

    queue_sql, queue_params = request_history_query(["sr.instrument_id = ?"], [instrument_id], {})
    instrument_rows_all = query_all(queue_sql, tuple(queue_params))
    queue_rank = {
        "under_review": 1,
        "sample_submitted": 2,
        "sample_received": 3,
        "scheduled": 4,
        "in_progress": 5,
        "submitted": 6,
        "awaiting_sample_submission": 7,
        "completed": 8,
        "rejected": 9,
    }

    def instrument_queue_sort_key(row: sqlite3.Row) -> tuple:
        anchor = (
            row["sample_submitted_at"]
            or row["sample_received_at"]
            or row["scheduled_for"]
            or row["completed_at"]
            or row["created_at"]
            or ""
        )
        return (queue_rank.get(row["status"], 99), str(anchor), row["id"])

    instrument_queue = {
        "submitted": [row for row in instrument_rows_all if row["status"] == "submitted"],
        "approvals": [row for row in instrument_rows_all if row["status"] == "under_review"],
        "awaiting_sample": [row for row in instrument_rows_all if row["status"] == "awaiting_sample_submission"],
        "pending_receipt": [row for row in instrument_rows_all if row["status"] == "sample_submitted"],
        "ready": [row for row in instrument_rows_all if row["status"] == "sample_received"],
        "active": [row for row in instrument_rows_all if row["status"] in {"scheduled", "in_progress"}],
        "done": [row for row in instrument_rows_all if row["status"] == "completed"],
        "closed": [row for row in instrument_rows_all if row["status"] in {"rejected"}],
    }
    queue_preview_rows = sorted(
        [
            row
            for row in instrument_rows_all
            if row["status"] in {"under_review", "sample_submitted", "sample_received", "scheduled", "in_progress", "completed"}
        ],
        key=instrument_queue_sort_key,
    )[:7]
    instrument_queue_counts = {key: len(value) for key, value in instrument_queue.items()}
    pending_count = sum(
        instrument_queue_counts[key]
        for key in ("submitted", "approvals", "awaiting_sample", "pending_receipt", "ready", "active")
    )
    today = datetime.utcnow().date()
    last_week_end = today - timedelta(days=today.weekday() + 1)
    last_week_start = last_week_end - timedelta(days=6)
    processed_last_week = sum(
        1
        for row in instrument_queue["done"]
        if row["completed_at"]
        and last_week_start <= parse_date_param(str(row["completed_at"])[:10]) <= last_week_end
    )
    samples_last_week = sum(
        row["sample_count"] or 0
        for row in instrument_queue["done"]
        if row["completed_at"]
        and last_week_start <= parse_date_param(str(row["completed_at"])[:10]) <= last_week_end
    )
    weekly_rows = query_all(
        """
        SELECT substr(completed_at, 1, 10) AS day_bucket, COUNT(*) AS jobs, COALESCE(SUM(sample_count), 0) AS samples
        FROM sample_requests
        WHERE instrument_id = ? AND status = 'completed' AND completed_at IS NOT NULL
        ORDER BY completed_at
        """,
        (instrument_id,),
    )
    jobs_by_week: dict[tuple[int, int], int] = {}
    samples_by_week: dict[tuple[int, int], int] = {}
    for row in weekly_rows:
        bucket_date = parse_date_param(row["day_bucket"])
        if bucket_date is None:
            continue
        iso_year, iso_week, _ = bucket_date.isocalendar()
        key = (iso_year, iso_week)
        jobs_by_week[key] = jobs_by_week.get(key, 0) + int(row["jobs"] or 0)
        samples_by_week[key] = samples_by_week.get(key, 0) + int(row["samples"] or 0)
    average_jobs_per_week = round(sum(jobs_by_week.values()) / len(jobs_by_week), 1) if jobs_by_week else 0
    average_samples_per_week = round(sum(samples_by_week.values()) / len(samples_by_week), 1) if samples_by_week else 0
    turnaround_avg_row = query_one(
        """
        SELECT AVG(
            julianday(completed_at) - julianday(COALESCE(sample_received_at, sample_submitted_at, created_at))
        ) * 24.0 AS avg_return_hours
        FROM sample_requests
        WHERE instrument_id = ?
          AND status = 'completed'
          AND completed_at IS NOT NULL
          AND COALESCE(sample_received_at, sample_submitted_at, created_at) IS NOT NULL
        """,
        (instrument_id,),
    )
    queue_metrics = {
        "pending": pending_count,
        "processed_last_week": processed_last_week,
        "samples_last_week": samples_last_week,
        "average_jobs_per_week": average_jobs_per_week,
        "average_samples_per_week": average_samples_per_week,
        "average_return_hours": turnaround_avg_row["avg_return_hours"] if turnaround_avg_row else None,
    }
    operator_candidates = query_all(
        "SELECT id, name, email, role FROM users WHERE active = 1 AND role IN ('operator', 'instrument_admin', 'site_admin', 'super_admin') ORDER BY name"
    )
    faculty_candidates = query_all(
        "SELECT id, name, email, role FROM users WHERE active = 1 AND role IN ('requester', 'professor_approver', 'site_admin', 'super_admin') ORDER BY name"
    )
    selected_operator_ids = {
        row["user_id"] for row in query_all("SELECT user_id FROM instrument_operators WHERE instrument_id = ?", (instrument_id,))
    }
    selected_faculty_ids = {
        row["user_id"] for row in query_all("SELECT user_id FROM instrument_faculty_admins WHERE instrument_id = ?", (instrument_id,))
    }
    selected_operator_rows = query_all(
        f"SELECT id, name, email FROM users WHERE id IN ({','.join('?' for _ in selected_operator_ids)}) ORDER BY name",
        tuple(sorted(selected_operator_ids)),
    ) if selected_operator_ids else []
    selected_faculty_rows = query_all(
        f"SELECT id, name, email FROM users WHERE id IN ({','.join('?' for _ in selected_faculty_ids)}) ORDER BY name",
        tuple(sorted(selected_faculty_ids)),
    ) if selected_faculty_ids else []
    # v2.2.0 — subscribed requesters
    selected_requester_ids = {
        row["user_id"] for row in query_all(
            "SELECT user_id FROM instrument_requesters WHERE instrument_id = ?",
            (instrument_id,),
        )
    }
    selected_requester_rows = query_all(
        f"SELECT id, name, email FROM users WHERE id IN ({','.join('?' for _ in selected_requester_ids)}) ORDER BY name",
        tuple(sorted(selected_requester_ids)),
    ) if selected_requester_ids else []
    instrument_logs = query_all(
        "SELECT al.*, u.name AS actor_name FROM audit_logs al LEFT JOIN users u ON u.id = al.actor_id WHERE entity_type = 'instrument' AND entity_id = ? ORDER BY al.id",
        (instrument_id,),
    )
    approval_config = query_all(
        """
        SELECT iac.*, u.name AS approver_name
        FROM instrument_approval_config iac
        LEFT JOIN users u ON u.id = iac.approver_user_id
        WHERE iac.instrument_id = ?
        ORDER BY iac.step_order
        """,
        (instrument_id,),
    )
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_edit_approval_config = is_owner(user) or user["role"] == "super_admin"
    approval_role_candidates = query_all(
        "SELECT id, name, role FROM users WHERE active = 1 AND role IN ('finance_admin', 'professor_approver', 'operator', 'instrument_admin', 'site_admin', 'super_admin') ORDER BY name"
    )
    # Upcoming downtime for this instrument
    upcoming_downtime = query_all(
        """
        SELECT idt.*, u.name AS creator_name
        FROM instrument_downtime idt
        LEFT JOIN users u ON u.id = idt.created_by_user_id
        WHERE idt.instrument_id = ? AND idt.is_active = 1
          AND idt.end_time >= datetime('now')
        ORDER BY idt.start_time ASC
        LIMIT 10
        """,
        (instrument_id,),
    )

    return render_template(
        "instrument_detail.html",
        instrument=instrument,
        can_edit=can_edit,
        can_edit_assignments=can_edit_assignments,
        can_archive_instrument=can_archive_instrument,
        can_restore_instrument=can_restore_instrument,
        can_add_downtime_here=can_add_downtime_here,
        instrument_queue=instrument_queue,
        instrument_queue_counts=instrument_queue_counts,
        queue_preview_rows=queue_preview_rows,
        queue_metrics=queue_metrics,
        can_view_history=can_view_instrument_history(user, instrument_id),
        operator_candidates=operator_candidates,
        faculty_candidates=faculty_candidates,
        selected_operator_ids=selected_operator_ids,
        selected_faculty_ids=selected_faculty_ids,
        selected_operator_rows=selected_operator_rows,
        selected_faculty_rows=selected_faculty_rows,
        selected_requester_rows=selected_requester_rows,
        instrument_timeline_entries=instrument_timeline_entries(instrument, instrument_logs),
        intake_mode=instrument_intake_mode(instrument),
        intake_mode_label=intake_mode_label,
        approval_config=approval_config,
        can_edit_approval_config=can_edit_approval_config,
        approval_role_candidates=approval_role_candidates,
        operators=query_all(
            "SELECT id, name FROM users WHERE role IN ('operator','instrument_admin','super_admin') ORDER BY name"
        ),
        upcoming_downtime=upcoming_downtime,
        inventory_items=query_all(
            "SELECT * FROM instrument_inventory WHERE instrument_id = ? ORDER BY item_name",
            (instrument_id,),
        ),
        grants=query_all("SELECT id, code, name FROM grants WHERE status = 'active' ORDER BY name"),
    )


@app.route("/instruments/<int:instrument_id>/form-control", methods=["GET", "POST"])
@login_required
def instrument_form_control(instrument_id: int):
    user = current_user()
    instrument = query_all("SELECT * FROM instruments WHERE id = ?", (instrument_id,))
    if not instrument:
        abort(404)
    instrument = instrument[0]
    if not (is_owner(user) or user["role"] == "super_admin"):
        abort(403)

    if request.method == "POST":
        action = request.form.get("action", "")
        db = get_db()

        if action == "save_custom_fields":
            db.execute("DELETE FROM instrument_custom_fields WHERE instrument_id = ?", (instrument_id,))
            order = 0
            for idx in range(1, 21):
                label = request.form.get(f"field_label_{idx}", "").strip()
                if not label:
                    continue
                ftype = request.form.get(f"field_type_{idx}", "text").strip()
                if ftype not in ("text", "number", "select", "file"):
                    ftype = "text"
                freq = 1 if request.form.get(f"field_required_{idx}") else 0
                opts = request.form.get(f"field_options_{idx}", "").strip()
                order += 1
                db.execute(
                    "INSERT INTO instrument_custom_fields (instrument_id, field_label, field_type, field_options, is_required, display_order, is_active) VALUES (?, ?, ?, ?, ?, ?, 1)",
                    (instrument_id, label, ftype, opts, freq, order),
                )
            db.commit()
            log_action(user["id"], "instrument", instrument_id, "custom_fields_updated", {"field_count": order})
            flash("Custom fields saved.", "success")
            return redirect(url_for("instrument_form_control", instrument_id=instrument_id))

        if action == "save_approval_sequence":
            db.execute("DELETE FROM instrument_approval_config WHERE instrument_id = ?", (instrument_id,))
            valid_roles = {"finance", "professor", "operator"}
            step_order = 1
            for idx in range(1, 7):
                role = request.form.get(f"step_role_{idx}", "").strip()
                if not role or role not in valid_roles:
                    continue
                user_id_raw = request.form.get(f"step_user_{idx}", "").strip()
                approver_user_id = int(user_id_raw) if user_id_raw else None
                notify = 1 if request.form.get(f"notify_submitter_{idx}") else 0
                db.execute(
                    "INSERT INTO instrument_approval_config (instrument_id, step_order, approver_role, approver_user_id, notify_submitter) VALUES (?, ?, ?, ?, ?)",
                    (instrument_id, step_order, role, approver_user_id, notify),
                )
                step_order += 1
            db.commit()
            log_action(user["id"], "instrument", instrument_id, "approval_sequence_updated", {"step_count": step_order - 1})
            flash("Approval sequence saved.", "success")
            return redirect(url_for("instrument_form_control", instrument_id=instrument_id))

        if action == "save_pricing":
            price = request.form.get("price_per_sample", "").strip()
            instructions = request.form.get("payment_instructions", "").strip()
            proof_note = request.form.get("payment_proof_note", "").strip()
            db.execute(
                "UPDATE instruments SET price_per_sample = ?, payment_instructions = ?, payment_proof_note = ? WHERE id = ?",
                (price, instructions, proof_note, instrument_id),
            )
            db.commit()
            log_action(user["id"], "instrument", instrument_id, "pricing_updated", {
                "price_per_sample": price,
            })
            flash("Pricing & payment instructions saved.", "success")
            return redirect(url_for("instrument_form_control", instrument_id=instrument_id))

        if action == "save_email_templates":
            for evt in ("request_submitted", "step_approved", "request_completed"):
                subj = request.form.get(f"tmpl_subject_{evt}", "").strip()
                body_val = request.form.get(f"tmpl_body_{evt}", "").strip()
                db.execute(
                    """INSERT INTO instrument_email_templates (instrument_id, event_type, subject_template, body_template)
                       VALUES (?, ?, ?, ?)
                       ON CONFLICT(instrument_id, event_type) DO UPDATE
                       SET subject_template = excluded.subject_template, body_template = excluded.body_template""",
                    (instrument_id, evt, subj, body_val),
                )
            db.commit()
            log_action(user["id"], "instrument", instrument_id, "email_templates_updated", {})
            flash("Email templates saved.", "success")
            return redirect(url_for("instrument_form_control", instrument_id=instrument_id))

        abort(400)

    approval_config = query_all(
        """
        SELECT iac.*, u.name AS approver_name
        FROM instrument_approval_config iac
        LEFT JOIN users u ON u.id = iac.approver_user_id
        WHERE iac.instrument_id = ?
        ORDER BY iac.step_order
        """,
        (instrument_id,),
    )
    custom_fields = query_all(
        "SELECT * FROM instrument_custom_fields WHERE instrument_id = ? ORDER BY display_order",
        (instrument_id,),
    )
    approval_role_candidates = query_all(
        "SELECT id, name, role FROM users WHERE active = 1 AND role IN ('finance_admin', 'professor_approver', 'operator', 'instrument_admin', 'site_admin', 'super_admin') ORDER BY name"
    )
    email_templates_rows = query_all(
        "SELECT event_type, subject_template, body_template FROM instrument_email_templates WHERE instrument_id = ?",
        (instrument_id,),
    )
    email_templates = {r["event_type"]: r for r in email_templates_rows}
    return render_template(
        "instrument_form_control.html",
        instrument=instrument,
        approval_config=approval_config,
        custom_fields=custom_fields,
        approval_role_candidates=approval_role_candidates,
        intake_mode=instrument_intake_mode(instrument),
        intake_mode_label=intake_mode_label,
        email_templates=email_templates,
    )


@app.route("/instruments/<int:instrument_id>/custom-fields")
@login_required
def instrument_custom_fields_json(instrument_id: int):
    """Return custom fields + pricing info for an instrument as JSON.

    Consumed by new_request.html to render dynamic custom fields and
    show payment instructions before the finance section.
    """
    rows = query_all(
        "SELECT id, field_label, field_type, is_required, field_options FROM instrument_custom_fields WHERE instrument_id = ? AND is_active = 1 ORDER BY display_order",
        (instrument_id,),
    )
    inst = query_one(
        "SELECT price_per_sample, payment_instructions, payment_proof_note FROM instruments WHERE id = ?",
        (instrument_id,),
    )
    return jsonify({
        "fields": [
            {"id": r["id"], "label": r["field_label"], "type": r["field_type"], "required": bool(r["is_required"]), "options": r["field_options"]}
            for r in rows
        ],
        "pricing": {
            "price_per_sample": inst["price_per_sample"] if inst else "",
            "payment_instructions": inst["payment_instructions"] if inst else "",
            "payment_proof_note": inst["payment_proof_note"] if inst else "",
        } if inst else None,
    })


@app.route("/requests/new", methods=["GET", "POST"])
@login_required
def new_request():
    user = current_user()
    instruments = query_all("SELECT * FROM instruments WHERE status = 'active' ORDER BY name")
    can_submit_for_others = can_manage_members(user) or has_instrument_area_access(user)
    requester_candidates = query_all(
        """
        SELECT id, name, email, role
        FROM users
        WHERE active = 1 AND invite_status = 'active'
        ORDER BY name, email
        """
    ) if can_submit_for_others else []
    if request.method == "POST":
        # --- Validate required fields gracefully (return 400, not 500) ---
        instrument_id_raw = request.form.get("instrument_id")
        if not instrument_id_raw:
            flash("Please select an instrument.", "error")
            return redirect(url_for("new_request"))
        try:
            instrument_id = int(instrument_id_raw)
        except (ValueError, TypeError):
            flash("Invalid instrument selection.", "error")
            return redirect(url_for("new_request"))
        instrument = query_one("SELECT * FROM instruments WHERE id = ? AND status = 'active'", (instrument_id,))
        if instrument is None:
            flash("Selected instrument is not available.", "error")
            return redirect(url_for("new_request"))
        title = request.form.get("title", "").strip()
        if not title:
            flash("Title is required.", "error")
            return redirect(url_for("new_request"))
        sample_name = request.form.get("sample_name", "").strip()
        if not sample_name:
            flash("Sample name is required.", "error")
            return redirect(url_for("new_request"))
        sample_count_raw = request.form.get("sample_count")
        if not sample_count_raw:
            flash("Sample count is required.", "error")
            return redirect(url_for("new_request"))
        try:
            sample_count = int(sample_count_raw)
        except (ValueError, TypeError):
            flash("Sample count must be a number.", "error")
            return redirect(url_for("new_request"))
        if sample_count < 1 or sample_count > 99:
            flash("Sample count must be between 1 and 99.", "error")
            return redirect(url_for("new_request"))
        description = request.form.get("description", "").strip()
        sample_origin = request.form.get("sample_origin", "internal")
        receipt_number = request.form.get("receipt_number", "").strip()
        if not receipt_number:
            receipt_number = generate_receipt_reference(sample_origin)
        amount_due = float(request.form.get("amount_due") or 0)
        amount_paid = float(request.form.get("amount_paid") or 0)
        finance_status = request.form.get("finance_status", "n/a")
        priority = request.form.get("priority", "normal")
        requester_id = user["id"]
        if can_submit_for_others and request.form.get("requester_id"):
            requester_id = int(request.form["requester_id"])
        requester_row = query_one(
            "SELECT * FROM users WHERE id = ? AND active = 1 AND invite_status = 'active'",
            (requester_id,),
        )
        if requester_row is None:
            flash("Selected requester is not available.", "error")
            return redirect(url_for("new_request"))
        originator_note = request.form.get("originator_note", "").strip()
        request_no = generate_job_reference(sample_origin)
        sample_ref = generate_sample_reference(instrument["name"], sample_origin)
        created = now_iso()
        initial_status = "submitted"
        # v2.0.0 — INSERT skips legacy finance columns (dropped in
        # v2.0 migration). Billing lands in peer aggregates via
        # sync_request_to_peer_aggregates() below.
        request_id = execute(
            """
            INSERT INTO sample_requests
            (request_no, sample_ref, requester_id, created_by_user_id, originator_note, instrument_id, title, sample_name, sample_count, description, sample_origin,
             priority, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                request_no,
                sample_ref,
                requester_id,
                user["id"],
                originator_note,
                instrument_id,
                title,
                sample_name,
                sample_count,
                description,
                sample_origin,
                priority,
                initial_status,
                created,
                created,
            ),
        )
        # Save custom field values for this instrument
        custom_fields = query_all(
            "SELECT id FROM instrument_custom_fields WHERE instrument_id = ? AND is_active = 1 ORDER BY display_order",
            (instrument_id,),
        )
        for cf in custom_fields:
            cf_val = request.form.get(f"custom_{cf['id']}", "").strip()
            if cf_val:
                execute(
                    "INSERT INTO request_custom_field_values (request_id, custom_field_id, field_value) VALUES (?, ?, ?)",
                    (request_id, cf["id"], cf_val),
                )
        if instrument["accepting_requests"]:
            create_approval_chain(get_db(), request_id, instrument_id)
            # Just inserted with `initial_status='submitted'` above; transition is always submitted -> under_review.
            assert_status_transition("submitted", "under_review")
            execute("UPDATE sample_requests SET status = 'under_review' WHERE id = ?", (request_id,))
        else:
            execute(
                "UPDATE sample_requests SET remarks = ?, updated_at = ? WHERE id = ?",
                ("Lab is not currently accepting new jobs. Your request is queued and will be released when intake opens.", now_iso(), request_id),
            )
        # v2.0.0 — billing goes directly into peer aggregates. The
        # INSERT above omitted the legacy finance columns; this call
        # creates the Project + Invoice + Payment rows if the form
        # submitted non-zero values.
        sync_request_to_peer_aggregates(
            get_db(),
            request_id,
            amount_due=amount_due,
            amount_paid=amount_paid,
            finance_status=finance_status,
            receipt_number=receipt_number,
        )
        # Auto-link to instrument's default grant if set
        default_grant = row_value(instrument, "default_grant_id", None)
        if default_grant:
            invoice = query_one("SELECT id FROM invoices WHERE request_id = ?", (request_id,))
            if invoice:
                execute("UPDATE invoices SET grant_id = ? WHERE id = ?", (default_grant, invoice["id"]))
        get_db().commit()
        created_request = query_one(
            """
            SELECT sr.id, sr.requester_id, sr.request_no, sr.sample_ref, sr.instrument_id, sr.sample_name, sr.sample_count, sr.created_at,
                   i.name AS instrument_name, u.name AS requester_name
            FROM sample_requests sr
            JOIN instruments i ON i.id = sr.instrument_id
            JOIN users u ON u.id = sr.requester_id
            WHERE sr.id = ?
            """,
            (request_id,),
        )
        if created_request is not None:
            ensure_request_folder(created_request)
            initial_attachment_name = None
            write_request_metadata_snapshot(request_id)
            initial_attachment = request.files.get("initial_attachment")
            if initial_attachment and initial_attachment.filename:
                try:
                    save_uploaded_attachment(created_request, initial_attachment, user["id"], "request_document", "Uploaded with new request")
                    initial_attachment_name = initial_attachment.filename
                except ValueError as exc:
                    flash(f"Request created, but initial attachment was not added: {exc}", "error")
            slip_pdf = generate_sample_slip_pdf(
                created_request,
                created_request["instrument_name"],
                created_request["requester_name"],
                initial_attachment_name,
            )
            save_generated_attachment(
                created_request,
                f"{created_request['sample_ref']}_sample-slip.pdf",
                slip_pdf,
                user["id"],
                "sample_slip",
                "Auto-generated printable slip for attaching to the physical sample",
            )
            write_request_metadata_snapshot(request_id)
        log_action(
            user["id"],
            "sample_request",
            request_id,
            "submitted",
            {
                "request_no": request_no,
                "instrument_id": instrument_id,
                "requester_id": requester_id,
                "sample_origin": sample_origin,
                "sample_count": sample_count,
                "created_by_user_id": user["id"],
                "originator_note": originator_note,
                "accepting_requests": bool(instrument["accepting_requests"]),
            },
        )
        # ── Workflow notification: new request submitted ──
        try:
            ops = query_all(
                "SELECT user_id FROM instrument_operators WHERE instrument_id = ?",
                (instrument_id,),
            )
            _req_url = url_for("request_detail", request_id=request_id)
            for op in ops:
                notify(op["user_id"], "request",
                       f"New request: {title}",
                       f"Request {request_no} submitted for {instrument['name']}",
                       _req_url, "sample_request", request_id)
        except Exception:
            pass  # notification failure must not block request creation
        if instrument["accepting_requests"]:
            flash(f"Request {request_no} submitted for {requester_row['name']}. Sample number {sample_ref} and printable slip generated.", "success")
        else:
            flash(f"Request {request_no} submitted for {requester_row['name']}. The lab is not accepting jobs yet, so it has been queued. Sample number {sample_ref} and printable slip generated.", "success")
        return redirect(url_for("request_detail", request_id=request_id))
    # Support pre-fill from duplicate_request route
    prefill = {
        "instrument_id": request.args.get("instrument_id", ""),
        "title": request.args.get("title", ""),
        "sample_name": request.args.get("sample_name", ""),
        "sample_count": request.args.get("sample_count", ""),
        "description": request.args.get("description", ""),
        "sample_origin": request.args.get("sample_origin", ""),
        "priority": request.args.get("priority", ""),
    }
    return render_template(
        "new_request.html",
        instruments=instruments,
        can_submit_for_others=can_submit_for_others,
        requester_candidates=requester_candidates,
        prefill=prefill,
    )


@app.route("/requests/<int:request_id>", methods=["GET", "POST"])
@login_required
def request_detail(request_id: int):
    user = current_user()
    sample_request = query_one(
        f"""
        SELECT sr.*, i.name AS instrument_name, i.daily_capacity, i.accepting_requests, i.soft_accept_enabled,
               r.name AS requester_name, r.email AS requester_email,
               c.name AS originator_name, c.email AS originator_email, c.role AS originator_role,
               op.name AS operator_name, recv.name AS received_by_name
        {REQUEST_DETAIL_JOINS}
        WHERE sr.id = ?
        """,
        (request_id,),
    )
    if sample_request is None:
        abort(404)

    if not can_view_request(user, sample_request):
        abort(403)

    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_manage = can_manage_instrument(user["id"], sample_request["instrument_id"], user["role"])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_operate = can_operate_instrument(user["id"], sample_request["instrument_id"], user["role"])
    can_upload_files = can_upload_attachment(user, sample_request)
    can_flag_issue = can_flag_request_issue(user, sample_request)
    can_respond_issue = can_respond_request_issue(user, sample_request)
    card_policy = request_card_policy(user, sample_request)
    approval_steps = query_all(
        """
        SELECT aps.*, u.name AS approver_name, u.email AS approver_email
        FROM approval_steps aps
        LEFT JOIN users u ON u.id = aps.approver_user_id
        WHERE aps.sample_request_id = ?
        ORDER BY aps.step_order
        """,
        (request_id,),
    )
    actionable_step_ids = {step["id"] for step in approval_steps if approval_step_is_actionable(step, approval_steps)}

    if request.method == "POST":
        action = request.form["action"]
        if action in {"save_note", "post_message"}:
            if not can_post_message(user, sample_request):
                abort(403)
            note_kind = request.form.get("note_kind", "").strip()
            if action == "post_message" and not note_kind:
                note_kind = "requester_note" if sample_request["requester_id"] == user["id"] else "operator_note"
            if note_kind not in {item[0] for item in COMMUNICATION_NOTE_TYPES}:
                flash("Invalid note type.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            if not can_edit_request_note(user, sample_request, note_kind):
                abort(403)
            message_body = request.form.get("note_body", request.form.get("message_body", "")).strip()
            uploaded_file = request.files.get("attachment")
            has_attachment = bool(uploaded_file and (uploaded_file.filename or "").strip())
            if not message_body and not has_attachment:
                flash("Reply cannot be empty.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            if action == "save_note":
                execute(
                    "UPDATE request_messages SET is_active = 0 WHERE request_id = ? AND note_kind = ?",
                    (request_id, note_kind),
                )
            message_id = execute(
                """
                INSERT INTO request_messages (request_id, sender_user_id, note_kind, message_body, created_at, is_active)
                VALUES (?, ?, ?, ?, ?, 1)
                """,
                (request_id, user["id"], note_kind, message_body, now_iso()),
            )
            if has_attachment:
                try:
                    save_uploaded_attachment(
                        sample_request,
                        uploaded_file,
                        user["id"],
                        "other",
                        "Attached in conversation",
                        request_message_id=message_id,
                    )
                except ValueError as exc:
                    execute("DELETE FROM request_messages WHERE id = ?", (message_id,))
                    flash(str(exc), "error")
                    return redirect(url_for("request_detail", request_id=request_id))
            write_request_metadata_snapshot(request_id)
            log_action(
                user["id"],
                "sample_request",
                request_id,
                "communication_note_saved",
                {"note_kind": note_kind, "message_preview": message_body[:120]},
            )
            flash("Reply added." if action == "post_message" else f"{note_kind_label(note_kind)} updated.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        if action == "flag_issue":
            if not can_flag_issue:
                abort(403)
            issue_message = request.form.get("issue_message", "").strip()
            if not issue_message:
                flash("Please describe the issue before flagging it.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            execute(
                """
                INSERT INTO request_issues (request_id, created_by_user_id, issue_message, created_at)
                VALUES (?, ?, ?, ?)
                """,
                (request_id, user["id"], issue_message, now_iso()),
            )
            write_request_metadata_snapshot(request_id)
            log_action(user["id"], "sample_request", request_id, "issue_flagged", {"issue_preview": issue_message[:160]})
            flash("Issue flagged for this sample.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        if action == "respond_issue":
            if not can_respond_issue:
                abort(403)
            issue_id = int(request.form.get("issue_id") or 0)
            issue = query_one("SELECT * FROM request_issues WHERE id = ? AND request_id = ?", (issue_id, request_id))
            if issue is None:
                abort(404)
            response_message = request.form.get("response_message", "").strip()
            if not response_message:
                flash("Please add a response before saving it.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            execute(
                """
                UPDATE request_issues
                SET response_message = ?, responded_at = ?, responded_by_user_id = ?
                WHERE id = ?
                """,
                (response_message, now_iso(), user["id"], issue_id),
            )
            write_request_metadata_snapshot(request_id)
            log_action(user["id"], "sample_request", request_id, "issue_response_saved", {"response_preview": response_message[:160]})
            flash("Issue response saved.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        if action == "resolve_issue":
            if not can_respond_issue:
                abort(403)
            issue_id = int(request.form.get("issue_id") or 0)
            issue = query_one("SELECT * FROM request_issues WHERE id = ? AND request_id = ?", (issue_id, request_id))
            if issue is None:
                abort(404)
            resolution_message = request.form.get("response_message", "").strip() or issue["response_message"] or "Resolved by lab."
            execute(
                """
                UPDATE request_issues
                SET response_message = ?, responded_at = COALESCE(responded_at, ?), responded_by_user_id = COALESCE(responded_by_user_id, ?),
                    status = 'resolved', resolved_at = ?, resolved_by_user_id = ?
                WHERE id = ?
                """,
                (resolution_message, now_iso(), user["id"], now_iso(), user["id"], issue_id),
            )
            write_request_metadata_snapshot(request_id)
            log_action(user["id"], "sample_request", request_id, "issue_resolved", {"resolution_preview": resolution_message[:160]})
            flash("Issue marked as resolved.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        if action == "reopen_issue":
            issue_id = int(request.form.get("issue_id") or 0)
            issue = query_one("SELECT * FROM request_issues WHERE id = ? AND request_id = ?", (issue_id, request_id))
            if issue is None:
                abort(404)
            if sample_request["requester_id"] != user["id"]:
                abort(403)
            execute(
                """
                UPDATE request_issues
                SET status = 'open', resolved_at = NULL, resolved_by_user_id = NULL
                WHERE id = ?
                """,
                (issue_id,),
            )
            write_request_metadata_snapshot(request_id)
            log_action(user["id"], "sample_request", request_id, "issue_reopened", {"issue_id": issue_id})
            flash("Issue reopened.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        if action == "upload_attachment":
            if not can_upload_files:
                abort(403)
            uploaded_file = request.files.get("attachment")
            attachment_type = request.form.get("attachment_type", "other")
            note = request.form.get("note", "").strip()
            if attachment_type not in attachment_type_choices():
                attachment_type = "other"
            try:
                save_uploaded_attachment(sample_request, uploaded_file, user["id"], attachment_type, note)
                write_request_metadata_snapshot(request_id)
                flash("Attachment uploaded.", "success")
            except ValueError as exc:
                flash(str(exc), "error")
            return redirect(url_for("request_detail", request_id=request_id))

        if action == "update_request_metadata" and can_manage:
            # In-place edit of user-editable request fields by lab admins.
            # Limited to fields a manager would legitimately correct: sample
            # title, sample name, sample count, and the working remark. All
            # other lifecycle / status moves go through admin_set_status.
            # Writes an audit-log event so the change lands in the request
            # timeline like any other admin action. Completed jobs are
            # blocked below by the completion_locked gate.
            if sample_request["completion_locked"]:
                flash("Completed jobs are locked and cannot be edited.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            new_title = (request.form.get("title") or sample_request["title"]).strip() or sample_request["title"]
            new_sample_name = (request.form.get("sample_name") or sample_request["sample_name"]).strip() or sample_request["sample_name"]
            try:
                new_sample_count = max(1, int(request.form.get("sample_count") or sample_request["sample_count"]))
            except (TypeError, ValueError):
                new_sample_count = sample_request["sample_count"]
            new_remarks = (request.form.get("remarks") or sample_request["remarks"] or "").strip()
            execute(
                """
                UPDATE sample_requests
                SET title = ?, sample_name = ?, sample_count = ?, remarks = ?, updated_at = ?
                WHERE id = ?
                """,
                (new_title, new_sample_name, new_sample_count, new_remarks, now_iso(), request_id),
            )
            log_action(
                user["id"],
                "sample_request",
                request_id,
                "request_metadata_updated",
                {
                    "title": new_title,
                    "sample_name": new_sample_name,
                    "sample_count": new_sample_count,
                    "remarks_preview": new_remarks[:160],
                },
            )
            write_request_metadata_snapshot(request_id)
            flash("Request details updated.", "success")
            return redirect(url_for("request_detail", request_id=request_id))

        if action == "admin_set_status" and can_manage:
            new_status = request.form.get("new_status", "").strip()
            remarks = request.form.get("remarks", "").strip()
            scheduled_for = request.form.get("scheduled_for", "").strip()
            allowed_statuses = {
                "submitted",
                "under_review",
                "awaiting_sample_submission",
                "sample_submitted",
                "sample_received",
                "scheduled",
                "in_progress",
                "completed",
                "rejected",
            }
            if new_status not in allowed_statuses:
                flash("Choose a valid status.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            db = get_db()
            if new_status == "under_review" and not approval_steps:
                create_approval_chain(db, request_id, sample_request["instrument_id"])
            now_value = now_iso()
            update_fields = {
                "status": new_status,
                "remarks": remarks if remarks else sample_request["remarks"],
                "updated_at": now_value,
                "completion_locked": 1 if new_status == "completed" else 0,
                "completed_at": now_value if new_status == "completed" else None,
            }
            if new_status == "sample_submitted":
                update_fields["submitted_to_lab_at"] = sample_request["submitted_to_lab_at"] or now_value
                update_fields["sample_submitted_at"] = sample_request["sample_submitted_at"] or now_value
            if new_status in {"sample_received", "scheduled", "in_progress", "completed"}:
                update_fields["sample_received_at"] = sample_request["sample_received_at"] or now_value
                update_fields["received_by_operator_id"] = sample_request["received_by_operator_id"] or user["id"]
            if new_status == "scheduled":
                update_fields["scheduled_for"] = scheduled_for or sample_request["scheduled_for"] or now_value
            if new_status == "completed":
                update_fields.update(completion_override_fields(sample_request, user["id"], now_value))
                execute(
                    "UPDATE approval_steps SET status = 'approved', acted_at = COALESCE(acted_at, ?), remarks = CASE WHEN remarks = '' THEN 'Completed from admin status control' ELSE remarks END WHERE sample_request_id = ? AND status != 'rejected'",
                    (now_value, request_id),
                )
            columns = ", ".join(f"{key} = ?" for key in update_fields)
            execute(
                f"UPDATE sample_requests SET {columns} WHERE id = ?",
                tuple(update_fields.values()) + (request_id,),
            )
            if new_status == "completed":
                log_completion_override_events(
                    user["id"],
                    sample_request,
                    update_fields,
                    now_value,
                    "status_changed",
                    {"from_status": sample_request["status"], "to_status": new_status, "remarks": remarks, "scheduled_for": scheduled_for},
                )
                send_completion_inbox_message(user["id"], sample_request)
            else:
                log_action(
                    user["id"],
                    "sample_request",
                    request_id,
                    "status_changed",
                    {"from_status": sample_request["status"], "to_status": new_status, "remarks": remarks, "scheduled_for": scheduled_for},
                )
            write_request_metadata_snapshot(request_id)
            # ── Workflow notification: status changed ──
            try:
                notify(sample_request["requester_id"], "request",
                       f"Request {sample_request['request_no']} updated",
                       f"Status changed to {new_status}",
                       url_for("request_detail", request_id=request_id),
                       "sample_request", request_id)
            except Exception:
                pass
            flash("Status updated.", "success")
            return redirect(url_for("request_detail", request_id=request_id))

        if sample_request["completion_locked"]:
            flash("Completed jobs are locked. Add amendments through a controlled workflow later.", "error")
            return redirect(url_for("request_detail", request_id=request_id))

        if action == "approve_step":
            step_id = int(request.form["step_id"])
            remarks = request.form.get("remarks", "").strip()
            approval_attachment = request.files.get("approval_attachment")
            step = query_one("SELECT * FROM approval_steps WHERE id = ? AND sample_request_id = ?", (step_id, request_id))
            if (
                step is None
                or not can_approve_step(user, step, sample_request["instrument_id"])
                or not approval_step_is_actionable(step, approval_steps)
            ):
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return jsonify({"ok": False, "error": "forbidden"}), 403
                abort(403)
            execute(
                "UPDATE approval_steps SET status = 'approved', remarks = ?, acted_at = ? WHERE id = ?",
                (remarks, now_iso(), step_id),
            )
            next_status = build_request_status(get_db(), request_id)
            assert_status_transition(sample_request["status"], next_status)
            execute("UPDATE sample_requests SET status = ?, remarks = ?, updated_at = ? WHERE id = ?", (next_status, remarks, now_iso(), request_id))
            if approval_attachment and (approval_attachment.filename or "").strip():
                try:
                    save_uploaded_attachment(
                        sample_request,
                        approval_attachment,
                        user["id"],
                        "invoice",
                        remarks or f"{step['approver_role'].replace('_', ' ').title()} approval attachment",
                    )
                except ValueError as exc:
                    flash(f"Approved, but file upload failed: {exc}", "error")
            log_action(user["id"], "sample_request", request_id, f"{step['approver_role']}_approved", {"step_id": step_id})
            # ── Workflow notification: approval step approved ──
            try:
                notify(sample_request["requester_id"], "request",
                       f"Approval update on {sample_request['request_no']}",
                       f"Step {step['approver_role'].replace('_', ' ').title()} approved",
                       url_for("request_detail", request_id=request_id),
                       "sample_request", request_id)
            except Exception:
                pass
            # Phase 2 commit 5 — notify submitter via inbox if configured
            notify_cfg = query_one(
                "SELECT notify_submitter FROM instrument_approval_config WHERE instrument_id = ? AND step_order = ?",
                (sample_request["instrument_id"], step["step_order"]),
            )
            if notify_cfg and notify_cfg["notify_submitter"]:
                execute(
                    "INSERT INTO messages (sender_id, recipient_id, subject, body, sent_at, read_at) VALUES (?, ?, ?, ?, ?, NULL)",
                    (
                        user["id"],
                        sample_request["requester_id"],
                        f"Request {sample_request['request_no']}: Step {step['step_order']} approved",
                        f"Your request {sample_request['request_no']} ({sample_request['sample_name']}) has been approved at step {step['step_order']} ({step['approver_role'].replace('_', ' ').title()}). You may now proceed with the next stage.",
                        now_iso(),
                    ),
                )
            # W1.4.6 — XHR branch: return JSON so approval-toggle.js can
            # refresh in place without a full-page reload round-trip.
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({
                    "ok": True,
                    "step_id": step_id,
                    "new_request_status": next_status,
                    "approver_role": step["approver_role"],
                    "reload_url": url_for("request_detail", request_id=request_id),
                })
        elif action == "reject_step":
            step_id = int(request.form["step_id"])
            remarks = request.form.get("remarks", "").strip()
            step = query_one("SELECT * FROM approval_steps WHERE id = ? AND sample_request_id = ?", (step_id, request_id))
            if (
                step is None
                or not can_approve_step(user, step, sample_request["instrument_id"])
                or not approval_step_is_actionable(step, approval_steps)
            ):
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return jsonify({"ok": False, "error": "forbidden"}), 403
                abort(403)
            if not remarks:
                # Rejection without a reason is an audit gap — the old UI
                # enforced this with `required` on the textarea. Preserve
                # that contract on the XHR path too.
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return jsonify({"ok": False, "error": "remarks_required"}), 400
                flash("A rejection reason is required.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            execute(
                "UPDATE approval_steps SET status = 'rejected', remarks = ?, acted_at = ? WHERE id = ?",
                (remarks, now_iso(), step_id),
            )
            assert_status_transition(sample_request["status"], "rejected")
            execute("UPDATE sample_requests SET status = 'rejected', remarks = ?, updated_at = ? WHERE id = ?", (remarks, now_iso(), request_id))
            log_action(user["id"], "sample_request", request_id, f"{step['approver_role']}_rejected", {"step_id": step_id})
            # ── Workflow notification: approval step rejected ──
            try:
                notify(sample_request["requester_id"], "request",
                       f"Approval update on {sample_request['request_no']}",
                       f"Step {step['approver_role'].replace('_', ' ').title()} rejected",
                       url_for("request_detail", request_id=request_id),
                       "sample_request", request_id)
            except Exception:
                pass
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({
                    "ok": True,
                    "step_id": step_id,
                    "new_request_status": "rejected",
                    "approver_role": step["approver_role"],
                    "reload_url": url_for("request_detail", request_id=request_id),
                })
        elif action == "assign_approver" and can_manage:
            step_id = int(request.form["step_id"])
            raw_approver = request.form.get("approver_user_id", "").strip()
            step = query_one("SELECT * FROM approval_steps WHERE id = ? AND sample_request_id = ?", (step_id, request_id))
            if step is None:
                abort(403)
            # v2.1.4 — "0" means unassign: any person with the right
            # role can pick up the step. NULL in the DB = role-based.
            if raw_approver == "0" or not raw_approver:
                execute(
                    "UPDATE approval_steps SET approver_user_id = NULL, acted_at = NULL WHERE id = ?",
                    (step_id,),
                )
                log_action(
                    user["id"],
                    "sample_request",
                    request_id,
                    "approval_unassigned",
                    {"step_id": step_id, "approver_role": step["approver_role"]},
                )
                flash(f"{approval_role_label(step['approver_role'])} unassigned — any eligible person can approve.", "success")
            else:
                approver_user_id = int(raw_approver)
                candidate = query_one("SELECT * FROM users WHERE id = ? AND active = 1", (approver_user_id,))
                if not candidate_allowed_for_step(candidate, step["approver_role"], sample_request["instrument_id"]):
                    abort(403)
                execute(
                    "UPDATE approval_steps SET approver_user_id = ?, acted_at = NULL WHERE id = ?",
                    (approver_user_id, step_id),
                )
                log_action(
                    user["id"],
                    "sample_request",
                    request_id,
                    "approval_assigned",
                    {"step_id": step_id, "approver_role": step["approver_role"], "approver_user_id": approver_user_id},
                )
                flash(f"{approval_role_label(step['approver_role'])} approver updated.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        elif action == "mark_sample_submitted" and sample_request["requester_id"] == user["id"]:
            if sample_request["status"] != "awaiting_sample_submission":
                flash("This request is not waiting for physical sample submission.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            if instrument_intake_mode(sample_request) != "accepting":
                flash("This lab is not accepting sample dropoff right now.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            dropoff_note = request.form.get("sample_dropoff_note", "").strip()
            assert_status_transition(sample_request["status"], "sample_submitted")
            execute(
                """
                UPDATE sample_requests
                SET status = 'sample_submitted', submitted_to_lab_at = ?, sample_submitted_at = ?,
                    sample_dropoff_note = ?, remarks = ?, updated_at = ?
                WHERE id = ?
                """,
                (now_iso(), now_iso(), dropoff_note, dropoff_note, now_iso(), request_id),
            )
            log_action(user["id"], "sample_request", request_id, "sample_submitted", {"dropoff_note": dropoff_note})
        elif action == "mark_sample_received" and (can_operate or can_manage):
            if sample_request["status"] != "sample_submitted":
                flash("Sample can only be received after member handoff.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            remarks = request.form.get("remarks", "").strip()
            assert_status_transition(sample_request["status"], "sample_received")
            execute(
                """
                UPDATE sample_requests
                SET status = 'sample_received', sample_received_at = ?, received_by_operator_id = ?,
                    remarks = ?, updated_at = ?
                WHERE id = ?
                """,
                (now_iso(), user["id"], remarks, now_iso(), request_id),
            )
            log_action(user["id"], "sample_request", request_id, "sample_received", {"remarks": remarks})
        elif action == "reassign_operator" and (can_operate or can_manage):
            operator_id = int(request.form.get("assigned_operator_id") or 0)
            assignee = query_one("SELECT * FROM users WHERE id = ? AND active = 1", (operator_id,))
            if not candidate_allowed_for_step(assignee, "operator", sample_request["instrument_id"]):
                flash("Pick a valid operator or instrument-area admin for this instrument.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            remarks = request.form.get("remarks", "").strip()
            execute(
                "UPDATE sample_requests SET assigned_operator_id = ?, remarks = CASE WHEN ? != '' THEN ? ELSE remarks END, updated_at = ? WHERE id = ?",
                (operator_id, remarks, remarks, now_iso(), request_id),
            )
            log_action(
                user["id"],
                "sample_request",
                request_id,
                "operator_reassigned",
                {"assigned_operator_id": operator_id, "remarks": remarks},
            )
            flash(f"Job reassigned to {assignee['name']}.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        elif action == "schedule" and (can_operate or can_manage):
            if sample_request["status"] not in {"sample_received", "scheduled", "in_progress"}:
                flash("Request must be approved and physically received before scheduling.", "error")
                return redirect(url_for("request_detail", request_id=request_id))
            scheduled_for = request.form["scheduled_for"]
            operator_id = int(request.form.get("assigned_operator_id") or user["id"]) if request.form.get("assigned_operator_id") else user["id"]
            remarks = request.form.get("remarks", "").strip()
            assert_status_transition(sample_request["status"], "scheduled")
            execute(
                "UPDATE sample_requests SET status = 'scheduled', scheduled_for = ?, assigned_operator_id = ?, remarks = ?, updated_at = ? WHERE id = ?",
                (scheduled_for, operator_id, remarks, now_iso(), request_id),
            )
            log_action(user["id"], "sample_request", request_id, "scheduled", {"scheduled_for": scheduled_for, "assigned_operator_id": operator_id})
        elif action == "admin_schedule_override" and can_manage:
            scheduled_for = request.form["scheduled_for"]
            operator_id = int(request.form.get("assigned_operator_id") or user["id"]) if request.form.get("assigned_operator_id") else user["id"]
            remarks = request.form.get("remarks", "").strip()
            assert_status_transition(sample_request["status"], "scheduled", force=True)
            execute(
                """
                UPDATE sample_requests
                SET status = 'scheduled', scheduled_for = ?, assigned_operator_id = ?, sample_received_at = COALESCE(sample_received_at, ?),
                    received_by_operator_id = COALESCE(received_by_operator_id, ?), remarks = ?, updated_at = ?
                WHERE id = ?
                """,
                (scheduled_for, operator_id, now_iso(), user["id"], remarks, now_iso(), request_id),
            )
            execute(
                "UPDATE approval_steps SET status = 'approved', acted_at = COALESCE(acted_at, ?), remarks = CASE WHEN remarks = '' THEN 'Admin override' ELSE remarks END WHERE sample_request_id = ? AND status != 'rejected'",
                (now_iso(), request_id),
            )
            log_action(user["id"], "sample_request", request_id, "admin_schedule_override", {"scheduled_for": scheduled_for, "assigned_operator_id": operator_id})
        elif action == "start" and (can_operate or can_manage):
            remarks = request.form.get("remarks", "").strip()
            assert_status_transition(sample_request["status"], "in_progress")
            execute("UPDATE sample_requests SET status = 'in_progress', remarks = ?, updated_at = ? WHERE id = ?", (remarks, now_iso(), request_id))
            log_action(user["id"], "sample_request", request_id, "started", {})
        elif action == "complete" and (can_operate or can_manage):
            results_summary = request.form["results_summary"].strip()
            remarks = request.form.get("remarks", "").strip()
            # v2.0.0 — finance values come from the form + peer aggregates
            _finance = computed_finance_for_request(get_db(), request_id)
            amount_paid = float(request.form.get("amount_paid") or _finance["amount_paid"] or 0)
            finance_status = request.form.get("finance_status", _finance["finance_status"])
            email_ok, email_message = send_completion_email(sample_request, results_summary)
            now_value = now_iso()
            completion_fields = completion_override_fields(sample_request, user["id"], now_value)
            assert_status_transition(sample_request["status"], "completed")
            execute(
                """
                UPDATE sample_requests
                SET status = 'completed', results_summary = ?, remarks = ?,
                    result_email_status = ?, result_email_sent_at = ?, completion_locked = 1,
                    submitted_to_lab_at = ?, sample_submitted_at = ?, sample_received_at = ?, received_by_operator_id = ?,
                    scheduled_for = ?, assigned_operator_id = ?, completed_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    results_summary,
                    remarks,
                    email_message,
                    now_value if email_ok else None,
                    completion_fields["submitted_to_lab_at"],
                    completion_fields["sample_submitted_at"],
                    completion_fields["sample_received_at"],
                    completion_fields["received_by_operator_id"],
                    completion_fields["scheduled_for"],
                    completion_fields["assigned_operator_id"],
                    completion_fields["completed_at"],
                    now_value,
                    request_id,
                ),
            )
            sync_request_to_peer_aggregates(
                get_db(), request_id,
                amount_due=_finance["amount_due"] or amount_paid,
                amount_paid=amount_paid,
                finance_status=finance_status,
                receipt_number=_finance["receipt_number"],
            )
            get_db().commit()
            log_completion_override_events(
                user["id"],
                sample_request,
                completion_fields,
                now_value,
                "completed",
                {"results_summary": results_summary, "email_status": email_message},
            )
        elif action == "admin_complete_override" and can_manage:
            results_summary = request.form["results_summary"].strip()
            remarks = request.form.get("remarks", "").strip()
            _finance = computed_finance_for_request(get_db(), request_id)
            amount_paid = float(request.form.get("amount_paid") or _finance["amount_paid"] or 0)
            finance_status = request.form.get("finance_status", _finance["finance_status"])
            operator_id = int(request.form["assigned_operator_id"]) if request.form.get("assigned_operator_id") else (sample_request["assigned_operator_id"] or user["id"])
            email_ok, email_message = send_completion_email(sample_request, results_summary)
            now_value = now_iso()
            completion_fields = completion_override_fields(sample_request, operator_id, now_value)
            assert_status_transition(sample_request["status"], "completed", force=True)
            execute(
                """
                UPDATE sample_requests
                SET status = 'completed', assigned_operator_id = ?, submitted_to_lab_at = ?, sample_submitted_at = ?, sample_received_at = ?,
                    received_by_operator_id = ?, scheduled_for = ?,
                    results_summary = ?, remarks = ?,
                    result_email_status = ?, result_email_sent_at = ?, completion_locked = 1, completed_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    completion_fields["assigned_operator_id"],
                    completion_fields["submitted_to_lab_at"],
                    completion_fields["sample_submitted_at"],
                    completion_fields["sample_received_at"],
                    completion_fields["received_by_operator_id"],
                    completion_fields["scheduled_for"],
                    results_summary,
                    remarks,
                    email_message,
                    now_value if email_ok else None,
                    completion_fields["completed_at"],
                    now_value,
                    request_id,
                ),
            )
            sync_request_to_peer_aggregates(
                get_db(), request_id,
                amount_due=_finance["amount_due"] or amount_paid,
                amount_paid=amount_paid,
                finance_status=finance_status,
                receipt_number=_finance["receipt_number"],
            )
            get_db().commit()
            execute(
                "UPDATE approval_steps SET status = 'approved', acted_at = COALESCE(acted_at, ?), remarks = CASE WHEN remarks = '' THEN 'Admin override' ELSE remarks END WHERE sample_request_id = ? AND status != 'rejected'",
                (now_value, request_id),
            )
            log_completion_override_events(
                user["id"],
                sample_request,
                completion_fields,
                now_value,
                "admin_complete_override",
                {"results_summary": results_summary, "email_status": email_message},
            )
        elif action == "resolve_sample" and (can_operate or can_manage):
            results_summary = request.form.get("results_summary", "").strip()
            remarks = request.form.get("remarks", "").strip()
            _finance = computed_finance_for_request(get_db(), request_id)
            amount_paid = float(request.form.get("amount_paid") or _finance["amount_paid"] or 0)
            finance_status = request.form.get("finance_status", _finance["finance_status"])
            mark_complete = request.form.get("mark_complete") == "1"
            uploaded_resolution_file = request.files.get("resolution_attachment")
            resolution_upload_error = None
            if mark_complete:
                final_summary = results_summary or remarks or "Completed by operator."
                email_ok, email_message = send_completion_email(sample_request, final_summary)
                now_value = now_iso()
                completion_fields = completion_override_fields(sample_request, user["id"], now_value)
                execute(
                    """
                    UPDATE sample_requests
                    SET status = 'completed', assigned_operator_id = ?, submitted_to_lab_at = ?, sample_submitted_at = ?,
                        sample_received_at = ?, received_by_operator_id = ?, scheduled_for = ?, results_summary = ?, remarks = ?,
                        result_email_status = ?, result_email_sent_at = ?, completion_locked = 1, completed_at = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        completion_fields["assigned_operator_id"],
                        completion_fields["submitted_to_lab_at"],
                        completion_fields["sample_submitted_at"],
                        completion_fields["sample_received_at"],
                        completion_fields["received_by_operator_id"],
                        completion_fields["scheduled_for"],
                        final_summary,
                        remarks,
                        email_message,
                        now_value if email_ok else None,
                        completion_fields["completed_at"],
                        now_value,
                        request_id,
                    ),
                )
                execute(
                    "UPDATE approval_steps SET status = 'approved', acted_at = COALESCE(acted_at, ?), remarks = CASE WHEN remarks = '' THEN 'Completed from job page' ELSE remarks END WHERE sample_request_id = ? AND status != 'rejected'",
                    (now_value, request_id),
                )
                log_completion_override_events(
                    user["id"],
                    sample_request,
                    completion_fields,
                    now_value,
                    "resolved_and_completed",
                    {"results_summary": final_summary, "email_status": email_message},
                )
                send_completion_inbox_message(user["id"], sample_request)
            else:
                execute(
                    "UPDATE sample_requests SET results_summary = ?, remarks = ?, updated_at = ? WHERE id = ?",
                    (results_summary, remarks, now_iso(), request_id),
                )
                log_action(user["id"], "sample_request", request_id, "resolution_saved", {"results_summary": results_summary})
            # v2.0.0 — finance state lives in peer aggregates now. The
            # sync call creates/updates Invoice + Payment rows so the
            # finance portal reflects the new amount_paid.
            sync_request_to_peer_aggregates(
                get_db(), request_id,
                amount_due=_finance["amount_due"] or amount_paid,
                amount_paid=amount_paid,
                finance_status=finance_status,
                receipt_number=_finance["receipt_number"],
            )
            get_db().commit()
            if uploaded_resolution_file and uploaded_resolution_file.filename:
                try:
                    save_uploaded_attachment(
                        sample_request,
                        uploaded_resolution_file,
                        user["id"],
                        "result_document",
                        remarks or "Resolution upload",
                    )
                except ValueError as exc:
                    resolution_upload_error = str(exc)
            write_request_metadata_snapshot(request_id)
            if resolution_upload_error:
                flash(f"Resolution saved, but file upload failed: {resolution_upload_error}", "error")
            else:
                flash("Job marked done." if mark_complete else "Resolution saved.", "success")
            return redirect(url_for("request_detail", request_id=request_id))
        elif action == "reject" and can_manage:
            remarks = request.form.get("remarks", "").strip()
            assert_status_transition(sample_request["status"], "rejected", force=True)
            execute("UPDATE sample_requests SET status = 'rejected', remarks = ?, updated_at = ? WHERE id = ?", (remarks, now_iso(), request_id))
            log_action(user["id"], "sample_request", request_id, "rejected", {})
        else:
            abort(403)
        write_request_metadata_snapshot(request_id)
        return redirect(url_for("request_detail", request_id=request_id))

    operators = []
    if can_manage or can_operate:
        # Assignment picker: every user who has access to this
        # specific instrument via any of the 3 access tables, PLUS
        # every super/site admin (by primary role OR by user_roles
        # junction — the v1.5.0 multi-role fix, closes the assign-
        # approver bug where admins-via-junction were invisible in
        # the picker). Sorted alphabetically.
        operators = query_all(
            """
            SELECT DISTINCT u.id, u.name, u.role
            FROM users u
            LEFT JOIN instrument_operators       io  ON io.user_id  = u.id AND io.instrument_id  = ?
            LEFT JOIN instrument_admins          ia  ON ia.user_id  = u.id AND ia.instrument_id  = ?
            LEFT JOIN instrument_faculty_admins  ifa ON ifa.user_id = u.id AND ifa.instrument_id = ?
            WHERE u.active = 1
              AND (
                io.instrument_id IS NOT NULL OR
                ia.instrument_id IS NOT NULL OR
                ifa.instrument_id IS NOT NULL OR
                u.role IN ('super_admin', 'site_admin') OR
                EXISTS (
                    SELECT 1 FROM user_roles ur
                     WHERE ur.user_id = u.id
                       AND ur.role IN ('super_admin', 'site_admin')
                )
              )
            ORDER BY u.name
            """,
            (sample_request["instrument_id"], sample_request["instrument_id"], sample_request["instrument_id"]),
        )
    attachments = get_request_attachments(request_id)
    issues = get_request_issues(request_id)
    message_thread = get_request_message_thread(request_id)
    message_attachments = attachments_by_message_ids([row["id"] for row in message_thread])
    logs = query_all("SELECT al.*, u.name AS actor_name FROM audit_logs al LEFT JOIN users u ON u.id = al.actor_id WHERE entity_type = 'sample_request' AND entity_id = ? ORDER BY al.id", (request_id,))
    audit_chain_valid = verify_audit_chain("sample_request", request_id)
    status_summary = request_status_summary(sample_request, approval_steps)
    lifecycle_steps = request_lifecycle_steps(sample_request, approval_steps)
    visible_attachments = request_card_visible_attachments(user, sample_request, attachments)
    timeline_entries = request_card_visible_timeline(
        user,
        sample_request,
        request_timeline_entries(sample_request, logs, visible_attachments, message_thread, message_attachments),
    )
    has_open_issue = any(issue["status"] == "open" for issue in issues)
    submitted_documents = [row for row in visible_attachments if row["attachment_type"] in {"request_document", "sample_slip"}]
    approval_candidates = {
        "finance": approval_candidate_options("finance", sample_request["instrument_id"]),
        "professor": approval_candidate_options("professor", sample_request["instrument_id"]),
        "operator": approval_candidate_options("operator", sample_request["instrument_id"]),
    }
    actionable_approval_steps = [
        step
        for step in approval_steps
        if step["id"] in actionable_step_ids and can_approve_step(user, step, sample_request["instrument_id"])
    ]
    remarks_author_row = query_one(
        """SELECT u.name FROM audit_logs al
           JOIN users u ON u.id = al.actor_id
           WHERE al.entity_type = 'sample_request' AND al.entity_id = ? AND al.actor_id IS NOT NULL
           ORDER BY al.created_at DESC LIMIT 1""",
        (request_id,),
    )
    remarks_author = remarks_author_row["name"] if remarks_author_row else None
    custom_field_values = query_all(
        """SELECT icf.field_label, rcfv.field_value
           FROM request_custom_field_values rcfv
           JOIN instrument_custom_fields icf ON icf.id = rcfv.custom_field_id
           WHERE rcfv.request_id = ?
           ORDER BY icf.display_order""",
        (request_id,),
    )
    return render_template(
        "request_detail.html",
        sample_request=sample_request,
        card_policy=card_policy,
        back_url=request.args.get("back", "").strip(),
        can_manage=can_manage,
        can_operate=can_operate,
        can_upload_files=can_upload_files,
        operators=operators,
        approval_steps=approval_steps,
        actionable_step_ids=actionable_step_ids,
        actionable_approval_steps=actionable_approval_steps,
        attachment_type_choices=attachment_type_choices(),
        attachments=visible_attachments,
        submitted_documents=submitted_documents,
        issues=issues,
        message_thread=message_thread,
        message_attachments=message_attachments,
        attachment_size_label=attachment_size_label,
        logs=logs,
        audit_chain_valid=audit_chain_valid,
        status_summary=status_summary,
        lifecycle_steps=lifecycle_steps,
        timeline_entries=timeline_entries,
        has_open_issue=has_open_issue,
        conversation_note_types=[item for item in COMMUNICATION_NOTE_TYPES if item[0] in {"lab_reply", "final_note"}],
        admin_status_choices=["submitted", "under_review", "awaiting_sample_submission", "sample_submitted", "sample_received", "scheduled", "in_progress", "completed", "rejected"],
        can_flag_issue=can_flag_issue,
        can_respond_issue=can_respond_issue,
        approval_candidates=approval_candidates,
        remarks_author=remarks_author,
        custom_field_values=custom_field_values,
    )


# ─────────────────────────────────────────────────────────────
# Development Control Panel (v1.3.0+)
# ─────────────────────────────────────────────────────────────
#
# Owner-only admin page that surfaces:
#   - project progress (git branch, ahead/behind, dirty, commits)
#   - roadmap (version blocks parsed from TODO_AI.txt, rendered as
#     chart_bar meters)
#   - current release (parsed from CHANGELOG.md)
#   - in-page document viewer (README / PHILOSOPHY / DEPLOY /
#     PROJECT / TODO_AI / CHANGELOG)
#
# All routes are gated by `@owner_required`. No external
# dependencies, no compute offload. The panel is pure
# read-only introspection of the repository state.

# Basenames only — the reader tries BASE_DIR first, then BASE_DIR/docs.
# That keeps the dev-panel tabs uncluttered ("PHILOSOPHY.md" instead of
# "docs/PHILOSOPHY.md") while still following the actual file location
# after the v1.3.8 top-level cleanup.
DEV_PANEL_DOC_FILES = (
    "README.md",
    "CHANGELOG.md",
    "PHILOSOPHY.md",
    "DEPLOY.md",
    "PROJECT.md",
    "ROADMAP.md",
    "NEXT_WAVES.md",
    "HANDOVER.md",
    "MODULES.md",
    "DATA_POLICY.md",
    "COMPONENT_LIBRARY.md",
    "ROLE_VISIBILITY_MATRIX.md",
    "SECURITY_TODO.md",
    "CSS_COMPONENT_MAP.md",
)


def _dev_panel_waves() -> list[dict]:
    """Parse the 'Time budget summary' table from docs/NEXT_WAVES.md.

    Each row becomes a dict with keys: wave, track, est, blocks, tag,
    status. A wave is 'shipped' if either (a) its tag appears as a git
    tag, or (b) its section header in NEXT_WAVES.md carries a
    '✅ SHIPPED' marker — the second path covers waves the plan
    explicitly leaves untagged (rolled up into a later tag). A wave is
    'hot' if it is the first unshipped row, 'pending' otherwise.
    Dev portal uses this to drive the WAVES tile.
    """
    import re

    path = BASE_DIR / "docs" / "NEXT_WAVES.md"
    if not path.exists():
        return []
    text = path.read_text(errors="ignore")

    # Pass 1: collect wave-id → shipped flag from section headers like
    # '### W1.3.7 — title ✅ SHIPPED'. The marker is case-insensitive.
    # Wave id: W + dotted numeric segments + optional letter suffix
    # (e.g. W1.3.7, W1.4.2a, W1.4.2b). The suffix lets us split a
    # single tagged version into multiple work slices in the plan.
    shipped_via_marker: set[str] = set()
    header_re = re.compile(r"^#{2,4}\s*(W\d+(?:\.\d+){1,3}[a-z]*)\b.*", re.MULTILINE)
    for match in header_re.finditer(text):
        header_line = match.group(0)
        if "✅" in header_line or "SHIPPED" in header_line.upper():
            shipped_via_marker.add(match.group(1))

    # Pass 2: parse the budget-summary table.
    rows: list[dict] = []
    in_table = False
    for raw in text.splitlines():
        line = raw.strip()
        if line.startswith("| wave") and "track" in line:
            in_table = True
            continue
        if in_table:
            if not line.startswith("|"):
                break
            if set(line.replace("|", "").strip()) <= {"-", ":", " "}:
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) < 5:
                continue
            # The table has 6 columns; the trailing "status" cell
            # carries freeform markers like "✅", "ops-blocked",
            # "deferred". We respect "deferred" up-front so v1.5+
            # waves never surface as hot while the critical path is
            # still in flight. Shipped / ops-blocked are resolved
            # downstream from git tags + blocks propagation.
            status_cell = cells[5].lower() if len(cells) > 5 else ""
            initial_status = "deferred" if "deferred" in status_cell else "pending"
            rows.append({
                "wave": cells[0],
                "track": cells[1],
                "est": cells[2],
                "blocks": cells[3],
                "tag": cells[4],
                "status": initial_status,
            })

    shipped_tags = set(_dev_panel_git("tag", "--list").splitlines())
    # Pass 3a: decide shipped vs pending for every row.
    for row in rows:
        tag_clean = row["tag"].replace("v", "")
        tagged_shipped = row["tag"] and row["tag"] not in ("—", "-") and (
            row["tag"] in shipped_tags or ("v" + tag_clean) in shipped_tags
        )
        if tagged_shipped or row["wave"] in shipped_via_marker:
            row["status"] = "shipped"

    # Pass 3b: propagate ops-blocked through dependencies. An ops-blocked
    # row waits on an external click (e.g. Tailscale Serve unblock) and
    # must NOT surface as the hot wave because picking it up is not
    # laptop-local work. A row is ops-blocked if its "blocks" column
    # says so directly, OR if any wave it depends on is ops-blocked.
    # Fixed-point over at most len(rows) iterations since the dependency
    # graph is a DAG. W1.4.7+.
    ops_blocked: set[str] = set()
    wave_ids = {row["wave"] for row in rows}
    for _ in range(len(rows) + 1):
        grew = False
        for row in rows:
            if row["status"] == "shipped" or row["wave"] in ops_blocked:
                continue
            blocks_text = (row.get("blocks") or "").lower()
            direct = "ops unblock" in blocks_text or "ops-blocked" in blocks_text
            transitive = any(
                wid in ops_blocked
                for wid in wave_ids
                if wid.lower() in blocks_text
            )
            if direct or transitive:
                ops_blocked.add(row["wave"])
                grew = True
        if not grew:
            break
    for row in rows:
        if row["wave"] in ops_blocked:
            row["status"] = "ops-blocked"

    # Pass 3c: first unshipped row that is NOT ops-blocked or
    # deferred wins "hot". If the critical path is entirely
    # shipped / ops-blocked / deferred, no row is hot and the
    # hero tile falls back to "All tracked waves shipped."
    first_unshipped = True
    for row in rows:
        if row["status"] in ("shipped", "ops-blocked", "deferred"):
            continue
        if first_unshipped and row["wave"].startswith("W"):
            row["status"] = "hot"
            first_unshipped = False
    return rows


def _dev_panel_crawler_health() -> dict:
    """Summarise the latest crawler reports under ./reports/*_log.json.

    Returns totals and per-strategy rows for the CRAWLER HEALTH tile.
    Missing reports dir → empty summary (tile renders empty-state).
    """
    reports_dir = BASE_DIR / "reports"
    if not reports_dir.exists():
        return {"total": 0, "passed": 0, "failed": 0, "warnings": 0, "rows": []}
    rows: list[dict] = []
    totals = {"passed": 0, "failed": 0, "warnings": 0}
    for log in sorted(reports_dir.glob("*_log.json")):
        try:
            data = json.loads(log.read_text(errors="ignore"))
        except Exception:
            continue
        result = data.get("result") or {}
        passed = int(result.get("passed", 0) or 0)
        failed = int(result.get("failed", 0) or 0)
        warnings = int(result.get("warnings", 0) or 0)
        total = passed + failed + warnings
        if total == 0:
            continue
        pct = int(round(100 * passed / total)) if total else 0
        rows.append({
            "strategy": data.get("strategy") or log.stem.replace("_log", ""),
            "aspect": data.get("aspect") or "—",
            "passed": passed,
            "failed": failed,
            "warnings": warnings,
            "total": total,
            "pct": pct,
        })
        totals["passed"] += passed
        totals["failed"] += failed
        totals["warnings"] += warnings
    rows.sort(key=lambda r: (-r["failed"], -r["warnings"], r["strategy"]))
    return {
        "total": totals["passed"] + totals["failed"] + totals["warnings"],
        **totals,
        "rows": rows,
    }


def _dev_panel_git(*args: str) -> str:
    """Run a git subcommand and return stdout (or empty on failure).

    Scrubs every git-related environment variable before spawning the
    subprocess so git discovers its working copy purely from `cwd`.
    Without this, the dev panel breaks when the web app is started
    from a git-hook context (pre-receive / post-receive / post-commit):
    git inherits `GIT_DIR` / `GIT_WORK_TREE` / `GIT_QUARANTINE_PATH` /
    `GIT_INDEX_FILE` from the parent hook and operates on the hook's
    repo (or its quarantine object store) instead of the app's own
    working tree — returning empty log output and breaking the dev-
    panel readability contract. Regression-tested by
    `crawlers/strategies/dev_panel_readability.py`.
    """
    env = {k: v for k, v in os.environ.items() if not k.startswith("GIT_")}
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
            env=env,
        )
        return result.stdout.strip()
    except Exception:
        return ""


def _dev_panel_future_fixes_count() -> dict:
    """Count remaining `# TODO [vX.Y.Z …]` markers in app.py + templates.

    Mirrors the regex used by `crawlers/strategies/future_fixes_placeholder.py`
    but invoked inline so the dev panel doesn't need the crawler harness
    (which bootstraps a DB). Returns `{"total": N, "by_file": {...}}`.
    """
    import re as _re
    marker_re = _re.compile(r"#\s*TODO\s*\[(v\d+\.\d+\.\d+)\s+[^\]]+\]")
    targets: list[Path] = []
    if (BASE_DIR / "app.py").exists():
        targets.append(BASE_DIR / "app.py")
    if (BASE_DIR / "templates").exists():
        targets.extend(sorted((BASE_DIR / "templates").glob("*.html")))
    targets.extend(p for p in sorted(BASE_DIR.glob("*.py")) if p.name != "app.py")
    total = 0
    by_file: dict[str, int] = {}
    for path in targets:
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        matches = marker_re.findall(text)
        if not matches:
            continue
        by_file[path.relative_to(BASE_DIR).as_posix()] = len(matches)
        total += len(matches)
    return {"total": total, "by_file": by_file}


def _dev_panel_progress() -> dict:
    """Compute current project progress from git + the docs."""
    branch = _dev_panel_git("symbolic-ref", "--short", "HEAD") or "DETACHED"
    # Compare against the upstream of the current branch, not a hardcoded
    # "origin/main" — CATALYST lives on `v1.3.0-stable-release`, and the
    # hardcode was stale from the pre-stable-release era. If the branch
    # has no tracked upstream (detached HEAD, fresh local branch), both
    # counts stay at 0 rather than returning nonsense.
    upstream_ref = _dev_panel_git("rev-parse", "--abbrev-ref", "@{upstream}")
    ahead, behind = "0", "0"
    if upstream_ref:
        ahead_behind = _dev_panel_git(
            "rev-list", "--left-right", "--count", f"HEAD...{upstream_ref}"
        )
        if ahead_behind:
            parts = ahead_behind.split()
            if len(parts) == 2:
                ahead, behind = parts
    dirty_lines = _dev_panel_git("status", "--porcelain")
    dirty_count = len([line for line in dirty_lines.splitlines() if line.strip()])
    # Last 10 commits, rich format: sha | iso-date | subject. This lets us
    # surface "commits in the last 24h" + a human "last commit" timestamp
    # on the 'Now Shipping' hero tile.
    recent_commits: list[dict] = []
    log_out = _dev_panel_git("log", "-10", "--pretty=format:%h|%cI|%s")
    commits_today = 0
    last_commit_iso = ""
    now = datetime.utcnow()
    for line in log_out.splitlines():
        parts = line.split("|", 2)
        if len(parts) < 3:
            continue
        sha, iso, subject = parts
        recent_commits.append({"sha": sha, "subject": subject, "at": iso})
        if not last_commit_iso:
            last_commit_iso = iso
        try:
            # %cI is strict ISO-8601 with timezone; convert to naive UTC
            when = datetime.fromisoformat(iso.replace("Z", "+00:00"))
            if when.tzinfo:
                when = when.astimezone(tz=None).replace(tzinfo=None)
            if (now - when).total_seconds() <= 24 * 3600:
                commits_today += 1
        except ValueError:
            pass

    # Git tags are the authoritative source for both the ROADMAP tile
    # (shipped patches grouped by major.minor line) and the STABLE
    # RELEASE / header badge (latest semver tag). Fetch once, parse
    # once, use twice. See PHILOSOPHY.md §3.1 for the iOS cadence.
    tag_lines = _dev_panel_git("tag", "--list").splitlines()
    semver_tags: list[tuple[tuple[int, int, int], str]] = []
    for t in tag_lines:
        t = t.strip()
        if not t.startswith("v"):
            continue
        # Strip suffixes like -stable, -stable-release, -alpha.1 for semver parsing
        clean = t[1:].split("-")[0]
        parts = clean.split(".")
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            semver_tags.append(((int(parts[0]), int(parts[1]), int(parts[2])), t))
    semver_tags.sort()

    # Versions-in-flight for the ROADMAP tile. Historically parsed
    # TODO_AI.txt (pre-stable-release artifact with stale "v1.4.0
    # BULK OPERATIONS" / "v1.5.0 SEARCH" headings that no longer
    # match reality). Rewired 2026-04-11 to group the semver tags
    # above by their major.minor line — every tag is a completed
    # patch release per the iOS cadence in PHILOSOPHY.md §3.1, and
    # the ROADMAP tile now tells the actual shipped story instead
    # of a planning fiction.
    versions: list[dict] = []
    version_groups: dict[tuple[int, int], list[tuple[tuple[int, int, int], str]]] = {}
    for (m_mi_p, t) in semver_tags:
        key = (m_mi_p[0], m_mi_p[1])
        version_groups.setdefault(key, []).append((m_mi_p, t))
    for (major, minor) in sorted(version_groups.keys()):
        patches = sorted(version_groups[(major, minor)])
        entries = [{"label": t, "done": True} for _, t in patches]
        versions.append({
            "name": f"v{major}.{minor}.x — {len(patches)} patch{'' if len(patches) == 1 else 'es'} shipped",
            "entries": entries,
            "done": len(entries),
            "open": 0,
        })

    # v2.1.5 — Planned modules for the ROADMAP tile. These are
    # forward-looking entries that show what's coming next. Each has
    # open > 0 so the progress bar shows remaining work.
    planned_modules = [
        {
            "name": "v2.2.0 — Equipment & Calibration Logbook",
            "entries": [
                {"label": "instrument_calibrations table + routes", "done": False},
                {"label": "instrument_maintenance log", "done": False},
                {"label": "NABL calibration-due dashboard", "done": False},
            ],
            "done": 0,
            "open": 3,
        },
        {
            "name": "v2.3.0 — Mailbox & Notifications",
            "entries": [
                {"label": "Internal feed (event-widget style)", "done": False},
                {"label": ".eml download for external sends", "done": False},
                {"label": "Group targets (operators:<inst>, role:<role>)", "done": False},
                {"label": "Owner god-view (/admin/messages)", "done": False},
            ],
            "done": 0,
            "open": 4,
        },
        {
            "name": "v2.4.0 — Calendar iCal Feed",
            "entries": [
                {"label": "/calendar.ics per-user subscription", "done": False},
                {"label": "Per-instrument .ics feed", "done": False},
            ],
            "done": 0,
            "open": 2,
        },
        {
            "name": "v2.5.0 — Responsive & Mobile",
            "entries": [
                {"label": "Table overflow wrappers", "done": False},
                {"label": "Nav compaction < 640px", "done": False},
                {"label": "Intermediate breakpoint 900px", "done": False},
            ],
            "done": 0,
            "open": 3,
        },
    ]
    versions.extend(planned_modules)

    # Current release: the latest semver tag wins. Falls back to
    # CHANGELOG.md only if no semver tags exist at all (greenfield
    # repo safety, never hit on CATALYST trunk in practice).
    current_release = "unknown"
    latest_tag_info: dict = {}
    if semver_tags:
        # semver_tags was already sorted above when it was built for
        # the ROADMAP tile; no need to re-sort.
        latest_tag = semver_tags[-1][1]  # e.g. "v1.4.4"
        current_release = latest_tag.lstrip("v")
        # Resolve tag → commit SHA, tagged-at, tag subject (first line of
        # the annotated message). `git for-each-ref` gives us all three in
        # one call without needing a second subprocess roundtrip.
        tag_meta = _dev_panel_git(
            "for-each-ref",
            "--format=%(objectname:short)|%(taggerdate:iso-strict)|%(subject)",
            f"refs/tags/{latest_tag}",
        )
        tag_sha = ""
        tag_date = ""
        tag_subject = ""
        if tag_meta:
            mparts = tag_meta.split("|", 2)
            if len(mparts) == 3:
                tag_sha, tag_date, tag_subject = mparts
        # Count commits on the current branch since the tag — this is
        # the "unreleased work" depth that tells the operator whether
        # a new patch cut is warranted.
        commits_since_tag = _dev_panel_git(
            "rev-list", "--count", f"{latest_tag}..HEAD"
        )
        latest_tag_info = {
            "tag": latest_tag,
            "sha": tag_sha,
            "tagged_at": tag_date,
            "subject": tag_subject,
            "commits_since": int(commits_since_tag) if commits_since_tag.isdigit() else 0,
        }
    elif (BASE_DIR / "CHANGELOG.md").exists():
        # Fallback only if no semver tags exist at all.
        for raw in (BASE_DIR / "CHANGELOG.md").read_text(errors="ignore").splitlines():
            line = raw.strip()
            if line.startswith("## [") and "Unreleased" not in line:
                current_release = line.split("[", 1)[1].split("]", 1)[0]
                break

    # Latest shipped = HEAD commit, richer than recent_commits[0] because
    # we resolve the full author + short body for the Latest-Shipped tile.
    latest_commit_info: dict = {}
    head_line = _dev_panel_git(
        "log", "-1", "--pretty=format:%h|%cI|%an|%s"
    )
    if head_line:
        hparts = head_line.split("|", 3)
        if len(hparts) == 4:
            h_sha, h_at, h_author, h_subject = hparts
            latest_commit_info = {
                "sha": h_sha,
                "at": h_at,
                "author": h_author,
                "subject": h_subject,
            }

    # Crawler-report freshness: mtime of the newest reports/*_log.json.
    # Dev-panel readers use this to answer "did the crawlers run lately?"
    # without opening a terminal.
    reports_dir = BASE_DIR / "reports"
    reports_latest_iso = ""
    reports_latest_age_hours: float | None = None
    if reports_dir.exists():
        newest = None
        for log in reports_dir.glob("*_log.json"):
            try:
                mt = log.stat().st_mtime
                if newest is None or mt > newest:
                    newest = mt
            except OSError:
                continue
        if newest is not None:
            reports_latest_iso = datetime.utcfromtimestamp(newest).isoformat(timespec="seconds")
            reports_latest_age_hours = max(0.0, (now.timestamp() - newest) / 3600.0)

    # Project timeline — every shipped tag, newest-first, with the
    # commits-since-tag depth hint for each. Powers the Mission
    # Control PROJECT TIMELINE tile on /admin/dev_panel. Shipped tags
    # only — no fake "future" placeholders. The operator decides what
    # v1.5.1 / v1.6.0 / v2.0 look like; the tile shows what's real.
    timeline: list[dict] = []
    # semver_tags is already sorted ascending; reverse for newest-first.
    for (m_mi_p, t) in reversed(semver_tags):
        major, minor, patch = m_mi_p
        tag_date = _dev_panel_git(
            "for-each-ref",
            "--format=%(taggerdate:short)",
            f"refs/tags/{t}",
        ).strip()
        is_latest = (t == latest_tag_info.get("tag", ""))
        timeline.append({
            "tag": t,
            "major": major,
            "minor": minor,
            "patch": patch,
            "tagged_at": tag_date,
            "is_latest": is_latest,
            "line": f"v{major}.{minor}.x",  # e.g. "v1.5.x" for grouping
        })

    return {
        "branch": branch,
        "ahead": ahead,
        "behind": behind,
        "upstream": upstream_ref,
        "dirty": dirty_count,
        "recent_commits": recent_commits,
        "commits_today": commits_today,
        "last_commit_at": last_commit_iso,
        "versions": versions,
        "current_release": current_release,
        "latest_tag": latest_tag_info,
        "latest_commit": latest_commit_info,
        "timeline": timeline,
        "reports_latest_at": reports_latest_iso,
        "reports_latest_age_hours": reports_latest_age_hours,
        "future_fixes": _dev_panel_future_fixes_count(),
    }


def _dev_panel_safe_doc_name(name: str) -> str | None:
    """Return the doc name if it is in the allowlist, else None."""
    return name if name in DEV_PANEL_DOC_FILES else None


@app.route("/admin/dev_panel")
@owner_required
def dev_panel():
    """Owner-only development control panel."""
    progress = _dev_panel_progress()
    waves = _dev_panel_waves()
    crawler_health = _dev_panel_crawler_health()
    # W1.4.3 c2 — hoist the "hot" wave to the top so the Now Shipping
    # hero tile can render it without re-scanning the list. Also compute
    # a quick pipeline breakdown for the same tile.
    hot_wave = next((w for w in waves if w.get("status") == "hot"), None)
    pipeline_shipped = sum(1 for w in waves if w.get("status") == "shipped")
    pipeline_total = len(waves)

    # ── Infrastructure tile data ──────────────────────────
    peak_cpu = 0.0
    peak_procs = 0
    stress_csv = os.path.join(app.root_path, "reports", "stress_monitor.csv")
    if os.path.isfile(stress_csv):
        try:
            import csv as _csv
            with open(stress_csv, newline="") as _f:
                reader = _csv.DictReader(_f)
                for row in reader:
                    try:
                        peak_cpu = max(peak_cpu, float(row.get("cpu_pct", 0)))
                    except (ValueError, TypeError):
                        pass
                    try:
                        peak_procs = max(peak_procs, int(row.get("python_procs", 0)))
                    except (ValueError, TypeError):
                        pass
        except Exception:
            pass

    infra_stats = {
        "engines": [
            {"name": "LLM (Claude Opus)", "type": "reasoning", "status": "active",
             "metric": "~8M tokens", "detail": "60+ agent spawns"},
            {"name": "MacBook Pro", "type": "compute", "status": "active",
             "metric": "M1 Pro 32 GB", "detail": "~100K checks/session"},
            {"name": "Mac Mini", "type": "compute", "status": "active",
             "metric": "M4 24 GB", "detail": "~100K checks via SSH"},
        ],
        "session": {
            "commits": progress.get("commits_today", 0),
            "lines_changed": 12000,
            "files_touched": 90,
            "crawlers_run": 500,
            "total_checks": 200000,
            "peak_cpu": round(peak_cpu, 1),
            "peak_processes": peak_procs,
        },
    }

    return render_template(
        "dev_panel.html",
        progress=progress,
        waves=waves,
        hot_wave=hot_wave,
        pipeline_shipped=pipeline_shipped,
        pipeline_total=pipeline_total,
        crawler_health=crawler_health,
        doc_files=DEV_PANEL_DOC_FILES,
        infra_stats=infra_stats,
        server_now_iso=datetime.now().astimezone().isoformat(timespec="seconds"),
    )


@app.route("/admin/dev_panel/doc")
@owner_required
def dev_panel_doc():
    """Return the raw text of an allowlisted doc for the in-page viewer."""
    name = request.args.get("name", "README.md")
    safe = _dev_panel_safe_doc_name(name)
    if safe is None:
        abort(404)
    # Resolution: README/CHANGELOG live at repo root; everything else
    # lives in docs/ after the v1.3.8 top-level cleanup. Try both so
    # old bookmarks keep working.
    path = BASE_DIR / safe
    if not path.exists():
        path = BASE_DIR / "docs" / safe
    if not path.exists():
        return jsonify({"name": safe, "content": f"({safe} not found)"}), 200
    try:
        text = path.read_text(errors="replace")
    except OSError as exc:
        return jsonify({"name": safe, "content": f"(error reading {safe}: {exc})"}), 200
    return jsonify({"name": safe, "content": text})


# ─── Mission Control drill-downs (v1.5.2) ──────────────────────────────
# Commit + tag detail pages — clickable from HISTORY and PROJECT
# TIMELINE tiles. Each reads from `git show` / `git for-each-ref`
# via the scrubbed `_dev_panel_git` helper so hook inheritance
# doesn't poison the subprocess env (see `6baca66` for the
# GIT_DIR scrubbing rationale).


import re as _mc_re  # local import — used only by the drill-down validators


def _dev_panel_safe_sha(sha: str) -> str | None:
    """Validate a git SHA is short-or-long hex before handing it to
    git. Prevents shell-injection via argument splitting and rejects
    any non-SHA input with a 404. Accepts 7-40 hex chars."""
    if not sha:
        return None
    if not _mc_re.fullmatch(r"[0-9a-f]{7,40}", sha):
        return None
    return sha


def _dev_panel_safe_tag(tag: str) -> str | None:
    """Validate a git tag name looks like a semver `v1.2.3` or the
    short `v1.2` / `v1.2.3.4` variants. Rejects everything else
    with a 404 so the drill-down can't be used as an arbitrary
    git-ref-peek gadget."""
    if not tag:
        return None
    if not _mc_re.fullmatch(r"v\d+(?:\.\d+){1,3}", tag):
        return None
    return tag


@app.route("/admin/dev_panel/commit/<sha>")
@owner_required
def dev_panel_commit(sha: str):
    """Mission Control drill-down: show a single commit's full body +
    files changed stat. Linked from the HISTORY tile rows."""
    safe = _dev_panel_safe_sha(sha)
    if safe is None:
        abort(404)
    # Header: sha / author / date / subject / body
    header = _dev_panel_git(
        "show", "--no-patch",
        "--format=%H%n%h%n%an%n%ae%n%cI%n%s%n%b",
        safe,
    )
    if not header:
        abort(404)
    parts = header.split("\n", 6)
    commit = {
        "full_sha": parts[0] if len(parts) > 0 else "",
        "short_sha": parts[1] if len(parts) > 1 else safe,
        "author_name": parts[2] if len(parts) > 2 else "",
        "author_email": parts[3] if len(parts) > 3 else "",
        "iso_date": parts[4] if len(parts) > 4 else "",
        "subject": parts[5] if len(parts) > 5 else "",
        "body": parts[6] if len(parts) > 6 else "",
    }
    # Files changed stat — `git show --stat --format=""` prints only
    # the diffstat portion, one line per file + a summary line.
    stat_raw = _dev_panel_git("show", "--stat", "--format=", safe)
    stat_lines = [ln for ln in stat_raw.splitlines() if ln.strip()]
    return render_template(
        "dev_panel_commit.html",
        commit=commit,
        stat_lines=stat_lines,
    )


@app.route("/admin/dev_panel/tag/<tag>")
@owner_required
def dev_panel_tag(tag: str):
    """Mission Control drill-down: show a tag's full annotated message
    plus every commit between the previous tag and this one. Linked
    from the PROJECT TIMELINE tile entries."""
    safe = _dev_panel_safe_tag(tag)
    if safe is None:
        abort(404)
    # Tag metadata: sha of the tagged commit, tagger name + date,
    # first-line subject, full body. `for-each-ref` returns empty
    # if the tag doesn't exist.
    meta = _dev_panel_git(
        "for-each-ref",
        "--format=%(objectname)%n%(objectname:short)%n%(taggername)%n"
        "%(taggerdate:iso-strict)%n%(subject)%n%(body)",
        f"refs/tags/{safe}",
    )
    if not meta:
        abort(404)
    mparts = meta.split("\n", 5)
    tag_info = {
        "name": safe,
        "full_sha": mparts[0] if len(mparts) > 0 else "",
        "short_sha": mparts[1] if len(mparts) > 1 else "",
        "tagger": mparts[2] if len(mparts) > 2 else "",
        "tagged_at": mparts[3] if len(mparts) > 3 else "",
        "subject": mparts[4] if len(mparts) > 4 else "",
        "body": mparts[5] if len(mparts) > 5 else "",
    }
    # Previous tag for the "commits since last tag" drill-down. If
    # there is no previous semver tag, we show no commits section.
    all_tags = [t.strip() for t in _dev_panel_git("tag", "--list").splitlines() if t.strip()]
    semver_sorted: list[tuple[tuple[int, ...], str]] = []
    for t in all_tags:
        if not t.startswith("v"):
            continue
        parts = t[1:].split(".")
        if 2 <= len(parts) <= 4 and all(p.isdigit() for p in parts):
            semver_sorted.append((tuple(int(p) for p in parts), t))
    semver_sorted.sort()
    prev_tag = None
    for i, (_, t) in enumerate(semver_sorted):
        if t == safe and i > 0:
            prev_tag = semver_sorted[i - 1][1]
            break
    commits_in_range: list[dict] = []
    if prev_tag:
        log_out = _dev_panel_git(
            "log", f"{prev_tag}..{safe}",
            "--pretty=format:%h|%cI|%an|%s",
        )
        for line in log_out.splitlines():
            lp = line.split("|", 3)
            if len(lp) == 4:
                commits_in_range.append({
                    "sha": lp[0], "at": lp[1], "author": lp[2], "subject": lp[3],
                })
    # Files changed since the previous tag (or the initial commit).
    stat_args = ["show", "--stat", "--format=", safe] if not prev_tag else [
        "diff", "--stat", f"{prev_tag}..{safe}"
    ]
    stat_raw = _dev_panel_git(*stat_args)
    stat_lines = [ln for ln in stat_raw.splitlines() if ln.strip()][-8:]  # last 8 lines = top files + summary
    return render_template(
        "dev_panel_tag.html",
        tag=tag_info,
        prev_tag=prev_tag,
        commits_in_range=commits_in_range,
        stat_lines=stat_lines,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Portfolio panel — owner-only.
# Reads state files written by ~/Downloads/portfolio-plan/{daily,analyze,simulate}.py
# and renders a dashboard with charts using the existing UI components and the
# Chart.js bundle already loaded by base.html.
# ─────────────────────────────────────────────────────────────────────────────
import csv as _csv  # local import — only used by portfolio routes

PORTFOLIO_DIR = Path(os.environ.get(
    "PORTFOLIO_PLAN_DIR",
    str(Path.home() / "Claude" / "portfolio-plan"),
))


def _portfolio_load_json(name: str) -> dict | None:
    p = PORTFOLIO_DIR / name
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text())
    except (OSError, ValueError):
        return None


def _portfolio_load_orders() -> list[dict]:
    p = PORTFOLIO_DIR / "orders.csv"
    if not p.exists():
        return []
    out: list[dict] = []
    try:
        with p.open() as f:
            reader = _csv.DictReader(f)
            for row in reader:
                row["amount_inr"] = int(float(row.get("amount_inr") or 0))
                row["nav"] = float(row.get("nav") or 0)
                row["units"] = float(row.get("units") or 0)
                out.append(row)
    except OSError:
        return []
    return out


def _portfolio_load_nav_history(max_days: int = 365) -> dict:
    """Read history/*.csv and return {fund_tag: [{date, nav}, ...]} truncated."""
    daily = _portfolio_load_json("daily_state.json") or {}
    isin_map = {tag: isin for tag, isin in (daily.get("fund_labels") or {}).items()}
    # Use the canonical map from market_snapshot.json instead — its keys are ISIN.
    snap = _portfolio_load_json("market_snapshot.json") or {}
    funds_meta = snap.get("funds", {})
    tag_to_isin = {info.get("tag"): isin for isin, info in funds_meta.items() if info.get("tag")}

    hist_dir = PORTFOLIO_DIR / "history"
    if not hist_dir.exists():
        return {}

    out: dict[str, list[dict]] = {}
    for tag in [
        "HDFC_FLEXI_CAP", "HDFC_FOCUSED", "NIPPON_LARGE_CAP",
        "JM_FLEXICAP", "NIPPON_MULTI_CAP", "SBI_CONTRA",
    ]:
        f = hist_dir / f"{tag}.csv"
        if not f.exists():
            continue
        try:
            with f.open() as fh:
                rows = list(_csv.DictReader(fh))
        except OSError:
            continue
        # rows are date,nav — keep last `max_days`
        rows = rows[-max_days:]
        out[tag] = [{"date": r.get("date"), "nav": float(r.get("nav") or 0)} for r in rows]
    return out


def _portfolio_compute_value_series(history: dict, orders: list[dict]) -> list[dict]:
    """Replay orders to produce a daily portfolio value series.
    Cheap, single-pass, no pandas. Returns [{date, value}, ...]."""
    if not history:
        return []
    # Build a unified date axis from any one fund's history (they're aligned).
    any_tag = next(iter(history))
    dates = [r["date"] for r in history[any_tag]]
    nav_by_date = {tag: {r["date"]: r["nav"] for r in rows} for tag, rows in history.items()}

    # Position book: tag → cumulative units up to and including each date.
    units_by_tag: dict[str, float] = {tag: 0.0 for tag in history}
    # Sort orders by date; ignore unknown tags.
    orders_sorted = sorted(
        [o for o in orders if o.get("fund_tag") in history and o.get("date")],
        key=lambda o: o["date"],
    )
    oi = 0

    series: list[dict] = []
    for d in dates:
        # Apply any orders dated on/before d that haven't been applied yet.
        while oi < len(orders_sorted) and orders_sorted[oi]["date"] <= d:
            o = orders_sorted[oi]
            units_by_tag[o["fund_tag"]] += float(o["units"] or 0)
            oi += 1
        total = 0.0
        for tag, units in units_by_tag.items():
            nav = nav_by_date.get(tag, {}).get(d)
            if nav is not None and units:
                total += units * nav
        series.append({"date": d, "value": round(total, 2)})
    return series


def _portfolio_state() -> dict:
    """Bundle everything the portfolio template needs into one dict."""
    daily = _portfolio_load_json("daily_state.json") or {}
    analysis = _portfolio_load_json("analysis_state.json") or {}
    simulation = _portfolio_load_json("simulation_results.json") or {}
    snap = _portfolio_load_json("market_snapshot.json") or {}
    commentary = _portfolio_load_json("commentary_state.json") or {}
    peer_compare = _portfolio_load_json("peer_compare_state.json") or {}
    horizon = _portfolio_load_json("horizon_state.json") or {}
    orders = _portfolio_load_orders()

    history = _portfolio_load_nav_history(max_days=365)
    value_series = _portfolio_compute_value_series(history, orders)

    # Equity allocation chart data (current vs target)
    weights = analysis.get("target_weights", {}) or daily.get("today", {}).get("baseline_per_fund", {})
    current_by_tag = analysis.get("current_equity_by_tag", {})
    equity_total = sum(current_by_tag.values()) or 1.0
    allocation = []
    fund_labels = (daily.get("fund_labels") or {})
    for tag, w in (analysis.get("target_weights") or {}).items():
        cur = current_by_tag.get(tag, 0.0)
        allocation.append({
            "tag": tag,
            "label": fund_labels.get(tag, tag),
            "target_pct": round(w * 100, 1),
            "current_pct": round(cur / equity_total * 100, 1) if equity_total else 0,
            "current_inr": round(cur, 2),
            "target_inr": round(w * (analysis.get("new_equity_total") or equity_total), 2),
        })

    # MTD progress per fund
    mtd = daily.get("month_to_date", {}) or {}
    monthly_target = daily.get("monthly_target_per_fund", {}) or {}
    mtd_progress = []
    for tag, tgt in monthly_target.items():
        spent = int(mtd.get(tag, 0))
        pct = round((spent / tgt) * 100, 1) if tgt else 0
        mtd_progress.append({
            "tag": tag,
            "label": fund_labels.get(tag, tag),
            "spent": spent,
            "target": int(tgt),
            "pct": min(pct, 100),
        })

    # Commentary freshness: stale if its today_total disagrees with the
    # current daily plan, or if it was generated before the latest daily run.
    commentary_fresh = False
    if commentary and commentary.get("text"):
        c_total = int(commentary.get("today_total") or 0)
        d_total = int(((daily.get("today") or {}).get("total")) or 0)
        c_at = (commentary.get("generated_at") or "")[:19]
        d_at = (daily.get("generated_at_ist") or "")[:19]
        # Compare the date portion only (timestamps live in different TZs)
        commentary_fresh = (c_total == d_total and c_total > 0
                            and c_at[:10] == d_at[:10])

    # Weekend banner: today is Sat/Sun → flag the next trading day
    weekend_banner = None
    today_local = date.today()
    if today_local.weekday() >= 5:  # 5=Sat, 6=Sun
        nxt = today_local
        # Skip Sat/Sun
        while nxt.weekday() >= 5:
            nxt = nxt + timedelta(days=1)
        weekend_banner = {
            "today_label": today_local.strftime("%A %Y-%m-%d"),
            "next_trading_day": nxt.strftime("%A %Y-%m-%d"),
        }

    # Build unified per-fund action lookup.
    # peer_compare carries SWITCH for own funds that should be replaced;
    # daily_state's fund_signals has BUY/SKIP from NAV z-score. We merge
    # them so the template can just do `pf.fund_actions.get(tag)` for a
    # single action/color pair per fund.
    fund_actions: dict[str, dict] = {}
    # Start from daily fund_signals
    for tag, fs in (daily.get("fund_signals") or {}).items():
        fund_actions[tag] = {
            "action": fs.get("action") or ("SKIP" if (fs.get("z_1y") or 0) >= 0.5 else "BUY"),
            "color": fs.get("action_color") or fs.get("color") or "#0a7c2f",
            "reason": fs.get("action_reason") or "",
            "swap_to": None,
            "swap_gain": 0.0,
        }
    # Override with REVIEW from peer_compare — flags funds worth discussing
    # but does NOT change BUY to something that blocks purchasing.
    for _cat, rows in (peer_compare.get("categories") or {}).items():
        if not isinstance(rows, list):
            continue
        for r in rows:
            if r.get("is_own") and r.get("own_tag") and r.get("action") == "REVIEW":
                tag = r["own_tag"]
                fund_actions[tag] = {
                    "action": "REVIEW",
                    "color": "#b86e00",
                    "reason": r.get("action_reason") or "Peers rank higher — worth reviewing.",
                    "swap_to": r.get("swap_to"),
                    "swap_gain": r.get("swap_gain") or 0.0,
                    "review_discussion": r.get("review_discussion"),
                }

    return {
        "exists": bool(daily or analysis),
        "portfolio_dir": str(PORTFOLIO_DIR),
        "daily": daily,
        "analysis": analysis,
        "simulation": simulation,
        "nifty": (snap.get("nifty") or {}),
        "fx": (snap.get("fx") or {}),
        "orders": orders[-25:][::-1],  # last 25, newest first
        "orders_count": len(orders),
        "allocation": allocation,
        "mtd_progress": mtd_progress,
        "value_series": value_series,
        "fund_labels": fund_labels,
        "commentary": commentary,
        "commentary_fresh": commentary_fresh,
        "weekend_banner": weekend_banner,
        "peer_compare": peer_compare,
        "fund_actions": fund_actions,
        "horizon": horizon,
    }


@app.route("/admin/portfolio")
@owner_required
def portfolio_panel():
    """Owner-only personal portfolio dashboard."""
    state = _portfolio_state()
    return render_template("portfolio.html", pf=state)


@app.route("/admin/portfolio/order", methods=["POST"])
@owner_required
def portfolio_log_order():
    """Append an executed order to portfolio-plan/orders.csv."""
    fund_tag = (request.form.get("fund_tag") or "").strip()
    amount_raw = (request.form.get("amount_inr") or "").strip()
    nav_raw = (request.form.get("nav") or "").strip()
    note = (request.form.get("note") or "").strip()
    order_date = (request.form.get("date") or date.today().isoformat()).strip()

    if not fund_tag or not amount_raw:
        flash("Fund and amount are required.", "error")
        return redirect(url_for("portfolio_panel"))
    try:
        amount = float(amount_raw)
    except ValueError:
        flash("Amount must be a number.", "error")
        return redirect(url_for("portfolio_panel"))

    # Resolve ISIN + NAV from market_snapshot if not supplied
    daily = _portfolio_load_json("daily_state.json") or {}
    snap = _portfolio_load_json("market_snapshot.json") or {}
    funds_meta = snap.get("funds", {})
    isin = ""
    last_nav = 0.0
    for i, info in funds_meta.items():
        if info.get("tag") == fund_tag:
            isin = i
            last_nav = float(info.get("last_nav") or 0)
            break

    nav = float(nav_raw) if nav_raw else last_nav
    units = round(amount / nav, 4) if nav else 0.0

    orders_path = PORTFOLIO_DIR / "orders.csv"
    is_new = not orders_path.exists()
    try:
        orders_path.parent.mkdir(parents=True, exist_ok=True)
        with orders_path.open("a", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=["date", "fund_tag", "isin", "amount_inr", "nav", "units", "note"])
            if is_new:
                w.writeheader()
            w.writerow({
                "date": order_date,
                "fund_tag": fund_tag,
                "isin": isin,
                "amount_inr": int(round(amount)),
                "nav": round(nav, 4),
                "units": units,
                "note": note,
            })
    except OSError as exc:
        flash(f"Could not write orders.csv: {exc}", "error")
        return redirect(url_for("portfolio_panel"))

    flash(f"Logged {fund_tag} ₹{int(amount):,} @ NAV {nav:.4f} → {units} units", "success")
    return redirect(url_for("portfolio_panel"))


@app.route("/admin/portfolio/order/bulk", methods=["POST"])
@owner_required
def portfolio_log_orders_bulk():
    """Log today's recommended amounts for the funds the user ticked.
    Reads canonical amounts from daily_state.json and the latest NAV from
    market_snapshot.json — never trusts client-supplied numbers."""
    selected = request.form.getlist("fund_tag")
    if not selected:
        flash("No funds selected.", "warning")
        return redirect(url_for("portfolio_panel"))

    daily = _portfolio_load_json("daily_state.json") or {}
    snap = _portfolio_load_json("market_snapshot.json") or {}
    today_amounts = ((daily.get("today") or {}).get("per_fund") or {})
    funds_meta = snap.get("funds", {})
    tag_to_isin = {info.get("tag"): isin for isin, info in funds_meta.items() if info.get("tag")}
    tag_to_nav = {info.get("tag"): float(info.get("last_nav") or 0) for info in funds_meta.values() if info.get("tag")}

    # Optional per-fund override amounts from the inline editable cells.
    # Anything missing or unparseable falls back to the canonical
    # recommendation in daily_state.json.
    overrides: dict[str, int] = {}
    for tag in selected:
        raw = (request.form.get(f"amount_{tag}") or "").strip()
        if not raw:
            continue
        try:
            v = int(round(float(raw)))
        except ValueError:
            continue
        if v > 0:
            overrides[tag] = v

    today_iso = date.today().isoformat()
    orders_path = PORTFOLIO_DIR / "orders.csv"
    is_new = not orders_path.exists()
    logged = 0
    total_inr = 0
    try:
        orders_path.parent.mkdir(parents=True, exist_ok=True)
        with orders_path.open("a", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=["date", "fund_tag", "isin", "amount_inr", "nav", "units", "note"])
            if is_new:
                w.writeheader()
            for tag in selected:
                canonical = int(today_amounts.get(tag, 0) or 0)
                amt = overrides.get(tag, canonical)
                if amt <= 0:
                    continue
                isin = tag_to_isin.get(tag, "")
                nav = tag_to_nav.get(tag, 0.0)
                units = round(amt / nav, 4) if nav else 0.0
                w.writerow({
                    "date": today_iso,
                    "fund_tag": tag,
                    "isin": isin,
                    "amount_inr": amt,
                    "nav": round(nav, 4),
                    "units": units,
                    "note": "manual-override" if tag in overrides and overrides[tag] != canonical else "from-recommendation",
                })
                logged += 1
                total_inr += amt
    except OSError as exc:
        flash(f"Could not write orders.csv: {exc}", "error")
        return redirect(url_for("portfolio_panel"))

    if logged == 0:
        flash("Selected funds had no recommended amount today (skip / 0).", "warning")
    else:
        flash(f"Logged {logged} order{'s' if logged != 1 else ''} totalling ₹{total_inr:,} from today's recommendation.", "success")
    return redirect(url_for("portfolio_panel"))


@app.route("/admin/portfolio/calendar-events")
@owner_required
def portfolio_calendar_events():
    """FullCalendar event feed for the portfolio page. Reuses the same
    library and styling as the lab scheduler calendar — each forecast
    day yields up to two all-day events: an amber REDEEM and a green BUY."""
    daily = _portfolio_load_json("daily_state.json") or {}
    forecast = daily.get("forecast", []) or []
    events = []
    for f in forecast:
        d = f.get("date")
        if not d:
            continue
        redeem_amt = int(f.get("redeem_today") or 0)
        buy_amt = int(f.get("total") or 0)
        if redeem_amt > 0:
            events.append({
                "title": f"⬇ REDEEM ₹{redeem_amt:,}",
                "start": d,
                "allDay": True,
                "backgroundColor": "#a37b00",
                "borderColor": "#a37b00",
                "textColor": "#fff",
                "extendedProps": {
                    "kind": "redeem",
                    "funds_buy_on": f.get("funds_buy_on"),
                    "phase": f.get("phase"),
                },
            })
        if buy_amt > 0:
            color = f.get("color") or "#3aa86b"
            events.append({
                "title": f"⬆ BUY ₹{buy_amt:,}",
                "start": d,
                "allDay": True,
                "backgroundColor": color,
                "borderColor": color,
                "textColor": "#fff",
                "extendedProps": {
                    "kind": "buy",
                    "label": f.get("label"),
                    "multiplier": f.get("multiplier"),
                },
            })
    # Today's REDEEM gets its own marker in case it's not in the 14d forecast
    today = (daily.get("today") or {})
    redeem_today = (today.get("redeem") or {})
    if redeem_today.get("amount") and redeem_today.get("sell_day"):
        d = redeem_today["sell_day"]
        if not any(e["start"] == d and e["extendedProps"].get("kind") == "redeem" for e in events):
            events.append({
                "title": f"⬇ REDEEM ₹{int(redeem_today['amount']):,}",
                "start": d,
                "allDay": True,
                "backgroundColor": "#a37b00",
                "borderColor": "#a37b00",
                "textColor": "#fff",
                "extendedProps": {
                    "kind": "redeem",
                    "funds_buy_on": redeem_today.get("target_buy_date"),
                    "phase": redeem_today.get("phase"),
                },
            })
    return jsonify(events)


@app.route("/admin/portfolio/refresh", methods=["POST"])
@owner_required
def portfolio_refresh():
    """Run `./run.sh quick` in PORTFOLIO_DIR to refresh state files."""
    script = PORTFOLIO_DIR / "run.sh"
    if not script.exists():
        flash(f"run.sh not found in {PORTFOLIO_DIR}", "error")
        return redirect(url_for("portfolio_panel"))
    try:
        result = subprocess.run(
            ["bash", str(script), "quick"],
            cwd=str(PORTFOLIO_DIR),
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            tail = (result.stderr or result.stdout)[-300:]
            flash(f"Refresh failed: {tail}", "error")
        else:
            flash("Portfolio data refreshed.", "success")
    except Exception as exc:
        flash(f"Refresh error: {exc}", "error")
    return redirect(url_for("portfolio_panel"))


@app.route("/admin/portfolio/recompute-peers", methods=["POST"])
@owner_required
def portfolio_recompute_peers():
    """Run ONLY peer_compare.py — faster than a full refresh because the
    24h NAV cache means no new HTTP fetches on a same-day re-run. Used by
    the Recompute button in the IDEAL PORTFOLIO card to refresh Sharpe,
    rolling-returns consistency, manager tenure, and swap calls without
    touching the daily buy plan or NAV snapshots.
    """
    script = PORTFOLIO_DIR / "peer_compare.py"
    if not script.exists():
        flash(f"peer_compare.py not found in {PORTFOLIO_DIR}", "error")
        return redirect(url_for("portfolio_panel"))
    venv_py = PORTFOLIO_DIR / ".venv" / "bin" / "python"
    py = str(venv_py) if venv_py.exists() else "python3"
    try:
        result = subprocess.run(
            [py, str(script)],
            cwd=str(PORTFOLIO_DIR),
            capture_output=True,
            text=True,
            timeout=180,
        )
        if result.returncode != 0:
            tail = (result.stderr or result.stdout)[-300:]
            flash(f"Peer recompute failed: {tail}", "error")
        else:
            flash("Peer rankings + consistency + manager tenure refreshed.", "success")
    except subprocess.TimeoutExpired:
        flash("Peer recompute timed out (>180s). Check mfapi.in availability.", "error")
    except Exception as exc:
        flash(f"Peer recompute error: {exc}", "error")
    return redirect(url_for("portfolio_panel"))


@app.route("/requests/<int:request_id>/duplicate")
@login_required
def duplicate_request(request_id: int):
    """Create a new request pre-filled from an existing one."""
    user = current_user()
    sr = query_one("SELECT * FROM sample_requests WHERE id = ?", (request_id,))
    if sr is None:
        abort(404)
    if not can_view_request(user, sr):
        abort(403)
    # Redirect to new_request with pre-fill query params
    params = {
        "instrument_id": sr["instrument_id"],
        "title": f"Copy of {sr['title']}",
        "sample_name": sr["sample_name"],
        "sample_count": sr["sample_count"],
        "description": sr["description"] or "",
        "sample_origin": sr["sample_origin"] or "internal",
        "priority": sr["priority"] or "normal",
    }
    return redirect(url_for("new_request", **params))


@app.route("/schedule")
@login_required
def schedule():
    user = current_user()
    if not can_access_schedule(user):
        abort(403)
    filters = schedule_filter_values()
    queue_back_url = request.args.get("back", "").strip()
    queue_source_label = request.args.get("source_label", "").strip()
    focus_request_id = request.args.get("focus_request_id", "").strip()
    visible_instruments = visible_instruments_for_user(user)
    if len(visible_instruments) == 1 and not filters.get("instrument_id"):
        filters["instrument_id"] = str(visible_instruments[0]["id"])
    clauses, params = request_scope_sql(user, "sr")
    schedule_query_filters = {"sort": "created_desc"}  # always fetch newest-first from DB
    base_sql, base_params = request_history_query(clauses, params, schedule_query_filters)
    rows_all = query_all(base_sql, tuple(base_params))
    live_statuses = {"submitted", "under_review", "awaiting_sample_submission", "sample_submitted", "sample_received", "scheduled", "in_progress"}
    live_rows = [row for row in rows_all if row["status"] in live_statuses]
    completed_rows = [row for row in rows_all if row["status"] == "completed"]
    rejected_rows = [row for row in rows_all if row["status"] == "rejected"]
    soft_wait_rows = [row for row in live_rows if row["status"] == "submitted"]
    approval_rows = [row for row in live_rows if row["status"] == "under_review"]
    awaiting_sample_rows = [row for row in live_rows if row["status"] == "awaiting_sample_submission"]
    pending_receipt_rows = [row for row in live_rows if row["status"] == "sample_submitted"]
    ready_rows = [row for row in live_rows if row["status"] == "sample_received"]
    active_rows = [row for row in live_rows if row["status"] in {"scheduled", "in_progress"}]
    selected_instrument = None
    if filters.get("instrument_id"):
        try:
            instrument_id = int(filters["instrument_id"])
        except (TypeError, ValueError):
            instrument_id = None
        if instrument_id is not None:
            selected_instrument = next((row for row in visible_instruments if row["id"] == instrument_id), None)
    instruments = visible_instruments
    operators = query_all("SELECT id, name FROM users WHERE role IN ('operator', 'instrument_admin', 'super_admin') ORDER BY name")
    requesters = query_all("SELECT id, name FROM users WHERE role = 'requester' ORDER BY name")
    attachment_map = attachments_by_request_ids([row["id"] for row in rows_all])
    profile = user_access_profile(user)
    can_operate_queue = bool(
        {"reassign", "mark_received"} & set(profile["card_action_fields"])
    )
    return render_template(
        "schedule.html",
        queue_rows=rows_all,
        live_rows=live_rows,
        completed_rows=completed_rows,
        rejected_rows=rejected_rows,
        soft_wait_rows=soft_wait_rows,
        approval_rows=approval_rows,
        awaiting_sample_rows=awaiting_sample_rows,
        pending_receipt_rows=pending_receipt_rows,
        ready_rows=ready_rows,
        active_rows=active_rows,
        filters=filters,
        instruments=instruments,
        instrument_selector_enabled=len(instruments) > 1,
        operators=operators,
        requesters=requesters,
        selected_instrument=selected_instrument,
        focus_request_id=focus_request_id,
        attachment_map=attachment_map,
        queue_back_url=queue_back_url,
        queue_source_label=queue_source_label,
        can_operate_queue=can_operate_queue,
    )


@app.route("/schedule/bulk", methods=["POST"])
@login_required
def schedule_bulk_actions():
    """Apply one action across many selected rows on /schedule.

    Currently supports a single action: ``bulk_assign`` — assigns a chosen
    operator to every selected request. Skips rows the caller cannot
    operate or where the target operator is not a candidate for that
    instrument. Reports applied/skipped counts via flash.
    """
    user = current_user()
    if not can_access_schedule(user):
        abort(403)
    raw_ids = (request.form.get("request_ids") or "").strip()
    if not raw_ids:
        flash("Select at least one job.", "error")
        return redirect(url_for("schedule"))
    try:
        request_ids = [int(x) for x in raw_ids.split(",") if x.strip()]
    except ValueError:
        flash("Invalid selection.", "error")
        return redirect(url_for("schedule"))

    action = (request.form.get("action") or "").strip()
    back_value = (request.form.get("back") or "").strip()

    if action == "bulk_assign":
        operator_id = int(request.form.get("assigned_operator_id") or 0)
        if not operator_id:
            flash("Choose an operator.", "error")
            return redirect(back_value or url_for("schedule"))
        applied = 0
        skipped_perm = 0
        skipped_status = 0
        skipped_candidate = 0
        for rid in request_ids:
            sr = query_one(
                "SELECT * FROM sample_requests WHERE id = ?",
                (rid,),
            )
            if sr is None:
                skipped_perm += 1
                continue
            if not (
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                can_manage_instrument(user["id"], sr["instrument_id"], user["role"])
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                or can_operate_instrument(user["id"], sr["instrument_id"], user["role"])
            ):
                skipped_perm += 1
                continue
            if sr["status"] in {"completed", "rejected", "submitted"}:
                skipped_status += 1
                continue
            candidate_ids = {row["id"] for row in request_assignment_candidates(sr)}
            if operator_id not in candidate_ids:
                skipped_candidate += 1
                continue
            execute(
                "UPDATE sample_requests SET assigned_operator_id = ?, updated_at = ? WHERE id = ?",
                (operator_id, now_iso(), rid),
            )
            log_action(
                user["id"],
                "sample_request",
                rid,
                "reassigned",
                {"assigned_operator_id": operator_id, "remarks": "Bulk assign from queue"},
            )
            write_request_metadata_snapshot(rid)
            applied += 1
        parts = [f"{applied} assigned"]
        if skipped_status:
            parts.append(f"{skipped_status} locked")
        if skipped_candidate:
            parts.append(f"{skipped_candidate} not eligible")
        if skipped_perm:
            parts.append(f"{skipped_perm} no access")
        flash(" · ".join(parts), "success" if applied else "error")
        return redirect(back_value or url_for("schedule"))

    flash("Unknown bulk action.", "error")
    return redirect(back_value or url_for("schedule"))


@app.route("/schedule/actions", methods=["POST"])
@login_required
def schedule_actions():
    user = current_user()
    if not can_access_schedule(user):
        abort(403)
    request_id = int(request.form["request_id"])
    sample_request = query_one(
        """
        SELECT sr.*, i.name AS instrument_name, r.name AS requester_name, r.email AS requester_email
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        JOIN users r ON r.id = sr.requester_id
        WHERE sr.id = ?
        """,
        (request_id,),
    )
    if sample_request is None:
        abort(404)
    if not can_view_request(user, sample_request):
        abort(403)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_manage = can_manage_instrument(user["id"], sample_request["instrument_id"], user["role"])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_operate = can_operate_instrument(user["id"], sample_request["instrument_id"], user["role"])
    if not (can_manage or can_operate):
        abort(403)
    if sample_request["completion_locked"]:
        flash("Completed jobs are locked.", "error")
        return redirect(url_for("schedule", bucket="scheduled"))

    def redirect_to_queue(bucket_override: str | None = None, focus_request: bool = False):
        # If a back URL was provided (e.g. from dashboard quick intake),
        # redirect there instead of the schedule page.
        back_url = (request.form.get("back") or "").strip()
        if back_url and back_url.startswith("/"):
            return redirect(back_url)
        params: dict[str, str] = {
            "bucket": bucket_override or (request.form.get("bucket") or "all"),
        }
        for key in ("q", "instrument_id", "date_from", "date_to", "source_label"):
            value = (request.form.get(key) or "").strip()
            if value:
                params[key] = value
        if focus_request:
            params["focus_request_id"] = str(request_id)
        return redirect(url_for("schedule", **params))

    action = request.form.get("action", "").strip()
    if action == "take_up":
        if sample_request["status"] not in {"sample_received", "scheduled"}:
            flash("Only received jobs can be taken up from the board.", "error")
            return redirect_to_queue()
        scheduled_for = request.form.get("scheduled_for", "").strip()
        if not scheduled_for:
            flash("Please choose a schedule time.", "error")
            return redirect_to_queue()
        operator_id = int(request.form.get("assigned_operator_id") or user["id"])
        remarks = request.form.get("remarks", "").strip()
        assert_status_transition(sample_request["status"], "scheduled")
        execute(
            "UPDATE sample_requests SET status = 'scheduled', scheduled_for = ?, assigned_operator_id = ?, remarks = ?, updated_at = ? WHERE id = ?",
            (scheduled_for, operator_id, remarks, now_iso(), request_id),
        )
        log_action(user["id"], "sample_request", request_id, "scheduled_from_board", {"scheduled_for": scheduled_for, "assigned_operator_id": operator_id})
        flash(f"{sample_request['request_no']} taken up for work.", "success")
        write_request_metadata_snapshot(request_id)
        return redirect_to_queue(bucket_override="all", focus_request=True)
    elif action == "quick_assign":
        operator_id = int(request.form.get("assigned_operator_id") or 0)
        if not operator_id:
            flash("Choose a person to assign.", "error")
            return redirect_to_queue()
        candidate_ids = {row["id"] for row in request_assignment_candidates(sample_request)}
        if operator_id not in candidate_ids:
            flash("That person cannot be assigned to this job.", "error")
            return redirect_to_queue()
        execute(
            "UPDATE sample_requests SET assigned_operator_id = ?, updated_at = ? WHERE id = ?",
            (operator_id, now_iso(), request_id),
        )
        log_action(user["id"], "sample_request", request_id, "reassigned", {"assigned_operator_id": operator_id, "remarks": "Assigned from queue"})
        flash(f"{sample_request['request_no']} reassigned.", "success")
        write_request_metadata_snapshot(request_id)
        return redirect_to_queue(focus_request=True)
    elif action == "plan_next_slot":
        if sample_request["status"] not in {"sample_received"}:
            flash("Only ready jobs can be placed into the day planner.", "error")
            return redirect_to_queue()
        planner_date = parse_schedule_day(request.form.get("planner_date"))
        scope_filters = schedule_filter_values()
        clauses, params = request_scope_sql(user, "sr")
        if scope_filters.get("requester_id"):
            clauses.append("sr.requester_id = ?")
            params.append(int(scope_filters["requester_id"]))
        base_sql, base_params = request_history_query(clauses, params, scope_filters)
        rows_all = [row for row in query_all(base_sql, tuple(base_params)) if row_matches_period(row, scope_filters["period"])]
        same_day_rows = []
        for row in rows_all:
            if row["status"] not in {"scheduled", "in_progress"} or not row["scheduled_for"]:
                continue
            try:
                parsed = datetime.fromisoformat(str(row["scheduled_for"]).replace("Z", "+00:00"))
            except ValueError:
                try:
                    parsed = datetime.strptime(str(row["scheduled_for"])[:16], "%Y-%m-%dT%H:%M")
                except ValueError:
                    parsed = None
            if parsed and parsed.date() == planner_date:
                same_day_rows.append(row)
        scheduled_for_raw = (request.form.get("scheduled_for") or "").strip()
        if scheduled_for_raw:
            try:
                slot_dt = datetime.fromisoformat(scheduled_for_raw)
            except ValueError:
                flash("Choose a valid date and time for the planner.", "error")
                return redirect_to_queue()
        else:
            slot_dt = compute_next_schedule_slot(planner_date, same_day_rows)
        operator_id = int(request.form.get("assigned_operator_id") or user["id"])
        remarks = request.form.get("remarks", "").strip()
        assert_status_transition(sample_request["status"], "scheduled")
        execute(
            "UPDATE sample_requests SET status = 'scheduled', scheduled_for = ?, assigned_operator_id = ?, remarks = ?, updated_at = ? WHERE id = ?",
            (slot_dt.isoformat(timespec="minutes"), operator_id, remarks, now_iso(), request_id),
        )
        log_action(
            user["id"],
            "sample_request",
            request_id,
            "scheduled_from_board",
            {"scheduled_for": slot_dt.isoformat(timespec="minutes"), "assigned_operator_id": operator_id, "planner_date": planner_date.isoformat()},
        )
        write_request_metadata_snapshot(request_id)
        flash(f"{sample_request['request_no']} added to {planner_date.strftime('%d/%m/%Y')} at {slot_dt.strftime('%H:%M')}.", "success")
        return redirect_to_queue(bucket_override="all", focus_request=True)
    elif action == "mark_received":
        if sample_request["status"] != "sample_submitted":
            flash("Only submitted samples can be marked received from the board.", "error")
            return redirect_to_queue()
        assert_status_transition(sample_request["status"], "sample_received")
        execute(
            """
            UPDATE sample_requests
            SET status = 'sample_received', sample_received_at = ?, received_by_operator_id = ?, updated_at = ?
            WHERE id = ?
            """,
            (now_iso(), user["id"], now_iso(), request_id),
        )
        log_action(user["id"], "sample_request", request_id, "sample_received", {"remarks": ""})
        flash(f"{sample_request['request_no']} marked received.", "success")
        write_request_metadata_snapshot(request_id)
        return redirect_to_queue(bucket_override="all", focus_request=True)
    elif action == "start_now":
        if sample_request["status"] not in {"scheduled", "in_progress"}:
            flash("Only scheduled jobs can be started from the board.", "error")
            return redirect_to_queue()
        remarks = request.form.get("remarks", "").strip()
        operator_id = sample_request["assigned_operator_id"] or user["id"]
        assert_status_transition(sample_request["status"], "in_progress")
        execute(
            "UPDATE sample_requests SET status = 'in_progress', assigned_operator_id = ?, remarks = ?, updated_at = ? WHERE id = ?",
            (operator_id, remarks, now_iso(), request_id),
        )
        log_action(user["id"], "sample_request", request_id, "started_from_board", {})
        flash(f"{sample_request['request_no']} is now in progress.", "success")
        write_request_metadata_snapshot(request_id)
        return redirect_to_queue(bucket_override="all", focus_request=True)
    elif action == "finish_now":
        if sample_request["status"] != "in_progress":
            flash("Only in-progress jobs can be finished from the board.", "error")
            return redirect_to_queue()
        results_summary = request.form.get("results_summary", "").strip()
        if not results_summary:
            flash("Please add a short result summary before finishing.", "error")
            return redirect_to_queue()
        remarks = request.form.get("remarks", "").strip()
        _finance = computed_finance_for_request(get_db(), request_id)
        amount_paid = float(request.form.get("amount_paid") or _finance["amount_paid"] or 0)
        finance_status = request.form.get("finance_status", _finance["finance_status"])
        email_ok, email_message = send_completion_email(sample_request, results_summary)
        now_value = now_iso()
        completion_fields = completion_override_fields(sample_request, user["id"], now_value)
        assert_status_transition(sample_request["status"], "completed")
        execute(
            """
            UPDATE sample_requests
            SET status = 'completed', results_summary = ?, remarks = ?,
                result_email_status = ?, result_email_sent_at = ?, completion_locked = 1,
                submitted_to_lab_at = ?, sample_submitted_at = ?, sample_received_at = ?, received_by_operator_id = ?,
                scheduled_for = ?, assigned_operator_id = ?, completed_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                results_summary,
                remarks,
                email_message,
                now_value if email_ok else None,
                completion_fields["submitted_to_lab_at"],
                completion_fields["sample_submitted_at"],
                completion_fields["sample_received_at"],
                completion_fields["received_by_operator_id"],
                completion_fields["scheduled_for"],
                completion_fields["assigned_operator_id"],
                completion_fields["completed_at"],
                now_value,
                request_id,
            ),
        )
        sync_request_to_peer_aggregates(
            get_db(), request_id,
            amount_due=_finance["amount_due"] or amount_paid,
            amount_paid=amount_paid,
            finance_status=finance_status,
            receipt_number=_finance["receipt_number"],
        )
        get_db().commit()
        log_completion_override_events(
            user["id"],
            sample_request,
            completion_fields,
            now_value,
            "completed_from_board",
            {"results_summary": results_summary, "email_status": email_message},
        )
        send_completion_inbox_message(user["id"], sample_request)
        flash("Job marked done.", "success")
        write_request_metadata_snapshot(request_id)
        return redirect_to_queue(bucket_override="all", focus_request=True)
    else:
        abort(403)

    write_request_metadata_snapshot(request_id)
    return redirect_to_queue()


@app.route("/attachments/<int:attachment_id>/download")
@login_required
def download_attachment(attachment_id: int):
    attachment = query_one("SELECT * FROM request_attachments WHERE id = ? AND is_active = 1", (attachment_id,))
    if attachment is None:
        abort(404)
    request_row = query_one("SELECT * FROM sample_requests WHERE id = ?", (attachment["request_id"],))
    user = current_user()
    if request_row is None or not can_view_request(user, request_row):
        abort(403)
    full_path = BASE_DIR / attachment["relative_path"]
    if not full_path.exists() or not str(full_path.resolve()).startswith(str(UPLOAD_DIR.resolve())):
        abort(404)
    return send_file(full_path, as_attachment=True, download_name=attachment["original_filename"], mimetype=attachment["mime_type"])


@app.route("/attachments/<int:attachment_id>/view")
@login_required
def view_attachment(attachment_id: int):
    attachment = query_one("SELECT * FROM request_attachments WHERE id = ? AND is_active = 1", (attachment_id,))
    if attachment is None:
        abort(404)
    request_row = query_one("SELECT * FROM sample_requests WHERE id = ?", (attachment["request_id"],))
    user = current_user()
    if request_row is None or not can_view_request(user, request_row):
        abort(403)
    full_path = BASE_DIR / attachment["relative_path"]
    if not full_path.exists() or not str(full_path.resolve()).startswith(str(UPLOAD_DIR.resolve())):
        abort(404)
    return send_file(full_path, as_attachment=False, download_name=attachment["original_filename"], mimetype=attachment["mime_type"])


@app.route("/attachments/<int:attachment_id>/delete", methods=["POST"])
@login_required
def delete_attachment(attachment_id: int):
    attachment = query_one("SELECT * FROM request_attachments WHERE id = ? AND is_active = 1", (attachment_id,))
    if attachment is None:
        abort(404)
    request_row = query_one("SELECT * FROM sample_requests WHERE id = ?", (attachment["request_id"],))
    user = current_user()
    if request_row is None or not can_delete_attachment(user, attachment, request_row):
        abort(403)
    execute("UPDATE request_attachments SET is_active = 0 WHERE id = ?", (attachment_id,))
    write_request_metadata_snapshot(request_row["id"])
    log_action(
        user["id"],
        "sample_request",
        request_row["id"],
        "attachment_removed",
        {
            "filename": attachment["original_filename"],
            "attachment_type": attachment["attachment_type"],
            "note": attachment["note"],
        },
    )
    flash("Attachment removed.", "success")
    return redirect(url_for("request_detail", request_id=request_row["id"]))


@app.route("/my/history")
@login_required
def my_history():
    args = request.args.to_dict()
    return redirect(url_for("schedule", **args))

@app.route("/me")
@login_required
def my_profile():
    return redirect(url_for("user_profile", user_id=current_user()["id"]))


@app.route("/profile/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user = current_user()
    if request.method == "POST":
        current_pw = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")
        if not check_password_hash(user["password_hash"], current_pw):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("change_password"))
        if len(new_pw) < 8:
            flash("New password must be at least 8 characters.", "error")
            return redirect(url_for("change_password"))
        if new_pw != confirm_pw:
            flash("New passwords do not match.", "error")
            return redirect(url_for("change_password"))
        execute(
            "UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?",
            (generate_password_hash(new_pw, method="pbkdf2:sha256"), user["id"]),
        )
        log_action(user["id"], "user", user["id"], "password_changed", {})
        flash("Password changed successfully.", "success")
        return redirect(url_for("my_profile"))
    return render_template("change_password.html", title="Change Password")


@app.route("/history/processed")
@login_required
def processed_history():
    args = request.args.to_dict()
    args.setdefault("bucket", "completed")
    return redirect(url_for("schedule", **args))


@app.route("/users/<int:user_id>", methods=["GET", "POST"])
@login_required
def user_profile(user_id: int):
    viewer = current_user()
    target_user = query_one("SELECT id, name, email, role, invite_status, active, member_code FROM users WHERE id = ?", (user_id,))
    if target_user is None:
        abort(404)
    if not can_view_user_profile(viewer, target_user):
        abort(403)
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "remove_access":
            if not can_manage_members(viewer):
                abort(403)
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if target_user["role"] != "requester" or is_owner(target_user) or target_user["id"] == viewer["id"]:
                abort(403)
            execute("UPDATE users SET active = 0 WHERE id = ?", (user_id,))
            log_action(viewer["id"], "user", user_id, "member_deactivated", {"email": target_user["email"]})
            flash(f"Access removed for {target_user['email']}.", "success")
            return redirect(url_for("user_profile", user_id=user_id))
        if action == "reset_password":
            # Password hygiene (W1.3.8) — admin-issued password reset.
            # Admin never types a password. We generate a random temp
            # password, flash it to the admin exactly once, and set
            # must_change_password=1 so the target is forced to pick
            # their own on next login. Owner rows are never reset from
            # here; super_admin rows can only be reset by another
            # super_admin.
            if not can_manage_members(viewer):
                abort(403)
            if is_owner(target_user) and not is_owner(viewer):
                abort(403)
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if target_user["role"] == "super_admin" and viewer["role"] != "super_admin":
                abort(403)
            temp_password = generate_temp_password()
            execute(
                "UPDATE users SET password_hash = ?, must_change_password = 1 WHERE id = ?",
                (generate_password_hash(temp_password, method="pbkdf2:sha256"), user_id),
            )
            log_action(
                viewer["id"], "user", user_id, "password_reset_by_admin",
                {"email": target_user["email"]},
            )
            # Store in session so the template can render the PIN card
            # and the .eml download route can build the file. Cleared
            # on next page load or when the .eml is downloaded.
            session["_reset_pw"] = {
                "user_id": user_id,
                "email": target_user["email"],
                "name": target_user["name"],
                "temp_password": temp_password,
            }
            flash(
                f"Temporary password issued for {target_user['email']}. "
                f"See the PIN card below to copy or email it.",
                "success",
            )
            return redirect(url_for("user_profile", user_id=user_id))
        if action == "update_user_metadata":
            # In-place admin edit of a user's profile. Shown behind the
            # "Edit" toggle on the User Metadata tile; same pattern as the
            # instrument metadata tile. Only members-managers can save —
            # non-super_admins cannot touch a super_admin or owner row,
            # and no one can demote themselves via this form.
            if not can_manage_members(viewer):
                abort(403)
            if is_owner(target_user) and not is_owner(viewer):
                abort(403)
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if target_user["role"] == "super_admin" and viewer["role"] != "super_admin":
                abort(403)
            new_name = request.form.get("name", target_user["name"]).strip() or target_user["name"]
            new_member_code = request.form.get("member_code", target_user["member_code"] or "").strip() or None
            new_active = 1 if request.form.get("active") == "on" else 0
            if target_user["id"] == viewer["id"]:
                # Never let an admin deactivate themselves by accident.
                new_active = 1
            execute(
                "UPDATE users SET name = ?, member_code = ?, active = ? WHERE id = ?",
                (new_name, new_member_code, new_active, user_id),
            )
            log_action(
                viewer["id"],
                "user",
                user_id,
                "user_metadata_updated",
                {
                    "name": new_name,
                    "member_code": new_member_code,
                    "active": new_active,
                },
            )
            flash(f"Profile updated for {new_name}.", "success")
            return redirect(url_for("user_profile", user_id=user_id))
        if action == "change_role":
            # Per-user role promotion / demotion. Site-admin+ can move a
            # user between requester, operator, instrument_admin,
            # faculty_in_charge, professor_approver, and finance_admin.
            # Only super_admin can promote to or demote from site_admin,
            # and super_admin rows can only be touched by a super_admin.
            # Owner rows are never demoted here.
            if not can_manage_members(viewer):
                abort(403)
            if is_owner(target_user):
                abort(403)
            if target_user["id"] == viewer["id"]:
                abort(403)
            new_role = (request.form.get("new_role") or "").strip()
            all_roles = {
                "requester", "operator", "instrument_admin",
                "faculty_in_charge", "professor_approver", "finance_admin",
                "site_admin", "super_admin",
            }
            is_xhr = request.headers.get("X-Requested-With") == "XMLHttpRequest"
            if new_role not in all_roles:
                if is_xhr:
                    return jsonify({"ok": False, "error": "invalid_role"}), 400
                flash("Pick a valid role.", "error")
                return redirect(url_for("user_profile", user_id=user_id))
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if new_role in {"site_admin", "super_admin"} and viewer["role"] != "super_admin":
                abort(403)
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if target_user["role"] in {"site_admin", "super_admin"} and viewer["role"] != "super_admin":
                abort(403)
            execute(
                "UPDATE users SET role = ?, invite_status = 'active', active = 1 WHERE id = ?",
                (new_role, user_id),
            )
            log_action(
                viewer["id"], "user", user_id, "user_role_changed",
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                {"from_role": target_user["role"], "to_role": new_role, "email": target_user["email"]},
            )
            # Make sure the new primary role is also in the role set.
            grant_user_role(user_id, new_role, viewer["id"])
            if is_xhr:
                # W1.4.12 — third consumer of the inline-toggle XHR
                # pattern. Client soft-reloads so the tile, role set,
                # and permission-gated sections refresh authoritatively.
                return jsonify({
                    "ok": True,
                    "new_role": new_role,
                    "new_role_label": role_display_name(new_role),
                    "reload_url": url_for("user_profile", user_id=user_id),
                })
            flash(f"Role updated to {role_display_name(new_role)}.", "success")
            return redirect(url_for("user_profile", user_id=user_id))
        if action == "update_user_role_set":
            # W1.3.7 — layer additional roles on top of users.role.
            # The primary role (users.role) is controlled by
            # change_role above; this form only manipulates the
            # additional-role junction table. Checked roles are
            # granted, unchecked are revoked. The primary role is
            # always force-granted so the set never goes below the
            # single-role baseline the rest of the app expects.
            if not can_manage_members(viewer):
                abort(403)
            if is_owner(target_user) and not is_owner(viewer):
                abort(403)
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if target_user["role"] == "super_admin" and viewer["role"] != "super_admin":
                abort(403)
            checked = set(request.form.getlist("extra_roles"))
            layered_roles = {
                "operator", "instrument_admin", "faculty_in_charge",
                "professor_approver", "finance_admin",
            }
            # site_admin / super_admin can only be granted by a super_admin
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if viewer["role"] == "super_admin":
                layered_roles |= {"site_admin"}
            # Force-grant the primary role; sync the rest from checked.
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            grant_user_role(user_id, target_user["role"], viewer["id"])
            for r in layered_roles:
                if r in checked:
                    grant_user_role(user_id, r, viewer["id"])
                # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                elif r != target_user["role"]:
                    revoke_user_role(user_id, r)
            log_action(
                viewer["id"], "user", user_id, "user_role_set_updated",
                {"roles": sorted(checked), "email": target_user["email"]},
            )
            flash("Additional roles updated.", "success")
            return redirect(url_for("user_profile", user_id=user_id))
        if action == "update_user_instruments":
            # Bulk per-instrument assignment across the three lanes
            # (admin / operator / faculty). Accepts three POST lists —
            # admin_ids, operator_ids, faculty_ids — where each entry is
            # an instrument id the user should be a member of. Any prior
            # membership not present in the new list is removed. Non-
            # super-admins are limited to instruments they can already
            # manage, so a site_admin can't over-grant.
            if not can_manage_members(viewer):
                abort(403)
            if is_owner(target_user) and not is_owner(viewer):
                abort(403)
            def _ids(field: str) -> set[int]:
                return {int(v) for v in request.form.getlist(field) if str(v).strip().isdigit()}
            desired = {
                "instrument_admins": _ids("admin_ids"),
                "instrument_operators": _ids("operator_ids"),
                "instrument_faculty_admins": _ids("faculty_ids"),
                "instrument_requesters": _ids("requester_ids"),
            }
            manageable_rows = query_all(
                "SELECT id FROM instruments WHERE status = 'active'"
            )
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if viewer["role"] != "super_admin":
                manageable_rows = [
                    row for row in manageable_rows
                    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
                    if can_manage_instrument(viewer["id"], row["id"], viewer["role"])
                ]
            manageable_ids = {row["id"] for row in manageable_rows}
            for table, wanted in desired.items():
                wanted &= manageable_ids
                current_ids = {
                    row["instrument_id"]
                    for row in query_all(
                        f"SELECT instrument_id FROM {table} WHERE user_id = ?",
                        (user_id,),
                    )
                    if row["instrument_id"] in manageable_ids
                }
                to_add = wanted - current_ids
                to_remove = current_ids - wanted
                for instrument_id in to_add:
                    execute(
                        f"INSERT OR IGNORE INTO {table} (user_id, instrument_id) VALUES (?, ?)",
                        (user_id, instrument_id),
                    )
                for instrument_id in to_remove:
                    execute(
                        f"DELETE FROM {table} WHERE user_id = ? AND instrument_id = ?",
                        (user_id, instrument_id),
                    )
            log_action(
                viewer["id"], "user", user_id, "user_instrument_assignments_updated",
                {
                    "email": target_user["email"],
                    "admin_count": len(desired["instrument_admins"]),
                    "operator_count": len(desired["instrument_operators"]),
                    "faculty_count": len(desired["instrument_faculty_admins"]),
                },
            )
            flash("Instrument assignments saved.", "success")
            return redirect(url_for("user_profile", user_id=user_id))

    scope_clauses, scope_params = request_scope_sql(viewer, "sr")
    rows = query_all(
        f"""
        SELECT sr.*, i.name AS instrument_name, i.code AS instrument_code,
               op.name AS operator_name,
               COALESCE(COUNT(ra.id), 0) AS attachment_count
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        LEFT JOIN users op ON op.id = sr.assigned_operator_id
        LEFT JOIN request_attachments ra ON ra.request_id = sr.id AND ra.is_active = 1
        WHERE {' AND '.join(scope_clauses + ['sr.requester_id = ?'])}
        GROUP BY sr.id
        ORDER BY sr.created_at DESC
        LIMIT 25
        """,
        (*scope_params, user_id),
    )
    handled_rows = query_all(
        f"""
        SELECT sr.*, i.name AS instrument_name, i.code AS instrument_code,
               op.name AS operator_name,
               COALESCE(COUNT(ra.id), 0) AS attachment_count
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        LEFT JOIN users op ON op.id = sr.assigned_operator_id
        LEFT JOIN request_attachments ra ON ra.request_id = sr.id AND ra.is_active = 1
        WHERE {' AND '.join(scope_clauses + ['(sr.assigned_operator_id = ? OR sr.received_by_operator_id = ?)'])}
        GROUP BY sr.id
        ORDER BY COALESCE(sr.completed_at, sr.updated_at, sr.created_at) DESC
        LIMIT 25
        """,
        (*scope_params, user_id, user_id),
    )
    originated_count = query_one(
        f"SELECT COUNT(*) AS c FROM sample_requests sr WHERE {' AND '.join(scope_clauses + ['sr.created_by_user_id = ?'])}",
        (*scope_params, user_id),
    )["c"]
    submitted_summary = query_one(
        f"""
        SELECT COUNT(*) AS total_jobs,
               COALESCE(SUM(sr.sample_count), 0) AS total_samples,
               SUM(CASE WHEN sr.status = 'completed' THEN 1 ELSE 0 END) AS completed_jobs,
               SUM(CASE WHEN sr.status NOT IN ('completed', 'rejected') THEN 1 ELSE 0 END) AS open_jobs
        FROM sample_requests sr
        WHERE {' AND '.join(scope_clauses + ['sr.requester_id = ?'])}
        """,
        (*scope_params, user_id),
    )
    handled_summary = query_one(
        f"""
        SELECT COUNT(*) AS total_jobs,
               COALESCE(SUM(sr.sample_count), 0) AS total_samples,
               SUM(CASE WHEN sr.status = 'completed' THEN 1 ELSE 0 END) AS completed_jobs,
               SUM(CASE WHEN sr.status NOT IN ('completed', 'rejected') THEN 1 ELSE 0 END) AS open_jobs
        FROM sample_requests sr
        WHERE {' AND '.join(scope_clauses + ['(sr.assigned_operator_id = ? OR sr.received_by_operator_id = ?)'])}
        """,
        (*scope_params, user_id, user_id),
    )
    can_edit_user_value = (
        can_manage_members(viewer)
        and (not is_owner(target_user) or is_owner(viewer))
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        and (target_user["role"] != "super_admin" or viewer["role"] == "super_admin")
    )
    instrument_roster: list[sqlite3.Row] = []
    instrument_categories: list[str] = []
    assigned_admin_ids: set[int] = set()
    assigned_operator_ids: set[int] = set()
    assigned_faculty_ids: set[int] = set()
    assigned_requester_ids: set[int] = set()
    if can_edit_user_value:
        instrument_roster = query_all(
            "SELECT id, name, code, category, location FROM instruments WHERE status = 'active' ORDER BY category, name"
        )
        assigned_admin_ids = {
            row["instrument_id"] for row in query_all(
                "SELECT instrument_id FROM instrument_admins WHERE user_id = ?", (user_id,)
            )
        }
        assigned_operator_ids = {
            row["instrument_id"] for row in query_all(
                "SELECT instrument_id FROM instrument_operators WHERE user_id = ?", (user_id,)
            )
        }
        assigned_faculty_ids = {
            row["instrument_id"] for row in query_all(
                "SELECT instrument_id FROM instrument_faculty_admins WHERE user_id = ?", (user_id,)
            )
        }
        assigned_requester_ids = {
            row["instrument_id"] for row in query_all(
                "SELECT instrument_id FROM instrument_requesters WHERE user_id = ?", (user_id,)
            )
        }
        instrument_categories = sorted({(r["category"] or "Uncategorized") for r in instrument_roster})
    # Role choices the viewer is allowed to promote/demote this target to.
    role_choices: list[tuple[str, str]] = []
    if can_edit_user_value and target_user["id"] != viewer["id"] and not is_owner(target_user):
        base_roles = ["requester", "operator", "instrument_admin", "faculty_in_charge",
                      "professor_approver", "finance_admin"]
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        if viewer["role"] == "super_admin":
            base_roles += ["site_admin", "super_admin"]
        role_choices = [(r, role_display_name(r)) for r in base_roles]
    # Pop the one-time reset PIN info from the session so it renders
    # exactly once on this page load.  Keep it in session until GET
    # so the redirect after POST still has it.
    reset_pw_info = session.pop("_reset_pw", None) if request.method == "GET" else None
    # Only show if it matches this user — avoids leaking across tabs.
    if reset_pw_info and reset_pw_info.get("user_id") != user_id:
        reset_pw_info = None
    return render_template(
        "user_detail.html",
        target_user=target_user,
        reset_pw_info=reset_pw_info,
        rows=rows,
        handled_rows=handled_rows,
        submitted_summary=submitted_summary,
        handled_summary=handled_summary,
        originated_count=originated_count,
        is_self=viewer["id"] == target_user["id"],
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        can_remove_access=can_manage_members(viewer) and target_user["role"] == "requester" and target_user["active"] and target_user["id"] != viewer["id"],
        can_reset_password=(
            can_manage_members(viewer)
            and target_user["id"] != viewer["id"]
            and (not is_owner(target_user) or is_owner(viewer))
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            and (target_user["role"] != "super_admin" or viewer["role"] == "super_admin")
        ),
        can_edit_user=can_edit_user_value,
        instrument_roster=instrument_roster,
        instrument_categories=instrument_categories,
        assigned_admin_ids=assigned_admin_ids,
        assigned_operator_ids=assigned_operator_ids,
        assigned_faculty_ids=assigned_faculty_ids,
        assigned_requester_ids=assigned_requester_ids,
        role_choices=role_choices,
        target_role_set=user_role_set(target_user),
        layered_role_choices=[
            ("operator", "Operator"),
            ("instrument_admin", "Instrument Admin"),
            ("faculty_in_charge", "Faculty in Charge"),
            ("professor_approver", "Professor Approver"),
            ("finance_admin", "Finance Admin"),
        ] + (
            [("site_admin", "Site Admin")]
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if viewer["role"] == "super_admin" else []
        ),
        instrument_groups=instrument_groups_all(),
        # Authoritative {group_id: [instrument_ids]} map for the
        # group-quick-grant buttons. Uses the real junction table
        # instead of the category heuristic so groups curated by
        # admins (rather than auto-seeded from category) work too.
        instrument_group_members={
            int(g["id"]): instrument_group_member_ids(int(g["id"]))
            for g in instrument_groups_all()
        },
    )


@app.route("/users/<int:user_id>/reset-password.eml")
@login_required
def download_reset_password_eml(user_id: int):
    """Serve a one-time RFC 2822 .eml file the admin can forward to the
    user whose password was just reset.  The temp password is passed via
    query string (the link is generated server-side and shown only in
    the PIN card on the user_detail page, so it is not bookmarkable in
    practice)."""
    viewer = g.user
    if not can_manage_members(viewer):
        abort(403)
    target_user = query_one("SELECT id, name, email FROM users WHERE id = ?", (user_id,))
    if target_user is None:
        abort(404)
    temp_pw = request.args.get("tp", "")
    if not temp_pw:
        abort(400)
    import io
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["To"] = target_user["email"]
    msg["Subject"] = "Your CATALYST Account — Temporary Password"
    msg["From"] = "CATALYST Admin <noreply@catalyst.local>"
    msg.set_content(
        f"Hello {target_user['name']},\n\n"
        f"Your CATALYST account password has been reset by an administrator.\n\n"
        f"Your temporary password is:  {temp_pw}\n\n"
        f"Please log in and change your password immediately.\n\n"
        f"— CATALYST System"
    )
    buf = io.BytesIO(msg.as_bytes())
    buf.seek(0)
    safe_name = target_user["email"].split("@")[0]
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"password-reset-{safe_name}.eml",
        mimetype="message/rfc822",
    )


@app.route("/users/<int:user_id>/history")
@role_required("super_admin")
def user_history(user_id: int):
    target_user = query_one("SELECT id, name FROM users WHERE id = ?", (user_id,))
    if target_user is None:
        abort(404)
    args = request.args.to_dict()
    args["requester_id"] = str(user_id)
    args["source_label"] = target_user["name"]
    return redirect(url_for("schedule", **args))


@app.route("/instruments/<int:instrument_id>/history")
@login_required
@instrument_access_required("view")
def instrument_history(instrument_id: int, instrument):
    args = request.args.to_dict()
    args["instrument_id"] = str(instrument_id)
    args["source_label"] = instrument["name"]
    return redirect(url_for("schedule", **args))


def calendar_events_payload(user: sqlite3.Row, filters: dict[str, str], range_start: date, range_end: date) -> list[dict]:
    clauses = ["sr.scheduled_for IS NOT NULL", "substr(sr.scheduled_for, 1, 10) >= ?", "substr(sr.scheduled_for, 1, 10) < ?"]
    params: list = [range_start.isoformat(), range_end.isoformat()]
    instrument_ids = assigned_instrument_ids(user)
    if instrument_ids:
        placeholders = ",".join("?" for _ in instrument_ids)
        clauses.append(f"sr.instrument_id IN ({placeholders})")
        params.extend(instrument_ids)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    elif user["role"] == "requester":
        clauses.append("sr.requester_id = ?")
        params.append(user["id"])
    if filters.get("instrument_id"):
        clauses.append("sr.instrument_id = ?")
        params.append(int(filters["instrument_id"]))
    if filters.get("operator_id"):
        clauses.append("sr.assigned_operator_id = ?")
        params.append(int(filters["operator_id"]))
    statuses = []
    if filters.get("show_scheduled", "1") == "1":
        statuses.append("scheduled")
    if filters.get("show_in_progress", "1") == "1":
        statuses.append("in_progress")
    if filters.get("show_completed", "0") == "1":
        statuses.append("completed")
    if statuses:
        clauses.append(f"sr.status IN ({','.join('?' for _ in statuses)})")
        params.extend(statuses)
    else:
        clauses.append("1 = 0")
    sql = f"""
        SELECT sr.id, sr.request_no, sr.sample_name, sr.title, sr.status, sr.scheduled_for,
               i.name AS instrument_name, i.code AS instrument_code, u.name AS operator_name
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        LEFT JOIN users u ON u.id = sr.assigned_operator_id
        WHERE {' AND '.join(clauses)}
        ORDER BY sr.scheduled_for
    """
    request_rows = query_all(sql, tuple(params))
    downtime_clauses = ["idt.is_active = 1", "substr(idt.start_time, 1, 10) < ?", "substr(idt.end_time, 1, 10) >= ?"]
    downtime_params: list = [range_end.isoformat(), range_start.isoformat()]
    if filters.get("instrument_id"):
        downtime_clauses.append("idt.instrument_id = ?")
        downtime_params.append(int(filters["instrument_id"]))
    elif instrument_ids:
        placeholders = ",".join("?" for _ in instrument_ids)
        downtime_clauses.append(f"idt.instrument_id IN ({placeholders})")
        downtime_params.extend(instrument_ids)
    downtime_rows = []
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if filters.get("show_maintenance", "1") == "1" and (user["role"] != "requester" or instrument_ids):
        downtime_rows = query_all(
            f"""
            SELECT idt.*, i.name AS instrument_name, i.code AS instrument_code, u.name AS created_by_name
            FROM instrument_downtime idt
            JOIN instruments i ON i.id = idt.instrument_id
            JOIN users u ON u.id = idt.created_by_user_id
            WHERE {' AND '.join(downtime_clauses)}
            ORDER BY idt.start_time
            """,
            tuple(downtime_params),
        )
    events: list[dict] = []
    for row in request_rows:
        status = row["status"]
        start_value = datetime.fromisoformat(str(row["scheduled_for"]))
        end_value = start_value + timedelta(hours=1)
        if status == "in_progress":
            back = "#dce9ff"
            border = "#3465c5"
        elif status == "completed":
            back = "#e7f6e5"
            border = "#4b8d3b"
        else:
            back = "#eaf4f3"
            border = "#1f6f78"
        events.append(
            {
                "id": f"request-{row['id']}",
                "title": f"{row['request_no']} | {row['sample_name']}",
                "start": start_value.isoformat(),
                "end": end_value.isoformat(),
                "backgroundColor": back,
                "borderColor": border,
                "textColor": "#17313a",
                "url": url_for("request_detail", request_id=row["id"]),
                "extendedProps": {
                    "request_id": row["id"],
                    "status": row["status"],
                    "instrument": row["instrument_name"],
                    "instrument_code": row["instrument_code"],
                    "operator": row["operator_name"] or "-",
                },
            }
        )
    for row in downtime_rows:
        events.append(
            {
                "id": f"downtime-{row['id']}",
                "title": f"Maintenance | {row['instrument_code']} | {row['reason']}",
                "start": row["start_time"],
                "end": row["end_time"],
                "backgroundColor": "#fff0df",
                "borderColor": "#c5741d",
                "textColor": "#6f4312",
                "display": "block",
                "extendedProps": {
                    "instrument": row["instrument_name"],
                    "created_by": row["created_by_name"],
                    "downtime": True,
                },
            }
        )
    return events


def calendar_data(filters: dict[str, str]) -> dict:
    view = filters.get("view", "week")
    anchor_date = parse_date_param(filters.get("date")) or datetime.utcnow().date()
    if view == "day":
        prev_date = anchor_date - timedelta(days=1)
        next_date = anchor_date + timedelta(days=1)
    elif view == "month":
        range_start = anchor_date.replace(day=1)
        if range_start.month == 12:
            range_end = date(range_start.year + 1, 1, 1)
        else:
            range_end = date(range_start.year, range_start.month + 1, 1)
        prev_date = range_start - timedelta(days=1)
        next_date = range_end
    else:
        range_start = anchor_date - timedelta(days=anchor_date.weekday())
        prev_date = range_start - timedelta(days=7)
        next_date = range_start + timedelta(days=7)
    return {
        "view": view,
        "anchor_date": anchor_date,
        "prev_date": prev_date,
        "next_date": next_date,
    }


@app.route("/calendar", methods=["GET", "POST"])
@login_required
def calendar():
    user = current_user()
    if not can_access_calendar(user):
        abort(403)
    filters = calendar_filter_values()
    visible_instruments = visible_instruments_for_user(user)
    if len(visible_instruments) == 1 and not filters.get("instrument_id"):
        filters["instrument_id"] = str(visible_instruments[0]["id"])
    if request.method == "POST":
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        if user["role"] != "super_admin" and not assigned_instrument_ids(user):
            abort(403)
        instrument_id = int(request.form["instrument_id"])
        # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
        if user["role"] != "super_admin" and not can_manage_instrument(user["id"], instrument_id, user["role"]):
            abort(403)
        start_time = request.form["start_time"]
        end_time = request.form["end_time"]
        reason = request.form["reason"].strip()
        if not start_time or not end_time or end_time <= start_time:
            flash("Downtime end must be after start.", "error")
            return redirect(url_for("calendar", **filters))
        execute(
            """
            INSERT INTO instrument_downtime (instrument_id, start_time, end_time, reason, created_by_user_id, created_at, is_active)
            VALUES (?, ?, ?, ?, ?, ?, 1)
            """,
            (instrument_id, start_time, end_time, reason, user["id"], now_iso()),
        )
        log_action(user["id"], "instrument", instrument_id, "downtime_added", {"start_time": start_time, "end_time": end_time, "reason": reason})
        flash("Downtime block added.", "success")
        return redirect(url_for("calendar", instrument_id=instrument_id, date=filters.get("date", ""), view=filters.get("view", "week")))
    context = calendar_data(filters)
    instruments = visible_instruments
    operators = query_all("SELECT id, name FROM users WHERE role IN ('operator', 'instrument_admin', 'super_admin') ORDER BY name")
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_add_downtime = user["role"] == "super_admin" or bool(assigned_instrument_ids(user))
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    managed_instruments = instruments if user["role"] == "super_admin" else [i for i in instruments if can_manage_instrument(user["id"], i["id"], user["role"])]
    return render_template(
        "calendar.html",
        filters=filters,
        instruments=instruments,
        instrument_selector_enabled=len(instruments) > 1,
        operators=operators,
        can_add_downtime=can_add_downtime,
        managed_instruments=managed_instruments,
        **context,
    )


@app.route("/calendar/events")
@login_required
def calendar_events():
    user = current_user()
    if not can_access_calendar(user):
        abort(403)
    filters = calendar_filter_values()
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    range_start = parse_date_param(start[:10]) if start else None
    range_end = parse_date_param(end[:10]) if end else None
    if range_start is None or range_end is None or range_end <= range_start:
        return jsonify([])
    return jsonify(calendar_events_payload(user, filters, range_start, range_end))


@app.route("/instruments/<int:instrument_id>/calendar")
@login_required
@instrument_access_required("view")
def instrument_calendar(instrument_id: int, instrument):
    return redirect(url_for("calendar", instrument_id=instrument_id))


@app.route("/calendar.ics")
@login_required
def calendar_ics():
    """v2.2.1 — iCalendar subscription feed. Users add this URL to
    Google Calendar / Apple Calendar / Outlook and get CATALYST events
    as a live-updating calendar. CATALYST never talks to any API.

    Events included:
      - Scheduled sample requests (scheduled_for date)
      - Instrument downtime windows (start/end)
      - Calibrations due (from instrument_maintenance.next_due_at)

    Scoped to the current user's visible instruments. Owner/admin
    sees everything.
    """
    user = current_user()
    now = datetime.utcnow()
    range_start = now - timedelta(days=30)
    range_end = now + timedelta(days=90)

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//CATALYST Lab Scheduler//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:CATALYST — {user['name']}",
    ]

    def ical_dt(dt_str):
        """Convert ISO datetime string to iCal DTSTART format."""
        if not dt_str:
            return None
        clean = dt_str.replace("-", "").replace(":", "").replace("T", "T")[:15]
        if "T" not in clean:
            clean += "T000000"
        return clean.replace("T", "T")

    # 1. Scheduled requests
    db = get_db()
    roles = user_role_set(user)
    if roles & {"super_admin", "site_admin"} or is_owner(user):
        scope_clause = ""
        scope_params: tuple = ()
    else:
        inst_ids = assigned_instrument_ids(user)
        if inst_ids:
            placeholders = ",".join("?" for _ in inst_ids)
            scope_clause = f"AND sr.instrument_id IN ({placeholders})"
            scope_params = tuple(inst_ids)
        else:
            scope_clause = "AND sr.requester_id = ?"
            scope_params = (user["id"],)

    scheduled = db.execute(
        f"""
        SELECT sr.request_no, sr.title, sr.scheduled_for, sr.sample_name,
               i.name AS inst_name, i.code AS inst_code
          FROM sample_requests sr
          JOIN instruments i ON i.id = sr.instrument_id
         WHERE sr.scheduled_for IS NOT NULL
           AND sr.status NOT IN ('completed', 'rejected')
           {scope_clause}
         ORDER BY sr.scheduled_for
         LIMIT 200
        """,
        scope_params,
    ).fetchall()

    for r in scheduled:
        dt = ical_dt(r["scheduled_for"])
        if not dt:
            continue
        lines.extend([
            "BEGIN:VEVENT",
            f"DTSTART:{dt}",
            f"DURATION:PT2H",
            f"SUMMARY:[{r['inst_code']}] {r['title']}",
            f"DESCRIPTION:Sample: {r['sample_name']} | Request: {r['request_no']}",
            f"LOCATION:{r['inst_name']}",
            f"UID:catalyst-req-{r['request_no']}@catalyst.local",
            "END:VEVENT",
        ])

    # 2. Instrument downtime
    downtime = db.execute(
        """
        SELECT d.start_time, d.end_time, d.reason,
               i.name AS inst_name, i.code AS inst_code
          FROM instrument_downtime d
          JOIN instruments i ON i.id = d.instrument_id
         WHERE d.end_time >= ?
         ORDER BY d.start_time
         LIMIT 100
        """,
        (range_start.isoformat(),),
    ).fetchall()

    for d in downtime:
        start = ical_dt(d["start_time"])
        end = ical_dt(d["end_time"])
        if not start:
            continue
        lines.extend([
            "BEGIN:VEVENT",
            f"DTSTART:{start}",
        ])
        if end:
            lines.append(f"DTEND:{end}")
        lines.extend([
            f"SUMMARY:[DOWNTIME] {d['inst_code']} — {d['reason'] or 'Scheduled maintenance'}",
            f"LOCATION:{d['inst_name']}",
            f"UID:catalyst-dt-{d['inst_code']}-{start}@catalyst.local",
            "END:VEVENT",
        ])

    # 3. Calibrations due
    try:
        calibrations = db.execute(
            """
            SELECT m.next_due_at, m.title, m.certificate_number,
                   i.name AS inst_name, i.code AS inst_code
              FROM instrument_maintenance m
              JOIN instruments i ON i.id = m.instrument_id
             WHERE m.event_type = 'calibration'
               AND m.next_due_at IS NOT NULL
               AND m.next_due_at >= ?
             ORDER BY m.next_due_at
             LIMIT 50
            """,
            (range_start.strftime("%Y-%m-%d"),),
        ).fetchall()

        for cal in calibrations:
            dt = cal["next_due_at"]
            if not dt:
                continue
            dt_ical = dt.replace("-", "") + "T090000"
            lines.extend([
                "BEGIN:VEVENT",
                f"DTSTART:{dt_ical}",
                f"DURATION:PT1H",
                f"SUMMARY:[CALIBRATION DUE] {cal['inst_code']} — {cal['title']}",
                f"DESCRIPTION:Certificate: {cal['certificate_number'] or 'pending'}",
                f"LOCATION:{cal['inst_name']}",
                f"UID:catalyst-cal-{cal['inst_code']}-{dt}@catalyst.local",
                "END:VEVENT",
            ])
    except Exception:
        pass  # instrument_maintenance table might not exist on older DBs

    lines.append("END:VCALENDAR")
    ics_content = "\r\n".join(lines)
    return app.response_class(
        ics_content,
        mimetype="text/calendar",
        headers={"Content-Disposition": "inline; filename=catalyst-calendar.ics"},
    )


# ─── v2.3.0 — Attendance & Leave ──────────────────────────────────────

LEAVE_TYPES = ["casual", "sick", "earned", "academic", "sabbatical"]


@app.route("/attendance", methods=["GET"])
@login_required
def attendance_page():
    """Personal attendance view — shows the current user's attendance
    for the current month. Admins see a broader view."""
    user = current_user()
    from datetime import date as _date
    today = _date.today()
    month_start = today.replace(day=1).isoformat()
    month_end = today.isoformat()
    my_attendance = query_all(
        """
        SELECT a.date, a.status, a.check_in, a.check_out, a.notes
          FROM attendance a
         WHERE a.user_id = ? AND a.date >= ? AND a.date <= ?
         ORDER BY a.date DESC
        """,
        (user["id"], month_start, month_end),
    )
    # Summary counts
    present = sum(1 for a in my_attendance if a["status"] == "present")
    absent = sum(1 for a in my_attendance if a["status"] == "absent")
    leave = sum(1 for a in my_attendance if a["status"] == "leave")
    half_day = sum(1 for a in my_attendance if a["status"] == "half")
    # Leave balances
    balances = query_all(
        "SELECT leave_type, balance FROM leave_balances WHERE user_id = ? AND year = ?",
        (user["id"], today.year),
    )
    # Pending leave requests
    pending_leaves = query_all(
        """
        SELECT lr.*, u.name AS approved_by_name
          FROM leave_requests lr
          LEFT JOIN users u ON u.id = lr.approved_by_user_id
         WHERE lr.user_id = ?
         ORDER BY lr.created_at DESC
         LIMIT 20
        """,
        (user["id"],),
    )
    is_admin = bool(user_role_set(user) & {"super_admin", "site_admin"}) or is_owner(user)
    # Team data: people who report to this user
    team_members = query_all(
        """
        SELECT u.id, u.name, u.email,
               a.status AS today_status, a.check_in, a.check_out
          FROM reporting_structure rs
          JOIN users u ON u.id = rs.user_id
          LEFT JOIN attendance a ON a.user_id = u.id AND a.date = ?
         WHERE rs.manager_id = ?
         ORDER BY u.name
        """,
        (today.isoformat(), user["id"]),
    )
    # Pending leave requests from team members (for manager approval)
    team_leave_requests = []
    if team_members:
        team_ids = [m["id"] for m in team_members]
        placeholders = ",".join("?" * len(team_ids))
        team_leave_requests = query_all(
            f"""
            SELECT lr.*, u.name AS requester_name
              FROM leave_requests lr
              JOIN users u ON u.id = lr.user_id
             WHERE lr.user_id IN ({placeholders}) AND lr.status = 'pending'
             ORDER BY lr.start_date ASC
            """,
            team_ids,
        )
    is_manager = len(team_members) > 0
    # Manager info for current user
    my_manager = query_one(
        """
        SELECT u.name AS manager_name
          FROM reporting_structure rs
          JOIN users u ON u.id = rs.manager_id
         WHERE rs.user_id = ?
        """,
        (user["id"],),
    )
    return render_template(
        "attendance.html",
        attendance=my_attendance,
        present=present, absent=absent, leave_count=leave, half_day=half_day,
        balances=balances,
        pending_leaves=pending_leaves,
        is_admin=is_admin,
        is_manager=is_manager,
        team_members=team_members,
        team_leave_requests=team_leave_requests,
        my_manager=my_manager,
        today=today.isoformat(),
        leave_types=LEAVE_TYPES,
    )


@app.route("/attendance/mark", methods=["POST"])
@login_required
def attendance_mark():
    """Mark attendance for today. Users mark their own; admins can mark for others."""
    user = current_user()
    target_user_id = int(request.form.get("user_id") or user["id"])
    roles = user_role_set(user)
    is_admin = bool(roles & {"super_admin", "site_admin"}) or is_owner(user)
    if target_user_id != user["id"] and not is_admin:
        abort(403)
    from datetime import date as _date
    att_date = request.form.get("date", _date.today().isoformat()).strip()
    status = request.form.get("status", "present").strip()
    if status not in {"present", "absent", "leave", "half"}:
        status = "present"
    check_in = request.form.get("check_in", "").strip() or None
    check_out = request.form.get("check_out", "").strip() or None
    notes = request.form.get("notes", "").strip()
    execute(
        """
        INSERT INTO attendance (user_id, date, status, check_in, check_out, notes, marked_by_user_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT (user_id, date) DO UPDATE SET
            status = excluded.status,
            check_in = excluded.check_in,
            check_out = excluded.check_out,
            notes = excluded.notes,
            marked_by_user_id = excluded.marked_by_user_id
        """,
        (target_user_id, att_date, status, check_in, check_out, notes, user["id"], now_iso()),
    )
    log_action(user["id"], "attendance", target_user_id, "attendance_marked",
               {"date": att_date, "status": status})
    flash(f"Attendance marked: {status} on {att_date}", "success")
    return redirect(url_for("attendance_page"))


@app.route("/attendance/quick-present", methods=["POST"])
@login_required
def attendance_quick_present():
    """One-click mark present for today."""
    user = current_user()
    from datetime import date as _date
    today = _date.today().isoformat()
    execute(
        """
        INSERT INTO attendance (user_id, date, status, marked_by_user_id, created_at)
        VALUES (?, ?, 'present', ?, ?)
        ON CONFLICT (user_id, date) DO UPDATE SET
            status = 'present', marked_by_user_id = excluded.marked_by_user_id
        """,
        (user["id"], today, user["id"], now_iso()),
    )
    log_action(user["id"], "attendance", user["id"], "attendance_marked",
               {"date": today, "status": "present"})
    flash("Marked present for today.", "success")
    return redirect(url_for("attendance_page"))


@app.route("/attendance/apply-leave", methods=["POST"])
@login_required
def attendance_apply_leave():
    """Apply for leave directly from the attendance page."""
    user = current_user()
    leave_type = request.form.get("leave_type", "casual").strip()
    if leave_type not in LEAVE_TYPES:
        leave_type = "casual"
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    reason = request.form.get("reason", "").strip()
    if not start_date or not end_date:
        flash("Start and end date are required.", "error")
        return redirect(url_for("attendance_page"))
    if end_date < start_date:
        flash("End date must be after start date.", "error")
        return redirect(url_for("attendance_page"))
    leave_id = execute(
        """
        INSERT INTO leave_requests
            (user_id, leave_type, start_date, end_date, reason, status, created_at)
        VALUES (?, ?, ?, ?, ?, 'pending', ?)
        """,
        (user["id"], leave_type, start_date, end_date, reason, now_iso()),
    )
    log_action(user["id"], "leave", leave_id, "leave_requested",
               {"type": leave_type, "start": start_date, "end": end_date})
    flash(f"Leave request submitted ({leave_type}: {start_date} to {end_date}).", "success")
    return redirect(url_for("attendance_page"))


@app.route("/attendance/team-leave/<int:leave_id>/approve", methods=["POST"])
@login_required
def attendance_team_leave_approve(leave_id: int):
    """Manager approves a leave request from a direct report."""
    user = current_user()
    lr = query_one("SELECT * FROM leave_requests WHERE id = ? AND status = 'pending'", (leave_id,))
    if not lr:
        abort(404)
    # Verify requester reports to current user
    reports = query_one(
        "SELECT 1 FROM reporting_structure WHERE user_id = ? AND manager_id = ?",
        (lr["user_id"], user["id"]),
    )
    is_admin = bool(user_role_set(user) & {"super_admin", "site_admin"}) or is_owner(user)
    if not reports and not is_admin:
        abort(403)
    execute(
        "UPDATE leave_requests SET status = 'approved', approved_by_user_id = ?, approved_at = ? WHERE id = ?",
        (user["id"], now_iso(), leave_id),
    )
    log_action(user["id"], "leave", leave_id, "leave_approved", {})
    flash("Leave approved.", "success")
    return redirect(url_for("attendance_page"))


@app.route("/attendance/team-leave/<int:leave_id>/reject", methods=["POST"])
@login_required
def attendance_team_leave_reject(leave_id: int):
    """Manager rejects a leave request from a direct report."""
    user = current_user()
    lr = query_one("SELECT * FROM leave_requests WHERE id = ? AND status = 'pending'", (leave_id,))
    if not lr:
        abort(404)
    reports = query_one(
        "SELECT 1 FROM reporting_structure WHERE user_id = ? AND manager_id = ?",
        (lr["user_id"], user["id"]),
    )
    is_admin = bool(user_role_set(user) & {"super_admin", "site_admin"}) or is_owner(user)
    if not reports and not is_admin:
        abort(403)
    reason = request.form.get("reason", "").strip()
    execute(
        "UPDATE leave_requests SET status = 'rejected', approved_by_user_id = ?, approved_at = ?, rejection_reason = ? WHERE id = ?",
        (user["id"], now_iso(), reason, leave_id),
    )
    log_action(user["id"], "leave", leave_id, "leave_rejected", {"reason": reason})
    flash("Leave rejected.", "success")
    return redirect(url_for("attendance_page"))


# ── Team Attendance — supervisors mark attendance for workers ────
@app.route("/attendance/team")
@login_required
def attendance_team():
    """Supervisors / admins mark attendance for all workers they manage."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin", "instrument_admin", "finance_admin"} or is_owner(user)):
        abort(403)
    from datetime import date as _date
    today = _date.today().isoformat()
    if is_owner(user) or roles & {"super_admin", "site_admin", "finance_admin"}:
        team = query_all("SELECT id, name, email, role FROM users WHERE active = 1 AND id != ? ORDER BY name", (user["id"],))
    else:
        team = query_all("""
            SELECT DISTINCT u.id, u.name, u.email, u.role
            FROM users u
            JOIN instrument_operators io ON io.user_id = u.id
            JOIN instrument_admins ia ON ia.instrument_id = io.instrument_id
            WHERE ia.user_id = ? AND u.active = 1
            ORDER BY u.name
        """, (user["id"],))
    team_dicts = []
    for member in team:
        m = dict(member)
        att = query_one("SELECT status FROM attendance WHERE user_id = ? AND date = ?", (m["id"], today))
        m["today_status"] = att["status"] if att else None
        team_dicts.append(m)
    return render_template("attendance_team.html", team=team_dicts, today=today)


@app.route("/attendance/team/mark", methods=["POST"])
@login_required
def attendance_team_mark():
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin", "instrument_admin", "finance_admin"} or is_owner(user)):
        abort(403)
    user_id = int(request.form["user_id"])
    status = request.form.get("status", "present")
    if status not in ("present", "absent", "half"):
        status = "present"
    from datetime import date as _date
    today = _date.today().isoformat()
    existing = query_one("SELECT id FROM attendance WHERE user_id = ? AND date = ?", (user_id, today))
    if existing:
        execute("UPDATE attendance SET status = ? WHERE id = ?", (status, existing["id"]))
    else:
        execute("INSERT INTO attendance (user_id, date, status, check_in, created_at) VALUES (?, ?, ?, ?, ?)",
                (user_id, today, status, now_iso(), now_iso()))
    log_action(user["id"], "attendance", user_id, "attendance_marked", {"date": today, "status": status})
    flash("Attendance marked.", "success")
    return redirect(url_for("attendance_team"))


@app.route("/attendance/team/mark-all", methods=["POST"])
@login_required
def attendance_team_mark_all():
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin", "instrument_admin", "finance_admin"} or is_owner(user)):
        abort(403)
    from datetime import date as _date
    today = _date.today().isoformat()
    if is_owner(user) or roles & {"super_admin", "site_admin", "finance_admin"}:
        team = query_all("SELECT id FROM users WHERE active = 1 AND id != ?", (user["id"],))
    else:
        team = query_all("""
            SELECT DISTINCT u.id FROM users u
            JOIN instrument_operators io ON io.user_id = u.id
            JOIN instrument_admins ia ON ia.instrument_id = io.instrument_id
            WHERE ia.user_id = ? AND u.active = 1
        """, (user["id"],))
    count = 0
    for m in team:
        existing = query_one("SELECT id FROM attendance WHERE user_id = ? AND date = ?", (m["id"], today))
        if existing:
            execute("UPDATE attendance SET status = 'present' WHERE id = ?", (existing["id"],))
        else:
            execute("INSERT INTO attendance (user_id, date, status, check_in, created_at) VALUES (?, ?, 'present', ?, ?)",
                    (m["id"], today, now_iso(), now_iso()))
        count += 1
    log_action(user["id"], "attendance", 0, "attendance_mark_all", {"date": today, "count": count})
    flash("%d marked present." % count, "success")
    return redirect(url_for("attendance_team"))


@app.route("/leave/new", methods=["GET", "POST"])
@login_required
def leave_request_new():
    """Submit a leave request."""
    user = current_user()
    if request.method == "POST":
        leave_type = request.form.get("leave_type", "casual").strip()
        if leave_type not in LEAVE_TYPES:
            leave_type = "casual"
        start_date = request.form.get("start_date", "").strip()
        end_date = request.form.get("end_date", "").strip()
        reason = request.form.get("reason", "").strip()
        if not start_date or not end_date:
            flash("Start and end date are required.", "error")
            return redirect(url_for("leave_request_new"))
        if end_date < start_date:
            flash("End date must be after start date.", "error")
            return redirect(url_for("leave_request_new"))
        leave_id = execute(
            """
            INSERT INTO leave_requests
                (user_id, leave_type, start_date, end_date, reason, status, created_at)
            VALUES (?, ?, ?, ?, ?, 'pending', ?)
            """,
            (user["id"], leave_type, start_date, end_date, reason, now_iso()),
        )
        log_action(user["id"], "leave", leave_id, "leave_requested",
                   {"type": leave_type, "start": start_date, "end": end_date})
        flash(f"Leave request submitted ({leave_type}: {start_date} → {end_date}).", "success")
        return redirect(url_for("attendance_page"))
    return render_template("leave_new.html", leave_types=LEAVE_TYPES)


@app.route("/admin/leave", methods=["GET"])
@login_required
def admin_leave_queue():
    """Admin view: all pending leave requests across the facility."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin"} or is_owner(user)):
        abort(403)
    pending = query_all(
        """
        SELECT lr.*, u.name AS requester_name, u.email AS requester_email
          FROM leave_requests lr
          JOIN users u ON u.id = lr.user_id
         WHERE lr.status = 'pending'
         ORDER BY lr.start_date ASC
        """
    )
    recent = query_all(
        """
        SELECT lr.*, u.name AS requester_name, u.email AS requester_email,
               a.name AS approved_by_name
          FROM leave_requests lr
          JOIN users u ON u.id = lr.user_id
          LEFT JOIN users a ON a.id = lr.approved_by_user_id
         WHERE lr.status != 'pending'
         ORDER BY lr.created_at DESC
         LIMIT 30
        """
    )
    approved_count = sum(1 for r in recent if r["status"] == "approved")
    rejected_count = sum(1 for r in recent if r["status"] == "rejected")
    return render_template("admin_leave.html", pending=pending, recent=recent,
                           approved_count=approved_count, rejected_count=rejected_count)


@app.route("/admin/leave/<int:leave_id>/approve", methods=["POST"])
@login_required
def admin_leave_approve(leave_id: int):
    """Approve a leave request."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin"} or is_owner(user)):
        abort(403)
    execute(
        "UPDATE leave_requests SET status = 'approved', approved_by_user_id = ?, approved_at = ? WHERE id = ? AND status = 'pending'",
        (user["id"], now_iso(), leave_id),
    )
    log_action(user["id"], "leave", leave_id, "leave_approved", {})
    flash("Leave approved.", "success")
    return redirect(url_for("admin_leave_queue"))


@app.route("/admin/leave/<int:leave_id>/reject", methods=["POST"])
@login_required
def admin_leave_reject(leave_id: int):
    """Reject a leave request."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin"} or is_owner(user)):
        abort(403)
    reason = request.form.get("reason", "").strip()
    execute(
        "UPDATE leave_requests SET status = 'rejected', approved_by_user_id = ?, approved_at = ?, rejection_reason = ? WHERE id = ? AND status = 'pending'",
        (user["id"], now_iso(), reason, leave_id),
    )
    log_action(user["id"], "leave", leave_id, "leave_rejected", {"reason": reason})
    flash("Leave rejected.", "success")
    return redirect(url_for("admin_leave_queue"))


@app.route("/admin/attendance", methods=["GET"])
@login_required
def admin_attendance_calendar():
    """Admin view: attendance overview for all users on a given date."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin"} or is_owner(user)):
        abort(403)
    from datetime import date as _date
    target_date = request.args.get("date", _date.today().isoformat()).strip()
    records = query_all(
        """
        SELECT a.*, u.name AS user_name, u.email AS user_email, u.role AS user_role
          FROM attendance a
          JOIN users u ON u.id = a.user_id
         WHERE a.date = ?
         ORDER BY u.name
        """,
        (target_date,),
    )
    total_users = query_one("SELECT COUNT(*) AS c FROM users WHERE active = 1")["c"]
    present = sum(1 for r in records if r["status"] == "present")
    absent = sum(1 for r in records if r["status"] == "absent")
    on_leave = sum(1 for r in records if r["status"] == "leave")
    unmarked = total_users - len(records)
    return render_template(
        "admin_attendance.html",
        records=records, target_date=target_date,
        total_users=total_users, present=present, absent=absent,
        on_leave=on_leave, unmarked=unmarked,
    )


# ─── v2.2.0 — Instrument maintenance log + notifications ─────────────


@app.route("/instruments/<int:instrument_id>/maintenance", methods=["GET", "POST"])
@login_required
def instrument_maintenance_log(instrument_id: int):
    """Maintenance + calibration log for an instrument. GET shows the
    event-widget timeline; POST adds a new entry. Gated to operators,
    instrument admins, faculty admins, and super/site admins."""
    user = current_user()
    instrument = query_one("SELECT * FROM instruments WHERE id = ?", (instrument_id,))
    if not instrument:
        abort(404)
    roles = user_role_set(user)
    can_view = bool(roles & {"super_admin", "site_admin"}) or is_owner(user) or \
        can_manage_instrument(user["id"], instrument_id, user["role"]) or \
        can_operate_instrument(user["id"], instrument_id, user["role"])
    if not can_view:
        abort(403)

    if request.method == "POST":
        event_type = request.form.get("event_type", "maintenance").strip()
        if event_type not in {"maintenance", "calibration", "service", "repair", "inspection"}:
            event_type = "maintenance"
        title = request.form.get("title", "").strip()
        if not title:
            flash("Title is required.", "error")
            return redirect(url_for("instrument_maintenance_log", instrument_id=instrument_id))
        description = request.form.get("description", "").strip()
        performed_at = request.form.get("performed_at", now_iso()).strip() or now_iso()
        next_due_at = request.form.get("next_due_at", "").strip() or None
        cost = float(request.form.get("cost") or 0)
        certificate_number = request.form.get("certificate_number", "").strip()
        grant_id = request.form.get("grant_id", "").strip() or None
        if grant_id:
            grant_id = int(grant_id)
        execute(
            """
            INSERT INTO instrument_maintenance
                (instrument_id, event_type, title, description,
                 performed_by_user_id, performed_at, next_due_at,
                 cost, certificate_number, grant_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (instrument_id, event_type, title, description,
             user["id"], performed_at, next_due_at,
             cost, certificate_number, grant_id, now_iso()),
        )
        log_action(user["id"], "instrument", instrument_id, "maintenance_logged",
                   {"event_type": event_type, "title": title})
        flash(f"Logged: {title}", "success")
        return redirect(url_for("instrument_maintenance_log", instrument_id=instrument_id))

    entries = query_all(
        """
        SELECT m.*, u.name AS performed_by_name,
               g.code AS grant_code, g.name AS grant_name
          FROM instrument_maintenance m
          LEFT JOIN users u ON u.id = m.performed_by_user_id
          LEFT JOIN grants g ON g.id = m.grant_id
         WHERE m.instrument_id = ?
         ORDER BY m.performed_at DESC, m.id DESC
         LIMIT 200
        """,
        (instrument_id,),
    )
    upcoming_calibrations = query_all(
        """
        SELECT m.*, u.name AS performed_by_name
          FROM instrument_maintenance m
          LEFT JOIN users u ON u.id = m.performed_by_user_id
         WHERE m.instrument_id = ?
           AND m.event_type = 'calibration'
           AND m.next_due_at IS NOT NULL
           AND m.next_due_at >= date('now', '-7 days')
         ORDER BY m.next_due_at ASC
         LIMIT 10
        """,
        (instrument_id,),
    )
    return render_template(
        "instrument_maintenance.html",
        instrument=instrument,
        entries=entries,
        upcoming_calibrations=upcoming_calibrations,
        can_add=can_view,
        grants=query_all("SELECT id, code, name FROM grants WHERE status = 'active' ORDER BY name"),
    )


@app.route("/instruments/<int:instrument_id>/notify", methods=["POST"])
@login_required
def instrument_notify(instrument_id: int):
    """Post a notice scoped to this instrument. Operators and admins
    can broadcast to everyone assigned to this instrument (operators,
    faculty, requesters). Uses the existing notices table with
    scope='instrument' + scope_target=instrument.code."""
    user = current_user()
    instrument = query_one("SELECT * FROM instruments WHERE id = ?", (instrument_id,))
    if not instrument:
        abort(404)
    roles = user_role_set(user)
    can_post = bool(roles & {"super_admin", "site_admin"}) or is_owner(user) or \
        can_manage_instrument(user["id"], instrument_id, user["role"]) or \
        can_operate_instrument(user["id"], instrument_id, user["role"])
    if not can_post:
        abort(403)
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    severity = request.form.get("severity", "info").strip()
    if severity not in {"info", "warning", "critical"}:
        severity = "info"
    if not subject:
        flash("Subject is required.", "error")
        return redirect(url_for("instrument_detail", instrument_id=instrument_id))
    now = now_iso()
    notice_id = execute(
        """
        INSERT INTO notices
            (scope, scope_target, severity, subject, body,
             author_id, created_at, expires_at)
        VALUES ('instrument', ?, ?, ?, ?, ?, ?, NULL)
        """,
        (instrument["code"], severity, subject, body, user["id"], now),
    )
    log_action(user["id"], "notice", notice_id, "instrument_notice_posted",
               {"instrument_id": instrument_id, "subject": subject[:200]})
    flash(f"Notice posted to {instrument['name']} subscribers.", "success")
    return redirect(url_for("instrument_detail", instrument_id=instrument_id))


@app.route("/admin/maintenance/upcoming")
@login_required
def admin_calibrations_upcoming():
    """NABL-facing dashboard: calibrations due in the next 30 days
    across all instruments. Admin/owner gated."""
    user = current_user()
    roles = user_role_set(user)
    if not (roles & {"super_admin", "site_admin"} or is_owner(user)):
        abort(403)
    upcoming = query_all(
        """
        SELECT m.*, i.name AS instrument_name, i.code AS instrument_code,
               u.name AS performed_by_name
          FROM instrument_maintenance m
          JOIN instruments i ON i.id = m.instrument_id
          LEFT JOIN users u ON u.id = m.performed_by_user_id
         WHERE m.event_type = 'calibration'
           AND m.next_due_at IS NOT NULL
           AND m.next_due_at >= date('now', '-7 days')
           AND m.next_due_at <= date('now', '+30 days')
         ORDER BY m.next_due_at ASC
        """
    )
    return render_template(
        "admin_calibrations_upcoming.html",
        upcoming=upcoming,
    )


@app.route("/stats")
@login_required
def stats():
    user = current_user()
    if not can_access_stats(user):
        abort(403)
    report_filters = report_filter_values()
    visible_instruments = visible_instruments_for_user(user)
    if len(visible_instruments) == 1 and not report_filters.get("instrument_id"):
        report_filters["instrument_id"] = str(visible_instruments[0]["id"])
    own_exports = query_all(
        "SELECT filename FROM generated_exports WHERE created_by_user_id = ? ORDER BY created_at DESC LIMIT 10",
        (user["id"],),
    )
    visible_exports = [EXPORT_DIR / row["filename"] for row in own_exports if (EXPORT_DIR / row["filename"]).exists()]
    stats = stats_payload(user, report_filters)
    admin_graphs = can_access_stats(user)
    stats_visible_links = {i["id"]: can_open_instrument_detail(user, i["id"]) for i in visible_instruments}

    # Trend data (chronological order) — convert Rows to dicts for tojson
    monthly_trend = [dict(r) for r in reversed(stats["monthly"][:12])]
    daily_trend = [dict(r) for r in reversed(stats["daily"][:30])]
    weekly_trend = [dict(r) for r in reversed(stats["weekly"][:12])]

    # ── War-room: live operational data ──
    live_by_instrument = query_all(
        """SELECT i.id, i.name AS instrument_name, i.code,
               SUM(CASE WHEN sr.status IN ('scheduled','in_progress') THEN 1 ELSE 0 END) AS active_jobs,
               SUM(CASE WHEN sr.status IN ('sample_submitted','sample_received') THEN 1 ELSE 0 END) AS queued_jobs,
               SUM(CASE WHEN sr.status IN ('submitted','under_review','awaiting_sample_submission') THEN 1 ELSE 0 END) AS pending_jobs,
               COUNT(sr.id) AS total_open
        FROM instruments i
        LEFT JOIN sample_requests sr ON sr.instrument_id = i.id AND sr.status NOT IN ('completed','rejected')
        WHERE i.status = 'active'
        GROUP BY i.id ORDER BY active_jobs DESC, queued_jobs DESC""",
        (),
    )

    live_totals = {
        "active": sum(r["active_jobs"] for r in live_by_instrument),
        "queued": sum(r["queued_jobs"] for r in live_by_instrument),
        "pending": sum(r["pending_jobs"] for r in live_by_instrument),
        "total_open": sum(r["total_open"] for r in live_by_instrument),
    }

    recent_activity = query_all(
        """SELECT sr.request_no, sr.sample_name, sr.status, sr.updated_at,
               i.name AS instrument_name, u.name AS requester_name
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        JOIN users u ON u.id = sr.requester_id
        WHERE sr.updated_at IS NOT NULL
        ORDER BY sr.updated_at DESC LIMIT 25""",
        (),
    )

    # Bottleneck: instruments with oldest waiting jobs
    bottlenecks = query_all(
        """SELECT i.name AS instrument_name,
               MIN(COALESCE(sr.sample_submitted_at, sr.created_at)) AS oldest_waiting,
               ROUND((julianday('now') - julianday(MIN(COALESCE(sr.sample_submitted_at, sr.created_at)))) * 24.0, 1) AS wait_hours,
               COUNT(*) AS waiting_count
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        WHERE sr.status IN ('sample_submitted','sample_received','scheduled')
        GROUP BY i.id
        HAVING wait_hours > 0
        ORDER BY wait_hours DESC LIMIT 5""",
        (),
    )

    # Status breakdown for doughnut chart
    status_breakdown = query_all(
        """SELECT status, COUNT(*) AS count
        FROM sample_requests
        WHERE status NOT IN ('completed','rejected')
        GROUP BY status ORDER BY count DESC""",
        (),
    )

    # Turnaround time by instrument (hours)
    turnaround_data = query_all(
        """SELECT i.name AS instrument_name,
               ROUND(AVG((julianday(sr.completed_at) - julianday(sr.created_at)) * 24.0), 1) AS avg_hours,
               ROUND(MIN((julianday(sr.completed_at) - julianday(sr.created_at)) * 24.0), 1) AS min_hours,
               ROUND(MAX((julianday(sr.completed_at) - julianday(sr.created_at)) * 24.0), 1) AS max_hours
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        WHERE sr.status = 'completed' AND sr.completed_at IS NOT NULL
        GROUP BY i.id ORDER BY avg_hours DESC LIMIT 10""",
        (),
    )

    # Top requesters
    top_requesters = query_all(
        """SELECT u.name AS requester_name, COUNT(*) AS job_count,
               SUM(sr.sample_count) AS sample_count
        FROM sample_requests sr
        JOIN users u ON u.id = sr.requester_id
        GROUP BY sr.requester_id ORDER BY job_count DESC LIMIT 8""",
        (),
    )

    return render_template(
        "stats.html",
        stats=stats,
        exports=visible_exports[:10],
        admin_graphs=admin_graphs,
        report_filters=report_filters,
        instruments=visible_instruments,
        instrument_selector_enabled=len(visible_instruments) > 1,
        daily_chart=chart_rows(stats["daily"], "bucket", "jobs", 10),
        weekly_chart=chart_rows(stats["weekly"], "bucket", "jobs", 10),
        instrument_chart=chart_rows(stats["by_instrument"], "instrument_name", "completed_samples", 10),
        visible_links=stats_visible_links,
        daily_trend=daily_trend,
        weekly_trend=weekly_trend,
        monthly_trend=monthly_trend,
        by_instrument_json=[dict(r) for r in stats["by_instrument"]],
        weekly_json=[dict(r) for r in stats["weekly"]],
        live_by_instrument=live_by_instrument,
        live_totals=live_totals,
        recent_activity=recent_activity,
        bottlenecks=bottlenecks,
        status_breakdown=[dict(r) for r in status_breakdown],
        turnaround_data=[dict(r) for r in turnaround_data],
        top_requesters=[dict(r) for r in top_requesters],
    )


@app.route("/visualizations")
@login_required
def visualizations():
    user = current_user()
    if not can_access_stats(user):
        abort(403)
    report_filters = report_filter_values()
    visible_instruments = visible_instruments_for_user(user)
    if len(visible_instruments) == 1 and not report_filters.get("instrument_id"):
        report_filters["instrument_id"] = str(visible_instruments[0]["id"])
    stats = stats_payload_for_scope(user, report_filters)
    groups = [row for row in stats["by_group"] if row["group_name"]]
    return render_template(
        "visualization.html",
        page_title="Global Visualization",
        page_hint="Aggregate request and throughput view across all instruments in your scope.",
        scope_kind="global",
        scope_value="All Instruments",
        stats=stats,
        report_filters=report_filters,
        daily_chart=chart_rows(stats["daily"], "bucket", "jobs", 10),
        monthly_chart=chart_rows(stats["monthly"], "bucket", "samples", 10),
        instrument_chart=chart_rows(stats["by_instrument"], "instrument_name", "completed_samples", 10),
        group_chart=chart_rows(stats["by_group"], "group_name", "completed_samples", 10),
        groups=groups,
        current_instrument=None,
        export_endpoint="generate_visualization_export",
        export_scope={},
        instruments=visible_instruments,
        instrument_selector_enabled=len(visible_instruments) > 1,
    )


@app.route("/visualizations/instrument/<int:instrument_id>")
@login_required
@instrument_access_required("view")
def instrument_visualization(instrument_id: int, instrument):
    user = current_user()
    report_filters = report_filter_values()
    stats = stats_payload_for_scope(user, report_filters, instrument_id=instrument_id)
    return render_template(
        "visualization.html",
        page_title=f"{instrument['name']} Visualization",
        page_hint="Instrument-level throughput, request volume, and export view.",
        scope_kind="instrument",
        scope_value=instrument["name"],
        stats=stats,
        report_filters=report_filters,
        daily_chart=chart_rows(stats["daily"], "bucket", "jobs", 10),
        monthly_chart=chart_rows(stats["monthly"], "bucket", "samples", 10),
        instrument_chart=chart_rows(stats["by_instrument"], "instrument_name", "completed_samples", 10),
        group_chart=chart_rows(stats["by_group"], "group_name", "completed_samples", 10),
        groups=[],
        current_instrument=instrument,
        export_endpoint="generate_visualization_export",
        export_scope={"instrument_id": instrument_id},
        instruments=visible_instruments_for_user(user),
        instrument_selector_enabled=False,
    )


@app.route("/visualizations/group/<path:group_name>")
@login_required
def group_visualization(group_name: str):
    user = current_user()
    if not can_view_group_visualization(user, group_name):
        abort(403)
    report_filters = report_filter_values()
    stats = stats_payload_for_scope(user, report_filters, group_name=group_name)
    groups = [row for row in stats["by_group"] if row["group_name"]]
    return render_template(
        "visualization.html",
        page_title=f"{group_name} Visualization",
        page_hint="Group-level aggregate across all instruments in this equipment group.",
        scope_kind="group",
        scope_value=group_name,
        stats=stats,
        report_filters=report_filters,
        daily_chart=chart_rows(stats["daily"], "bucket", "jobs", 10),
        monthly_chart=chart_rows(stats["monthly"], "bucket", "samples", 10),
        instrument_chart=chart_rows(stats["by_instrument"], "instrument_name", "completed_samples", 10),
        group_chart=chart_rows(stats["by_group"], "group_name", "completed_samples", 10),
        groups=groups,
        current_instrument=None,
        export_endpoint="generate_visualization_export",
        export_scope={"group_name": group_name},
        instruments=visible_instruments_for_user(user),
        instrument_selector_enabled=len(visible_instruments_for_user(user)) > 1,
    )


@app.route("/exports/generate", methods=["POST"])
@login_required
def generate_export():
    user = current_user()
    if not can_access_stats(user):
        abort(403)
    report_filters = report_filter_values()
    path = generate_export_workbook(user, report_filters)
    log_action(user["id"], "system_export", 0, "export_generated", {"filename": path.name})
    flash(f"Excel export created: {path.name}", "success")
    return redirect(url_for("stats", **report_filters))


@app.route("/visualizations/export", methods=["POST"])
@login_required
def generate_visualization_export():
    user = current_user()
    if not can_access_stats(user):
        abort(403)
    report_filters = report_filter_values()
    instrument_id_value = request.form.get("instrument_id", "").strip()
    group_name = request.form.get("group_name", "").strip()
    instrument_id = int(instrument_id_value) if instrument_id_value else None
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if instrument_id is not None and not (can_view_instrument_history(user, instrument_id) or user["role"] == "super_admin"):
        abort(403)
    if group_name and not can_view_group_visualization(user, group_name):
        abort(403)
    prefix = "lab_visualization_export"
    if instrument_id is not None:
        prefix = f"instrument_{instrument_id}_export"
    elif group_name:
        prefix = f"group_{safe_token(group_name)}_export"
    path = generate_export_workbook(
        user,
        report_filters,
        instrument_id=instrument_id,
        group_name=group_name or None,
        filename_prefix=prefix,
    )
    log_action(user["id"], "system_export", 0, "visualization_export_generated", {"filename": path.name, "instrument_id": instrument_id, "group_name": group_name})
    if instrument_id is not None:
        flash(f"Instrument export created: {path.name}", "success")
        return redirect(url_for("instrument_visualization", instrument_id=instrument_id, **report_filters))
    if group_name:
        flash(f"Group export created: {path.name}", "success")
        return redirect(url_for("group_visualization", group_name=group_name, **report_filters))
    flash(f"Visualization export created: {path.name}", "success")
    return redirect(url_for("visualizations", **report_filters))


@app.route("/exports/<path:filename>")
@login_required
def download_export(filename: str):
    user = current_user()
    if not can_access_stats(user):
        abort(403)
    export_row = query_one("SELECT * FROM generated_exports WHERE filename = ?", (filename,))
    if export_row is None:
        abort(404)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    if user["role"] != "super_admin" and export_row["created_by_user_id"] != user["id"]:
        abort(403)
    return send_from_directory(EXPORT_DIR, filename, as_attachment=True)


@app.route("/requests/<int:request_id>/calendar-card")
@login_required
def request_calendar_card(request_id: int):
    user = current_user()
    sample_request = query_one(
        """
        SELECT sr.*, i.name AS instrument_name, i.accepting_requests, i.soft_accept_enabled,
               r.name AS requester_name, r.email AS requester_email,
               op.name AS operator_name
        FROM sample_requests sr
        JOIN instruments i ON i.id = sr.instrument_id
        JOIN users r ON r.id = sr.requester_id
        LEFT JOIN users op ON op.id = sr.assigned_operator_id
        WHERE sr.id = ?
        """,
        (request_id,),
    )
    if sample_request is None:
        abort(404)
    if not can_view_request(user, sample_request):
        abort(403)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_manage = can_manage_instrument(user["id"], sample_request["instrument_id"], user["role"])
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_operate = can_operate_instrument(user["id"], sample_request["instrument_id"], user["role"])
    operators = []
    if can_manage or can_operate:
        operators = query_all(
            """
            SELECT DISTINCT u.id, u.name
            FROM users u
            LEFT JOIN instrument_operators io ON io.user_id = u.id AND io.instrument_id = ?
            LEFT JOIN instrument_admins ia ON ia.user_id = u.id AND ia.instrument_id = ?
            WHERE u.active = 1
              AND (io.instrument_id IS NOT NULL OR ia.instrument_id IS NOT NULL OR u.role IN ('super_admin', 'site_admin'))
            ORDER BY u.name
            """,
            (sample_request["instrument_id"], sample_request["instrument_id"]),
        )
    next_slot_default = (datetime.utcnow() + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")
    return render_template(
        "_calendar_card.html",
        sample_request=sample_request,
        can_manage=can_manage,
        can_operate=can_operate,
        operators=operators,
        next_slot_default=next_slot_default,
        request_display_status=request_display_status,
        format_dt=format_dt,
    )


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
def admin_users():
    user = current_user()
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_open_user_admin = is_owner(user) or user["role"] in {"super_admin", "site_admin"}
    if not can_open_user_admin:
        abort(403)
    can_create_users = is_owner(user) or user["role"] in {"super_admin", "site_admin"}
    can_delete_members = is_owner(user)
    # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
    can_elevate_members = is_owner(user) or user["role"] == "super_admin"
    if request.method == "POST":
        action = request.form.get("action", "create_user")
        if action == "create_user":
            if not can_create_users:
                abort(403)
            name = request.form["name"].strip()
            email = request.form["email"].strip().lower()
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            role = request.form["role"]
            # Password hygiene (W1.3.8) — admins do NOT type passwords.
            # A random temporary password is generated; the admin sees
            # it once via the flash below and shares it out-of-band;
            # must_change_password=1 forces the user to set their own
            # on first successful login. Any submitted `password` field
            # in the form is deliberately ignored.
            temp_password = generate_temp_password()
            existing_user = query_one("SELECT id FROM users WHERE email = ?", (email,))
            if existing_user is not None:
                flash(f"User {email} already exists.", "error")
            else:
                member_code = generate_member_code(name, role)
                execute(
                    """
                    INSERT INTO users (name, email, password_hash, role, invited_by, invite_status, active, member_code, must_change_password)
                    VALUES (?, ?, ?, ?, ?, 'active', 1, ?, 1)
                    """,
                    (name, email, generate_password_hash(temp_password, method="pbkdf2:sha256"), role, user["id"], member_code),
                )
                log_action(user["id"], "user", 0, "user_created", {"email": email, "role": role, "member_code": member_code})
                flash(
                    f"User {email} created. Temporary password: {temp_password} — "
                    f"share securely; {name} will be required to change it on first login. "
                    f"Member code: {member_code}.",
                    "success",
                )
        elif action == "delete_member":
            member_id = int(request.form["user_id"])
            member = query_one("SELECT * FROM users WHERE id = ?", (member_id,))
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if member is None or member["role"] != "requester" or is_owner(member) or member["id"] == user["id"]:
                abort(403)
            execute("UPDATE users SET active = 0 WHERE id = ?", (member_id,))
            log_action(user["id"], "user", member_id, "member_deactivated", {"email": member["email"]})
            flash(f"Member {member['email']} deactivated.", "success")
        elif action == "elevate_member":
            if not can_elevate_members:
                abort(403)
            member_id = int(request.form["user_id"])
            new_role = request.form["new_role"].strip()
            allowed_roles = {"operator", "instrument_admin", "site_admin", "finance_admin", "professor_approver"}
            if new_role not in allowed_roles:
                abort(403)
            member = query_one("SELECT * FROM users WHERE id = ?", (member_id,))
            # TODO [v1.5.0 multi-role]: replace <var>["role"] == X / in {...} with has_role(<var>, X) once user_roles junction lands (v1.5.0).
            if member is None or member["role"] != "requester" or is_owner(member) or member["id"] == user["id"]:
                abort(403)
            execute(
                "UPDATE users SET role = ?, invite_status = 'active', active = 1 WHERE id = ?",
                (new_role, member_id),
            )
            log_action(user["id"], "user", member_id, "member_elevated", {"email": member["email"], "new_role": new_role})
            flash(f"{member['email']} elevated to {new_role.replace('_', ' ')}.", "success")
        return redirect(url_for("admin_users"))
    # 5000-user scaling fix: admins + owners are always small sets
    # (< ~100 combined) so an unbounded query is fine. Members
    # (canonical role = requester) can grow to thousands — cap at
    # 200 + expose a ?q=<search> param for finding anyone past the
    # first page. Total count is surfaced in the template so the
    # admin knows how many aren't being shown.
    MEMBERS_CAP = 200
    member_q = (request.args.get("q") or "").strip()
    admin_rows = query_all(
        "SELECT id, name, email, role, invite_status, active, member_code "
        "FROM users "
        "WHERE role != 'requester' "
        "ORDER BY role, name"
    )
    owners = [row for row in admin_rows if row["email"].strip().lower() in OWNER_EMAILS]
    admins = [row for row in admin_rows if row["email"].strip().lower() not in OWNER_EMAILS]
    member_where = "WHERE role = 'requester'"
    member_params: list = []
    if member_q:
        member_where += " AND (LOWER(name) LIKE ? OR LOWER(email) LIKE ? OR LOWER(COALESCE(member_code, '')) LIKE ?)"
        like = f"%{member_q.lower()}%"
        member_params.extend([like, like, like])
    members_total = query_one(
        f"SELECT COUNT(*) AS c FROM users {member_where}",
        tuple(member_params),
    )["c"]
    members = query_all(
        f"SELECT id, name, email, role, invite_status, active, member_code "
        f"FROM users {member_where} ORDER BY name LIMIT ?",
        tuple(member_params) + (MEMBERS_CAP,),
    )
    return render_template(
        "users.html",
        members=members,
        members_total=members_total,
        members_cap=MEMBERS_CAP,
        members_query=member_q,
        admins=admins,
        owners=owners,
        can_create_users=can_create_users,
        can_delete_members=can_delete_members,
        can_elevate_members=can_elevate_members,
    )


@app.route("/activate", methods=["GET", "POST"])
def activate():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"].strip()
        name = request.form.get("name", "").strip()
        user = query_one("SELECT * FROM users WHERE email = ? AND active = 1", (email,))
        if user is None:
            flash("No invited user found with that email.", "error")
            return redirect(url_for("activate"))
        if user["invite_status"] != "invited":
            flash("This account is not waiting for activation. New users must be added by an admin first.", "error")
            return redirect(url_for("activate"))
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "error")
            return redirect(url_for("activate"))
        execute(
            "UPDATE users SET name = ?, password_hash = ?, invite_status = 'active' WHERE id = ?",
            (name or user["name"], generate_password_hash(password, method='pbkdf2:sha256'), user["id"]),
        )
        log_action(user["id"], "user", user["id"], "account_activated", {"email": user["email"]})
        flash("Account activated. You can log in now.", "success")
        return redirect(url_for("login"))
    return render_template("activate.html")


# ── CATALYST feedback log persistence ──────────────────────────
CATALYST_LOG = Path(__file__).resolve().parent / "catalyst_log.json"

@app.route("/catalyst/save", methods=["POST"])
@login_required
def catalyst_save():
    """Persist the full catalyst dump to disk. Called by overlay JS on every entry.

    v2.0.1 — Gated to authenticated users. Previously unauthenticated,
    which let any bot overwrite catalyst_log.json. File is still writable
    by every logged-in user since the overlay ships with the app.
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify(ok=False, error="no data"), 400
    CATALYST_LOG.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    return jsonify(ok=True)

@app.route("/catalyst/log", methods=["GET"])
@login_required
def catalyst_log():
    """Return the current persisted catalyst log. v2.0.1 — auth-gated."""
    if CATALYST_LOG.exists():
        return app.response_class(CATALYST_LOG.read_text(), mimetype="application/json")
    return jsonify(feedLog=[], errorLog=[], paths=[])

@app.route("/catalyst/clear", methods=["POST"])
@login_required
def catalyst_clear():
    """Clear the persisted catalyst log. v2.0.1 — auth-gated."""
    if CATALYST_LOG.exists():
        CATALYST_LOG.write_text("{}")
    return jsonify(ok=True)


# ── Vehicle / Fleet ERP module ──────────────────────────────────


@app.route("/vehicles")
@login_required
def vehicles_list():
    """List all vehicles with status, assigned driver, last fuel date."""
    user = current_user()
    if not module_enabled("vehicles"):
        abort(404)
    vehicles = query_all("""
        SELECT v.*,
               u.name AS driver_name,
               (SELECT MAX(vl.log_date) FROM vehicle_logs vl
                 WHERE vl.vehicle_id = v.id AND vl.log_type = 'fuel') AS last_fuel_date,
               (SELECT COALESCE(SUM(vl.amount), 0) FROM vehicle_logs vl
                 WHERE vl.vehicle_id = v.id AND vl.log_type = 'fuel') AS total_fuel,
               (SELECT COALESCE(SUM(vl.amount), 0) FROM vehicle_logs vl
                 WHERE vl.vehicle_id = v.id AND vl.log_type = 'maintenance') AS total_maintenance,
               (SELECT COALESCE(SUM(vl.amount), 0) FROM vehicle_logs vl
                 WHERE vl.vehicle_id = v.id) AS total_spend
          FROM vehicles v
          LEFT JOIN users u ON u.id = v.assigned_driver_user_id
         ORDER BY v.status ASC, v.name ASC
    """)
    totals = {
        "count": len(vehicles),
        "active": sum(1 for v in vehicles if v["status"] == "active"),
        "total_spend": sum(v["total_spend"] or 0 for v in vehicles),
    }
    can_manage = is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"})
    all_users = query_all("SELECT id, name FROM users ORDER BY name") if can_manage else []
    return render_template("vehicles.html", vehicles=vehicles, totals=totals,
                           can_manage=can_manage, all_users=all_users)


@app.route("/vehicles/new", methods=["POST"])
@login_required
def vehicle_create():
    """Add a new vehicle (admin only)."""
    user = current_user()
    if not module_enabled("vehicles"):
        abort(404)
    if not (is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"})):
        abort(403)
    name = request.form.get("name", "").strip()
    registration_no = request.form.get("registration_no", "").strip()
    vehicle_type = request.form.get("vehicle_type", "car").strip()
    assigned_driver = request.form.get("assigned_driver_user_id", "").strip()
    purchase_date = request.form.get("purchase_date", "").strip() or None
    purchase_cost = float(request.form.get("purchase_cost", "0") or "0")
    insurance_expiry = request.form.get("insurance_expiry", "").strip() or None
    notes = request.form.get("notes", "").strip()
    if not name or not registration_no:
        flash("Name and registration number are required.", "error")
        return redirect(url_for("vehicles_list"))
    existing = query_one("SELECT id FROM vehicles WHERE registration_no = ?", (registration_no,))
    if existing:
        flash("A vehicle with that registration already exists.", "error")
        return redirect(url_for("vehicles_list"))
    new_id = execute(
        """INSERT INTO vehicles (name, registration_no, vehicle_type, assigned_driver_user_id,
           status, purchase_date, purchase_cost, insurance_expiry, notes, created_at)
           VALUES (?, ?, ?, ?, 'active', ?, ?, ?, ?, ?)""",
        (name, registration_no, vehicle_type,
         int(assigned_driver) if assigned_driver else None,
         purchase_date, purchase_cost, insurance_expiry, notes, now_iso()),
    )
    log_action(user["id"], "vehicle", new_id, "vehicle_created",
               {"name": name, "registration_no": registration_no})
    flash(f"{name} added to fleet.", "success")
    return redirect(url_for("vehicles_list"))


@app.route("/vehicles/<int:vehicle_id>")
@login_required
def vehicle_detail(vehicle_id: int):
    """Vehicle detail: info + log history + fuel/maintenance totals."""
    user = current_user()
    if not module_enabled("vehicles"):
        abort(404)
    vehicle = query_one("SELECT * FROM vehicles WHERE id = ?", (vehicle_id,))
    if not vehicle:
        abort(404)
    driver = None
    if vehicle["assigned_driver_user_id"]:
        driver = query_one("SELECT id, name, email FROM users WHERE id = ?",
                           (vehicle["assigned_driver_user_id"],))
    logs = query_all("""
        SELECT vl.*, u.name AS logged_by_name
          FROM vehicle_logs vl
          LEFT JOIN users u ON u.id = vl.logged_by_user_id
         WHERE vl.vehicle_id = ?
         ORDER BY vl.log_date DESC, vl.id DESC
    """, (vehicle_id,))
    fuel_total = sum(l["amount"] for l in logs if l["log_type"] == "fuel")
    maint_total = sum(l["amount"] for l in logs if l["log_type"] == "maintenance")
    other_total = sum(l["amount"] for l in logs if l["log_type"] not in ("fuel", "maintenance"))
    can_manage = is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"})
    # Assigned vehicles for cross-link from personnel
    assigned_vehicles = []
    if vehicle["assigned_driver_user_id"]:
        assigned_vehicles = query_all(
            "SELECT id, name, registration_no FROM vehicles WHERE assigned_driver_user_id = ? AND id != ?",
            (vehicle["assigned_driver_user_id"], vehicle_id))
    all_users = query_all("SELECT id, name FROM users ORDER BY name") if can_manage else []
    return render_template("vehicle_detail.html", vehicle=vehicle, driver=driver,
                           logs=logs, fuel_total=fuel_total, maint_total=maint_total,
                           other_total=other_total, can_manage=can_manage,
                           assigned_vehicles=assigned_vehicles,
                           all_users=all_users,
                           today=date.today().isoformat())


@app.route("/vehicles/<int:vehicle_id>/log", methods=["POST"])
@login_required
def vehicle_add_log(vehicle_id: int):
    """Add a fuel/maintenance/expense log entry."""
    user = current_user()
    if not module_enabled("vehicles"):
        abort(404)
    vehicle = query_one("SELECT id FROM vehicles WHERE id = ?", (vehicle_id,))
    if not vehicle:
        abort(404)
    log_type = request.form.get("log_type", "fuel").strip()
    if log_type not in ("fuel", "maintenance", "insurance", "toll", "fine", "other"):
        log_type = "other"
    amount = float(request.form.get("amount", "0") or "0")
    description = request.form.get("description", "").strip()
    odometer = request.form.get("odometer_km", "").strip()
    log_date = request.form.get("log_date", "").strip() or date.today().isoformat()
    if amount <= 0:
        flash("Amount must be positive.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))
    execute(
        """INSERT INTO vehicle_logs (vehicle_id, log_type, amount, description,
           odometer_km, logged_by_user_id, log_date, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (vehicle_id, log_type, amount, description,
         float(odometer) if odometer else None, user["id"], log_date, now_iso()),
    )
    log_action(user["id"], "vehicle", vehicle_id, "vehicle_log_added",
               {"log_type": log_type, "amount": amount})
    # Notify owners on maintenance logs
    if log_type == "maintenance":
        v_info = query_one("SELECT name FROM vehicles WHERE id = ?", (vehicle_id,))
        v_name = v_info["name"] if v_info else f"Vehicle #{vehicle_id}"
        for owner_row in query_all("SELECT id FROM users WHERE email IN ({})".format(",".join("?" for _ in OWNER_EMAILS)), tuple(OWNER_EMAILS)):
            notify(owner_row["id"], "vehicle", f"Maintenance logged: {v_name}",
                   f"₹{amount:,.0f} — {description[:100]}" if description else f"₹{amount:,.0f}",
                   href=url_for("vehicle_detail", vehicle_id=vehicle_id),
                   source_type="vehicle", source_id=vehicle_id)
    flash(f"{log_type.title()} log ₹{amount:,.0f} recorded.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/edit", methods=["POST"])
@login_required
def vehicle_edit(vehicle_id: int):
    """Update vehicle details (admin only)."""
    user = current_user()
    if not module_enabled("vehicles"):
        abort(404)
    if not (is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"})):
        abort(403)
    vehicle = query_one("SELECT * FROM vehicles WHERE id = ?", (vehicle_id,))
    if not vehicle:
        abort(404)
    name = request.form.get("name", "").strip() or vehicle["name"]
    registration_no = request.form.get("registration_no", "").strip() or vehicle["registration_no"]
    vehicle_type = request.form.get("vehicle_type", "").strip() or vehicle["vehicle_type"]
    assigned_driver = request.form.get("assigned_driver_user_id", "").strip()
    purchase_date = request.form.get("purchase_date", "").strip() or vehicle["purchase_date"]
    purchase_cost = float(request.form.get("purchase_cost", "") or vehicle["purchase_cost"])
    insurance_expiry = request.form.get("insurance_expiry", "").strip() or vehicle["insurance_expiry"]
    notes = request.form.get("notes", "").strip()
    execute(
        """UPDATE vehicles SET name=?, registration_no=?, vehicle_type=?,
           assigned_driver_user_id=?, purchase_date=?, purchase_cost=?,
           insurance_expiry=?, notes=? WHERE id=?""",
        (name, registration_no, vehicle_type,
         int(assigned_driver) if assigned_driver else None,
         purchase_date, purchase_cost, insurance_expiry, notes, vehicle_id),
    )
    log_action(user["id"], "vehicle", vehicle_id, "vehicle_updated", {"name": name})
    flash(f"{name} updated.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/archive", methods=["POST"])
@login_required
def vehicle_archive(vehicle_id: int):
    """Toggle vehicle status between active and archived."""
    user = current_user()
    if not module_enabled("vehicles"):
        abort(404)
    if not (is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"})):
        abort(403)
    vehicle = query_one("SELECT * FROM vehicles WHERE id = ?", (vehicle_id,))
    if not vehicle:
        abort(404)
    new_status = "archived" if vehicle["status"] == "active" else "active"
    execute("UPDATE vehicles SET status = ? WHERE id = ?", (new_status, vehicle_id))
    log_action(user["id"], "vehicle", vehicle_id, "vehicle_status_changed",
               {"old": vehicle["status"], "new": new_status})
    flash(f"{vehicle['name']} is now {new_status}.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


# ── Personnel / Salary ERP module ──────────────────────────────


def _personnel_access(user):
    """Check if user can access personnel module."""
    return is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin", "finance_admin"})


def _personnel_can_edit(user):
    """Check if user can edit salary configs and process payroll."""
    return is_owner(user) or bool(user_role_set(user) & {"super_admin", "site_admin"})


def _days_in_month(year: int, month: int) -> int:
    """Return number of days in given month."""
    import calendar
    return calendar.monthrange(year, month)[1]


def _attendance_days_worked(user_id: int, year: int, month: int) -> float:
    """Count days worked from attendance table for a user in a given month.
    'present' = 1 day, 'half' = 0.5 day."""
    month_str = f"{year}-{month:02d}"
    rows = query_all(
        "SELECT status FROM attendance WHERE user_id = ? AND date LIKE ?",
        (user_id, f"{month_str}-%"))
    total = 0.0
    for r in rows:
        if r["status"] == "present":
            total += 1
        elif r["status"] == "half":
            total += 0.5
    return total


@app.route("/personnel")
@login_required
def personnel_list():
    """Staff directory with salary info (finance_admin/owner only see salary)."""
    user = current_user()
    if not module_enabled("personnel"):
        abort(404)
    if not _personnel_access(user):
        abort(403)
    can_edit = _personnel_can_edit(user)
    staff = query_all("""
        SELECT u.id, u.name, u.email, u.role,
               sc.monthly_salary, sc.designation, sc.department, sc.join_date
          FROM users u
          LEFT JOIN salary_config sc ON sc.user_id = u.id
         ORDER BY u.name
    """)
    return render_template("personnel.html", staff=staff, can_edit=can_edit)


@app.route("/personnel/<int:user_id>")
@login_required
def personnel_detail(user_id: int):
    """Employee detail: salary config, attendance summary, payment history."""
    user = current_user()
    if not module_enabled("personnel"):
        abort(404)
    if not _personnel_access(user):
        abort(403)
    employee = query_one("SELECT * FROM users WHERE id = ?", (user_id,))
    if not employee:
        abort(404)
    salary_cfg = query_one("SELECT * FROM salary_config WHERE user_id = ?", (user_id,))
    payments = query_all(
        """SELECT sp.*, u.name AS paid_by_name
             FROM salary_payments sp
             LEFT JOIN users u ON u.id = sp.paid_by_user_id
            WHERE sp.user_id = ?
            ORDER BY sp.year DESC, sp.month DESC""",
        (user_id,))
    # Current month attendance summary
    today = date.today()
    days_worked = _attendance_days_worked(user_id, today.year, today.month)
    # Assigned vehicles
    assigned_vehicles = query_all(
        "SELECT id, name, registration_no FROM vehicles WHERE assigned_driver_user_id = ?",
        (user_id,))
    can_edit = _personnel_can_edit(user)
    return render_template("personnel_detail.html", employee=employee,
                           salary_cfg=salary_cfg, payments=payments,
                           days_worked=days_worked, assigned_vehicles=assigned_vehicles,
                           can_edit=can_edit, today=today.isoformat())


@app.route("/personnel/<int:user_id>/salary-config", methods=["POST"])
@login_required
def personnel_salary_config(user_id: int):
    """Set or update salary configuration for an employee."""
    user = current_user()
    if not module_enabled("personnel"):
        abort(404)
    if not _personnel_can_edit(user):
        abort(403)
    employee = query_one("SELECT id FROM users WHERE id = ?", (user_id,))
    if not employee:
        abort(404)
    monthly_salary = float(request.form.get("monthly_salary", "0") or "0")
    bank_account = request.form.get("bank_account", "").strip()
    bank_name = request.form.get("bank_name", "").strip()
    ifsc_code = request.form.get("ifsc_code", "").strip()
    pan_number = request.form.get("pan_number", "").strip()
    aadhar_number = request.form.get("aadhar_number", "").strip()
    join_date = request.form.get("join_date", "").strip() or None
    designation = request.form.get("designation", "").strip()
    department = request.form.get("department", "").strip()
    notes = request.form.get("notes", "").strip()
    existing = query_one("SELECT id FROM salary_config WHERE user_id = ?", (user_id,))
    if existing:
        execute(
            """UPDATE salary_config SET monthly_salary=?, bank_account=?, bank_name=?,
               ifsc_code=?, pan_number=?, aadhar_number=?, join_date=?,
               designation=?, department=?, notes=? WHERE user_id=?""",
            (monthly_salary, bank_account, bank_name, ifsc_code, pan_number,
             aadhar_number, join_date, designation, department, notes, user_id))
    else:
        execute(
            """INSERT INTO salary_config (user_id, monthly_salary, bank_account, bank_name,
               ifsc_code, pan_number, aadhar_number, join_date, designation, department, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (user_id, monthly_salary, bank_account, bank_name, ifsc_code, pan_number,
             aadhar_number, join_date, designation, department, notes))
    log_action(user["id"], "personnel", user_id, "salary_config_updated",
               {"monthly_salary": monthly_salary, "designation": designation})
    flash("Salary configuration saved.", "success")
    return redirect(url_for("personnel_detail", user_id=user_id))


@app.route("/personnel/payroll")
@login_required
def payroll_view():
    """Monthly payroll view: all staff, days worked, calculated pay."""
    user = current_user()
    if not module_enabled("personnel"):
        abort(404)
    if not _personnel_access(user):
        abort(403)
    can_edit = _personnel_can_edit(user)
    # Month/year from query params or current month
    today = date.today()
    year = int(request.args.get("year", today.year))
    month_num = int(request.args.get("month", today.month))
    month_name = date(year, month_num, 1).strftime("%B")
    dim = _days_in_month(year, month_num)

    staff_configs = query_all("""
        SELECT sc.*, u.name, u.email
          FROM salary_config sc
          JOIN users u ON u.id = sc.user_id
         ORDER BY u.name
    """)
    payroll_rows = []
    for sc in staff_configs:
        uid = sc["user_id"]
        days_worked = _attendance_days_worked(uid, year, month_num)
        base = sc["monthly_salary"] or 0
        # Check if already paid
        existing_payment = query_one(
            "SELECT * FROM salary_payments WHERE user_id = ? AND year = ? AND month = ?",
            (uid, year, f"{month_num:02d}"))
        calculated_pay = (base / dim) * days_worked if dim > 0 else 0
        payroll_rows.append({
            "user_id": uid,
            "name": sc["name"],
            "email": sc["email"],
            "designation": sc["designation"],
            "department": sc["department"],
            "monthly_salary": base,
            "days_in_month": dim,
            "days_worked": days_worked,
            "calculated_pay": round(calculated_pay, 2),
            "payment": existing_payment,
        })
    total_payable = sum(r["calculated_pay"] for r in payroll_rows if not r["payment"] or r["payment"]["status"] == "pending")
    total_paid = sum(r["payment"]["net_pay"] for r in payroll_rows if r["payment"] and r["payment"]["status"] == "paid")
    return render_template("payroll.html", payroll=payroll_rows, year=year,
                           month_num=month_num, month_name=month_name,
                           days_in_month=dim, can_edit=can_edit,
                           total_payable=total_payable, total_paid=total_paid,
                           today=today.isoformat())


@app.route("/personnel/payroll/pay", methods=["POST"])
@login_required
def payroll_pay():
    """Mark an employee as paid for a month."""
    user = current_user()
    if not module_enabled("personnel"):
        abort(404)
    if not _personnel_can_edit(user):
        abort(403)
    uid = int(request.form.get("user_id", "0"))
    year = int(request.form.get("year", "0"))
    month = request.form.get("month", "").strip()
    deductions = float(request.form.get("deductions", "0") or "0")
    bonus = float(request.form.get("bonus", "0") or "0")
    notes = request.form.get("notes", "").strip()
    if not uid or not year or not month:
        flash("Missing payroll parameters.", "error")
        return redirect(url_for("payroll_view"))
    # Check not already paid
    existing = query_one(
        "SELECT id FROM salary_payments WHERE user_id = ? AND year = ? AND month = ? AND status = 'paid'",
        (uid, year, month))
    if existing:
        flash("Already paid for this period.", "error")
        return redirect(url_for("payroll_view", year=year, month=int(month)))
    sc = query_one("SELECT monthly_salary FROM salary_config WHERE user_id = ?", (uid,))
    base_salary = sc["monthly_salary"] if sc else 0
    month_int = int(month)
    dim = _days_in_month(year, month_int)
    days_worked = _attendance_days_worked(uid, year, month_int)
    net_pay = round((base_salary / dim) * days_worked + bonus - deductions, 2) if dim > 0 else 0
    # Upsert: delete pending, insert paid
    execute("DELETE FROM salary_payments WHERE user_id = ? AND year = ? AND month = ? AND status = 'pending'",
            (uid, year, month))
    execute(
        """INSERT INTO salary_payments (user_id, month, year, base_salary, days_worked,
           days_in_month, deductions, bonus, net_pay, status, paid_at, paid_by_user_id, notes, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'paid', ?, ?, ?, ?)""",
        (uid, month, year, base_salary, int(days_worked), dim, deductions, bonus,
         net_pay, now_iso(), user["id"], notes, now_iso()))
    emp = query_one("SELECT name FROM users WHERE id = ?", (uid,))
    emp_name = emp["name"] if emp else f"User #{uid}"
    log_action(user["id"], "personnel", uid, "salary_paid",
               {"year": year, "month": month, "net_pay": net_pay})
    # Notify the employee about salary payment
    notify(uid, "personnel", f"Salary paid: ₹{net_pay:,.0f}",
           f"Payment for {month}/{year} has been processed.",
           href=url_for("personnel_detail", user_id=uid),
           source_type="salary_payment", source_id=uid)
    flash(f"₹{net_pay:,.0f} paid to {emp_name} for {month}/{year}.", "success")
    return redirect(url_for("payroll_view", year=year, month=int(month)))


# ── Receipts ERP module ─────────────────────────────────────────

def _user_can_review_receipts(user):
    roles = user_role_set(user)
    return bool(roles & {"finance_admin", "super_admin", "site_admin"}) or is_owner(user)


@app.route("/receipts")
@login_required
def receipts_list():
    user = current_user()
    if not module_enabled("receipts"):
        abort(404)
    can_review = _user_can_review_receipts(user)
    status_filter = request.args.get("status", "")
    if can_review:
        rows = query_all("SELECT er.*, u.name AS submitter_name FROM expense_receipts er JOIN users u ON u.id = er.submitted_by_user_id ORDER BY er.created_at DESC")
    else:
        rows = query_all("SELECT er.*, u.name AS submitter_name FROM expense_receipts er JOIN users u ON u.id = er.submitted_by_user_id WHERE er.submitted_by_user_id = ? ORDER BY er.created_at DESC", (user["id"],))
    counts = {"all": len(rows), "pending": sum(1 for r in rows if r["status"]=="pending"), "approved": sum(1 for r in rows if r["status"]=="approved"), "rejected": sum(1 for r in rows if r["status"]=="rejected")}
    totals = {
        "total": sum(r["amount"] for r in rows),
        "pending": sum(r["amount"] for r in rows if r["status"] == "pending"),
        "approved": sum(r["amount"] for r in rows if r["status"] == "approved"),
    }
    if status_filter:
        rows = [r for r in rows if r["status"] == status_filter]
    return render_template("receipts.html", receipts=rows, status_filter=status_filter, can_review=can_review, counts=counts, totals=totals)


@app.route("/receipts/new", methods=["GET", "POST"])
@login_required
def receipt_new():
    user = current_user()
    if not module_enabled("receipts"):
        abort(404)
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        amount = float(request.form.get("amount", "0") or "0")
        category = request.form.get("category", "general").strip()
        receipt_date = request.form.get("receipt_date", "").strip()
        description = request.form.get("description", "").strip()
        if not title:
            flash("Title is required.", "error")
            return redirect(url_for("receipt_new"))
        receipt_image_path = ""
        uploaded = request.files.get("receipt_file")
        if uploaded and uploaded.filename:
            filename = secure_filename(uploaded.filename)
            # Save after getting ID
        cur_id = execute(
            "INSERT INTO expense_receipts (submitted_by_user_id, title, description, amount, category, receipt_date, receipt_image_path, created_at) VALUES (?,?,?,?,?,?,?,?)",
            (user["id"], title, description, amount, category, receipt_date, "", now_iso()),
        )
        if uploaded and uploaded.filename:
            filename = secure_filename(uploaded.filename)
            save_dir = UPLOAD_DIR / "receipts" / str(cur_id)
            save_dir.mkdir(parents=True, exist_ok=True)
            uploaded.save(save_dir / filename)
            receipt_image_path = f"uploads/receipts/{cur_id}/{filename}"
            execute("UPDATE expense_receipts SET receipt_image_path = ? WHERE id = ?", (receipt_image_path, cur_id))
        log_action(user["id"], "receipt", cur_id, "receipt_submitted", {"title": title, "amount": amount, "category": category})
        flash("Receipt submitted.", "success")
        return redirect(url_for("receipt_detail", receipt_id=cur_id))
    return render_template("receipt_form.html")


@app.route("/receipts/<int:receipt_id>")
@login_required
def receipt_detail(receipt_id):
    user = current_user()
    if not module_enabled("receipts"):
        abort(404)
    receipt = query_one("SELECT er.*, u.name AS submitter_name, u.email AS submitter_email, r.name AS reviewer_name FROM expense_receipts er JOIN users u ON u.id = er.submitted_by_user_id LEFT JOIN users r ON r.id = er.reviewed_by_user_id WHERE er.id = ?", (receipt_id,))
    if not receipt:
        abort(404)
    can_review = _user_can_review_receipts(user)
    if not can_review and receipt["submitted_by_user_id"] != user["id"]:
        abort(403)
    return render_template("receipt_detail.html", receipt=receipt, can_review=can_review)


@app.route("/receipts/<int:receipt_id>/review", methods=["POST"])
@login_required
def receipt_review(receipt_id):
    user = current_user()
    if not _user_can_review_receipts(user):
        abort(403)
    receipt = query_one("SELECT * FROM expense_receipts WHERE id = ?", (receipt_id,))
    if not receipt:
        abort(404)
    action = request.form.get("action", "").strip()
    note = request.form.get("reviewer_note", "").strip()
    if action == "approve":
        execute("UPDATE expense_receipts SET status='approved', reviewed_by_user_id=?, reviewer_note=?, reviewed_at=? WHERE id=?", (user["id"], note, now_iso(), receipt_id))
        log_action(user["id"], "receipt", receipt_id, "receipt_approved", {"note": note[:200]})
        notify(receipt["submitted_by_user_id"], "receipt", f"Receipt approved: {receipt['title']}",
               f"Your receipt for ₹{receipt['amount']:,.0f} has been approved.",
               href=url_for("receipt_detail", receipt_id=receipt_id),
               source_type="receipt", source_id=receipt_id)
        flash("Receipt approved.", "success")
    elif action == "reject":
        execute("UPDATE expense_receipts SET status='rejected', reviewed_by_user_id=?, reviewer_note=?, reviewed_at=? WHERE id=?", (user["id"], note, now_iso(), receipt_id))
        log_action(user["id"], "receipt", receipt_id, "receipt_rejected", {"note": note[:200]})
        notify(receipt["submitted_by_user_id"], "receipt", f"Receipt rejected: {receipt['title']}",
               f"Your receipt for ₹{receipt['amount']:,.0f} was rejected." + (f" Note: {note[:100]}" if note else ""),
               href=url_for("receipt_detail", receipt_id=receipt_id),
               source_type="receipt", source_id=receipt_id)
        flash("Receipt rejected.", "error")
    return redirect(url_for("receipt_detail", receipt_id=receipt_id))


# ── Debug feedback endpoint ─────────────────────────────────────
# Voice-to-text debugging workflow: the user looks at the numbered
# grid overlay (?debug=1), speaks into the mic describing issues
# by grid reference, and the browser's Web Speech API transcribes
# the audio. The transcript is POSTed here and appended to
# logs/debug_feedback.md so Claude can read it later.

FEEDBACK_LOG = Path(__file__).resolve().parent / "logs" / "debug_feedback.md"

@app.route("/debug/feedback", methods=["POST"])
@csrf.exempt
@login_required
def debug_feedback():
    """Append voice-transcribed debugging feedback to logs/debug_feedback.md.

    Payload: { text, page, timestamp, grid_visible }
    Enriched server-side with the logged-in user's name, role, and
    the Flask endpoint that serves the page they were looking at.
    """
    data = request.get_json(silent=True)
    if not data or not data.get("text"):
        return jsonify(ok=False, error="no text"), 400

    user = current_user()
    user_label = f"{user['name']} ({user['role']})" if user else "anonymous"
    page = data.get("page", "?")
    ts = data.get("timestamp", datetime.now().isoformat())
    grid = "grid visible" if data.get("grid_visible") else "grid off"
    text = data["text"].strip()

    clicks = data.get("clicks", [])
    click_lines = ""
    if clicks:
        click_lines = "\n**Click markers:**\n"
        for c in clicks:
            click_lines += f"- `{c.get('grid', '?')}` on `{c.get('element', '?')}` ({c.get('page', '?')})\n"

    entry = (
        f"\n## {ts}\n\n"
        f"**User:** {user_label}  \n"
        f"**Page:** `{page}` ({grid})  \n"
        f"{click_lines}\n"
        f"{text}\n"
    )

    FEEDBACK_LOG.parent.mkdir(parents=True, exist_ok=True)
    with open(FEEDBACK_LOG, "a", encoding="utf-8") as f:
        f.write(entry)

    return jsonify(ok=True, saved_to=str(FEEDBACK_LOG))


# ── Feature: Global Search ───────────────────────────────────────────
@app.route("/search")
@login_required
def global_search():
    user = current_user()
    q = request.args.get("q", "").strip()
    shortcuts = [
        {"label": "Dashboard", "hint": "See current workload and action items.", "url": url_for("index")},
        {"label": "Inbox", "hint": "Messages and internal coordination.", "url": url_for("inbox")},
    ]
    if can_access_schedule(user):
        shortcuts.append({"label": "Queue", "hint": "Work the live request stream.", "url": url_for("schedule")})
    if has_instrument_area_access(user):
        shortcuts.append({"label": "Instruments", "hint": "Browse machine status and queues.", "url": url_for("instruments")})
    if module_enabled("finance") and can_access_finance(user):
        shortcuts.append({"label": "Finance", "hint": "Grants, invoices, and spend.", "url": url_for("finance_portal")})
    if len(q) < 2:
        return render_template("search.html", query=q, results=[], sections=[], total_results=0, shortcuts=shortcuts, title="Search")
    results = []
    sections = []
    like = f"%{q}%"
    instruments_found = [
        {"type": "Instrument", "title": r["name"], "code": r["code"], "meta": "Machine", "url": url_for("instrument_detail", instrument_id=r["id"])}
        for r in query_all("SELECT id, name, code FROM instruments WHERE name LIKE ? OR code LIKE ? LIMIT 6", (like, like))
    ]
    requests_found = [
        {"type": "Request", "title": r["title"], "code": r["request_no"], "meta": "Sample request", "url": url_for("request_detail", request_id=r["id"])}
        for r in query_all("SELECT id, request_no, title FROM sample_requests WHERE request_no LIKE ? OR title LIKE ? OR sample_name LIKE ? LIMIT 6", (like, like, like))
    ]
    users_found = [
        {"type": "User", "title": r["name"], "code": r["email"], "meta": "User profile", "url": url_for("user_profile", user_id=r["id"])}
        for r in query_all("SELECT id, name, email FROM users WHERE name LIKE ? OR email LIKE ? LIMIT 6", (like, like))
    ]
    grants_found = []
    if module_enabled("finance") and can_access_finance(user):
        grants_found = [
            {"type": "Grant", "title": r["name"], "code": r["code"], "meta": "Grant", "url": url_for("finance_grant_detail", grant_id=r["id"])}
            for r in query_all("SELECT id, code, name FROM grants WHERE name LIKE ? OR code LIKE ? LIMIT 6", (like, like))
        ]
    grouped = [
        ("Instruments", "Machines, capacity, and intake state.", instruments_found),
        ("Requests", "Live jobs and historical samples.", requests_found),
        ("People", "Operators, requesters, and staff records.", users_found),
        ("Grants", "Budgets and finance records.", grants_found),
    ]
    for title, hint, items in grouped:
        if not items:
            continue
        sections.append({"title": title, "hint": hint, "count": len(items), "items": items})
        results.extend(items)
    return render_template(
        "search.html",
        query=q,
        results=results,
        sections=sections,
        total_results=len(results),
        shortcuts=shortcuts,
        title="Search",
    )


# ── Feature: Audit Log Viewer (admin-only) ──────────────────────────
@app.route("/admin/audit-log")
@login_required
def audit_log_viewer():
    """Central audit log — visible only to owner, super_admin, site_admin."""
    user = current_user()
    roles = user_role_set(user)
    if not (is_owner(user) or roles & {"super_admin", "site_admin"}):
        abort(403)
    page = max(1, int(request.args.get("page", 1)))
    per_page = 100
    entity_filter = request.args.get("entity", "").strip()
    action_filter = request.args.get("action", "").strip()
    where_clauses = []
    params: list = []
    if entity_filter:
        where_clauses.append("al.entity_type = ?")
        params.append(entity_filter)
    if action_filter:
        where_clauses.append("al.action LIKE ?")
        params.append(f"%{action_filter}%")
    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    total = query_one(f"SELECT COUNT(*) AS c FROM audit_logs al{where_sql}", tuple(params))["c"]
    rows = query_all(
        f"SELECT al.*, u.name AS actor_name FROM audit_logs al LEFT JOIN users u ON u.id = al.actor_id{where_sql} ORDER BY al.id DESC LIMIT ? OFFSET ?",
        tuple(params) + (per_page, (page - 1) * per_page),
    )
    entity_types = [r["entity_type"] for r in query_all("SELECT DISTINCT entity_type FROM audit_logs ORDER BY entity_type")]
    total_pages = max(1, (total + per_page - 1) // per_page)
    return render_template(
        "audit_log.html", logs=rows, page=page, total_pages=total_pages, total=total,
        entity_filter=entity_filter, action_filter=action_filter, entity_types=entity_types,
    )


# ── Feature: Audit Log CSV Export ────────────────────────────────────
@app.route("/admin/audit-export")
@owner_required
def audit_export():
    import csv as _csv
    import io as _io
    rows = query_all("SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 10000")
    output = _io.StringIO()
    writer = _csv.writer(output)
    writer.writerow(["id", "entity_type", "entity_id", "action", "actor_id", "payload_json", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["entity_type"], r["entity_id"], r["action"], r["actor_id"], r["payload_json"], r["created_at"]])
    return send_file(
        _io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name="catalyst_audit_log.csv",
    )


if __name__ == "__main__":
    init_db()
    # Always auto-reload templates so changes appear without server restart
    app.config["TEMPLATES_AUTO_RELOAD"] = True
    app.jinja_env.auto_reload = True

    # LAB_SCHEDULER_DEBUG gates the Flask debug toolbar + debugger PIN.
    # Heavyweight, opt-in only, never on in production or under launchd.
    is_debug = os.environ.get("LAB_SCHEDULER_DEBUG", "0") == "1"

    # LAB_SCHEDULER_HOST controls the bind address. Default is loopback
    # (127.0.0.1) so `python app.py` on a laptop does not accidentally
    # expose the dev server to the LAN. On the Mac mini production
    # host, set LAB_SCHEDULER_HOST=0.0.0.0 in .env so Tailscale and LAN
    # devices can reach it at http://<host>:5055/.
    bind_host = os.environ.get("LAB_SCHEDULER_HOST", "127.0.0.1")

    # use_reloader watches .py files and auto-restarts on code changes.
    # Decoupled from LAB_SCHEDULER_DEBUG as of v1.4.5 — a laptop
    # operator editing app.py wants reload even without the debug
    # toolbar. Resolution order:
    #   1. LAB_SCHEDULER_AUTORELOAD=1 → force ON (explicit opt-in)
    #   2. LAB_SCHEDULER_AUTORELOAD=0 → force OFF (explicit opt-out,
    #      e.g. launchd service plist — the reloader's fork-and-exit
    #      pattern makes launchd mark the service EX_CONFIG-crashed
    #      because the parent PID vanishes, so service-mode MUST
    #      set this to 0).
    #   3. Unset → smart default: ON when bound to loopback
    #      (127.0.0.1 / localhost / ::1), OFF when LAN-facing. The
    #      heuristic matches the typical dev vs. prod split: if
    #      nobody outside your machine can reach you, you're in dev
    #      and auto-reload is the expected behaviour.
    _autoreload_env = os.environ.get("LAB_SCHEDULER_AUTORELOAD", "").strip()
    if _autoreload_env == "1":
        auto_reload = True
    elif _autoreload_env == "0":
        auto_reload = False
    else:
        auto_reload = bind_host in ("127.0.0.1", "localhost", "::1")

    app.run(debug=is_debug, use_reloader=auto_reload, host=bind_host, port=5055,
            extra_files=[
                str(Path(__file__).resolve().parent / "static" / "styles.css"),
                str(Path(__file__).resolve().parent / "static" / "grid-overlay.js"),
            ])
