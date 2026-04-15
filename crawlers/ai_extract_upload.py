"""
ai_extract_upload — parse an uploaded spreadsheet from `data/<lane>/ai_uploads/`
and emit a typed proposal JSON suitable for human review.

Read-only. Does NOT write to any database. Does NOT call any app endpoint.
Intended as the EXTRACT stage of the four-stage AI ingestion pipeline:

    INGEST  →  EXTRACT  →  REVIEW  →  APPLY

The output of this module is the input to the (future) REVIEW UI / API,
which in turn feeds the canonical APPLY paths in app.py
(`bulk_create_users`, instrument admin, `instrument_operators`).

ERP-AWARE: takes a mandatory `--erp lab|ravikiran` flag that selects
the right schema interpretation. There is no default — per the ERP
silo policy ("separate gits which serve specializations of the
wrappers, separate data store"), wrong-target writes must be a loud
error, not a silent default.

  --erp lab        instrument-list-with-operator-name spreadsheets
                   (cols: SerialNo | Instrument | Operator(s))
                   matches against `instruments` table on --db

  --erp ravikiran  household-staff roster spreadsheets across
                   KITCHEN / PF / TUCKSHOP / FLEET sheets
                   (cols: SerialNo | Name | Designation | DateOfJoining
                          | Salary | TotalDay | PresentDays)
                   each row → a proposed user, role from designation,
                   department from sheet name

Usage:
    python -m crawlers.ai_extract_upload --erp <lab|ravikiran> <path-to-xlsx> [--db <sqlite>]

Examples:
    python -m crawlers.ai_extract_upload --erp lab \\
        data/demo/ai_uploads/9/2026-04-15_Instrument_list_with_operator_name.xlsx \\
        --db data/demo/lab_scheduler.db

    python -m crawlers.ai_extract_upload --erp ravikiran \\
        ~/ravikiran-services/data/ai_uploads/18/2026-04-15_attendance_sheet.xlsx

The `--db` flag is optional; without it, no match-against-existing is performed
(matched_db_code fields stay empty). Operational DB is intentionally not
defaulted — per WORKFLOW.md §3.5, real lab data is off-limits to dev tooling.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable

from crawlers.common import validate_erp_db_match

try:
    import openpyxl  # type: ignore
except ImportError as exc:  # pragma: no cover
    print(json.dumps({"error": f"openpyxl required: {exc}"}), file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# Data model — what the proposal looks like
# ---------------------------------------------------------------------------

@dataclass
class ProposedUser:
    name: str
    email: str
    role: str
    short_code: str
    source_row: int

@dataclass
class ProposedInstrument:
    sheet_name: str
    proposed_code: str  # blank if matched_db_code is set
    proposed_category: str
    matched_db_code: str = ""
    matched_db_id: int = 0
    match_confidence: float = 0.0

@dataclass
class ProposedLink:
    operator_short_code: str
    instrument_handle: str  # either matched DB code (INST-xxx) or proposed_code
    source_row: int

@dataclass
class Proposal:
    source_file: str
    detected_action_type: str
    sheet_dimensions: dict
    proposed_users: list[ProposedUser] = field(default_factory=list)
    proposed_instruments: list[ProposedInstrument] = field(default_factory=list)
    proposed_links: list[ProposedLink] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_TITLE_PREFIX = re.compile(r"^(mr|mrs|ms|miss|dr|prof|professor|shri|smt)\.?\s+", re.I)


def normalise_name(raw: str) -> str:
    """Strip stray whitespace + collapse 'Mr.Foo' → 'Mr. Foo'."""
    if not raw:
        return ""
    s = unicodedata.normalize("NFKC", raw).strip()
    s = re.sub(r"\s+", " ", s)
    # Insert space after a leading title that's missing one ('Mr.Ranjit' → 'Mr. Ranjit')
    s = re.sub(r"^(mr|mrs|ms|miss|dr|prof)\.(\S)", r"\1. \2", s, flags=re.I)
    return s


def short_code_for(name: str, taken: set[str]) -> str:
    """Generate a 3-letter short_code from a name. Bumps to <FirstLast> on collision."""
    cleaned = _TITLE_PREFIX.sub("", normalise_name(name)).strip()
    parts = re.findall(r"[A-Za-z]+", cleaned)
    if not parts:
        return ""
    if len(parts) == 1:
        base = (parts[0][:3]).upper()
    else:
        base = (parts[0][0] + parts[1][:2]).upper()
    candidate = base
    n = 2
    while candidate in taken:
        candidate = f"{base[:2]}{n}"
        n += 1
        if n > 9:
            candidate = f"{parts[0][:2].upper()}{parts[1][:1].upper()}"
            break
    return candidate


def email_for(name: str, domain: str = "mitwpu.edu.in") -> str:
    cleaned = _TITLE_PREFIX.sub("", normalise_name(name)).strip().lower()
    parts = re.findall(r"[a-z]+", cleaned)
    if not parts:
        return ""
    return f"{'.'.join(parts)}@{domain}"


def split_operators(cell: str) -> list[str]:
    """A cell like 'Mr.Ranjit Kate / Mrs.Aparna Potdar' → two names."""
    if not cell:
        return []
    # Common separators in this dataset: '/', '&', ',', ';'
    chunks = re.split(r"\s*[\/&,;]\s*", cell)
    return [normalise_name(c) for c in chunks if c.strip()]


def fuzzy_match_instrument(needle: str, candidates: Iterable[tuple]) -> tuple[str, int, float]:
    """Best-effort fuzzy match: returns (db_code, db_id, confidence in [0,1]).
    candidates is an iterable of (id, code, name). Empty match returns ('', 0, 0.0)."""
    needle_words = set(_tokenise(needle))
    if not needle_words:
        return "", 0, 0.0
    best = ("", 0, 0.0)
    for inst_id, code, name in candidates:
        cand_words = set(_tokenise(name))
        if not cand_words:
            continue
        overlap = needle_words & cand_words
        union = needle_words | cand_words
        if not union:
            continue
        score = len(overlap) / len(union)  # Jaccard
        # Boost if the needle is a substring of the candidate name
        if needle.strip().lower() in name.lower() or name.lower() in needle.strip().lower():
            score = max(score, 0.85)
        if score > best[2]:
            best = (code, inst_id, score)
    return best


_STOPWORDS = {"with", "and", "the", "for", "of", "system"}

def _tokenise(s: str) -> list[str]:
    if not s:
        return []
    s = s.lower()
    s = s.replace("-", " ").replace("/", " ").replace(",", " ")
    return [w for w in re.findall(r"[a-z0-9]+", s) if w not in _STOPWORDS and len(w) > 1]


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def read_sheet(path: Path) -> tuple[str, list[list[str]], dict]:
    """Open the first non-empty sheet, return (sheet_name, rows, dims)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [
            ["" if v is None else str(v).strip() for v in r]
            for r in ws.iter_rows(values_only=True)
        ]
        # Drop trailing empty rows
        while rows and not any(c.strip() for c in rows[-1]):
            rows.pop()
        if rows:
            return sn, rows, {"rows": ws.max_row, "cols": ws.max_column}
    return "(empty)", [], {"rows": 0, "cols": 0}


def read_all_sheets(path: Path) -> list[tuple[str, list[list[str]], dict]]:
    """Open every non-empty sheet, return [(sheet_name, rows, dims), ...]."""
    out = []
    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [
            ["" if v is None else str(v).strip() for v in r]
            for r in ws.iter_rows(values_only=True)
        ]
        while rows and not any(c.strip() for c in rows[-1]):
            rows.pop()
        if rows:
            out.append((sn, rows, {"rows": ws.max_row, "cols": ws.max_column}))
    return out


def find_header_row(rows: list[list[str]]) -> int:
    """Heuristic: the first row that contains 'name', 'operator', or 'instrument'
    in any cell (case-insensitive)."""
    for idx, row in enumerate(rows):
        joined = " ".join(c.lower() for c in row)
        if any(k in joined for k in ("operator", "instrument", "equipment", "name")):
            if "sr" in joined or "no" in joined or "name" in joined:
                return idx
    return 0


# ---------------------------------------------------------------------------
# Main extractor for the "operator list for instruments" pattern
# ---------------------------------------------------------------------------

def extract_operator_list(path: Path, db_path: Path | None) -> Proposal:
    sheet_name, rows, dims = read_sheet(path)
    proposal = Proposal(
        source_file=str(path),
        detected_action_type="operator_list_for_instruments",
        sheet_dimensions={"sheet_name": sheet_name, **dims},
    )
    if not rows:
        proposal.notes.append("Sheet is empty.")
        return proposal

    header_idx = find_header_row(rows)
    proposal.notes.append(f"Header row detected at index {header_idx}.")

    # Existing instruments for matching (read-only)
    db_instruments: list[tuple] = []
    if db_path and db_path.exists():
        try:
            with sqlite3.connect(f"file:{db_path}?mode=ro", uri=True) as conn:
                db_instruments = list(conn.execute("SELECT id, code, name FROM instruments"))
            proposal.notes.append(f"Loaded {len(db_instruments)} existing instruments from {db_path.name} for matching.")
        except sqlite3.Error as exc:
            proposal.notes.append(f"DB read failed ({exc}); proceeding without matches.")
    else:
        proposal.notes.append("No --db supplied; instrument matches will be empty.")

    seen_codes: set[str] = set()
    seen_users: dict[str, ProposedUser] = {}  # short_code -> user

    for ridx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        # Expecting columns: serial, instrument_name, operator(s)
        # Be lenient about column position.
        cells = [c for c in row if c.strip()]
        if len(cells) < 2:
            continue
        # Serial column may or may not be present
        if cells[0].rstrip(".").isdigit():
            cells = cells[1:]
        if len(cells) < 2:
            continue
        instrument_cell, operator_cell = cells[0], cells[1]
        if not instrument_cell or not operator_cell:
            continue

        # 1) Match instrument
        match_code, match_id, conf = ("", 0, 0.0)
        if db_instruments:
            match_code, match_id, conf = fuzzy_match_instrument(instrument_cell, db_instruments)
        proposed_code = "" if match_code else f"INST-NEW-{ridx:03d}"
        instrument_handle = match_code or proposed_code
        proposal.proposed_instruments.append(ProposedInstrument(
            sheet_name=instrument_cell,
            proposed_code=proposed_code,
            proposed_category="",
            matched_db_code=match_code,
            matched_db_id=match_id,
            match_confidence=round(conf, 3),
        ))

        # 2) Operators (one cell may list 2+)
        for op_name in split_operators(operator_cell):
            sc = short_code_for(op_name, seen_codes)
            if not sc:
                continue
            if sc not in seen_users:
                seen_codes.add(sc)
                seen_users[sc] = ProposedUser(
                    name=op_name,
                    email=email_for(op_name),
                    role="operator",
                    short_code=sc,
                    source_row=ridx,
                )
            proposal.proposed_links.append(ProposedLink(
                operator_short_code=sc,
                instrument_handle=instrument_handle,
                source_row=ridx,
            ))

    proposal.proposed_users = list(seen_users.values())
    return proposal


# ---------------------------------------------------------------------------
# Ravikiran extractor — household staff roster across multiple sheets
# ---------------------------------------------------------------------------

# Map sheet name → role / office_location prefix
RAVIKIRAN_SHEET_DEPT = {
    "KITCHEN":  ("operator", "Kitchen"),
    "PF":       ("operator", "Office (PF)"),
    "TUCKSHOP": ("operator", "Tuck Shop"),
    "TUCK":     ("operator", "Tuck Shop"),
    "FLEET":    ("operator", "Fleet · Driver"),
    "DRIVERS":  ("operator", "Fleet · Driver"),
    "LAUNDRY":  ("operator", "Laundry"),
    "OFFICE":   ("operator", "Office"),
}

# Designation → 3-letter prefix for short_code
_DESIGNATION_PREFIX = {
    "cook":      "CK",
    "helper":    "HL",
    "chapati":   "CH",
    "wash":      "WS",
    "bekar":     "BK",
    "utility":   "UT",
    "supervisor":"SP",
    "account":   "AC",
    "cashier":   "CS",
    "driver":    "DR",
    "waiter":    "WT",
    "cleaner":   "CL",
    "laundry":   "LD",
}


def _ravikiran_short_code(name: str, designation: str, taken: set[str]) -> str:
    cleaned = re.sub(r"[^A-Za-z]+", " ", name).upper().split()
    if not cleaned:
        return ""
    initials = "".join(p[0] for p in cleaned[:3])
    # Prefix from designation if recognised (improves readability across rows)
    desg = (designation or "").lower()
    desg_prefix = ""
    for k, p in _DESIGNATION_PREFIX.items():
        if k in desg:
            desg_prefix = p
            break
    base = (desg_prefix + initials)[:5] if desg_prefix else initials[:5]
    candidate = base
    n = 2
    while candidate in taken:
        candidate = f"{base[:4]}{n}"
        n += 1
        if n > 99:
            break
    return candidate


def _ravikiran_email(name: str, sheet_dept: str) -> str:
    cleaned = re.sub(r"[^A-Za-z]+", " ", name).strip().lower()
    parts = cleaned.split()
    if not parts:
        return ""
    return f"{'.'.join(parts)}@ravikiran.local"


def _ravikiran_find_header(rows: list[list[str]]) -> int:
    """Find the header row containing NAME and Designation (case-insensitive)."""
    for idx, row in enumerate(rows):
        joined = " ".join(c.lower() for c in row)
        if "name" in joined and ("designation" in joined or "desig" in joined):
            return idx
    # Fallback: first row with "name" and a serial column heading
    for idx, row in enumerate(rows):
        joined = " ".join(c.lower() for c in row)
        if "name" in joined and ("sr" in joined or "no" in joined):
            return idx
    return 0


def _column_index(header: list[str], *needles: str) -> int | None:
    """Find the first column whose header contains any of the needles (lowercase substring)."""
    for i, cell in enumerate(header):
        c = cell.lower()
        if any(n in c for n in needles):
            return i
    return None


def extract_ravikiran_roster(path: Path) -> Proposal:
    """Iterate every sheet, treat each row as a household-staff member."""
    proposal = Proposal(
        source_file=str(path),
        detected_action_type="household_staff_roster",
        sheet_dimensions={},
    )
    sheets = read_all_sheets(path)
    if not sheets:
        proposal.notes.append("Workbook has no non-empty sheets.")
        return proposal

    proposal.sheet_dimensions = {sn: dims for sn, _, dims in sheets}
    seen_codes: set[str] = set()
    seen_emails: dict[str, ProposedUser] = {}

    for sheet_name, rows, _ in sheets:
        sheet_key = sheet_name.strip().upper()
        # Pick role / dept defaults from sheet name
        default_role, dept_label = ("operator", sheet_name.strip())
        for k, (r, lbl) in RAVIKIRAN_SHEET_DEPT.items():
            if k in sheet_key:
                default_role, dept_label = r, lbl
                break

        header_idx = _ravikiran_find_header(rows)
        header = rows[header_idx] if header_idx < len(rows) else []
        col_name  = _column_index(header, "name")
        col_desig = _column_index(header, "designation", "desig")
        col_doj   = _column_index(header, "joining", "doj")
        col_sal   = _column_index(header, "salary", "wage")
        if col_name is None:
            proposal.notes.append(f"[{sheet_name}] could not locate NAME column — skipped.")
            continue

        rows_used = 0
        for ridx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            if not any(c.strip() for c in row):
                continue
            name = row[col_name] if col_name < len(row) else ""
            name = re.sub(r"\s+", " ", name).strip()
            if not name or name.lower() in {"name"}:
                continue
            # Filter out designation-only rows (no real personal name)
            if name.upper() == name and len(name.split()) == 1 and len(name) <= 4:
                continue
            designation = (row[col_desig] if (col_desig is not None and col_desig < len(row)) else "").strip()
            doj         = (row[col_doj]   if (col_doj   is not None and col_doj   < len(row)) else "").strip()
            salary      = (row[col_sal]   if (col_sal   is not None and col_sal   < len(row)) else "").strip()

            email = _ravikiran_email(name, dept_label)
            if email in seen_emails:
                continue  # duplicate name across sheets — keep first occurrence
            sc = _ravikiran_short_code(name, designation, seen_codes)
            seen_codes.add(sc)
            seen_emails[email] = ProposedUser(
                name=name.title() if name.upper() == name else name,
                email=email,
                role=default_role,
                short_code=sc,
                source_row=ridx,
            )
            rows_used += 1

        proposal.notes.append(f"[{sheet_name}] department={dept_label} → {rows_used} staff extracted.")

    proposal.proposed_users = list(seen_emails.values())
    return proposal


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

EXTRACTORS = {
    "lab": extract_operator_list,
    "ravikiran": extract_ravikiran_roster,
}


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(prog="ai_extract_upload", description=__doc__.split("\n\n")[0])
    ap.add_argument(
        "--erp",
        required=True,
        choices=sorted(EXTRACTORS.keys()),
        help="Which ERP wrapper this upload belongs to. NO DEFAULT — wrong-target writes "
             "must be a loud error per the silo policy. Choices: lab, ravikiran.",
    )
    ap.add_argument("file", type=Path, help="Path to the uploaded .xlsx")
    ap.add_argument(
        "--db",
        type=Path,
        default=None,
        help="Optional sqlite path for matching against existing entities. "
             "Operational DB is intentionally NOT defaulted (WORKFLOW.md §3.5).",
    )
    ap.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    ap.add_argument("--out", type=Path, default=None,
                    help="Write JSON to this file in addition to stdout.")
    args = ap.parse_args(argv)
    if args.db is not None:
        args.db = validate_erp_db_match(args.erp, args.db)

    if not args.file.exists():
        print(json.dumps({"error": f"file not found: {args.file}"}), file=sys.stderr)
        return 2
    if args.file.suffix.lower() not in (".xlsx", ".xlsm"):
        print(json.dumps({"error": f"only .xlsx/.xlsm supported, got {args.file.suffix}"}), file=sys.stderr)
        return 2

    extractor = EXTRACTORS[args.erp]
    if args.erp == "lab":
        proposal = extractor(args.file, args.db)
    else:
        proposal = extractor(args.file)
    proposal_dict = asdict(proposal)
    proposal_dict["erp"] = args.erp
    indent = 2 if args.pretty else None
    rendered = json.dumps(proposal_dict, indent=indent, ensure_ascii=False)
    print(rendered)
    if args.out:
        args.out.write_text(rendered, encoding="utf-8")
    return 0


if __name__ == "__main__":
    sys.exit(main())
