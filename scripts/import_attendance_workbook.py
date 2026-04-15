#!/usr/bin/env python3
"""Convert Ravikiran-group attendance workbooks into onboarding review files.

The workbook is a real-world operational source, not a clean HR export:
sheet names vary by company, headers move around, and salary / joining-date
fields are often sparse. This script normalizes the rows into a single CSV
that admins can review before creating accounts.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


COMPANY_MAP = {
    "KITCHEN": ("Ravikiran Services", "Kitchen"),
    "PF": ("Ravikiran Services", "PF / Accounts"),
    "TUCKSHOP": ("Suryajyoti Services", "Tuck Shop"),
    "WAGES": ("Ravikiran Services", "Operations / Wages"),
    "GOPAL DOODH DAIRY": ("Gopal Doodh Dairy", "Dairy"),
    "RK SERVICES": ("RK Services", "Laundry"),
}

HEADER_ALIASES = {
    "sr.no": "sr_no",
    "sr no": "sr_no",
    "name": "name",
    "employee name": "name",
    "designation": "designation",
    "date of joining": "date_of_joining",
    "date of joining ": "date_of_joining",
    "salary": "salary",
    "total day": "total_days",
    "total payment day": "total_payment_days",
    "total present days": "present_days",
}


@dataclass
class EmployeeRecord:
    source_sheet: str
    company_name: str
    business_unit: str
    sr_no: str
    employee_name: str
    normalized_name: str
    suggested_username: str
    designation: str
    suggested_role: str
    join_date_raw: str
    salary_raw: str
    total_days: str
    present_days: str
    attendance_ratio: str
    review_owner: str
    review_status: str
    onboarding_note: str


def canonical_sheet_name(name: str) -> str:
    return " ".join((name or "").replace("\n", " ").split()).upper().strip()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_name(name: str) -> str:
    text = clean_text(name).upper()
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\b(MR|MRS|MS|DR|PROF)\.?\b", "", text)
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def slug_username(name: str) -> str:
    base = normalize_name(name).lower()
    parts = [part for part in base.split() if part]
    if not parts:
        return "user"
    if len(parts) == 1:
        return parts[0][:20]
    return (parts[0] + parts[-1])[:20]


def normalize_header(value: object) -> str:
    key = clean_text(value).lower()
    key = re.sub(r"\s+", " ", key)
    return HEADER_ALIASES.get(key, "")


def find_header_row(rows: list[tuple[object, ...]]) -> tuple[int, dict[int, str]]:
    for idx, row in enumerate(rows):
        mapping: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            header = normalize_header(cell)
            if header:
                mapping[col_idx] = header
        if "name" in mapping.values() and (
            "designation" in mapping.values()
            or "present_days" in mapping.values()
        ):
            return idx, mapping
    raise ValueError("Could not detect a usable header row")


def suggest_role(designation: str, unit: str) -> str:
    text = f"{designation} {unit}".lower()
    if any(token in text for token in ("account", "accounts", "cashier", "finance")):
        return "finance_admin"
    if "driver" in text:
        return "operator"
    if any(token in text for token in ("manager", "supervisor", "head")):
        return "site_admin"
    return "requester"


def attendance_ratio(present: str, total: str) -> str:
    try:
        present_value = float(present)
        total_value = float(total)
        if total_value <= 0:
            return ""
        return f"{(present_value / total_value):.0%}"
    except Exception:
        return ""


def onboarding_note(record: EmployeeRecord) -> str:
    bits: list[str] = []
    if not record.join_date_raw:
        bits.append("Joining date missing in source.")
    if not record.salary_raw:
        bits.append("Salary missing in source.")
    if record.suggested_role == "finance_admin":
        bits.append("Review for finance/admin permissions before activation.")
    if record.company_name == "Gopal Doodh Dairy":
        bits.append("Keep under Ravikiran group but separate company bucket.")
    if record.company_name == "RK Services":
        bits.append("Laundry company; keep isolated from Ravikiran Services staff views.")
    return " ".join(bits) or "Ready for Nikita / Prashant review."


def parse_workbook(path: Path) -> list[EmployeeRecord]:
    wb = load_workbook(path, data_only=True)
    records: list[EmployeeRecord] = []
    seen_usernames: Counter[str] = Counter()

    for ws in wb.worksheets:
        sheet_key = canonical_sheet_name(ws.title)
        company_name, unit = COMPANY_MAP.get(sheet_key, (ws.title.strip(), ws.title.strip()))
        rows = list(ws.iter_rows(values_only=True))
        header_row_idx, mapping = find_header_row(rows[:12])

        for row in rows[header_row_idx + 1 :]:
            row_values = {field: clean_text(row[idx]) for idx, field in mapping.items()}
            employee_name = row_values.get("name", "")
            if not employee_name:
                continue
            if employee_name.lower().startswith("raj services wages"):
                continue
            normalized = normalize_name(employee_name)
            if not normalized:
                continue

            username = slug_username(employee_name)
            seen_usernames[username] += 1
            if seen_usernames[username] > 1:
                username = f"{username}{seen_usernames[username]}"

            record = EmployeeRecord(
                source_sheet=ws.title.strip(),
                company_name=company_name,
                business_unit=unit,
                sr_no=row_values.get("sr_no", ""),
                employee_name=clean_text(employee_name),
                normalized_name=normalized,
                suggested_username=username,
                designation=row_values.get("designation", ""),
                suggested_role=suggest_role(row_values.get("designation", ""), unit),
                join_date_raw=row_values.get("date_of_joining", ""),
                salary_raw=row_values.get("salary", ""),
                total_days=row_values.get("total_days", ""),
                present_days=row_values.get("present_days", ""),
                attendance_ratio=attendance_ratio(
                    row_values.get("present_days", ""),
                    row_values.get("total_days", ""),
                ),
                review_owner="nikita, prashant",
                review_status="pending_review",
                onboarding_note="",
            )
            record.onboarding_note = onboarding_note(record)
            records.append(record)

    return records


def write_csv(path: Path, records: Iterable[EmployeeRecord]) -> None:
    fieldnames = list(EmployeeRecord.__annotations__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(record.__dict__)


def write_markdown(path: Path, source_path: Path, records: list[EmployeeRecord]) -> None:
    company_counts = Counter(record.company_name for record in records)
    role_counts = Counter(record.suggested_role for record in records)
    unit_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for record in records:
        unit_counts[record.company_name][record.business_unit] += 1

    lines = [
        "# Ravikiran Group onboarding import",
        "",
        f"Source workbook: `{source_path}`",
        "",
        "This import keeps each company separate inside the Ravikiran group.",
        "It does not invent reporting lines or family ownership relationships.",
        "Every row is staged for Nikita / Prashant review before account creation.",
        "",
        "## Company totals",
        "",
        "| Company | People |",
        "|---|---:|",
    ]
    for company_name, count in sorted(company_counts.items()):
        lines.append(f"| {company_name} | {count} |")

    lines += [
        "",
        "## Suggested role totals",
        "",
        "| Suggested role | People |",
        "|---|---:|",
    ]
    for role, count in sorted(role_counts.items()):
        lines.append(f"| {role} | {count} |")

    lines += [
        "",
        "## Unit breakdown",
        "",
    ]
    for company_name in sorted(unit_counts):
        lines.append(f"### {company_name}")
        lines.append("")
        lines.append("| Unit | People |")
        lines.append("|---|---:|")
        for unit, count in sorted(unit_counts[company_name].items()):
            lines.append(f"| {unit} | {count} |")
        lines.append("")

    lines += [
        "## Review rules",
        "",
        "- Keep the companies separate even though they belong to the Ravikiran group.",
        "- Do not infer line managers from attendance sheets alone.",
        "- Review `finance_admin` suggestions carefully before enabling elevated permissions.",
        "- Default passwords should only be assigned at actual account-creation time.",
        "",
        "## Next step",
        "",
        f"Use the CSV export at `{path.with_suffix('.csv').name}` as the account-creation review sheet for Nikita and Prashant.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--csv-out", type=Path, required=True)
    parser.add_argument("--md-out", type=Path, required=True)
    args = parser.parse_args()

    records = parse_workbook(args.workbook)
    args.csv_out.parent.mkdir(parents=True, exist_ok=True)
    args.md_out.parent.mkdir(parents=True, exist_ok=True)
    write_csv(args.csv_out, records)
    write_markdown(args.md_out, args.workbook, records)
    print(f"Wrote {len(records)} records to {args.csv_out}")
    print(f"Wrote summary to {args.md_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
